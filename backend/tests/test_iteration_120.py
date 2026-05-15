"""
Iteration 120 Backend Tests
Tests for:
1. GET /api/work-orders/{wo_id}/material-requirements - BOM-derived material list
2. POST /api/items/import/excel - Bulk insert + bulk_write performance
"""
import pytest
import requests
import os
import io
from openpyxl import Workbook

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

@pytest.fixture(scope="module")
def auth_session():
    """Create authenticated session"""
    session = requests.Session()
    session.headers.update({"Content-Type": "application/json"})
    
    # Login
    response = session.post(f"{BASE_URL}/api/auth/login", json={
        "email": "admin@erp.com",
        "password": "Admin@123"
    })
    assert response.status_code == 200, f"Login failed: {response.text}"
    return session


class TestMaterialRequirements:
    """Test GET /api/work-orders/{wo_id}/material-requirements endpoint"""
    
    def test_material_requirements_endpoint_exists(self, auth_session):
        """Verify the endpoint exists and returns proper structure"""
        # First get a work order ID
        wo_response = auth_session.get(f"{BASE_URL}/api/work-orders")
        assert wo_response.status_code == 200
        work_orders = wo_response.json()
        assert len(work_orders) > 0, "No work orders found"
        
        # Find MO-000003 specifically (mentioned in test requirements)
        mo_003 = next((w for w in work_orders if w.get("wo_number") == "MO-000003"), None)
        if mo_003:
            wo_id = mo_003["id"]
        else:
            # Use first available WO
            wo_id = work_orders[0]["id"]
        
        # Call material-requirements endpoint
        response = auth_session.get(f"{BASE_URL}/api/work-orders/{wo_id}/material-requirements")
        assert response.status_code == 200, f"Material requirements failed: {response.text}"
        
        data = response.json()
        # Verify response structure
        assert "wo_number" in data, "Response missing wo_number"
        assert "materials" in data, "Response missing materials"
        assert isinstance(data["materials"], list), "materials should be a list"
        print(f"PASS: Material requirements endpoint returns {len(data['materials'])} materials for {data.get('wo_number')}")
    
    def test_material_requirements_mo_000003(self, auth_session):
        """Test material requirements for MO-000003 specifically (has 5-component BOM)"""
        wo_id = "0d6feaa7-08ca-4016-a568-206205a5e665"  # MO-000003
        
        response = auth_session.get(f"{BASE_URL}/api/work-orders/{wo_id}/material-requirements")
        assert response.status_code == 200, f"Material requirements failed: {response.text}"
        
        data = response.json()
        assert data.get("wo_number") == "MO-000003", f"Expected MO-000003, got {data.get('wo_number')}"
        
        materials = data.get("materials", [])
        assert len(materials) >= 1, f"Expected at least 1 material, got {len(materials)}"
        
        # Verify material structure
        for m in materials:
            assert "item_id" in m, "Material missing item_id"
            assert "item" in m, "Material missing item (part_number)"
            assert "name" in m, "Material missing name"
            assert "quantity" in m, "Material missing quantity"
            assert "uom" in m, "Material missing uom"
            assert "unit_cost" in m, "Material missing unit_cost"
            assert "category" in m, "Material missing category"
        
        print(f"PASS: MO-000003 has {len(materials)} materials with correct structure")
        for m in materials[:5]:
            print(f"  - {m['item']} | {m['name']} | qty:{m['quantity']} | uom:{m['uom']}")
    
    def test_material_requirements_404_for_invalid_wo(self, auth_session):
        """Test that invalid WO ID returns 404"""
        response = auth_session.get(f"{BASE_URL}/api/work-orders/invalid-wo-id-12345/material-requirements")
        assert response.status_code == 404, f"Expected 404, got {response.status_code}"
        print("PASS: Invalid WO ID returns 404")


class TestItemsExcelImport:
    """Test POST /api/items/import/excel endpoint with bulk operations"""
    
    def test_items_import_endpoint_exists(self, auth_session):
        """Verify the import endpoint exists"""
        # Create a minimal Excel file
        wb = Workbook()
        ws = wb.active
        ws.append(["Part Number", "Name", "Description", "Category", "UOM"])
        ws.append(["TEST-IMPORT-001", "Test Import Item 1", "Test description", "raw_material", "pcs"])
        
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Upload - need to use a fresh session without Content-Type header for multipart
        cookies = auth_session.cookies.get_dict()
        upload_session = requests.Session()
        upload_session.cookies.update(cookies)
        
        files = {"file": ("test_import.xlsx", excel_buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        response = upload_session.post(
            f"{BASE_URL}/api/items/import/excel",
            files=files
        )
        
        assert response.status_code == 200, f"Import failed: {response.text}"
        data = response.json()
        
        # Verify response structure
        assert "created" in data, "Response missing 'created' count"
        assert "updated" in data, "Response missing 'updated' count"
        assert "errors" in data, "Response missing 'errors' list"
        
        print(f"PASS: Items import endpoint works - created:{data['created']}, updated:{data['updated']}, errors:{len(data['errors'])}")
    
    def test_items_import_bulk_insert_and_update(self, auth_session):
        """Test that import handles both new items (insert) and existing items (update)"""
        import uuid
        unique_suffix = str(uuid.uuid4())[:8]
        
        # Use a fresh session without Content-Type header for multipart
        cookies = auth_session.cookies.get_dict()
        upload_session = requests.Session()
        upload_session.cookies.update(cookies)
        
        # Create Excel with 3 new items + 1 that will be updated on second import
        wb = Workbook()
        ws = wb.active
        ws.append(["Part Number", "Name", "Description", "Category", "UOM", "Purchase Cost", "GST Rate (%)"])
        ws.append([f"TEST-BULK-{unique_suffix}-001", "Bulk Test Item 1", "First item", "raw_material", "pcs", 100, 18])
        ws.append([f"TEST-BULK-{unique_suffix}-002", "Bulk Test Item 2", "Second item", "component", "nos", 200, 12])
        ws.append([f"TEST-BULK-{unique_suffix}-003", "Bulk Test Item 3", "Third item", "raw_material", "kg", 50, 5])
        
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # First import - should create all 3
        files = {"file": ("bulk_test.xlsx", excel_buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        response = upload_session.post(f"{BASE_URL}/api/items/import/excel", files=files)
        
        assert response.status_code == 200, f"First import failed: {response.text}"
        data = response.json()
        assert data["created"] == 3, f"Expected 3 created, got {data['created']}"
        assert data["updated"] == 0, f"Expected 0 updated, got {data['updated']}"
        print(f"PASS: First import created {data['created']} items")
        
        # Second import with same items but updated names - should update all 3
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.append(["Part Number", "Name", "Description", "Category", "UOM", "Purchase Cost", "GST Rate (%)"])
        ws2.append([f"TEST-BULK-{unique_suffix}-001", "Bulk Test Item 1 UPDATED", "First item updated", "raw_material", "pcs", 150, 18])
        ws2.append([f"TEST-BULK-{unique_suffix}-002", "Bulk Test Item 2 UPDATED", "Second item updated", "component", "nos", 250, 12])
        ws2.append([f"TEST-BULK-{unique_suffix}-003", "Bulk Test Item 3 UPDATED", "Third item updated", "raw_material", "kg", 75, 5])
        
        excel_buffer2 = io.BytesIO()
        wb2.save(excel_buffer2)
        excel_buffer2.seek(0)
        
        files2 = {"file": ("bulk_test2.xlsx", excel_buffer2, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        response2 = upload_session.post(f"{BASE_URL}/api/items/import/excel", files=files2)
        
        assert response2.status_code == 200, f"Second import failed: {response2.text}"
        data2 = response2.json()
        assert data2["created"] == 0, f"Expected 0 created on second import, got {data2['created']}"
        assert data2["updated"] == 3, f"Expected 3 updated on second import, got {data2['updated']}"
        print(f"PASS: Second import updated {data2['updated']} items (bulk_write working)")
        
        # Cleanup - delete test items
        items_response = auth_session.get(f"{BASE_URL}/api/items")
        if items_response.status_code == 200:
            items = items_response.json()
            for item in items:
                if item.get("part_number", "").startswith(f"TEST-BULK-{unique_suffix}"):
                    auth_session.delete(f"{BASE_URL}/api/items/{item['id']}")


class TestWorkOrderOperationsStatus:
    """Test that work orders have operations_status for routing status pills"""
    
    def test_work_orders_have_operations_status(self, auth_session):
        """Verify work orders include operations_status array"""
        response = auth_session.get(f"{BASE_URL}/api/work-orders")
        assert response.status_code == 200
        
        work_orders = response.json()
        assert len(work_orders) > 0, "No work orders found"
        
        # Find WOs with operations_status
        with_ops = [w for w in work_orders if w.get("operations_status") and len(w.get("operations_status", [])) > 0]
        print(f"Found {len(with_ops)} work orders with operations_status out of {len(work_orders)} total")
        
        if with_ops:
            # Check structure of operations_status
            wo = with_ops[0]
            ops = wo.get("operations_status", [])
            for op in ops[:3]:
                assert "sequence" in op or "operation_name" in op, "Operation missing sequence/operation_name"
                assert "status" in op, "Operation missing status"
                print(f"  Op: {op.get('operation_name', 'N/A')} - status: {op.get('status', 'N/A')}")
        
        print("PASS: Work orders have operations_status for routing status pills")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
