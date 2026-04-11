"""
Phase 2 Features Test Suite
Tests for:
1. MRP -> PO creation (POST /api/purchase-orders/from-mrp)
2. Excel Export/Import for Items, BOM, Routings
3. Work Order operation-level tracking (Job Card)
"""
import pytest
import requests
import os
import io
from datetime import datetime, timedelta

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

class TestPhase2Features:
    """Phase 2 feature tests"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup test session with admin login"""
        self.session = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})
        
        # Login as admin
        login_resp = self.session.post(f"{BASE_URL}/api/auth/login", json={
            "email": "admin@erp.com",
            "password": "Admin@123"
        })
        assert login_resp.status_code == 200, f"Login failed: {login_resp.text}"
        
        # Store cookies for subsequent requests
        self.cookies = login_resp.cookies
        yield
    
    # ==================== MRP -> PO Creation Tests ====================
    
    def test_mrp_suggestions_endpoint(self):
        """Test GET /api/mrp/suggestions returns purchase suggestions"""
        resp = self.session.get(f"{BASE_URL}/api/mrp/suggestions", cookies=self.cookies)
        assert resp.status_code == 200, f"MRP suggestions failed: {resp.text}"
        data = resp.json()
        assert isinstance(data, list), "Suggestions should be a list"
        print(f"MRP suggestions count: {len(data)}")
    
    def test_get_suppliers_for_po(self):
        """Test GET /api/suppliers returns suppliers for PO creation"""
        resp = self.session.get(f"{BASE_URL}/api/suppliers", cookies=self.cookies)
        assert resp.status_code == 200, f"Get suppliers failed: {resp.text}"
        data = resp.json()
        assert isinstance(data, list), "Suppliers should be a list"
        assert len(data) > 0, "Should have at least one supplier"
        print(f"Suppliers count: {len(data)}")
        return data
    
    def test_create_po_from_mrp(self):
        """Test POST /api/purchase-orders/from-mrp creates PO with GST calculation"""
        # Get suppliers
        suppliers = self.test_get_suppliers_for_po()
        supplier_id = suppliers[0]["id"]
        
        # Get items to order
        items_resp = self.session.get(f"{BASE_URL}/api/items?category=raw_material", cookies=self.cookies)
        assert items_resp.status_code == 200
        items = items_resp.json()
        
        if len(items) == 0:
            pytest.skip("No raw materials to create PO")
        
        # Create PO from MRP
        po_data = {
            "supplier_id": supplier_id,
            "items": [
                {"item_id": items[0]["id"], "quantity": 100, "unit_price": 45.00}
            ]
        }
        
        resp = self.session.post(f"{BASE_URL}/api/purchase-orders/from-mrp", 
                                 json=po_data, cookies=self.cookies)
        assert resp.status_code == 201, f"Create PO from MRP failed: {resp.text}"
        
        po = resp.json()
        assert "po_number" in po, "PO should have po_number"
        assert "total_amount" in po, "PO should have total_amount"
        assert "subtotal" in po, "PO should have subtotal"
        assert "total_tax" in po, "PO should have total_tax"
        assert "lines" in po, "PO should have lines"
        
        # Verify GST calculation
        assert po["subtotal"] == 4500.0, f"Subtotal should be 4500, got {po['subtotal']}"
        assert po["total_amount"] > po["subtotal"], "Total should include GST"
        
        # Check GST fields
        assert "total_cgst" in po or "total_igst" in po, "PO should have GST breakdown"
        
        print(f"Created PO: {po['po_number']}, Total: ${po['total_amount']}")
        return po
    
    def test_create_po_from_mrp_invalid_supplier(self):
        """Test POST /api/purchase-orders/from-mrp with invalid supplier returns 404"""
        po_data = {
            "supplier_id": "invalid-supplier-id",
            "items": [{"item_id": "some-item", "quantity": 10, "unit_price": 10}]
        }
        
        resp = self.session.post(f"{BASE_URL}/api/purchase-orders/from-mrp", 
                                 json=po_data, cookies=self.cookies)
        assert resp.status_code == 404, f"Should return 404 for invalid supplier: {resp.text}"
    
    # ==================== Excel Export Tests ====================
    
    def test_items_export_excel(self):
        """Test GET /api/items/export/excel returns xlsx file"""
        resp = self.session.get(f"{BASE_URL}/api/items/export/excel", cookies=self.cookies)
        assert resp.status_code == 200, f"Items export failed: {resp.text}"
        
        content_type = resp.headers.get("Content-Type", "")
        assert "spreadsheetml" in content_type or "octet-stream" in content_type, \
            f"Should return xlsx content type, got: {content_type}"
        
        content_disp = resp.headers.get("Content-Disposition", "")
        assert "items_master.xlsx" in content_disp, f"Should have filename in header: {content_disp}"
        
        assert len(resp.content) > 0, "Excel file should have content"
        print(f"Items export size: {len(resp.content)} bytes")
    
    def test_bom_export_excel(self):
        """Test GET /api/bom/export/excel returns xlsx file"""
        resp = self.session.get(f"{BASE_URL}/api/bom/export/excel", cookies=self.cookies)
        assert resp.status_code == 200, f"BOM export failed: {resp.text}"
        
        content_type = resp.headers.get("Content-Type", "")
        assert "spreadsheetml" in content_type or "octet-stream" in content_type, \
            f"Should return xlsx content type, got: {content_type}"
        
        content_disp = resp.headers.get("Content-Disposition", "")
        assert "bom_data.xlsx" in content_disp, f"Should have filename in header: {content_disp}"
        
        assert len(resp.content) > 0, "Excel file should have content"
        print(f"BOM export size: {len(resp.content)} bytes")
    
    def test_routings_export_excel(self):
        """Test GET /api/routings/export/excel returns xlsx file"""
        resp = self.session.get(f"{BASE_URL}/api/routings/export/excel", cookies=self.cookies)
        assert resp.status_code == 200, f"Routings export failed: {resp.text}"
        
        content_type = resp.headers.get("Content-Type", "")
        assert "spreadsheetml" in content_type or "octet-stream" in content_type, \
            f"Should return xlsx content type, got: {content_type}"
        
        content_disp = resp.headers.get("Content-Disposition", "")
        assert "routings.xlsx" in content_disp, f"Should have filename in header: {content_disp}"
        
        assert len(resp.content) > 0, "Excel file should have content"
        print(f"Routings export size: {len(resp.content)} bytes")
    
    # ==================== Excel Import Tests ====================
    
    def test_items_import_excel(self):
        """Test POST /api/items/import/excel accepts xlsx file"""
        from openpyxl import Workbook
        
        # Create test Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "Items Master"
        
        # Headers matching export format
        headers = ["Part Number", "Name", "Description", "Category", "UOM", "Unit Cost", 
                   "Lead Time (Days)", "Safety Stock", "Current Stock", "Reorder Point", 
                   "HSN Code", "GST Rate (%)"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Test data row
        test_data = ["TEST-IMPORT-001", "Test Import Item", "Test description", 
                     "raw_material", "pcs", 25.50, 7, 50, 100, 30, "7208", 18]
        for col, value in enumerate(test_data, 1):
            ws.cell(row=2, column=col, value=value)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Upload file - remove Content-Type header for multipart
        headers_backup = self.session.headers.copy()
        del self.session.headers["Content-Type"]
        
        files = {"file": ("test_items.xlsx", output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        resp = self.session.post(f"{BASE_URL}/api/items/import/excel", 
                                 files=files, cookies=self.cookies)
        
        # Restore headers
        self.session.headers = headers_backup
        
        assert resp.status_code == 200, f"Items import failed: {resp.text}"
        result = resp.json()
        
        assert "created" in result, "Should have created count"
        assert "updated" in result, "Should have updated count"
        assert result["created"] + result["updated"] >= 1, "Should create or update at least 1 item"
        
        print(f"Items import: created={result['created']}, updated={result['updated']}")
        
        # Verify item was created
        items_resp = self.session.get(f"{BASE_URL}/api/items?search=TEST-IMPORT-001", cookies=self.cookies)
        items = items_resp.json()
        assert len(items) > 0, "Imported item should exist"
        assert items[0]["name"] == "Test Import Item"
    
    def test_bom_import_excel(self):
        """Test POST /api/bom/import/excel accepts xlsx file"""
        from openpyxl import Workbook
        
        # First get existing items to use in BOM
        items_resp = self.session.get(f"{BASE_URL}/api/items", cookies=self.cookies)
        items = items_resp.json()
        
        # Find a finished good and a raw material
        parent_item = next((i for i in items if i["category"] in ["finished_good", "sub_assembly"]), None)
        component_item = next((i for i in items if i["category"] == "raw_material"), None)
        
        if not parent_item or not component_item:
            pytest.skip("Need finished good and raw material items for BOM import test")
        
        # Create test Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "BOM Data"
        
        # Headers matching export format
        headers = ["Parent Part Number", "Parent Name", "Revision", "Status", 
                   "Component Part Number", "Component Name", "Quantity", "Is Alternate",
                   "Effectivity Start", "Effectivity End"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Test data row
        test_data = [parent_item["part_number"], parent_item["name"], "Z", "draft",
                     component_item["part_number"], component_item["name"], 5, "No", "", ""]
        for col, value in enumerate(test_data, 1):
            ws.cell(row=2, column=col, value=value)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Upload file - remove Content-Type header for multipart
        headers_backup = self.session.headers.copy()
        del self.session.headers["Content-Type"]
        
        files = {"file": ("test_bom.xlsx", output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        resp = self.session.post(f"{BASE_URL}/api/bom/import/excel", 
                                 files=files, cookies=self.cookies)
        
        # Restore headers
        self.session.headers = headers_backup
        
        assert resp.status_code == 200, f"BOM import failed: {resp.text}"
        result = resp.json()
        
        assert "created" in result, "Should have created count"
        assert "updated" in result, "Should have updated count"
        
        print(f"BOM import: created={result['created']}, updated={result['updated']}")
    
    # ==================== Work Order Operation Tracking (Job Card) Tests ====================
    
    def test_get_work_orders_with_operations(self):
        """Test GET /api/work-orders returns WOs with operations_status"""
        resp = self.session.get(f"{BASE_URL}/api/work-orders", cookies=self.cookies)
        assert resp.status_code == 200, f"Get work orders failed: {resp.text}"
        
        work_orders = resp.json()
        assert isinstance(work_orders, list), "Work orders should be a list"
        
        # Find WO with operations
        wo_with_ops = [wo for wo in work_orders if wo.get("operations_status")]
        print(f"Work orders with operations: {len(wo_with_ops)}/{len(work_orders)}")
        
        if wo_with_ops:
            wo = wo_with_ops[0]
            assert "operations_status" in wo, "WO should have operations_status"
            ops = wo["operations_status"]
            assert isinstance(ops, list), "operations_status should be a list"
            if ops:
                assert "sequence" in ops[0], "Operation should have sequence"
                assert "status" in ops[0], "Operation should have status"
                assert "operation_name" in ops[0], "Operation should have operation_name"
        
        return work_orders
    
    def test_update_operation_status_start(self):
        """Test PUT /api/work-orders/{id}/operations/{seq} - Start operation"""
        # Get work orders with pending operations
        work_orders = self.test_get_work_orders_with_operations()
        
        # Find a WO with pending first operation
        target_wo = None
        for wo in work_orders:
            ops = wo.get("operations_status", [])
            if ops and ops[0].get("status") == "pending" and wo.get("status") in ["pending", "in_progress"]:
                target_wo = wo
                break
        
        if not target_wo:
            pytest.skip("No work order with pending first operation found")
        
        wo_id = target_wo["id"]
        first_op = target_wo["operations_status"][0]
        seq = first_op["sequence"]
        
        # Start the first operation
        resp = self.session.put(
            f"{BASE_URL}/api/work-orders/{wo_id}/operations/{seq}",
            json={"status": "in_progress"},
            cookies=self.cookies
        )
        
        assert resp.status_code == 200, f"Start operation failed: {resp.text}"
        updated_wo = resp.json()
        
        # Verify operation status changed
        updated_op = next((op for op in updated_wo["operations_status"] if op["sequence"] == seq), None)
        assert updated_op is not None, "Operation should exist"
        assert updated_op["status"] == "in_progress", f"Operation status should be in_progress, got {updated_op['status']}"
        
        # Verify WO status changed to in_progress
        assert updated_wo["status"] == "in_progress", f"WO status should be in_progress, got {updated_wo['status']}"
        
        print(f"Started operation {seq} on WO {target_wo['wo_number']}")
        return updated_wo
    
    def test_update_operation_status_complete(self):
        """Test PUT /api/work-orders/{id}/operations/{seq} - Complete operation"""
        # Get work orders with in_progress operations
        resp = self.session.get(f"{BASE_URL}/api/work-orders", cookies=self.cookies)
        work_orders = resp.json()
        
        # Find a WO with in_progress operation that is not completed
        target_wo = None
        target_op = None
        for wo in work_orders:
            if wo.get("status") == "completed":
                continue
            ops = wo.get("operations_status", [])
            for op in ops:
                if op.get("status") == "in_progress":
                    target_wo = wo
                    target_op = op
                    break
            if target_wo:
                break
        
        if not target_wo:
            pytest.skip("No work order with in_progress operation found")
        
        wo_id = target_wo["id"]
        seq = target_op["sequence"]
        
        # Complete the operation
        resp = self.session.put(
            f"{BASE_URL}/api/work-orders/{wo_id}/operations/{seq}",
            json={"status": "completed", "quantity_completed": target_wo.get("quantity", 1)},
            cookies=self.cookies
        )
        
        assert resp.status_code == 200, f"Complete operation failed: {resp.text}"
        updated_wo = resp.json()
        
        # Verify operation status changed
        updated_op = next((op for op in updated_wo["operations_status"] if op["sequence"] == seq), None)
        assert updated_op is not None, "Operation should exist"
        assert updated_op["status"] == "completed", f"Operation status should be completed, got {updated_op['status']}"
        
        print(f"Completed operation {seq} on WO {target_wo['wo_number']}")
    
    def test_operation_sequence_validation(self):
        """Test that starting operation fails if previous not completed"""
        # Get work orders with multiple operations
        resp = self.session.get(f"{BASE_URL}/api/work-orders", cookies=self.cookies)
        work_orders = resp.json()
        
        # Find a WO with at least 2 operations where first is completed and second is pending
        target_wo = None
        second_op_seq = None
        for wo in work_orders:
            if wo.get("status") == "completed":
                continue
            ops = wo.get("operations_status", [])
            if len(ops) >= 3:
                # Find a pending operation that has a pending operation before it
                for i in range(1, len(ops)):
                    if ops[i].get("status") == "pending" and ops[i-1].get("status") == "pending":
                        target_wo = wo
                        second_op_seq = ops[i]["sequence"]
                        break
            if target_wo:
                break
        
        if not target_wo:
            pytest.skip("No work order with suitable pending operations found")
        
        wo_id = target_wo["id"]
        
        # Try to start the operation (should fail because previous is pending)
        resp = self.session.put(
            f"{BASE_URL}/api/work-orders/{wo_id}/operations/{second_op_seq}",
            json={"status": "in_progress"},
            cookies=self.cookies
        )
        
        assert resp.status_code == 400, f"Should fail with 400, got {resp.status_code}: {resp.text}"
        assert "must be completed first" in resp.text.lower() or "previous" in resp.text.lower(), \
            f"Error should mention previous operation: {resp.text}"
        
        print(f"Correctly blocked starting operation {second_op_seq} before previous completed")
    
    def test_operation_not_found(self):
        """Test PUT /api/work-orders/{id}/operations/{seq} with invalid sequence returns 404"""
        resp = self.session.get(f"{BASE_URL}/api/work-orders", cookies=self.cookies)
        work_orders = resp.json()
        
        # Find a non-completed work order
        target_wo = None
        for wo in work_orders:
            if wo.get("status") in ["pending", "in_progress"]:
                target_wo = wo
                break
        
        if not target_wo:
            pytest.skip("No non-completed work orders found")
        
        wo_id = target_wo["id"]
        
        resp = self.session.put(
            f"{BASE_URL}/api/work-orders/{wo_id}/operations/99999",
            json={"status": "in_progress"},
            cookies=self.cookies
        )
        
        assert resp.status_code == 404, f"Should return 404 for invalid sequence: {resp.text}"


class TestPhase2Cleanup:
    """Cleanup test data"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        self.session = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})
        
        login_resp = self.session.post(f"{BASE_URL}/api/auth/login", json={
            "email": "admin@erp.com",
            "password": "Admin@123"
        })
        self.cookies = login_resp.cookies
        yield
    
    def test_cleanup_test_items(self):
        """Clean up test items created during import tests"""
        # Get test items
        resp = self.session.get(f"{BASE_URL}/api/items?search=TEST-IMPORT", cookies=self.cookies)
        if resp.status_code == 200:
            items = resp.json()
            for item in items:
                if item["part_number"].startswith("TEST-IMPORT"):
                    del_resp = self.session.delete(f"{BASE_URL}/api/items/{item['id']}", cookies=self.cookies)
                    print(f"Deleted test item: {item['part_number']}")
