from dotenv import load_dotenv
load_dotenv()

from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, UploadFile, File, Body
from fastapi.responses import StreamingResponse, JSONResponse
from starlette.middleware.cors import CORSMiddleware
import os
import logging
import io
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any, Tuple, Set
import uuid
from datetime import datetime, timezone, timedelta
from bson import ObjectId
import jwt  # used by /auth/refresh route below

# Shared core building blocks (DB, permissions, auth)
from core.db import client, db
from core.permissions import (
    ALL_MODULES, ALL_ACTIONS, MODULE_ACTIONS,
    allowed_actions_for, DEFAULT_PERMISSIONS, get_default_permissions,
)
from core.auth import (
    JWT_ALGORITHM, JWT_SECRET,
    hash_password, verify_password,
    get_jwt_secret, create_access_token, create_refresh_token,
    get_current_user, get_cookie_settings,
    require_roles, _require_access,
)

# Create the main app
app = FastAPI(title="Machinery Manufacturing ERP")

# Create routers
api_router = APIRouter(prefix="/api")
auth_router = APIRouter(prefix="/auth")
items_router = APIRouter(prefix="/items")
item_groups_router = APIRouter(prefix="/item-groups")
bom_router = APIRouter(prefix="/bom")
mrp_router = APIRouter(prefix="/mrp")
quality_router = APIRouter(prefix="/quality")
inventory_router = APIRouter(prefix="/inventory")
production_router = APIRouter(prefix="/production")
users_router = APIRouter(prefix="/users")
dashboard_router = APIRouter(prefix="/dashboard")
suppliers_router = APIRouter(prefix="/suppliers")
purchase_orders_router = APIRouter(prefix="/purchase-orders")
warehouses_router = APIRouter(prefix="/warehouses")
work_centers_router = APIRouter(prefix="/work-centers")
routings_router = APIRouter(prefix="/routings")
work_orders_router = APIRouter(prefix="/work-orders")
settings_router = APIRouter(prefix="/settings")
customers_router = APIRouter(prefix="/customers")
grn_router = APIRouter(prefix="/grn")
purchase_invoices_router = APIRouter(prefix="/purchase-invoices")
jobwork_router = APIRouter(prefix="/job-work")
crm_router = APIRouter(prefix="/crm")

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ================== MODELS ==================

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: str = "user"  # legacy field — real permissions now come from role_group_id
    permissions: Optional[dict] = None
    role_group_id: str  # REQUIRED — all users must be mapped to a group

class UserUpdate(BaseModel):
    email: Optional[str] = None  # Admins can update another user's email / login ID
    name: Optional[str] = None
    role: Optional[str] = None
    password: Optional[str] = None
    permissions: Optional[dict] = None
    status: Optional[str] = None
    role_group_id: Optional[str] = None
    signature_url: Optional[str] = None  # base64 PNG/JPG data-URL for digital signature
    assigned_customer_ids: Optional[List[str]] = None  # Customers visible to this user beyond their own (admin-managed)

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    id: str
    email: str
    name: str
    role: str
    created_at: datetime

class ItemCreate(BaseModel):
    part_number: str
    name: str
    description: Optional[str] = ""
    category: str  # raw_material, component, sub_assembly, finished_good
    group_id: Optional[str] = None  # NEW: reference to item_groups collection (optional)
    unit_of_measure: str = "pcs"
    unit_cost: float = 0.0
    purchase_price: float = 0.0  # Last PO price; auto-updates from new POs (RM only)
    sale_price: float = 0.0  # Selling price to customers (all categories)
    lead_time_days: int = 0
    safety_stock: int = 0
    current_stock: int = 0
    reorder_point: int = 0
    hsn_code: Optional[str] = ""
    gst_rate: Optional[float] = 18.0
    # ===== Phase 2: Attribute-driven variants =====
    # variant_attributes shape:
    #   [{name: str, values: [{value: str, short_code: str(max 4)}]}, ...]
    # Backward-compat: legacy values as list[str] still accepted; short_code defaults
    # to the first 4 chars of the value (uppercased, alnum only).
    variant_attributes: Optional[List[Dict[str, Any]]] = None
    auto_suffix_variant_sku: Optional[bool] = True
    # When the item is itself a generated variant of a parent.
    parent_item_id: Optional[str] = None      # link to the master item
    is_variant: Optional[bool] = False        # marks a generated variant child
    variant_short_codes: Optional[Dict[str, str]] = None   # { "Motor Power": "1HP", "Voltage": "220V" } — the chosen short codes
    variant_values: Optional[Dict[str, str]] = None        # { "Motor Power": "1HP", "Voltage": "220V" } — display labels
    is_active: Optional[bool] = True                       # variants get is_active=False when an attribute value is removed

class ItemUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    category: Optional[str] = None
    group_id: Optional[str] = None
    unit_of_measure: Optional[str] = None
    unit_cost: Optional[float] = None
    purchase_price: Optional[float] = None
    sale_price: Optional[float] = None
    lead_time_days: Optional[int] = None
    safety_stock: Optional[int] = None
    current_stock: Optional[int] = None
    reorder_point: Optional[int] = None
    hsn_code: Optional[str] = None
    gst_rate: Optional[float] = None
    variant_attributes: Optional[List[Dict[str, Any]]] = None
    auto_suffix_variant_sku: Optional[bool] = None
    parent_item_id: Optional[str] = None
    is_variant: Optional[bool] = None
    variant_short_codes: Optional[Dict[str, str]] = None
    variant_values: Optional[Dict[str, str]] = None
    is_active: Optional[bool] = None


class ItemGroupCreate(BaseModel):
    name: str  # e.g. "Motors", "Bearings", "V-Belts"
    parent_category: Optional[str] = None  # raw_material / component / sub_assembly / finished_good (optional: leave blank = any)
    default_hsn_code: Optional[str] = None  # When set, all items in this group inherit this HSN
    default_gst_rate: Optional[float] = None  # When set, all items in this group inherit this GST%
    description: Optional[str] = ""

class ItemGroupUpdate(BaseModel):
    name: Optional[str] = None
    parent_category: Optional[str] = None
    default_hsn_code: Optional[str] = None
    default_gst_rate: Optional[float] = None
    description: Optional[str] = None

class RoutingEntry(BaseModel):
    name: str
    cost: float = 0.0  # Cost per unit for this operation

class BOMComponentCreate(BaseModel):
    item_id: str
    quantity: float
    unit_of_measure: str = "pcs"
    is_alternate: bool = False
    alternate_for: Optional[str] = None
    position: Optional[int] = None
    routings: Optional[List[Any]] = []  # Each item: str (legacy) OR {name, cost}
    # Variant filter — empty / null = common to ALL variants.
    # Otherwise: AND-logic dict, only included when SO/MO variant selection matches.
    # Example: {"Motor Power": "2HP"} → component used only when selection contains Motor Power=2HP.
    applies_to: Optional[Dict[str, str]] = None

class BOMCreate(BaseModel):
    parent_item_id: str
    name: str
    description: Optional[str] = ""
    revision: str = "A"
    effectivity_date: Optional[datetime] = None
    expiry_date: Optional[datetime] = None
    status: str = "draft"  # draft, active, obsolete
    components: List[BOMComponentCreate] = []
    parent_routings: Optional[List[Any]] = []  # Each item: str (legacy) OR {name, cost}

class BOMUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    revision: Optional[str] = None
    effectivity_date: Optional[datetime] = None
    expiry_date: Optional[datetime] = None
    status: Optional[str] = None
    components: Optional[List[BOMComponentCreate]] = None
    parent_routings: Optional[List[Any]] = None


def normalize_routings(routings: Optional[List[Any]]) -> List[Dict[str, Any]]:
    """Normalize legacy routings (list of strings) to list of {name, cost}.
    Accepts: ["LC Cutting", {"name": "Bending", "cost": 50}] → [{"name": "LC Cutting", "cost": 0}, {"name": "Bending", "cost": 50}]"""
    if not routings:
        return []
    out = []
    for r in routings:
        if isinstance(r, str):
            out.append({"name": r, "cost": 0.0})
        elif isinstance(r, dict):
            out.append({"name": r.get("name", ""), "cost": float(r.get("cost", 0) or 0)})
    return out


def routings_total_cost(routings: Optional[List[Any]]) -> float:
    """Sum of cost across all routing entries."""
    return sum(r.get("cost", 0) for r in normalize_routings(routings))


def _build_variant_sku(part_number: str, variant_selection: Optional[Dict[str, str]]) -> str:
    """Build a display SKU by suffixing the part number with selected variant values.
    Example: ("FG-001", {"Motor Power":"2HP","Voltage":"440V"}) -> "FG-001-2HP-440V".
    Values are normalised (spaces stripped) and joined in insertion order so the
    same variant_selection always yields the same SKU.
    """
    if not variant_selection:
        return part_number
    parts = []
    for v in variant_selection.values():
        if v is None or str(v).strip() == "":
            continue
        parts.append(str(v).strip().replace(" ", ""))
    if not parts:
        return part_number
    return f"{part_number}-{'-'.join(parts)}"


def _filter_components_by_variant(components: List[Dict[str, Any]], variant_selection: Optional[Dict[str, str]]) -> List[Dict[str, Any]]:
    """Phase 2 — filter BOM components by variant_selection (AND-logic).

    Rules:
      • Components with no `applies_to` (None / empty dict) are ALWAYS included
        (common to all variants).
      • Components with an `applies_to` map are included ONLY if every (k, v)
        in the map matches `variant_selection[k]`.
      • If `variant_selection` is None / empty, only common-to-all components
        are kept (legacy non-variant behaviour preserved as long as no BOM
        component carries an `applies_to`).
    """
    if not components:
        return []
    if not variant_selection:
        variant_selection = {}
    out = []
    for c in components:
        applies = c.get("applies_to") or {}
        if not applies:
            out.append(c)
            continue
        match = True
        for k, v in applies.items():
            if str(variant_selection.get(k, "")).strip() != str(v).strip():
                match = False
                break
        if match:
            out.append(c)
    return out


def _merge_duplicate_components(components: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Merge duplicate component rows by `(item_id, is_alternate)` — sums
    quantities and unions routings. Server-side safety net mirroring the
    client's BOMPage.handleSubmit logic so malformed payloads (raw API,
    Excel import) can't write duplicate components either."""
    if not components:
        return components or []
    merged: Dict[str, Dict[str, Any]] = {}
    for c in components:
        item_id = c.get("item_id")
        if not item_id:
            continue
        key = f"{item_id}__{1 if c.get('is_alternate') else 0}"
        existing = merged.get(key)
        if existing is None:
            merged[key] = {**c, "routings": list(c.get("routings") or [])}
        else:
            existing["quantity"] = float(existing.get("quantity") or 0) + float(c.get("quantity") or 0)
            seen_ids = {(r.get("name"), r.get("cost")) for r in (existing.get("routings") or [])}
            for r in (c.get("routings") or []):
                key_r = (r.get("name"), r.get("cost"))
                if key_r not in seen_ids:
                    existing["routings"].append(r)
                    seen_ids.add(key_r)
    return list(merged.values())


async def compute_bom_costs(item_id: str, _depth: int = 0, _visited: Optional[set] = None) -> Dict[str, Any]:
    """Return {rm_cost, process_cost, process_names, fg_process_cost} for this item.
    Recursive: for each BOM component, if that component itself has a BOM, use its RECURSIVE
    Total/Unit (rm + process) as the child rate. Otherwise use item.unit_cost.
    This matches how the BOM viewer calculates Material & Total/Unit up the tree.
    
    Strategy 1: Item IS a BOM parent.
    Strategy 2: Item is a COMPONENT in another parent's BOM — use that component line's routings.
    """
    if _visited is None:
        _visited = set()
    if item_id in _visited or _depth > 25:  # cycle / depth guard
        return {"rm_cost": 0.0, "process_cost": 0.0, "process_names": [], "fg_process_cost": 0.0}
    _visited.add(item_id)
    
    strategy1 = None
    # Strategy 1: this item IS a BOM parent
    bom = await db.boms.find_one({"parent_item_id": item_id, "status": "active"}, {"_id": 0})
    if not bom:
        bom = await db.boms.find_one({"parent_item_id": item_id}, {"_id": 0})
    if bom:
        rm = 0.0
        comp_proc = 0.0
        names = []
        for r in normalize_routings(bom.get("parent_routings", [])):
            if r.get("name"):
                names.append(r["name"])
        for comp in bom.get("components", []):
            c_item = await db.items.find_one({"id": comp.get("item_id")}, {"_id": 0})
            if not c_item:
                continue
            comp_qty = comp.get("quantity", 0)
            # If the child has its own BOM → use recursive Total/Unit (material + process)
            # This captures nested assemblies (e.g., SG-1 contains PT-1 which has its own BOM).
            child_bom_exists = await db.boms.find_one(
                {"parent_item_id": comp["item_id"]},
                {"_id": 0, "id": 1}
            )
            if child_bom_exists and comp["item_id"] not in _visited:
                child_costs = await compute_bom_costs(comp["item_id"], _depth + 1, _visited)
                child_total_unit = (child_costs.get("rm_cost", 0) or 0) + (child_costs.get("process_cost", 0) or 0)
                rm += comp_qty * child_total_unit
                # Skip component-line routings — the child's own BOM parent_routings is already
                # rolled into child_total_unit. Adding comp.routings here would double-count.
            else:
                # Leaf item (raw material or no BOM) → use items master unit_cost
                rm += comp_qty * (c_item.get("unit_cost", 0) or 0)
                # Leaf: component-line routings ARE the source of truth for this item's process.
                comp_proc += routings_total_cost(comp.get("routings", []))
                for r in normalize_routings(comp.get("routings", [])):
                    if r.get("name") and r["name"] not in names:
                        names.append(r["name"])
        parent_proc = routings_total_cost(bom.get("parent_routings", []))
        strategy1 = {
            "rm_cost": round(rm, 2),
            "process_cost": round(parent_proc + comp_proc, 2),
            "process_names": names,
            "fg_process_cost": round(parent_proc, 2)
        }
        if strategy1["process_cost"] > 0 or strategy1["rm_cost"] > 0:
            return strategy1
    
    # Strategy 2: this item is a COMPONENT in some parent BOM — use that component's routings
    parent_bom = await db.boms.find_one({"components.item_id": item_id, "status": "active"}, {"_id": 0})
    if not parent_bom:
        parent_bom = await db.boms.find_one({"components.item_id": item_id}, {"_id": 0})
    if parent_bom:
        for comp in parent_bom.get("components", []):
            if comp.get("item_id") == item_id:
                proc = routings_total_cost(comp.get("routings", []))
                names = [r["name"] for r in normalize_routings(comp.get("routings", [])) if r.get("name")]
                if strategy1:
                    return {"rm_cost": strategy1["rm_cost"], "process_cost": round(proc, 2), "process_names": names, "fg_process_cost": strategy1.get("fg_process_cost", 0)}
                item_rec = await db.items.find_one({"id": item_id}, {"_id": 0})
                rm = item_rec.get("unit_cost", 0) if item_rec else 0
                return {"rm_cost": round(rm, 2), "process_cost": round(proc, 2), "process_names": names, "fg_process_cost": 0.0}
    
    if strategy1:
        return strategy1
    return {"rm_cost": 0.0, "process_cost": 0.0, "process_names": [], "fg_process_cost": 0.0}


async def compute_bom_total_unit_cost(item_id: str) -> float:
    """BOM Total/Unit = RM cost (material) + process cost (parent + component routings).
    Used as the RATE for Part/SA RM lines in SC (with material) and Job OS."""
    costs = await compute_bom_costs(item_id)
    return round(costs.get("rm_cost", 0) + costs.get("process_cost", 0), 2)


async def compute_bom_fg_process_only(item_id: str) -> float:
    """FG Process = ONLY the parent_routings cost of the item's own BOM. This is what the user
    sees in the BOM header as "FG Process: ₹X". Used as the processing_charges for SC (with RM)."""
    costs = await compute_bom_costs(item_id)
    return round(costs.get("fg_process_cost", 0), 2)


def routing_cost_for_process(bom: Dict[str, Any], item_id: str, process_name: str) -> float:
    """Find the cost for a specific process_name in a BOM component (or parent if item_id matches parent)."""
    if not bom or not process_name:
        return 0.0
    target_name = process_name.strip().lower()
    if bom.get("parent_item_id") == item_id:
        for r in normalize_routings(bom.get("parent_routings", [])):
            if r.get("name", "").strip().lower() == target_name:
                return r.get("cost", 0.0)
    for comp in bom.get("components", []):
        if comp.get("item_id") == item_id:
            for r in normalize_routings(comp.get("routings", [])):
                if r.get("name", "").strip().lower() == target_name:
                    return r.get("cost", 0.0)
    return 0.0


async def find_routing_cost(item_id: str, process_name: str) -> float:
    """Search any BOM (as parent or as component) for the given item + process and return its cost.

    Searches BOTH:
      (a) the item's own BOM(s) — `parent_routings` for the named op
      (b) every BOM where this item appears as a component — that component's `routings` for the named op

    Multiple matches: returns the FIRST non-zero match. Scans every candidate
    BOM rather than just the first (older versions only checked one), so the
    cost is found even when the routing op is defined on a non-default
    parent BOM (e.g., the FG BOM that uses this part).
    """
    if not item_id or not process_name:
        return 0.0
    # (a) item as parent — try every BOM, prefer active first.
    async for bom in db.boms.find({"parent_item_id": item_id, "status": "active"}, {"_id": 0}):
        c = routing_cost_for_process(bom, item_id, process_name)
        if c:
            return c
    async for bom in db.boms.find({"parent_item_id": item_id, "status": {"$ne": "active"}}, {"_id": 0}):
        c = routing_cost_for_process(bom, item_id, process_name)
        if c:
            return c
    # (b) item as component — try every BOM that lists it, prefer active.
    async for parent_bom in db.boms.find({"components.item_id": item_id, "status": "active"}, {"_id": 0}):
        c = routing_cost_for_process(parent_bom, item_id, process_name)
        if c:
            return c
    async for parent_bom in db.boms.find({"components.item_id": item_id, "status": {"$ne": "active"}}, {"_id": 0}):
        c = routing_cost_for_process(parent_bom, item_id, process_name)
        if c:
            return c
    return 0.0

class ProductionOrderLine(BaseModel):
    line_id: Optional[str] = None  # UUID per line (populated on create)
    line_no: Optional[int] = None  # 1-based sequence
    bom_id: str
    quantity: int
    due_date: Optional[datetime] = None
    order_type: str = "mts"  # mts | mto  (legacy "auto" treated as mts)
    notes: Optional[str] = ""
    # When the SO line originated from a Quotation line, retain the mapping
    # so the quotation balance can be tracked.
    source_quotation_line_no: Optional[int] = None
    # Phase 2 — variant selection for attribute-driven BOMs.
    # Keyed by attribute name, value = chosen attribute value.
    # Example: {"Motor Power": "2HP", "Voltage": "440V"}
    # Empty/None = no variant filtering (legacy non-variant BOM).
    variant_selection: Optional[Dict[str, str]] = None
    # Populated on confirm — how the qty was split:
    reserved_qty: Optional[int] = 0   # to be fulfilled from FG stock
    mo_qty: Optional[int] = 0         # to be manufactured via MO
    status: Optional[str] = "draft"   # draft | confirmed | in_progress | completed | cancelled

class ProductionOrderCreate(BaseModel):
    # New multi-line mode
    lines: Optional[List[ProductionOrderLine]] = None
    customer_id: Optional[str] = None
    # Optional source-quotation link (when the user creates the SO via the
    # "From Quotation" picker on the Production page).
    source_quotation_id: Optional[str] = None
    source_quotation_no: Optional[str] = None
    # Legacy single-line fields — used when `lines` is not provided
    bom_id: Optional[str] = None
    quantity: Optional[int] = None
    due_date: Optional[datetime] = None
    priority: str = "medium"  # low, medium, high, urgent
    notes: Optional[str] = ""

class ProductionOrderUpdate(BaseModel):
    quantity: Optional[int] = None
    due_date: Optional[datetime] = None
    priority: Optional[str] = None
    status: Optional[str] = None
    notes: Optional[str] = None
    lines: Optional[List[ProductionOrderLine]] = None
    customer_id: Optional[str] = None

class InspectionTemplateCreate(BaseModel):
    name: str
    description: Optional[str] = ""
    category: str  # incoming, in_process, final
    checklist_items: List[Dict[str, Any]]

class InspectionRecordCreate(BaseModel):
    template_id: str
    item_id: str
    production_order_id: Optional[str] = None
    lot_number: Optional[str] = None
    quantity_inspected: int
    results: List[Dict[str, Any]]
    overall_result: str  # pass, fail, conditional

class InventoryTransactionCreate(BaseModel):
    item_id: str
    transaction_type: str  # receive, issue, adjust, transfer
    quantity: int
    reference_type: Optional[str] = None  # production_order, purchase_order, adjustment
    reference_id: Optional[str] = None
    notes: Optional[str] = ""
    warehouse_id: Optional[str] = None
    from_warehouse_id: Optional[str] = None
    to_warehouse_id: Optional[str] = None

# ================== PROCUREMENT MODELS ==================

class SupplierCreate(BaseModel):
    code: Optional[str] = ""  # Auto-generated from number series if blank
    name: str
    contact_person: Optional[str] = ""
    email: Optional[str] = ""
    phone: Optional[str] = ""
    address: Optional[str] = ""
    address_line2: Optional[str] = ""
    city: Optional[str] = ""
    state: Optional[str] = ""
    pin_code: Optional[str] = ""
    gstin: Optional[str] = ""
    state_code: Optional[str] = ""
    payment_terms: Optional[str] = "Net 30"
    lead_time_days: int = 7
    rating: Optional[int] = 3  # 1-5 stars
    status: str = "active"  # active, inactive
    assigned_user_ids: Optional[List[str]] = []  # Per-user contact ownership (same model as customers)

class SupplierUpdate(BaseModel):
    name: Optional[str] = None
    contact_person: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    address_line2: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    pin_code: Optional[str] = None
    gstin: Optional[str] = None
    state_code: Optional[str] = None
    payment_terms: Optional[str] = None
    lead_time_days: Optional[int] = None
    rating: Optional[int] = None
    status: Optional[str] = None
    assigned_user_ids: Optional[List[str]] = None

class MRPCreatePORequest(BaseModel):
    supplier_id: str
    items: list  # [{"item_id": "...", "quantity": 100, "unit_price": 45}]

class PurchaseOrderLineCreate(BaseModel):
    item_id: str
    description: Optional[str] = ""
    quantity: float
    unit_price: float
    uom: Optional[str] = "pcs"
    hsn_code: Optional[str] = ""
    gst_rate: Optional[float] = 18.0
    discount_type: Optional[str] = "percentage"  # percentage or amount
    discount_value: Optional[float] = 0.0
    notes: Optional[str] = ""

class POAdditionalCharge(BaseModel):
    charge_type_id: Optional[str] = ""
    name: str
    hsn_code: Optional[str] = ""
    gst_rate: Optional[float] = 18.0
    amount: float = 0.0

class PurchaseOrderCreate(BaseModel):
    supplier_id: str
    expected_date: datetime
    delivery_warehouse_id: Optional[str] = ""
    quotation_ref: Optional[str] = ""
    quotation_date: Optional[datetime] = None
    lines: List[PurchaseOrderLineCreate]
    additional_charges: Optional[List[POAdditionalCharge]] = []
    notes: Optional[str] = ""
    terms_conditions: Optional[str] = None  # Overrides default PO T&C from company settings
    revision_label: Optional[str] = None    # Manual revision label ("A", "1", "R01")
    currency: Optional[str] = "INR"          # INR (default), USD, EUR, GBP, AED — non-INR = export/import (no GST)

class PurchaseOrderUpdate(BaseModel):
    supplier_id: Optional[str] = None
    expected_date: Optional[datetime] = None
    delivery_warehouse_id: Optional[str] = None
    quotation_ref: Optional[str] = None
    quotation_date: Optional[datetime] = None
    lines: Optional[List[PurchaseOrderLineCreate]] = None
    additional_charges: Optional[List[POAdditionalCharge]] = None
    status: Optional[str] = None
    notes: Optional[str] = None
    terms_conditions: Optional[str] = None
    revision_label: Optional[str] = None
    currency: Optional[str] = None

class GRNLineVerify(BaseModel):
    item_id: str
    received_quantity: float
    verified_price: float

class GRNCreate(BaseModel):
    po_id: str
    supplier_invoice_no: str = ""
    supplier_invoice_date: Optional[datetime] = None
    lines: List[GRNLineVerify]
    warehouse_id: Optional[str] = ""
    notes: Optional[str] = ""
    status: Optional[str] = "posted"  # "draft" => save without stock/PO updates; "posted" => commit immediately (legacy)

class GRNUpdate(BaseModel):
    """Patch payload for draft GRNs (posted GRNs are read-only)."""
    supplier_invoice_no: Optional[str] = None
    supplier_invoice_date: Optional[datetime] = None
    lines: Optional[List[GRNLineVerify]] = None
    warehouse_id: Optional[str] = None
    notes: Optional[str] = None

class ManualGRNLine(BaseModel):
    item_id: str
    received_quantity: float
    verified_price: float
    uom: Optional[str] = "pcs"
    hsn_code: Optional[str] = ""

class ManualGRNCreate(BaseModel):
    """GRN created without a preceding Purchase Order (direct receipt)."""
    supplier_id: str
    supplier_invoice_no: str = ""
    supplier_invoice_date: Optional[datetime] = None
    lines: List[ManualGRNLine]
    warehouse_id: Optional[str] = ""
    notes: Optional[str] = ""
    # Optional Manual-DC linkage. When set, the GRN bumps each matching
    # DC line's `received_qty` so the DC list shows accurate balance and
    # the DC picker filters out fully-received DCs.
    manual_dc_id: Optional[str] = None

class POChargeTypeCreate(BaseModel):
    name: str
    hsn_code: Optional[str] = ""
    gst_rate: Optional[float] = 18.0

class POChargeTypeUpdate(BaseModel):
    name: Optional[str] = None
    hsn_code: Optional[str] = None
    gst_rate: Optional[float] = None

# ===== Purchase Invoice Models =====
class PurchaseInvoiceLineItem(BaseModel):
    item_id: str
    quantity: float
    unit_price: float
    discount: Optional[float] = 0
    hsn_code: Optional[str] = ""
    gst_rate: Optional[float] = 18.0
    is_process_charge: Optional[bool] = False
    description: Optional[str] = ""

class PurchaseInvoiceCreate(BaseModel):
    supplier_id: str
    po_id: Optional[str] = ""
    grn_id: Optional[str] = ""
    grn_ids: Optional[List[str]] = None  # Multi-GRN invoicing: select multiple GRNs from the same supplier
    invoice_no: str
    invoice_date: datetime
    due_date: Optional[datetime] = None
    lines: List[PurchaseInvoiceLineItem]
    additional_charges: Optional[List[POAdditionalCharge]] = []  # Freight, packaging, insurance etc.
    notes: Optional[str] = ""
    is_manual: Optional[bool] = False  # Manual PI — no parent GRN, direct entry (services, freight etc.)

class PurchaseInvoiceUpdate(BaseModel):
    invoice_no: Optional[str] = None
    invoice_date: Optional[datetime] = None
    due_date: Optional[datetime] = None
    status: Optional[str] = None
    notes: Optional[str] = None
    lines: Optional[List[PurchaseInvoiceLineItem]] = None
    additional_charges: Optional[List[POAdditionalCharge]] = None

# ===== Job Work / Subcontracting Models =====
class JobWorkLineItem(BaseModel):
    item_id: str
    quantity: float
    rate: Optional[float] = 0
    processing_charges: Optional[float] = 0  # Per-unit processing charge (used on Job Card OS DC lines)
    item_description: Optional[str] = None  # Per-line description carried over to DC
    process_name: Optional[str] = None  # Specific outsourced routing op (Job Card OS lines)

class JobWorkPartItem(BaseModel):
    item_id: str
    quantity: float = 0
    charges: float = 0
    process_name: Optional[str] = None  # Specific routing op being outsourced (Job Card OS) — falsy/null on Full MO-SC.
    item_description: Optional[str] = None  # User-editable description / remarks shown on DC.

class SubcontractOrderCreate(BaseModel):
    supplier_id: str
    lines: List[JobWorkLineItem]
    job_work_parts: Optional[List[JobWorkPartItem]] = []
    expected_return_date: Optional[datetime] = None
    processing_charges: Optional[float] = 0
    notes: Optional[str] = ""

class SubcontractOrderUpdate(BaseModel):
    expected_return_date: Optional[datetime] = None
    processing_charges: Optional[float] = None
    status: Optional[str] = None
    notes: Optional[str] = None
    lines: Optional[List[JobWorkLineItem]] = None
    job_work_parts: Optional[List[JobWorkPartItem]] = None
    dc_created: Optional[bool] = None
    supplier_id: Optional[str] = None  # Allow changing vendor during edit (only when no DC sent yet)

class DCCreate(BaseModel):
    subcontract_order_id: str
    lines: List[JobWorkLineItem]
    warehouse_id: Optional[str] = ""
    notes: Optional[str] = ""
    skip_stock_deduct: Optional[bool] = False  # For Job Card outsource — parts go for processing, not consumed

class ManualDCLineItem(BaseModel):
    item_id: str
    quantity: float
    unit: Optional[str] = "pcs"
    unit_price: Optional[float] = 0  # Per-unit value of the goods being shipped (for value declaration on the DC)
    processing_charges: Optional[float] = 0
    notes: Optional[str] = ""
    # User-editable per-line description that prints under the part number/
    # name on the DC. Separate from `notes` so the description (a spec /
    # variant / colour) stays on the items table while notes are line-level
    # comments.
    item_description: Optional[str] = ""

class ManualDCCreate(BaseModel):
    """Standalone DC not tied to any Subcontract Order. Used for direct
    shipments where goods go out and a GRN will come back later (DC-GRN flow)."""
    supplier_id: str
    lines: List[ManualDCLineItem]
    warehouse_id: Optional[str] = ""
    dc_purpose: Optional[str] = "subcontract"  # subcontract | rework | repair | other
    notes: Optional[str] = ""
    skip_stock_deduct: Optional[bool] = False
    dc_date: Optional[str] = None  # YYYY-MM-DD; defaults to today server-side

class SubcontractReceiptLineItem(BaseModel):
    item_id: str
    received_quantity: float
    quality_result: Optional[str] = "accept"
    reject_qty: Optional[float] = 0
    rework_qty: Optional[float] = 0

class SubcontractReceiptCreate(BaseModel):
    subcontract_order_id: str
    dc_id: Optional[str] = ""
    lines: List[SubcontractReceiptLineItem]
    warehouse_id: Optional[str] = ""
    notes: Optional[str] = ""

# ================== STORES/WAREHOUSE MODELS ==================

class WarehouseCreate(BaseModel):
    code: str
    name: str
    location: Optional[str] = ""
    address: Optional[str] = ""
    is_default: bool = False
    status: str = "active"

class WarehouseUpdate(BaseModel):
    name: Optional[str] = None
    location: Optional[str] = None
    address: Optional[str] = None
    is_default: Optional[bool] = None
    status: Optional[str] = None

class StockTransferCreate(BaseModel):
    item_id: str
    from_warehouse_id: str
    to_warehouse_id: str
    quantity: int
    notes: Optional[str] = ""

# ================== MANUFACTURING PROCESS MODELS ==================

class WorkCenterCreate(BaseModel):
    code: str
    name: str
    description: Optional[str] = ""
    hourly_rate: float = 0.0
    capacity_per_hour: float = 1.0
    status: str = "active"

class WorkCenterUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    hourly_rate: Optional[float] = None
    capacity_per_hour: Optional[float] = None
    status: Optional[str] = None

class RoutingOperationCreate(BaseModel):
    sequence: int
    operation_name: str  # References a routing name (e.g., "LC Cutting", "Welding")
    description: Optional[str] = ""
    setup_time_minutes: int = 0
    run_time_minutes: int = 0

class RoutingCreate(BaseModel):
    name: str  # e.g., "LC Cutting", "Welding", "Assembly"
    description: Optional[str] = ""
    status: str = "active"

class RoutingUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    status: Optional[str] = None

class WorkOrderCreate(BaseModel):
    # MTS: pick an item directly (production_order_id not required).
    # MTO: pick an existing Sales Order (and optionally a specific line).
    order_type: Optional[str] = "mto"   # "mts" | "mto" (legacy callers default to MTO)
    production_order_id: Optional[str] = ""   # required for MTO
    source_so_line_id: Optional[str] = ""     # MTO line id when SO has multiple lines
    item_id: Optional[str] = ""               # required for MTS (item with active BOM)
    routing_id: Optional[str] = ""  # Optional - routing now comes from BOM
    quantity: int
    due_date: Optional[datetime] = None       # MO due date (used by both modes)
    scheduled_start: Optional[datetime] = None
    scheduled_end: Optional[datetime] = None
    notes: Optional[str] = ""
    is_subcontract: Optional[bool] = False
    subcontract_supplier_id: Optional[str] = ""
    subcontract_type: Optional[str] = "with_material"  # with_material | without_material
    # Phase 2 — variant selection. For MTO it's inherited from the SO line if not provided.
    variant_selection: Optional[Dict[str, str]] = None

class WorkOrderUpdate(BaseModel):
    status: Optional[str] = None
    actual_start: Optional[datetime] = None
    actual_end: Optional[datetime] = None
    quantity_completed: Optional[int] = None
    notes: Optional[str] = None
    is_subcontract: Optional[bool] = None
    subcontract_supplier_id: Optional[str] = None
    subcontract_type: Optional[str] = None  # with_material | without_material

class WorkOrderOperationUpdate(BaseModel):
    status: str  # pending, in_progress, completed, stopped
    actual_start: Optional[datetime] = None
    actual_end: Optional[datetime] = None
    quantity_completed: Optional[int] = None
    operator: Optional[str] = None
    quality_result: Optional[str] = None  # accept, reject, rework
    reject_qty: Optional[int] = None
    rework_qty: Optional[int] = None
    notes: Optional[str] = None
    is_outsource: Optional[bool] = None
    outsource_supplier_id: Optional[str] = None
    outsource_charges: Optional[float] = None
    # PARTIAL OS support — the user can outsource just a subset of the WO's
    # qty (e.g. 30 of 100 pcs go to vendor; the remaining 70 are still
    # available to Start in-house). When omitted/zero we fall back to the
    # full MO qty (legacy behaviour).
    outsource_quantity: Optional[float] = None
    work_center_id: Optional[str] = None
    process_cost_per_unit: Optional[float] = None
    run_number: Optional[int] = None  # Target specific run when stopping/completing a parallel operator

# ================== GST / INDIA COMPLIANCE MODELS ==================

INDIAN_STATES = {
    "01": "Jammu & Kashmir", "02": "Himachal Pradesh", "03": "Punjab", "04": "Chandigarh",
    "05": "Uttarakhand", "06": "Haryana", "07": "Delhi", "08": "Rajasthan",
    "09": "Uttar Pradesh", "10": "Bihar", "11": "Sikkim", "12": "Arunachal Pradesh",
    "13": "Nagaland", "14": "Manipur", "15": "Mizoram", "16": "Tripura",
    "17": "Meghalaya", "18": "Assam", "19": "West Bengal", "20": "Jharkhand",
    "21": "Odisha", "22": "Chhattisgarh", "23": "Madhya Pradesh", "24": "Gujarat",
    "26": "Dadra & Nagar Haveli and Daman & Diu", "27": "Maharashtra", "29": "Karnataka",
    "30": "Goa", "31": "Lakshadweep", "32": "Kerala", "33": "Tamil Nadu",
    "34": "Puducherry", "35": "Andaman & Nicobar Islands", "36": "Telangana",
    "37": "Andhra Pradesh"
}

GST_SLABS = [0, 5, 12, 18, 28]  # Seed values. Runtime values come from `tax_slabs` collection.
DEFAULT_UOMS = [
    {"code": "pcs", "name": "Pieces"},
    {"code": "nos", "name": "Numbers"},
    {"code": "kg",  "name": "Kilogram"},
    {"code": "g",   "name": "Gram"},
    {"code": "m",   "name": "Metre"},
    {"code": "mm",  "name": "Millimetre"},
    {"code": "cm",  "name": "Centimetre"},
    {"code": "ltr", "name": "Litre"},
    {"code": "ml",  "name": "Millilitre"},
    {"code": "box", "name": "Box"},
    {"code": "set", "name": "Set"},
    {"code": "sheet","name": "Sheet"},
    {"code": "roll","name": "Roll"},
    {"code": "pkt", "name": "Packet"},
]

class CompanySettingsUpdate(BaseModel):
    company_name: Optional[str] = None
    gstin: Optional[str] = None
    state_code: Optional[str] = None
    address: Optional[str] = None
    address_line2: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    pin_code: Optional[str] = None
    country: Optional[str] = None
    pan: Optional[str] = None
    cin: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    website: Optional[str] = None
    logo_data: Optional[str] = None
    tagline: Optional[str] = None
    primary_currency: Optional[str] = None
    secondary_currency: Optional[str] = None
    # Default Terms & Conditions printed on every Purchase Order. Multi-line string.
    po_terms_conditions: Optional[str] = None
    # Bank details printed on Proforma Invoice / Tax Invoice
    bank_name: Optional[str] = None
    bank_branch: Optional[str] = None
    bank_account: Optional[str] = None
    bank_ifsc: Optional[str] = None
    bank_upi: Optional[str] = None
    # Optional intro paragraph prepended as a cover page when printing Quotations.
    quotation_cover_intro: Optional[str] = None
    # Default Terms & Conditions printed on Tax Invoices (separate from quotation_cover_intro).
    invoice_terms_conditions: Optional[str] = None
    # Appyflow GSTIN lookup API key (override for backend APPYFLOW_API_KEY env var).
    appyflow_api_key: Optional[str] = None
    # GST e-Invoice (IRN) integration credentials (NIC Sandbox or GSP like MasterGST/ClearTax).
    gst_einvoice_enabled: Optional[bool] = None
    gst_einvoice_provider: Optional[str] = None   # "nic_sandbox" / "nic_prod" / "mastergst" / custom
    gst_einvoice_endpoint: Optional[str] = None   # e.g. https://einv-apisandbox.nic.in
    gst_einvoice_username: Optional[str] = None
    gst_einvoice_password: Optional[str] = None   # stored as-is; admin-only access
    gst_einvoice_api_key: Optional[str] = None    # some GSPs (MasterGST) require additional API key

class CustomerCreate(BaseModel):
    code: Optional[str] = ""  # Auto-generated from number series if blank
    name: str
    gstin: Optional[str] = ""
    state_code: Optional[str] = ""
    contact_person: Optional[str] = ""
    email: Optional[str] = ""
    phone: Optional[str] = ""
    address: Optional[str] = ""
    address_line2: Optional[str] = ""
    city: Optional[str] = ""
    state: Optional[str] = ""
    pin_code: Optional[str] = ""
    payment_terms: Optional[str] = "Net 30"
    status: str = "active"
    # Multi-salesperson assignment. A non-admin user only sees a customer if
    # they created it OR their id is in this list.
    assigned_user_ids: Optional[List[str]] = []

class CustomerUpdate(BaseModel):
    name: Optional[str] = None
    gstin: Optional[str] = None
    state_code: Optional[str] = None
    contact_person: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    address_line2: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    pin_code: Optional[str] = None
    payment_terms: Optional[str] = None
    status: Optional[str] = None
    assigned_user_ids: Optional[List[str]] = None

# ================== AUTH ROUTES ==================

@auth_router.post("/register")
async def register(user_data: UserCreate, response: Response):
    email = user_data.email.lower()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_doc = {
        "email": email,
        "password_hash": hash_password(user_data.password),
        "name": user_data.name,
        "role": user_data.role,
        "permissions": user_data.permissions or get_default_permissions(user_data.role),
        "status": "active",
        "created_at": datetime.now(timezone.utc)
    }
    result = await db.users.insert_one(user_doc)
    user_id = str(result.inserted_id)
    
    access_token = create_access_token(user_id, email)
    refresh_token = create_refresh_token(user_id)
    
    response.set_cookie(key="access_token", value=access_token, httponly=True, secure=get_cookie_settings()["secure"], samesite=get_cookie_settings()["samesite"], max_age=900, path="/")
    response.set_cookie(key="refresh_token", value=refresh_token, httponly=True, secure=get_cookie_settings()["secure"], samesite=get_cookie_settings()["samesite"], max_age=604800, path="/")
    
    return {"id": user_id, "email": email, "name": user_data.name, "role": user_data.role}

@auth_router.post("/login")
async def login(user_data: UserLogin, response: Response, request: Request):
    try:
        email = user_data.email.lower()
        
        user = await db.users.find_one({"email": email})
        if not user:
            raise HTTPException(status_code=401, detail="Invalid credentials")
        
        pw_hash = user.get("password_hash", "")
        if not pw_hash or not verify_password(user_data.password, pw_hash):
            raise HTTPException(status_code=401, detail="Invalid credentials")
        
        user_id = str(user["_id"])
        access_token = create_access_token(user_id, email)
        refresh_token = create_refresh_token(user_id)
        
        response.set_cookie(key="access_token", value=access_token, httponly=True, secure=get_cookie_settings()["secure"], samesite=get_cookie_settings()["samesite"], max_age=900, path="/")
        response.set_cookie(key="refresh_token", value=refresh_token, httponly=True, secure=get_cookie_settings()["secure"], samesite=get_cookie_settings()["samesite"], max_age=604800, path="/")
        
        # Overlay role group permissions + admin flag if assigned.
        effective_perms = user.get("permissions") or get_default_permissions(user["role"])
        is_admin_group = False
        view_all_parties = False
        role_group = None
        if user.get("role_group_id"):
            group = await db.role_groups.find_one({"id": user["role_group_id"]}, {"_id": 0})
            if group:
                role_group = group
                if group.get("permissions"):
                    effective_perms = group["permissions"]
                is_admin_group = bool(group.get("is_admin_group"))
                view_all_parties = bool(group.get("view_all_parties"))

        return {
            "id": user_id,
            "email": user["email"],
            "name": user["name"],
            "role": user["role"],
            "permissions": effective_perms,
            "role_group_id": user.get("role_group_id"),
            "role_group": role_group,
            "is_admin_group": is_admin_group,
            "view_all_parties": view_all_parties,
            "signature_url": user.get("signature_url", ""),
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Login error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Login failed: {str(e)}")

@auth_router.post("/logout")
async def logout(response: Response):
    response.delete_cookie(key="access_token", path="/")
    response.delete_cookie(key="refresh_token", path="/")
    return {"message": "Logged out successfully"}

@auth_router.get("/me")
async def get_me(request: Request):
    user = await get_current_user(request)
    # Ensure permissions exist (for older users without permissions)
    if "permissions" not in user or not user["permissions"]:
        user["permissions"] = get_default_permissions(user.get("role", "inventory_manager"))
    # If user is mapped to a role group, overlay its permissions + expose is_admin_group flag
    group_id = user.get("role_group_id")
    if group_id:
        group = await db.role_groups.find_one({"id": group_id}, {"_id": 0})
        if group:
            user["role_group"] = group
            if group.get("permissions"):
                user["permissions"] = group["permissions"]
            user["is_admin_group"] = bool(group.get("is_admin_group"))
            user["view_all_parties"] = bool(group.get("view_all_parties"))
    return user

@auth_router.post("/refresh")
async def refresh_token(request: Request, response: Response):
    token = request.cookies.get("refresh_token")
    if not token:
        raise HTTPException(status_code=401, detail="No refresh token")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user_id = payload["sub"]
        user = await db.users.find_one({"_id": ObjectId(user_id)})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        
        access_token = create_access_token(user_id, user["email"])
        response.set_cookie(key="access_token", value=access_token, httponly=True, secure=get_cookie_settings()["secure"], samesite=get_cookie_settings()["samesite"], max_age=900, path="/")
        return {"message": "Token refreshed"}
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Refresh token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# ================== ITEMS ROUTES ==================

@items_router.get("")
async def get_items(request: Request, category: Optional[str] = None, search: Optional[str] = None, group_id: Optional[str] = None, lite: Optional[bool] = False):
    await get_current_user(request)
    query = {}
    if category:
        query["category"] = category
    if group_id:
        query["group_id"] = group_id
    if search:
        query["$or"] = [
            {"part_number": {"$regex": search, "$options": "i"}},
            {"name": {"$regex": search, "$options": "i"}}
        ]
    # `lite=1` projects only the fields needed by item-picker dropdowns (BOM,
    # PO, Quotation, etc.) — id, part_number, name, description, category,
    # unit_of_measure, hsn_code, gst_rate, sale_price, purchase_price,
    # unit_cost. Cuts payload size ~3-5x for catalogues with thousands of SKUs
    # and shaves seconds off slow connections.
    if lite:
        projection = {
            "_id": 0, "id": 1, "part_number": 1, "name": 1, "description": 1,
            "category": 1, "group_id": 1, "unit_of_measure": 1,
            "hsn_code": 1, "gst_rate": 1,
            "unit_cost": 1, "sale_price": 1, "purchase_price": 1,
            "current_stock": 1, "safety_stock": 1, "reorder_point": 1,
            "lead_time_days": 1,
            "variant_attributes": 1, "auto_suffix_variant_sku": 1,
            "parent_item_id": 1, "is_variant": 1,
            "variant_short_codes": 1, "variant_values": 1, "is_active": 1,
        }
    else:
        projection = {"_id": 0}
    # No practical hard cap — ERPs routinely carry thousands of SKUs, and the
    # BOM / PO / SO / Quotation pickers load the full list then filter client-side.
    # Soft-retired variants (is_active=False — orphans of removed attribute
    # values that still have transaction history) are HIDDEN by default so
    # they don't clutter the picker / stock list. Pass `?include_inactive=1`
    # if a future flow needs them.
    if "is_active" not in query:
        query["is_active"] = {"$ne": False}
    items = await db.items.find(query, projection).to_list(50000)
    return items

@items_router.get("/{item_id}")
async def get_item(item_id: str, request: Request):
    await get_current_user(request)
    item = await db.items.find_one({"id": item_id}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    return item

async def _apply_item_group_overrides(item_doc: dict) -> dict:
    """If item_doc has group_id AND the referenced group has default_hsn_code/default_gst_rate set,
    OVERRIDE those fields on the item. This implements the user's rule:
    "If HSN/GST is assigned at group level, all items under this group consider the same HSN & GST%".
    """
    group_id = item_doc.get("group_id")
    if not group_id:
        return item_doc
    group = await db.item_groups.find_one({"id": group_id}, {"_id": 0})
    if not group:
        return item_doc
    if group.get("default_hsn_code"):
        item_doc["hsn_code"] = group["default_hsn_code"]
    if group.get("default_gst_rate") is not None:
        item_doc["gst_rate"] = float(group["default_gst_rate"])
    return item_doc


@items_router.post("", status_code=201)
async def create_item(item_data: ItemCreate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager", "inventory_manager"], module="items", action="create")
    existing = await db.items.find_one({"part_number": item_data.part_number})
    if existing:
        raise HTTPException(status_code=400, detail="Part number already exists")
    # UOM is mandatory and must reference a UOM in the master.
    uom_code = (item_data.unit_of_measure or "").strip().lower()
    if not uom_code:
        raise HTTPException(status_code=400, detail="Unit of Measure (UOM) is mandatory")
    uom_exists = await db.uoms.find_one({"code": uom_code}, {"_id": 0, "code": 1})
    if not uom_exists:
        raise HTTPException(status_code=400, detail=f"UOM '{uom_code}' is not configured. Add it under Settings → Units of Measure first.")
    
    item_doc = {
        "id": str(uuid.uuid4()),
        **item_data.model_dump(),
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"]
    }
    # Sync unit_cost with purchase_price on creation if unit_cost not explicitly set
    if item_doc.get("purchase_price") and not item_doc.get("unit_cost"):
        item_doc["unit_cost"] = item_doc["purchase_price"]
    # Apply group-level HSN/GST overrides
    item_doc = await _apply_item_group_overrides(item_doc)
    await db.items.insert_one(item_doc)
    item_doc.pop("_id", None)
    return item_doc

@items_router.put("/{item_id}")
async def update_item(item_id: str, item_data: ItemUpdate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager", "inventory_manager"], module="items", action="edit")
    update_data = {k: v for k, v in item_data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    # If group_id is being set/changed, apply group-level HSN/GST overrides
    if "group_id" in update_data and update_data["group_id"]:
        merged = dict(update_data)
        merged = await _apply_item_group_overrides(merged)
        update_data["hsn_code"] = merged.get("hsn_code", update_data.get("hsn_code"))
        update_data["gst_rate"] = merged.get("gst_rate", update_data.get("gst_rate"))

    update_data["updated_at"] = datetime.now(timezone.utc)
    # Normalize variant_attributes to canonical shape (List[{name, values:[{value,short_code}]}]).
    # We also opportunistically prune orphan variant children that no longer
    # match the new attribute set — this is the "delete a value and update
    # item" flow the user expects. Logic:
    #   * Variants whose SKU is no longer in the valid combos are checked
    #     against all transaction collections.
    #   * No references found → hard-deleted (removes the row outright).
    #   * Otherwise → soft-retired (is_active=False) so transaction history
    #     stays intact but they disappear from pickers / stock lists.
    prune_result = None
    if "variant_attributes" in update_data:
        new_attrs = _normalize_variant_attributes(update_data["variant_attributes"])
        update_data["variant_attributes"] = new_attrs
        parent = await db.items.find_one({"id": item_id}, {"_id": 0})
        if parent and parent.get("category") in ("component", "raw_material"):
            prune_result = await _prune_obsolete_variants(parent, new_attrs)
    result = await db.items.update_one({"id": item_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    
    item = await db.items.find_one({"id": item_id}, {"_id": 0})
    if prune_result is not None:
        item["_variant_prune"] = prune_result
    return item

@items_router.delete("/{item_id}")
async def delete_item(item_id: str, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin"], module="items", action="delete")
    # Referential integrity — block delete if the item is referenced in transactions
    checks = [
        ("boms", {"$or": [{"parent_item_id": item_id}, {"components.item_id": item_id}]}, "BOM(s)"),
        ("purchase_orders", {"line_items.item_id": item_id}, "Purchase Order(s)"),
        ("grns", {"line_items.item_id": item_id}, "GRN(s)"),
        ("subcontract_orders", {"$or": [{"lines.item_id": item_id}, {"job_work_parts.item_id": item_id}]}, "Subcontract Order(s)"),
        ("delivery_challans", {"lines.item_id": item_id}, "Delivery Challan(s)"),
        ("work_orders", {"item_id": item_id}, "Work/Manufacturing Order(s)"),
        ("production_orders", {"$or": [{"bom_id": {"$exists": True}}, {"lines.bom_id": {"$exists": True}}]}, None),  # SOs resolved via BOM link — handled separately below
        ("inventory_transactions", {"item_id": item_id}, "Inventory transaction(s)"),
        ("purchase_invoices", {"line_items.item_id": item_id}, "Purchase Invoice(s)"),
        ("crm_quotations", {"lines.item_id": item_id}, "Quotation(s)"),
        ("crm_tickets", {"product_ids": item_id}, "Support Ticket(s)"),
    ]
    blockers = []
    for coll_name, query, label in checks:
        if label is None:
            continue  # skip production_orders (resolved via BOM)
        try:
            coll = db[coll_name]
            cnt = await coll.count_documents(query)
            if cnt > 0:
                blockers.append(f"{cnt} {label}")
        except Exception:
            pass
    if blockers:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete — this item is referenced in: {', '.join(blockers)}. Remove those records first.",
        )

    result = await db.items.delete_one({"id": item_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    return {"message": "Item deleted"}



def _normalize_variant_attributes(raw: Optional[List[Dict[str, Any]]]) -> List[Dict[str, Any]]:
    """Normalize variant_attributes from frontend to the canonical shape:
       [{name: str, values: [{value: str, short_code: str(<=4)}]}, ...]

    Accepts legacy formats:
      • values as List[str] → coerce to [{value:'x', short_code:'X'}].
      • short_code missing → default to first 4 alnum-upper chars of value.
    """
    if not raw:
        return []
    out = []
    invalid_lengths: List[str] = []  # collect violations to raise a single, helpful error
    for attr in raw:
        if not isinstance(attr, dict):
            continue
        name = (attr.get("name") or "").strip()
        if not name:
            continue
        raw_vals = attr.get("values") or []
        norm_vals = []
        for v in raw_vals:
            if isinstance(v, str):
                val = v.strip()
                if not val:
                    continue
                sc = "".join(ch for ch in val.upper() if ch.isalnum())[:4]
                # Enforce exactly-4-character rule (matches frontend).
                if len(val) != 4:
                    invalid_lengths.append(f"{name}={val!r}")
                    continue
                norm_vals.append({"value": val, "short_code": sc or val[:4]})
            elif isinstance(v, dict):
                val = (v.get("value") or "").strip()
                if not val:
                    continue
                sc = (v.get("short_code") or "").strip()
                if not sc:
                    sc = "".join(ch for ch in val.upper() if ch.isalnum())[:4]
                sc = sc[:4]
                # Enforce exactly-4-character rule on the value AND the short_code.
                if len(val) != 4 or len(sc) != 4:
                    invalid_lengths.append(f"{name}={val!r}/{sc!r}")
                    continue
                norm_vals.append({"value": val, "short_code": sc})
        if norm_vals:
            out.append({"name": name, "values": norm_vals})
    if invalid_lengths:
        raise HTTPException(
            status_code=400,
            detail=f"Variant value(s) must be exactly 4 characters: {', '.join(invalid_lengths)}",
        )
    return out


def _all_variant_combinations(variant_attributes: List[Dict[str, Any]]) -> List[Dict[str, Dict[str, str]]]:
    """Return every possible (attribute → {value, short_code}) combination.
    Order follows the attribute order in `variant_attributes`."""
    if not variant_attributes:
        return []
    # Build a list of [(attr_name, [value_dict, ...]), ...]
    axes = [(a["name"], a["values"]) for a in variant_attributes if a.get("values")]
    if not axes:
        return []
    import itertools
    combos = []
    for tup in itertools.product(*[vals for _, vals in axes]):
        combo = {}
        for (attr_name, _), val in zip(axes, tup):
            combo[attr_name] = {"value": val["value"], "short_code": val["short_code"]}
        combos.append(combo)
    return combos


def _build_variant_sku_from_short_codes(parent_sku: str, combo: Dict[str, Dict[str, str]]) -> str:
    """Build SKU from parent + short_codes joined.
    Example: ('FG-001', {Motor Power:{short_code:'1HP'}, Voltage:{short_code:'220V'}}) -> 'FG-001-1HP-220V'."""
    parts = [c.get("short_code", "").strip() for c in combo.values()]
    parts = [p for p in parts if p]
    if not parts:
        return parent_sku
    return f"{parent_sku}-{'-'.join(parts)}"


async def _is_variant_referenced(variant_id: str) -> bool:
    """Returns True iff the given variant SKU appears in ANY transactional
    collection. Mirrors the safety check inside generate_item_variants so
    both the PUT /items path AND generate-variants treat orphans identically.

    NOTE: PO/Quotation/TaxInvoice/PI/PurchaseInvoice/DC all use `lines.item_id`
    (NOT `line_items.item_id`). GRNs and quotations may use either — we
    check both keys to stay safe."""
    ref_collections = [
        ("inventory_transactions", {"item_id": variant_id}),
        ("purchase_orders", {"lines.item_id": variant_id}),
        ("purchase_invoices", {"lines.item_id": variant_id}),
        ("grns", {"$or": [{"lines.item_id": variant_id}, {"line_items.item_id": variant_id}]}),
        ("subcontract_orders", {"$or": [{"lines.item_id": variant_id}, {"job_work_parts.item_id": variant_id}]}),
        ("delivery_challans", {"lines.item_id": variant_id}),
        ("work_orders", {"$or": [{"item_id": variant_id}, {"reserved_materials.item_id": variant_id}, {"consumed_materials.item_id": variant_id}]}),
        ("quotations", {"lines.item_id": variant_id}),
        ("tax_invoices", {"lines.item_id": variant_id}),
        ("proforma_invoices", {"lines.item_id": variant_id}),
        ("boms", {"components.item_id": variant_id}),
    ]
    for col_name, q in ref_collections:
        try:
            cnt = await db[col_name].count_documents(q, limit=1)
        except Exception:
            cnt = 0
        if cnt:
            return True
    return False


async def _prune_obsolete_variants(parent: Dict[str, Any], new_attrs: List[Dict[str, Any]]) -> Dict[str, List[str]]:
    """Given a parent item and its (already normalized) NEW variant_attributes,
    walk every existing variant child and either:
      * hard-delete it (no transactional footprint), or
      * RAISE HTTPException 400 when it IS referenced anywhere.

    User rule (strict): "Variant used anywhere in transaction should not allow
    me to delete the variant." We therefore refuse the entire update rather
    than silently soft-retiring it — the operator must roll back / replace
    the transactional usage before they can remove the attribute value.

    Returns: {"deleted_skus": [...]}  (retired_in_use_skus stays an empty
    list for back-compat with old callers).
    Empty `new_attrs` means "no variants any more" → every child is a
    candidate for pruning. This is the path that fires when the user removes
    the LAST attribute and clicks "Update Item"."""
    item_id = parent.get("id")
    if not item_id:
        return {"deleted_skus": [], "retired_in_use_skus": []}
    parent_sku = parent.get("part_number") or ""
    valid_skus: set = set()
    if new_attrs:
        try:
            combos = _all_variant_combinations(new_attrs)
            valid_skus = {_build_variant_sku_from_short_codes(parent_sku, c) for c in combos}
        except Exception:
            valid_skus = set()
    children = await db.items.find(
        {"parent_item_id": item_id, "is_variant": True},
        {"_id": 0, "id": 1, "part_number": 1, "is_active": 1},
    ).to_list(2000)
    # First pass — collect any obsolete variants that are referenced. If
    # ANY are referenced, refuse the whole prune (atomic — we don't want
    # to delete some while keeping others).
    blocked: List[str] = []
    candidates_for_delete: List[Dict[str, Any]] = []
    for ch in children:
        if ch.get("part_number") in valid_skus:
            # Still in the combination set — leave it alone (or reactivate if soft-retired).
            if ch.get("is_active") is False:
                await db.items.update_one({"id": ch["id"]}, {"$set": {"is_active": True, "updated_at": datetime.now(timezone.utc)}})
            continue
        if await _is_variant_referenced(ch["id"]):
            blocked.append(ch.get("part_number", ""))
        else:
            candidates_for_delete.append(ch)
    if blocked:
        raise HTTPException(
            status_code=400,
            detail=(
                "Cannot remove variant value(s): the following variant SKU(s) are used in "
                "Purchase Orders / GRNs / Invoices / BOMs / Inventory transactions and must "
                "be reversed there first → "
                + ", ".join(sorted(blocked))
            ),
        )
    # Safe to hard-delete the unused obsoletes.
    deleted: List[str] = []
    for ch in candidates_for_delete:
        await db.items.delete_one({"id": ch["id"]})
        deleted.append(ch.get("part_number", ""))
    return {"deleted_skus": deleted, "retired_in_use_skus": []}


async def _compute_inherited_variants(item_id: str) -> List[Dict[str, Any]]:
    """For an FG/SG, walk its active BOM and aggregate `variant_attributes`
    from every variant-bearing component. Attributes with the same name are
    merged: union of values, dedup by value, first-seen short_code wins.

    Rules for which component contributes:
    - CP / RM components: use their OWN `variant_attributes` (leaves).
    - Variant CHILD references (is_variant=True) walk up to their parent first.
    - SG / FG components: ALWAYS recurse into their own BOM first. If the
      recursion yields any variants, ONLY those count (treats the SG as a
      pass-through to its leaf components). Falls back to the SG's own
      `variant_attributes` ONLY when the recursion returns nothing — so
      legacy SGs with own variants but no variant-bearing children still work.

    Returns the canonical normalised list — same shape as Item.variant_attributes.
    """
    bom = await db.boms.find_one(
        {"parent_item_id": item_id, "status": "active"},
        {"_id": 0, "components": 1},
    ) or await db.boms.find_one(
        {"parent_item_id": item_id},
        {"_id": 0, "components": 1},
    )
    if not bom:
        return []
    merged: Dict[str, Dict[str, Dict[str, str]]] = {}
    order: List[str] = []
    for comp in (bom.get("components") or []):
        cid = comp.get("item_id")
        if not cid:
            continue
        citem = await db.items.find_one(
            {"id": cid},
            {"_id": 0, "id": 1, "variant_attributes": 1, "category": 1, "is_variant": 1, "parent_item_id": 1, "part_number": 1, "name": 1},
        )
        if not citem:
            continue
        if citem.get("is_variant") and citem.get("parent_item_id"):
            parent_doc = await db.items.find_one(
                {"id": citem["parent_item_id"]},
                {"_id": 0, "id": 1, "variant_attributes": 1, "category": 1, "part_number": 1, "name": 1},
            )
            if parent_doc:
                citem = parent_doc

        if citem.get("category") in ("sub_assembly", "finished_good"):
            # Pass-through: prefer recursion into the SG's own BOM. Use own
            # variants only when the recursion yields nothing.
            attrs = await _compute_inherited_variants(citem.get("id") or cid)
            if not attrs:
                attrs = _normalize_variant_attributes(citem.get("variant_attributes") or [])
        else:
            attrs = _normalize_variant_attributes(citem.get("variant_attributes") or [])

        if not attrs:
            continue
        for a in attrs:
            name = a["name"]
            if name not in merged:
                merged[name] = {}
                order.append(name)
            for v in a["values"]:
                key = v["value"]
                if key not in merged[name]:
                    merged[name][key] = {"value": v["value"], "short_code": v["short_code"]}
    return [{"name": n, "values": list(merged[n].values())} for n in order]


async def _compute_inherited_variants_breakdown(item_id: str) -> List[Dict[str, Any]]:
    """Per-component breakdown of inherited variants — used by BOM dialog
    to show which COMPONENT contributes which axis. Each entry:
      {component_id, component_part_number, component_name, variant_attributes: [...]}.

    Only LEAF variant-bearing components (CP/RM with own variant_attributes,
    or SG/FG that have own variants but no variant-bearing descendants)
    appear. SG/FG that delegate to their BOM are skipped — their leaves are
    shown directly.
    """
    bom = await db.boms.find_one(
        {"parent_item_id": item_id, "status": "active"},
        {"_id": 0, "components": 1},
    ) or await db.boms.find_one(
        {"parent_item_id": item_id},
        {"_id": 0, "components": 1},
    )
    if not bom:
        return []
    out: List[Dict[str, Any]] = []
    seen: Set[str] = set()
    for comp in (bom.get("components") or []):
        cid = comp.get("item_id")
        if not cid:
            continue
        citem = await db.items.find_one(
            {"id": cid},
            {"_id": 0, "id": 1, "variant_attributes": 1, "category": 1, "is_variant": 1, "parent_item_id": 1, "part_number": 1, "name": 1},
        )
        if not citem:
            continue
        if citem.get("is_variant") and citem.get("parent_item_id"):
            parent_doc = await db.items.find_one(
                {"id": citem["parent_item_id"]},
                {"_id": 0, "id": 1, "variant_attributes": 1, "category": 1, "part_number": 1, "name": 1},
            )
            if parent_doc:
                citem = parent_doc

        if citem.get("category") in ("sub_assembly", "finished_good"):
            # Pass-through: recurse for leaves. Only fall back to own variants
            # when this SG/FG has NO variant-bearing descendants.
            sub_leaves = await _compute_inherited_variants_breakdown(citem.get("id") or cid)
            if sub_leaves:
                for child in sub_leaves:
                    sig = (child.get("component_part_number") or "", tuple(sorted(a["name"] for a in (child.get("variant_attributes") or []))))
                    if sig not in seen:
                        seen.add(sig)
                        out.append(child)
                continue
            # No leaf variants below → use own (legacy SG case).
            own_attrs = _normalize_variant_attributes(citem.get("variant_attributes") or [])
            if own_attrs:
                sig = (citem.get("part_number") or "", tuple(sorted(a["name"] for a in own_attrs)))
                if sig not in seen:
                    seen.add(sig)
                    out.append({
                        "component_id": citem.get("id") or cid,
                        "component_part_number": citem.get("part_number") or "",
                        "component_name": citem.get("name") or "",
                        "variant_attributes": own_attrs,
                    })
        else:
            # CP / RM leaf.
            own_attrs = _normalize_variant_attributes(citem.get("variant_attributes") or [])
            if own_attrs:
                sig = (citem.get("part_number") or "", tuple(sorted(a["name"] for a in own_attrs)))
                if sig not in seen:
                    seen.add(sig)
                    out.append({
                        "component_id": citem.get("id") or cid,
                        "component_part_number": citem.get("part_number") or "",
                        "component_name": citem.get("name") or "",
                        "variant_attributes": own_attrs,
                    })
    return out


async def _get_effective_variants(item: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Return the effective variant_attributes used for variant generation,
    SO/MO selectors, and FG-stock-by-variant credit.

    - Component / Raw Material: their own `variant_attributes` (the leaf items
      ARE the source of truth — they have no BOM to inherit from).
    - Finished Good / Sub-Assembly: ALWAYS prefer the union INHERITED from
      variant-bearing BOM components. Falls back to legacy own
      `variant_attributes` only when the FG/SG has no BOM yet (so a
      half-set-up item still shows something). This avoids stale FG-level
      variants (e.g. 1-axis legacy values) shadowing the real 2+ axes
      contributed by the BOM components.
    """
    if item.get("category") in ("finished_good", "sub_assembly"):
        inherited = await _compute_inherited_variants(item["id"])
        if inherited:
            return inherited
        return _normalize_variant_attributes(item.get("variant_attributes") or [])
    return _normalize_variant_attributes(item.get("variant_attributes") or [])


async def _filter_variant_selection_for_item(item_id: str, variant_selection: Optional[Dict[str, str]]) -> Optional[Dict[str, str]]:
    """Return the subset of `variant_selection` whose attribute names actually
    appear in this item's effective variants (own for CP/RM, inherited for
    FG/SG via BOM walk). If none match, returns None — caller treats this
    as a "plain" WO (no variant_selection).

    Used when auto-creating child WOs from a parent MO: a child SG whose BOM
    tree has no variant-bearing components should run as a plain WO; a child
    SG whose tree DOES have variants gets only the matching axes.
    """
    if not variant_selection:
        return None
    item = await db.items.find_one({"id": item_id}, {"_id": 0})
    if not item:
        return None
    effective = await _get_effective_variants(item)
    if not effective:
        return None
    relevant_names = {a.get("name") for a in effective if a.get("name")}
    filtered = {k: v for k, v in variant_selection.items() if k in relevant_names}
    return filtered or None



async def _resolve_variant_child_item(comp_item: Dict[str, Any], variant_selection: Optional[Dict[str, str]]) -> Optional[Dict[str, Any]]:
    """For a BOM component whose item carries its OWN variant_attributes,
    pick the matching variant child item based on the MO's variant_selection.

    Example:
      comp_item = {part_number: 'CRW0E8000091', variant_attributes: [{name: 'Grit Size', values: [...] }]}
      variant_selection = {'Grit Size': '30GT', 'Sieve Slot': '1.0mm'}  # MO-level UNION
      → looks up variant child 'CRW0E8000091-30GT' and returns it.

    Returns None when the component has no variants, or when the MO's
    variant_selection doesn't pick any of this component's axes.
    """
    if not variant_selection:
        return None
    attrs = _normalize_variant_attributes(comp_item.get("variant_attributes") or [])
    if not attrs:
        return None
    combo: Dict[str, Dict[str, str]] = {}
    for a in attrs:
        v = variant_selection.get(a["name"])
        if v is None:
            continue
        match = next((x for x in a["values"] if x["value"] == v), None)
        if match:
            combo[a["name"]] = {"value": match["value"], "short_code": match["short_code"]}
    if not combo:
        return None
    sku = _build_variant_sku_from_short_codes(comp_item.get("part_number") or "", combo)
    child = await db.items.find_one({"part_number": sku}, {"_id": 0})
    return child



@items_router.get("/{item_id}/effective-variants")
async def get_effective_variants(item_id: str, request: Request):
    """Return the variant axes that drive variant SKUs for this item.
    For CP/RM this is the item's own `variant_attributes`. For FG/SG it
    is the UNION of variant_attributes from variant-bearing BOM components
    (recursively walking SG sub-BOMs), or the legacy own value if set.

    Response includes `variant_sources` — a per-component breakdown so the
    BOM dialog can show "CRW0E…: Grit Size: 16,24,30" rather than only the
    merged union (helps when multiple components contribute different axes).
    """
    await get_current_user(request)
    item = await db.items.find_one({"id": item_id}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    effective = await _get_effective_variants(item)
    breakdown: List[Dict[str, Any]] = []
    source = "none"
    if item.get("category") in ("finished_good", "sub_assembly"):
        breakdown = await _compute_inherited_variants_breakdown(item_id)
        if breakdown:
            source = "inherited"
        elif effective:
            source = "own"  # FG/SG with no BOM yet — falling back to legacy own
    elif effective:
        source = "own"
    return {
        "item_id": item_id,
        "category": item.get("category"),
        "source": source,
        "variant_attributes": effective,
        "variant_sources": breakdown,
    }


@items_router.get("/{item_id}/variants")
async def list_item_variants(item_id: str, request: Request):
    """List every generated variant child of a parent item."""
    await get_current_user(request)
    parent = await db.items.find_one({"id": item_id})
    if not parent:
        raise HTTPException(status_code=404, detail="Parent item not found")
    variants = await db.items.find(
        {"parent_item_id": item_id, "is_variant": True},
        {"_id": 0}
    ).sort("part_number", 1).to_list(500)
    return variants


@items_router.post("/{item_id}/preview-variants")
async def preview_item_variants(item_id: str, request: Request):
    """Compute every combination from the parent's *effective* variant_attributes
    (own for CP/RM, inherited from BOM components for FG/SG) and report which
    already exist as child items vs which would be NEW."""
    await get_current_user(request)
    parent = await db.items.find_one({"id": item_id})
    if not parent:
        raise HTTPException(status_code=404, detail="Parent item not found")
    norm_attrs = await _get_effective_variants(parent)
    if not norm_attrs:
        return {"combinations": [], "existing_skus": [], "summary": "No variant attributes defined or inherited for this item."}
    parent_sku = parent.get("part_number") or ""
    combos = _all_variant_combinations(norm_attrs)
    # Existing children: map by sku for fast lookup
    existing = await db.items.find(
        {"parent_item_id": item_id, "is_variant": True},
        {"_id": 0, "part_number": 1, "id": 1, "is_active": 1, "current_stock": 1}
    ).to_list(1000)
    existing_by_sku = {e["part_number"]: e for e in existing}
    rows = []
    for combo in combos:
        sku = _build_variant_sku_from_short_codes(parent_sku, combo)
        ex = existing_by_sku.get(sku)
        rows.append({
            "sku": sku,
            "combination": combo,
            "exists": bool(ex),
            "is_active": ex.get("is_active", True) if ex else True,
            "current_stock": ex.get("current_stock", 0) if ex else 0,
            "label": ", ".join(f"{k}: {v['value']}" for k, v in combo.items()),
        })
    return {
        "parent_item_id": item_id,
        "parent_sku": parent_sku,
        "combinations": rows,
        "existing_count": sum(1 for r in rows if r["exists"]),
        "new_count": sum(1 for r in rows if not r["exists"]),
    }


@items_router.post("/{item_id}/generate-variants")
async def generate_item_variants(item_id: str, payload: dict = Body(default={}), request: Request = None):
    """Generate variant child items based on `selected_skus` from /preview-variants.

    Payload shape:
      { "selected_skus": ["FG-001-1HP-220V", "FG-001-2HP-440V", ...] }

    For each requested SKU not already present, a new item record is created
    inheriting key fields from the parent. Existing variant items are left as-is.

    Also: any existing variant SKU NOT in the current combination set is marked
    `is_active=False` (so removing an attribute value safely retires variants
    without losing stock history).
    """
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="items", action="create")
    parent = await db.items.find_one({"id": item_id})
    if not parent:
        raise HTTPException(status_code=404, detail="Parent item not found")
    norm_attrs = await _get_effective_variants(parent)
    if not norm_attrs:
        raise HTTPException(status_code=400, detail="Parent item has no variant attributes (own or inherited from BOM components).")
    parent_sku = parent.get("part_number") or ""
    combos = _all_variant_combinations(norm_attrs)
    selected_skus = set(payload.get("selected_skus") or [])
    # Index all valid combos by sku.
    combo_by_sku = {
        _build_variant_sku_from_short_codes(parent_sku, c): c for c in combos
    }
    if not selected_skus:
        # Default to generating ALL combos.
        selected_skus = set(combo_by_sku.keys())
    # Fetch existing variants of this parent.
    existing = await db.items.find(
        {"parent_item_id": item_id, "is_variant": True},
        {"_id": 0, "part_number": 1, "id": 1}
    ).to_list(1000)
    existing_by_sku = {e["part_number"]: e for e in existing}
    created = []
    reactivated = []
    valid_skus_in_combos = set(combo_by_sku.keys())
    for sku in selected_skus:
        if sku not in combo_by_sku:
            continue  # SKU not in current combination set — ignore
        ex = existing_by_sku.get(sku)
        if ex:
            # Reactivate if it was retired.
            await db.items.update_one({"id": ex["id"]}, {"$set": {"is_active": True}})
            reactivated.append(sku)
            continue
        # Build child item from parent.
        combo = combo_by_sku[sku]
        short_codes = {k: v["short_code"] for k, v in combo.items()}
        values_map = {k: v["value"] for k, v in combo.items()}
        # Copy + override.
        child = {k: v for k, v in parent.items() if k not in ("_id", "id", "variant_attributes", "auto_suffix_variant_sku")}
        child["id"] = str(uuid.uuid4())
        child["part_number"] = sku
        child["name"] = parent.get("name", "") + " · " + ", ".join(f"{k}: {v}" for k, v in values_map.items())
        child["parent_item_id"] = item_id
        child["is_variant"] = True
        child["variant_short_codes"] = short_codes
        child["variant_values"] = values_map
        child["variant_attributes"] = None  # only the master holds attrs
        child["is_active"] = True
        child["current_stock"] = 0
        child["reserved_stock"] = 0
        child["safety_stock"] = parent.get("safety_stock", 0) or 0
        child["reorder_point"] = parent.get("reorder_point", 0) or 0
        child["created_at"] = datetime.now(timezone.utc)
        child["created_by"] = user["id"]
        await db.items.insert_one(child)
        child.pop("_id", None)
        created.append(child)
    # Retire any pre-existing variant whose SKU is no longer in the valid combos
    # (i.e. an attribute value was removed). HARD BLOCK if any referenced —
    # the user must reverse the transactional usage first. Otherwise hard-
    # delete (clean removal).
    blocked_skus = []
    candidates: List[Dict[str, Any]] = []
    for sku, ex in existing_by_sku.items():
        if sku in valid_skus_in_combos:
            continue
        if await _is_variant_referenced(ex["id"]):
            blocked_skus.append(sku)
        else:
            candidates.append(ex)
    if blocked_skus:
        raise HTTPException(
            status_code=400,
            detail=(
                "Cannot remove variant value(s): the following SKU(s) are used in "
                "Purchase Orders / GRNs / Invoices / BOMs / Inventory transactions and must "
                "be reversed there first → "
                + ", ".join(sorted(blocked_skus))
            ),
        )
    retired_skus: List[str] = []
    for ex in candidates:
        await db.items.delete_one({"id": ex["id"]})
        retired_skus.append(ex.get("part_number", ""))
    # Variant migration — once a parent has at least one variant child, its
    # OWN current_stock is no longer the source of truth (variants own the
    # stock balances). Zero out the parent and emit an 'adjust' inventory
    # transaction so the audit trail stays clean. Fixes the "parent shows
    # phantom stock above the variant rollup" bug.
    parent_after = await db.items.find_one({"id": item_id}, {"_id": 0})
    if parent_after and float(parent_after.get("current_stock") or 0) > 0:
        any_active_children = await db.items.count_documents({
            "parent_item_id": item_id, "is_variant": True, "is_active": {"$ne": False}
        })
        if any_active_children:
            prev = float(parent_after.get("current_stock") or 0)
            await db.items.update_one({"id": item_id}, {"$set": {"current_stock": 0, "updated_at": datetime.now(timezone.utc)}})
            await db.inventory_transactions.insert_one({
                "id": str(uuid.uuid4()),
                "item_id": item_id,
                "transaction_type": "adjust",
                "quantity": -prev,
                "reference_type": "variant_migration",
                "reference_id": item_id,
                "previous_stock": prev,
                "new_stock": 0,
                "notes": "Parent stock zeroed — variant SKUs now own balances. Allocate to a variant via Items & Stock → Edit.",
                "created_at": datetime.now(timezone.utc),
                "created_by": user["id"],
            })
    return {
        "message": f"Generated {len(created)} new variant(s)"
            + (f", reactivated {len(reactivated)}" if reactivated else "")
            + (f", deleted {len(retired_skus)} obsolete" if retired_skus else ""),
        "created": created,
        "reactivated_skus": reactivated,
        "deleted_skus": retired_skus,
        "retired_in_use_skus": [],
        "deactivated_skus": retired_skus,
    }


# ================== ITEM GROUPS ROUTES ==================
# User-managed taxonomy (Motors, Bearings, Valves, V-Belts, etc.)
# When a group has default_hsn_code / default_gst_rate, all items in that group inherit them.

@item_groups_router.get("")
async def list_item_groups(request: Request, parent_category: Optional[str] = None):
    await get_current_user(request)
    query = {}
    if parent_category:
        query["parent_category"] = parent_category
    groups = await db.item_groups.find(query, {"_id": 0}).sort("name", 1).to_list(500)
    # Attach item_count for UX (how many items belong to each group)
    for g in groups:
        g["item_count"] = await db.items.count_documents({"group_id": g["id"]})
    return groups

@item_groups_router.post("", status_code=201)
async def create_item_group(data: ItemGroupCreate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager", "inventory_manager"], module="items", action="create")
    existing = await db.item_groups.find_one({"name": data.name, "parent_category": data.parent_category})
    if existing:
        raise HTTPException(status_code=400, detail=f"Group '{data.name}' already exists under {data.parent_category or 'any category'}")
    doc = {
        "id": str(uuid.uuid4()),
        **data.model_dump(),
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"],
    }
    await db.item_groups.insert_one(doc)
    doc.pop("_id", None)
    doc["item_count"] = 0
    return doc

@item_groups_router.put("/{group_id}")
async def update_item_group(group_id: str, data: ItemGroupUpdate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager", "inventory_manager"], module="items", action="edit")
    update = {k: v for k, v in data.model_dump(exclude_none=True).items()}
    if not update:
        raise HTTPException(status_code=400, detail="No data to update")
    update["updated_at"] = datetime.now(timezone.utc)
    result = await db.item_groups.update_one({"id": group_id}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Item group not found")
    # If default_hsn_code or default_gst_rate changed, cascade to ALL items in the group
    if "default_hsn_code" in update or "default_gst_rate" in update:
        cascade = {}
        if "default_hsn_code" in update and update["default_hsn_code"]:
            cascade["hsn_code"] = update["default_hsn_code"]
        if "default_gst_rate" in update and update["default_gst_rate"] is not None:
            cascade["gst_rate"] = float(update["default_gst_rate"])
        if cascade:
            await db.items.update_many({"group_id": group_id}, {"$set": cascade})
    group = await db.item_groups.find_one({"id": group_id}, {"_id": 0})
    group["item_count"] = await db.items.count_documents({"group_id": group_id})
    return group

@item_groups_router.delete("/{group_id}")
async def delete_item_group(group_id: str, request: Request):
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can delete item groups")
    count = await db.items.count_documents({"group_id": group_id})
    if count > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete — {count} item(s) still belong to this group. Reassign or remove them first.")
    result = await db.item_groups.delete_one({"id": group_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Item group not found")
    return {"message": "Item group deleted"}


# ================== BOM ROUTES ==================

@bom_router.get("")
async def get_boms(request: Request, status: Optional[str] = None):
    await get_current_user(request)
    query = {}
    if status:
        query["status"] = status
    boms = await db.boms.find(query, {"_id": 0}).to_list(10000)

    # Batch-load parent items in ONE query instead of N find_one() calls.
    # Previously the per-BOM loop did `db.items.find_one({"id": parent_item_id})`
    # for every BOM, which on a 100-BOM list cost ~100 round-trips before the
    # response even started streaming.
    parent_ids = [b.get("parent_item_id") for b in boms if b.get("parent_item_id")]
    items_map = {}
    if parent_ids:
        async for it in db.items.find({"id": {"$in": parent_ids}}, {"_id": 0}):
            items_map[it["id"]] = it
    for bom in boms:
        bom["parent_item"] = items_map.get(bom.get("parent_item_id"))
    return boms


@bom_router.get("/rollup-costs")
async def get_all_bom_rollup_costs(request: Request, status: Optional[str] = "active"):
    """Return rollup costs for ALL BOMs (default: active) in a single response.

    Replaces the N+1 pattern where the BOM list page used to call
    `/api/bom/{id}/explode` for every active BOM (one HTTP round-trip each)
    just to render the inline `Total` / `FG Process` cost tags on the panel
    headers. This endpoint pre-loads all items + all BOMs into memory, then
    runs the same recursive cost rollup logic in-process with O(N) DB calls
    total instead of O(N × depth × children).

    Response shape:
        { bom_id: { fg_process_cost_per_unit, components_cost, total_rollup_cost } }
    """
    await get_current_user(request)

    # 1. Load every BOM (we need ALL of them, not just active, because child
    #    BOMs of an active parent might be in another state — but for rollup
    #    we only follow children that are status='active').
    all_boms = await db.boms.find({}, {"_id": 0}).to_list(5000)
    # Map: parent_item_id -> active BOM doc (single source of truth for child lookups)
    active_bom_by_parent = {b["parent_item_id"]: b for b in all_boms if b.get("status") == "active"}
    # Map: bom_id -> bom doc
    bom_by_id = {b["id"]: b for b in all_boms}

    # 2. Pre-load all items in one query so component cost lookups don't hit Mongo.
    all_items = {}
    async for it in db.items.find({}, {"_id": 0, "id": 1, "unit_cost": 1}):
        all_items[it["id"]] = it

    # 2b. Pre-load SC-order job-work-part charges (latest per item_id) and
    #     completed-WO operation cost (latest per item_id). The /explode
    #     endpoint falls back to these when a component has no routings, so
    #     rollup-costs must mirror that fallback or the panel total visibly
    #     changes when the user expands the panel.
    sc_charge_by_item: Dict[str, float] = {}
    async for so in db.subcontract_orders.find(
        {"status": {"$in": ["in_progress", "completed"]}, "job_work_parts": {"$exists": True, "$ne": []}},
        {"_id": 0, "job_work_parts": 1, "created_at": 1},
    ).sort("created_at", -1):
        for jwp in so.get("job_work_parts", []):
            iid = jwp.get("item_id")
            charge = float(jwp.get("charges") or 0)
            # Keep the FIRST hit per item (sort is created_at DESC → latest first).
            if iid and charge and iid not in sc_charge_by_item:
                sc_charge_by_item[iid] = charge
    wo_proc_by_item: Dict[str, float] = {}
    async for wo in db.work_orders.find(
        {"status": "completed", "operations_status": {"$exists": True}},
        {"_id": 0, "item_id": 1, "operations_status": 1, "actual_end": 1},
    ).sort("actual_end", -1):
        iid = wo.get("item_id")
        if not iid or iid in wo_proc_by_item:
            continue
        total = 0.0
        for op in wo.get("operations_status", []):
            total += float(op.get("process_cost_per_unit") or 0)
        if total:
            wo_proc_by_item[iid] = total

    # 3. In-memory recursive rollup. Each BOM's cost is memoized in `cache` so
    #    a shared sub-assembly used 10 times is computed once.
    cache = {}

    def rollup(bom_id: str, depth: int = 0):
        if bom_id in cache:
            return cache[bom_id]
        if depth > 12:
            return {"components_cost": 0.0, "fg_process_cost": 0.0, "total": 0.0}
        bom = bom_by_id.get(bom_id)
        if not bom:
            return {"components_cost": 0.0, "fg_process_cost": 0.0, "total": 0.0}

        components_cost = 0.0
        for comp in bom.get("components", []):
            cid = comp.get("item_id")
            qty = comp.get("quantity", 0) or 0
            child_bom = active_bom_by_parent.get(cid)
            if child_bom:
                child = rollup(child_bom["id"], depth + 1)
                # Match the /explode endpoint EXACTLY (single source of truth):
                #   Material Cost  = sum of child's children-only rollup (no FG process)
                #                  → that's child["components_cost"], NOT child["total"].
                #   Process Cost   = child's own FG-process (parent_routings).
                # The previous version used child["total"] for unit_cost AND added
                # child["fg_process_cost"] again, double-counting it. /rollup-costs
                # was therefore inflated vs. /explode, causing the panel total to
                # change after the user clicked to expand a BOM.
                unit_cost = child["components_cost"]
                process_cost_per_unit = child["fg_process_cost"]
            else:
                item = all_items.get(cid)
                unit_cost = (item or {}).get("unit_cost", 0) or 0
                process_cost_per_unit = routings_total_cost(comp.get("routings", []))
            # Same SC/WO fallback as /explode — only kicks in when no routing
            # has been configured for this leaf/component.
            if not process_cost_per_unit:
                process_cost_per_unit = sc_charge_by_item.get(cid, 0) or wo_proc_by_item.get(cid, 0)
            components_cost += (unit_cost + process_cost_per_unit) * qty

        fg_process_cost = routings_total_cost(bom.get("parent_routings", []))
        total = components_cost + fg_process_cost
        cache[bom_id] = {
            "components_cost": components_cost,
            "fg_process_cost": fg_process_cost,
            "total": total,
        }
        return cache[bom_id]

    # 4. Build the response. Filter by `status` query param (default: 'active').
    target_boms = [b for b in all_boms if (not status or b.get("status") == status)]
    out = {}
    for b in target_boms:
        r = rollup(b["id"])
        out[b["id"]] = {
            "fg_process_cost_per_unit": round(r["fg_process_cost"], 2),
            "components_cost": round(r["components_cost"], 2),
            "total_rollup_cost": round(r["total"], 2),
        }
    return out

@bom_router.get("/costs/{item_id}")
async def get_bom_costs_for_item(item_id: str, request: Request):
    """Return BOM-based RM cost, process cost, and process names for an item."""
    await get_current_user(request)
    return await compute_bom_costs(item_id)

@bom_router.get("/routing-cost")
async def get_specific_routing_cost(request: Request, item_id: str, process_name: str):
    """Return the per-unit cost of a SPECIFIC routing operation on an item.
    Used by Job Card OS flows where one particular routing op is outsourced
    and we need that op's cost (not the combined process_cost across all
    routings)."""
    await get_current_user(request)
    cost = await find_routing_cost(item_id, process_name)
    return {"item_id": item_id, "process_name": process_name, "cost": cost}

@bom_router.get("/{bom_id}")
async def get_bom(bom_id: str, request: Request):
    await get_current_user(request)
    bom = await db.boms.find_one({"id": bom_id}, {"_id": 0})
    if not bom:
        raise HTTPException(status_code=404, detail="BOM not found")
    
    # Enrich with item details
    item = await db.items.find_one({"id": bom.get("parent_item_id")}, {"_id": 0})
    bom["parent_item"] = item
    
    # Enrich components with item details
    for comp in bom.get("components", []):
        comp_item = await db.items.find_one({"id": comp.get("item_id")}, {"_id": 0})
        comp["item"] = comp_item
        
        # Check if component has its own BOM (for multi-level)
        child_bom = await db.boms.find_one({"parent_item_id": comp.get("item_id"), "status": "active"}, {"_id": 0})
        comp["has_bom"] = child_bom is not None
        if child_bom:
            comp["child_bom_id"] = child_bom.get("id")
    
    return bom

@bom_router.get("/by-item/{item_id}/preview")
async def get_bom_preview_by_item(item_id: str, request: Request):
    """Lightweight BOM preview by ITEM id (not bom id). Used by the Manual DC
    dialog and similar list-row "expand" UIs to show the constituent
    parts/RMs of a selected item with each child's unit cost. Single-level
    explode (children of the item's active BOM). Returns:
      { has_bom: bool, bom_id, components: [{item_id, part_number, name,
        category, quantity, uom, unit_cost, extended_cost}] }

    `unit_cost` is the rolled-up cost: for components with their own BOM
    (Parts / Sub-Assemblies), it's the deep rollup (children + parent_routings).
    For leaves it falls back to purchase_price → unit_cost on the master.
    Mirrors the cost chain used by JW-OS so manual SOs price RMs identically.
    """
    await get_current_user(request)
    bom = await db.boms.find_one({"parent_item_id": item_id, "status": "active"}, {"_id": 0})
    if not bom:
        return {"has_bom": False, "bom_id": None, "components": []}
    comp_ids = [c.get("item_id") for c in (bom.get("components") or []) if c.get("item_id")]
    items_map = {}
    if comp_ids:
        async for it in db.items.find({"id": {"$in": comp_ids}}, {"_id": 0, "id": 1, "part_number": 1, "name": 1, "category": 1, "unit_of_measure": 1, "purchase_price": 1, "unit_cost": 1, "description": 1}):
            items_map[it["id"]] = it
    # Build a child_bom map for the components that have their own BOM, so we
    # can resolve the deep-rolled unit_cost without round-tripping per item.
    child_boms_map: Dict[str, Dict[str, Any]] = {}
    if comp_ids:
        async for cb in db.boms.find({"parent_item_id": {"$in": comp_ids}, "status": "active"}, {"_id": 0}):
            child_boms_map[cb.get("parent_item_id")] = cb

    async def _rolled_unit_cost(iid: str, depth: int = 0) -> float:
        """Recursive rollup: components(qty × rolled cost) + parent_routings cost."""
        if depth > 12:  # cycle / runaway guard
            return 0.0
        bom_doc = child_boms_map.get(iid) if depth == 0 else await db.boms.find_one({"parent_item_id": iid, "status": "active"}, {"_id": 0})
        if not bom_doc:
            it = await db.items.find_one({"id": iid}, {"_id": 0, "purchase_price": 1, "unit_cost": 1})
            return float((it or {}).get("purchase_price") or (it or {}).get("unit_cost") or 0)
        total = 0.0
        for c in (bom_doc.get("components") or []):
            if c.get("is_alternate"):
                continue
            qty = float(c.get("quantity") or 0)
            child_cost = await _rolled_unit_cost(c.get("item_id"), depth + 1)
            total += qty * child_cost
        total += float(routings_total_cost(bom_doc.get("parent_routings", []) or []))
        return total

    components = []
    for c in (bom.get("components") or []):
        if c.get("is_alternate"):
            continue  # alternates aren't shipped — skip in preview
        cid = c.get("item_id")
        it = items_map.get(cid) or {}
        qty = float(c.get("quantity") or 0)
        if cid in child_boms_map:
            unit_cost = await _rolled_unit_cost(cid)
        else:
            unit_cost = float(it.get("purchase_price") or it.get("unit_cost") or 0)
        components.append({
            "item_id": cid,
            "part_number": it.get("part_number") or "",
            "name": it.get("name") or "",
            "category": it.get("category") or "",
            "description": it.get("description") or "",
            "quantity": qty,
            "uom": it.get("unit_of_measure") or "pcs",
            "unit_cost": unit_cost,
            "extended_cost": qty * unit_cost,
        })
    return {"has_bom": True, "bom_id": bom.get("id"), "components": components}


@bom_router.get("/{bom_id}/explode")
async def explode_bom(bom_id: str, request: Request, levels: int = 10):
    """Get full multi-level BOM explosion with rollup costing"""
    await get_current_user(request)
    
    async def explode_level(bom_id: str, level: int, max_levels: int):
        if level > max_levels:
            return []
        
        bom = await db.boms.find_one({"id": bom_id}, {"_id": 0})
        if not bom:
            return []
        
        result = []
        for comp in bom.get("components", []):
            item = await db.items.find_one({"id": comp.get("item_id")}, {"_id": 0})
            comp_data = {
                "level": level,
                "item": item,
                "quantity": comp.get("quantity"),
                "is_alternate": comp.get("is_alternate", False),
                "routings": comp.get("routings", []),
                "children": [],
                "unit_cost": 0,
                "extended_cost": 0
            }
            
            # Check for child BOM
            child_bom = await db.boms.find_one({"parent_item_id": comp.get("item_id"), "status": "active"}, {"_id": 0})
            child_fg_process = 0
            if child_bom:
                comp_data["child_bom_id"] = child_bom.get("id")
                comp_data["children"] = await explode_level(child_bom.get("id"), level + 1, max_levels)
                # Routings column — source from child BOM's parent_routings (single source of truth model)
                comp_data["routings"] = normalize_routings(child_bom.get("parent_routings", []))
                # Material Cost (unit_cost column) = children's rolled-up Total × qty ONLY.
                # The child's own parent_routings (its FG Process) goes into the Process column
                # below, so that editing a sub-BOM's parent_routings visibly updates the Process
                # Cost/Unit column on the parent tree (not buried into Material).
                children_rollup = sum(c.get("extended_cost", 0) for c in comp_data["children"])
                child_fg_process = routings_total_cost(child_bom.get("parent_routings", []))
                comp_data["unit_cost"] = children_rollup
            else:
                # Leaf node - use item unit_cost. Routings stay as comp-line routings (for leaves).
                comp_data["unit_cost"] = item.get("unit_cost", 0) if item else 0
                comp_data["routings"] = normalize_routings(comp.get("routings", []))
            
            comp_data["extended_cost"] = 0  # Will be recalculated after process cost
            
            # Calculate process cost — ONE source of truth model:
            #   • If the component has its own BOM → use THAT BOM's parent_routings (child_fg_process).
            #     Component-line routings on the parent's BOM line are ignored to prevent double counting.
            #     User edits PT-1's process in ONE place (PT-1's BOM parent_routings) and it flows up.
            #   • If the component is a leaf (no child BOM) → use component-line routings on parent's line.
            #   • Fallback: SC order charges, then completed WO actuals.
            if child_bom:
                process_cost_per_unit = child_fg_process
            else:
                process_cost_per_unit = routings_total_cost(comp.get("routings", []))
            
            if not process_cost_per_unit:
                # Check SC orders where this item was outsourced (job_work_parts charges)
                sc_orders = await db.subcontract_orders.find(
                    {"job_work_parts.item_id": comp.get("item_id"), "status": {"$in": ["in_progress", "completed"]}},
                    {"_id": 0, "job_work_parts": 1}
                ).sort("created_at", -1).to_list(1)
                if sc_orders:
                    for jwp in sc_orders[0].get("job_work_parts", []):
                        if jwp.get("item_id") == comp.get("item_id") and jwp.get("charges"):
                            process_cost_per_unit += jwp["charges"]
            
            if not process_cost_per_unit:
                # Check completed WO operations (inhouse process costs)
                latest_wo = await db.work_orders.find_one(
                    {"item_id": comp.get("item_id"), "status": "completed", "operations_status": {"$exists": True}},
                    {"_id": 0, "operations_status": 1},
                    sort=[("actual_end", -1)]
                )
                if latest_wo:
                    for op in latest_wo.get("operations_status", []):
                        if op.get("process_cost_per_unit"):
                            process_cost_per_unit += op["process_cost_per_unit"]
            
            comp_data["process_cost_per_unit"] = process_cost_per_unit
            comp_data["total_cost_per_unit"] = comp_data["unit_cost"] + process_cost_per_unit
            comp_data["extended_cost"] = comp_data["total_cost_per_unit"] * comp.get("quantity", 0)
            
            result.append(comp_data)
        return result
    
    bom = await db.boms.find_one({"id": bom_id}, {"_id": 0})
    if not bom:
        raise HTTPException(status_code=404, detail="BOM not found")
    
    parent_item = await db.items.find_one({"id": bom.get("parent_item_id")}, {"_id": 0})
    explosion = await explode_level(bom_id, 1, levels)
    
    # Components rollup (material + component process costs)
    components_cost = sum(c.get("extended_cost", 0) for c in explosion)
    # FG parent process cost (e.g., Assembly, Powder Coating done on the parent item)
    fg_process_cost = routings_total_cost(bom.get("parent_routings", []))
    total_cost = components_cost + fg_process_cost
    
    return {
        "bom": bom,
        "parent_item": parent_item,
        "explosion": explosion,
        "fg_process_cost_per_unit": round(fg_process_cost, 2),
        "components_cost": round(components_cost, 2),
        "total_rollup_cost": round(total_cost, 2)
    }

@bom_router.post("")
async def create_bom(bom_data: BOMCreate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="bom", action="create")
    # Verify parent item exists
    parent_item = await db.items.find_one({"id": bom_data.parent_item_id})
    if not parent_item:
        raise HTTPException(status_code=404, detail="Parent item not found")
    # Guard: Raw Material items are leaf parts — they cannot have a BOM of
    # their own (no children, no routing). Block at the API layer so this is
    # enforced whether the BOM was created via UI, Excel import, or a raw API
    # call.
    if (parent_item.get("category") or "").lower() == "raw_material":
        raise HTTPException(
            status_code=400,
            detail=f"Item '{parent_item.get('part_number')}' is a Raw Material — RM items cannot have a BOM.",
        )
    
    # Normalize component & parent routings to {name, cost}
    normalized_components = []
    for c in bom_data.components:
        cd = c.model_dump()
        cd["routings"] = normalize_routings(cd.get("routings", []))
        normalized_components.append(cd)
    # Server-side de-dupe safety net: collapse duplicate (item_id, is_alternate)
    # rows and SUM their quantities. Mirrors the client-side logic in
    # BOMPage.handleSubmit so duplicates can't slip in via Excel import or
    # raw API calls.
    normalized_components = _merge_duplicate_components(normalized_components)
    
    bom_doc = {
        "id": str(uuid.uuid4()),
        "parent_item_id": bom_data.parent_item_id,
        "name": bom_data.name,
        "description": bom_data.description,
        "revision": bom_data.revision,
        "effectivity_date": bom_data.effectivity_date or datetime.now(timezone.utc),
        "expiry_date": bom_data.expiry_date,
        "status": bom_data.status,
        "components": normalized_components,
        "parent_routings": normalize_routings(bom_data.parent_routings or []),
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"]
    }
    await db.boms.insert_one(bom_doc)
    bom_doc.pop("_id", None)
    return bom_doc


@bom_router.post("/dedupe-all")
async def dedupe_all_boms(request: Request):
    """One-off admin sweep — walks every BOM in the database and merges any
    duplicate component rows that were saved BEFORE client-side / server-side
    de-dupe was added. Idempotent: running twice is a no-op once everything
    is clean. Returns a summary of how many BOMs and rows were merged."""
    user = await get_current_user(request)
    _require_access(user, ["admin"], module="bom", action="edit")
    cursor = db.boms.find({}, {"_id": 0})
    boms_processed = 0
    boms_modified = 0
    rows_merged = 0
    async for bom in cursor:
        boms_processed += 1
        comps = bom.get("components") or []
        before = len(comps)
        cleaned = _merge_duplicate_components(comps)
        after = len(cleaned)
        if after < before:
            await db.boms.update_one(
                {"id": bom["id"]},
                {"$set": {"components": cleaned, "updated_at": datetime.now(timezone.utc)}},
            )
            boms_modified += 1
            rows_merged += (before - after)
    return {
        "boms_processed": boms_processed,
        "boms_modified": boms_modified,
        "rows_merged": rows_merged,
    }

@bom_router.put("/{bom_id}")
async def update_bom(bom_id: str, bom_data: BOMUpdate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="bom", action="edit")
    update_data = {}
    for k, v in bom_data.model_dump().items():
        if v is not None:
            if k == "components":
                normalized = []
                for c in v:
                    cd = c.model_dump() if hasattr(c, 'model_dump') else dict(c)
                    cd["routings"] = normalize_routings(cd.get("routings", []))
                    normalized.append(cd)
                # Server-side de-dupe safety net (same as create_bom).
                update_data[k] = _merge_duplicate_components(normalized)
            elif k == "parent_routings":
                update_data[k] = normalize_routings(v)
            else:
                update_data[k] = v
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    update_data["updated_at"] = datetime.now(timezone.utc)
    result = await db.boms.update_one({"id": bom_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="BOM not found")
    
    bom = await db.boms.find_one({"id": bom_id}, {"_id": 0})
    return bom

@bom_router.post("/{bom_id}/revise")
async def create_bom_revision(bom_id: str, new_revision: str, request: Request):
    """Create a new revision of an existing BOM"""
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="bom", action="create")
    bom = await db.boms.find_one({"id": bom_id}, {"_id": 0})
    if not bom:
        raise HTTPException(status_code=404, detail="BOM not found")
    
    # Mark old BOM as obsolete
    await db.boms.update_one({"id": bom_id}, {"$set": {"status": "obsolete"}})
    
    # Create new revision
    new_bom = {
        "id": str(uuid.uuid4()),
        "parent_item_id": bom["parent_item_id"],
        "name": bom["name"],
        "description": bom.get("description", ""),
        "revision": new_revision,
        "previous_revision_id": bom_id,
        "effectivity_date": datetime.now(timezone.utc),
        "expiry_date": None,
        "status": "active",
        "components": bom.get("components", []),
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"]
    }
    await db.boms.insert_one(new_bom)
    new_bom.pop("_id", None)
    return new_bom

@bom_router.delete("/{bom_id}")
async def delete_bom(bom_id: str, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin"], module="bom", action="delete")
    result = await db.boms.delete_one({"id": bom_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="BOM not found")
    return {"message": "BOM deleted"}

# ================== PRODUCTION ORDER ROUTES ==================

@production_router.get("")
async def get_production_orders(request: Request, status: Optional[str] = None):
    await get_current_user(request)
    query = {}
    if status:
        query["status"] = status
    orders = await db.production_orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    for order in orders:
        bom = await db.boms.find_one({"id": order.get("bom_id")}, {"_id": 0})
        if bom:
            order["bom"] = bom
            item = await db.items.find_one({"id": bom.get("parent_item_id")}, {"_id": 0})
            order["item"] = item
        # Per-line MO consumption: pull every non-cancelled MO referencing this SO and bucket by source_so_line_id.
        mos = await db.work_orders.find(
            {"production_order_id": order["id"], "parent_wo_id": None, "status": {"$ne": "cancelled"}},
            {"quantity": 1, "source_so_line_id": 1, "_id": 0}
        ).to_list(1000)
        per_line_mo = {}
        for mo in mos:
            lid = mo.get("source_so_line_id") or ""
            per_line_mo[lid] = per_line_mo.get(lid, 0) + int(mo.get("quantity", 0) or 0)
        # Hydrate each line with its BOM + parent_item for the multi-line display
        for ln in (order.get("lines") or []):
            ln_bom = await db.boms.find_one({"id": ln.get("bom_id")}, {"_id": 0})
            if ln_bom:
                ln["bom"] = ln_bom
                ln["item"] = await db.items.find_one({"id": ln_bom.get("parent_item_id")}, {"_id": 0})
            # mo_qty_created on this specific line (tracked via source_so_line_id on MOs).
            ln["mo_qty_created"] = per_line_mo.get(ln.get("line_id"), 0)
            ln_qty = int(ln.get("quantity", 0) or 0)
            ln_resv = int(ln.get("reserved_qty", 0) or 0)
            # available_for_mo = qty - reserved (already in FG stock) - MOs already created.
            ln["available_for_mo"] = max(0, ln_qty - ln_resv - ln["mo_qty_created"])
        order["mo_qty_created"] = sum(per_line_mo.values())
    return orders

@production_router.get("/{order_id}")
async def get_production_order(order_id: str, request: Request):
    await get_current_user(request)
    order = await db.production_orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Production order not found")
    
    bom = await db.boms.find_one({"id": order.get("bom_id")}, {"_id": 0})
    if bom:
        order["bom"] = bom
        item = await db.items.find_one({"id": bom.get("parent_item_id")}, {"_id": 0})
        order["item"] = item
    return order

@production_router.post("")
async def create_production_order(order_data: ProductionOrderCreate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="production", action="create")
    # Resolve lines — either provided directly (multi-line) or built from legacy fields (single-line).
    raw_lines = order_data.lines or []
    if not raw_lines:
        # Legacy single-line SO — synthesize one line from the top-level fields.
        if not order_data.bom_id or not order_data.quantity or not order_data.due_date:
            raise HTTPException(status_code=400, detail="Provide either `lines[]` or bom_id + quantity + due_date")
        raw_lines = [ProductionOrderLine(
            bom_id=order_data.bom_id,
            quantity=order_data.quantity,
            due_date=order_data.due_date,
            order_type="mts",
            notes=order_data.notes or ""
        )]

    # Validate + enrich each line (BOM exists, assign line_id + line_no).
    enriched_lines = []
    for idx, ln in enumerate(raw_lines, start=1):
        if ln.quantity is None or ln.quantity <= 0:
            raise HTTPException(status_code=400, detail=f"Line {idx}: quantity must be > 0")
        bom = await db.boms.find_one({"id": ln.bom_id})
        if not bom:
            raise HTTPException(status_code=404, detail=f"Line {idx}: BOM {ln.bom_id} not found")
        if ln.order_type not in ("auto", "mts", "mto"):
            raise HTTPException(status_code=400, detail=f"Line {idx}: order_type must be mts | mto")
        # Legacy "auto" is normalised to "mts" so downstream MRP/confirm flows
        # only ever see the two supported modes.
        normalized_order_type = "mts" if (ln.order_type or "mts") == "auto" else ln.order_type
        # Phase 2 — compute display SKU (parent part_number + variant suffix).
        variant_sku = None
        if ln.variant_selection:
            parent_item = await db.items.find_one({"id": bom.get("parent_item_id")}, {"_id": 0, "part_number": 1})
            if parent_item:
                variant_sku = _build_variant_sku(parent_item.get("part_number") or "", ln.variant_selection)
        enriched_lines.append({
            "line_id": str(uuid.uuid4()),
            "line_no": idx,
            "bom_id": ln.bom_id,
            "quantity": ln.quantity,
            "due_date": ln.due_date,
            "order_type": normalized_order_type,
            "notes": ln.notes or "",
            "source_quotation_line_no": ln.source_quotation_line_no,
            "variant_selection": ln.variant_selection or None,
            "variant_sku": variant_sku,
            "reserved_qty": 0,
            "mo_qty": 0,
            "status": "draft"
        })

    # Generate order number
    count = await db.production_orders.count_documents({})
    order_number = f"SO-{str(count + 1).zfill(6)}"

    # Top-level legacy fields mirror the first line (keeps MO create flow backward-compatible).
    first_line = enriched_lines[0]
    total_qty = sum(l["quantity"] for l in enriched_lines)
    latest_due = max((l["due_date"] for l in enriched_lines if l.get("due_date")), default=None)

    order_doc = {
        "id": str(uuid.uuid4()),
        "order_number": order_number,
        "customer_id": order_data.customer_id,
        "lines": enriched_lines,
        # Legacy mirror fields — kept for existing MO / MRP flows that still read them.
        "bom_id": first_line["bom_id"],
        "quantity": first_line["quantity"] if len(enriched_lines) == 1 else total_qty,
        "due_date": first_line.get("due_date") or latest_due,
        "priority": order_data.priority,
        "status": "draft",
        "notes": order_data.notes or "",
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"]
    }
    # Optional link back to the source quotation (when SO was built via the
    # "From Quotation" picker on the Production page).
    if order_data.source_quotation_id:
        order_doc["source_quotation_id"] = order_data.source_quotation_id
        order_doc["source_quotation_no"] = order_data.source_quotation_no or ""
    await db.production_orders.insert_one(order_doc)
    order_doc.pop("_id", None)
    # If this SO is sourced from a quotation, update the quotation balance and
    # mark it converted once every line is fully consumed.
    if order_data.source_quotation_id:
        await _refresh_quotation_conversion_status(order_data.source_quotation_id, order_doc["id"], order_number)
    return order_doc

@production_router.put("/{order_id}")
async def update_production_order(order_id: str, order_data: ProductionOrderUpdate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="production", action="edit")
    order = await db.production_orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Production order not found")
    
    # Block edit if full quantity is already covered by MOs
    mos = await db.work_orders.find(
        {"production_order_id": order_id, "parent_wo_id": None, "status": {"$ne": "cancelled"}},
        {"quantity": 1, "_id": 0}
    ).to_list(1000)
    mo_qty_total = sum(mo.get("quantity", 0) for mo in mos)
    if mo_qty_total >= order.get("quantity", 0):
        raise HTTPException(status_code=400, detail=f"Cannot edit: Manufacturing Orders already created for full quantity ({mo_qty_total}/{order['quantity']}). Cancel existing MOs first.")
    
    update_data = {k: v for k, v in order_data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    update_data["updated_at"] = datetime.now(timezone.utc)
    await db.production_orders.update_one({"id": order_id}, {"$set": update_data})
    
    updated = await db.production_orders.find_one({"id": order_id}, {"_id": 0})
    return updated

@production_router.post("/{order_id}/confirm")
async def confirm_production_order(order_id: str, request: Request):
    """Confirm a draft SO. For each line, apply the MTO/MTS/auto split:
      • mts  → reserve up to qty from FG stock (no MO)
      • mto  → create MO for full qty (no reservation)
      • auto → reserve available FG stock + create MO for any shortfall (smart split)
    The actual MO is still created via the Manufacturing page (using `mo_qty` per line as
    the authoritative to-manufacture amount). This endpoint only computes + stores the
    split so the floor team knows what to do per line.
    """
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="production", action="create")
    order = await db.production_orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Sales order not found")
    if order.get("status") != "draft":
        raise HTTPException(status_code=400, detail=f"Only draft orders can be confirmed. Current status: {order.get('status')}")

    # Ensure lines exist — older SOs may have only legacy fields.
    lines = order.get("lines") or []
    if not lines and order.get("bom_id"):
        lines = [{
            "line_id": str(uuid.uuid4()),
            "line_no": 1,
            "bom_id": order["bom_id"],
            "quantity": order.get("quantity", 0),
            "due_date": order.get("due_date"),
            "order_type": "auto",
            "notes": "",
            "reserved_qty": 0,
            "mo_qty": 0,
            "status": "draft"
        }]

    summary_lines = []
    for ln in lines:
        qty = int(ln.get("quantity", 0) or 0)
        order_type = ln.get("order_type", "mto")
        # Backward compatibility: legacy SOs may still have order_type='auto'.
        # Map it to mto (the closer match — uses FG stock to reduce MO qty).
        if order_type == "auto":
            order_type = "mto"
        # FG parent item from the BOM
        bom = await db.boms.find_one({"id": ln.get("bom_id")}, {"_id": 0})
        fg_item_id = bom.get("parent_item_id") if bom else None
        fg_item = await db.items.find_one({"id": fg_item_id}, {"_id": 0}) if fg_item_id else None
        available_stock = int(fg_item.get("current_stock", 0) or 0) if fg_item else 0
        already_reserved = int(fg_item.get("reserved_stock", 0) or 0) if fg_item else 0
        effective_stock = max(0, available_stock - already_reserved)

        reserved_qty = 0
        mo_qty = 0
        if order_type == "mts":
            # NEW MTS: never use FG stock — always produce full SO qty. Child
            # SG/parts are auto-reserved when the MO is created (see
            # `create_child_work_orders` below).
            mo_qty = qty
        elif order_type == "mto":
            # NEW MTO: FG stock IS used. Reserve what's available, MO covers
            # only the balance. Child SG/parts auto-reserve at MO create time.
            reserved_qty = min(qty, effective_stock)
            mo_qty = max(0, qty - reserved_qty)

        # Increment reserved_stock on the FG item for the reserved portion (so subsequent
        # SO confirmations see the lower effective_stock).
        if reserved_qty > 0 and fg_item_id:
            await db.items.update_one(
                {"id": fg_item_id},
                {"$inc": {"reserved_stock": reserved_qty}}
            )

        ln["reserved_qty"] = reserved_qty
        ln["mo_qty"] = mo_qty
        ln["order_type"] = order_type
        ln["status"] = "confirmed"
        summary_lines.append({
            "line_no": ln.get("line_no"),
            "bom_id": ln.get("bom_id"),
            "fg_part_number": fg_item.get("part_number") if fg_item else None,
            "fg_name": fg_item.get("name") if fg_item else None,
            "quantity": qty,
            "order_type": order_type,
            "reserved_qty": reserved_qty,
            "mo_qty": mo_qty,
            "available_stock_before": effective_stock
        })

    await db.production_orders.update_one(
        {"id": order_id},
        {"$set": {
            "status": "confirmed",
            "lines": lines,
            "confirmed_at": datetime.now(timezone.utc),
            "confirmed_by": user["id"],
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    updated = await db.production_orders.find_one({"id": order_id}, {"_id": 0})
    updated["confirm_summary"] = summary_lines
    return updated

@production_router.post("/{order_id}/cancel")
async def cancel_production_order(order_id: str, request: Request):
    """Cancel a sales order with full cascade: SO → MOs → reverse stock → cancel job cards"""
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="production", action="create")
    order = await db.production_orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Sales order not found")
    
    if order.get("status") == "cancelled":
        raise HTTPException(status_code=400, detail="Order is already cancelled")
    if order.get("status") == "completed":
        raise HTTPException(status_code=400, detail="Cannot cancel a completed order")
    
    cancelled_mos = []
    skipped_completed_mos = []
    reversed_materials = []
    released_reservations = []
    
    # Release any reserved_stock bookings for SO lines that were in MTS/auto-reserved state.
    for ln in (order.get("lines") or []):
        reserved = int(ln.get("reserved_qty", 0) or 0)
        if reserved <= 0:
            continue
        bom = await db.boms.find_one({"id": ln.get("bom_id")}, {"_id": 0})
        fg_item_id = bom.get("parent_item_id") if bom else None
        if fg_item_id:
            await db.items.update_one(
                {"id": fg_item_id},
                {"$inc": {"reserved_stock": -reserved}}
            )
            released_reservations.append({"line_no": ln.get("line_no"), "fg_item_id": fg_item_id, "qty_released": reserved})
    
    # Find all manufacturing orders linked to this sales order
    mos = await db.work_orders.find({"production_order_id": order_id}).to_list(1000)
    
    for mo in mos:
        mo_id = mo["id"]
        mo_status = mo.get("status", "pending")
        mo_number = mo.get("wo_number", "")
        
        # Preserve completed MOs — their FG stock remains, MO stays in "completed" state.
        # Only cancel MOs that are still pending/in_progress/draft/confirmed/released.
        if mo_status in ("completed", "cancelled"):
            if mo_status == "completed":
                skipped_completed_mos.append(mo_number)
            continue

        # Release child-level reservations recorded when this MO was created.
        # Each entry in `child_reservations` decrements the corresponding
        # item's `reserved_stock` by the booked qty so subsequent SOs/MOs
        # see the freed stock available again.
        for resv in (mo.get("child_reservations") or []):
            resv_item = resv.get("item_id")
            resv_qty = int(resv.get("qty", 0) or 0)
            if resv_item and resv_qty > 0:
                await db.items.update_one(
                    {"id": resv_item},
                    {"$inc": {"reserved_stock": -resv_qty}},
                )
        
        # 1. Reverse consumed materials (if MO was started and materials were consumed)
        if mo.get("materials_consumed") and mo.get("consumed_materials"):
            for mat in mo["consumed_materials"]:
                comp_item = await db.items.find_one({"id": mat["item_id"]})
                if comp_item:
                    current_stock = comp_item.get("current_stock", 0)
                    restore_qty = mat["quantity"]
                    new_stock = current_stock + restore_qty
                    
                    # Create reversal transaction
                    tx_doc = {
                        "id": str(uuid.uuid4()),
                        "item_id": mat["item_id"],
                        "transaction_type": "receive",
                        "quantity": restore_qty,
                        "reference_type": "cancellation",
                        "reference_id": mo_id,
                        "previous_stock": current_stock,
                        "new_stock": new_stock,
                        "notes": f"Reversal: SO {order.get('order_number')} cancelled - MO {mo_number}",
                        "created_at": datetime.now(timezone.utc),
                        "created_by": user["id"]
                    }
                    await db.inventory_transactions.insert_one(tx_doc)
                    await db.items.update_one({"id": mat["item_id"]}, {"$set": {"current_stock": new_stock}})
                    
                    reversed_materials.append({
                        "item": mat.get("item", ""),
                        "name": mat.get("name", ""),
                        "quantity": restore_qty,
                        "mo_number": mo_number
                    })
        
        # 3. Cancel the manufacturing order
        await db.work_orders.update_one(
            {"id": mo_id},
            {"$set": {
                "status": "cancelled",
                "cancelled_at": datetime.now(timezone.utc),
                "cancelled_reason": f"Parent SO {order.get('order_number')} cancelled",
                "updated_at": datetime.now(timezone.utc)
            }}
        )
        cancelled_mos.append(mo_number)
    
    # SO final state: if some MOs completed, mark as partially_cancelled; else cancelled.
    final_status = "partially_cancelled" if skipped_completed_mos else "cancelled"
    await db.production_orders.update_one(
        {"id": order_id},
        {"$set": {
            "status": final_status,
            "cancelled_at": datetime.now(timezone.utc),
            "cancelled_by": user["id"],
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    msg_parts = [f"Sales Order {order.get('order_number')}"]
    if skipped_completed_mos:
        msg_parts.append(f"partially cancelled — {len(skipped_completed_mos)} completed MO(s) preserved")
    else:
        msg_parts.append("cancelled successfully")
    
    return {
        "message": " ".join(msg_parts),
        "cancelled_mos": cancelled_mos,
        "preserved_completed_mos": skipped_completed_mos,
        "reversed_materials": reversed_materials,
        "released_reservations": released_reservations
    }


@production_router.post("/{order_id}/reserve-line")
async def reserve_so_line(order_id: str, payload: dict = Body(default={}), request: Request = None):
    """Reserve stock on a Sales Order line's PARENT item (FG/SG/CP).

    This is the user-driven 'Reserve' button on the SO line. It does NOT explode the
    BOM — it only increments `reserved_stock` on the BOM's parent item by the
    full line quantity. Pair with `/release-line` to undo.
    """
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="production", action="edit")
    order = await db.production_orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Sales order not found")
    # Reserve is only allowed AFTER the SO is confirmed (i.e. not still in draft).
    if order.get("status") in (None, "", "draft"):
        raise HTTPException(status_code=400, detail="Sales Order must be confirmed before stock can be reserved")
    line_id = payload.get("line_id")
    line_no = payload.get("line_no")
    target_line = None
    for ln in order.get("lines") or []:
        if line_id and ln.get("line_id") == line_id:
            target_line = ln
            break
        if line_no is not None and str(ln.get("line_no")) == str(line_no):
            target_line = ln
            break
    if not target_line:
        raise HTTPException(status_code=404, detail=f"Line not found (line_id={line_id}, line_no={line_no})")
    qty = int(target_line.get("quantity", 0) or 0)
    if qty <= 0:
        raise HTTPException(status_code=400, detail="Invalid line quantity")
    if int(target_line.get("reserved_qty", 0) or 0) >= qty:
        raise HTTPException(status_code=400, detail="Line is already fully reserved. Release it first to re-reserve.")
    bom = await db.boms.find_one({"id": target_line.get("bom_id")}, {"_id": 0, "parent_item_id": 1})
    if not bom or not bom.get("parent_item_id"):
        raise HTTPException(status_code=404, detail="BOM (or its parent item) not found for this line")
    parent_item_id = bom["parent_item_id"]
    parent_item = await db.items.find_one({"id": parent_item_id}, {"_id": 0, "part_number": 1, "name": 1, "current_stock": 1, "reserved_stock": 1})
    if not parent_item:
        raise HTTPException(status_code=404, detail="Parent item missing in master")
    current_stock = int(parent_item.get("current_stock", 0) or 0)
    already_reserved = int(parent_item.get("reserved_stock", 0) or 0)
    free_stock = max(0, current_stock - already_reserved)
    prev_line_reserved = int(target_line.get("reserved_qty", 0) or 0)
    qty_still_needed = qty - prev_line_reserved
    if free_stock <= 0:
        raise HTTPException(
            status_code=400,
            detail=f"No FG stock available to reserve for {parent_item.get('part_number')} (stock={current_stock}, already reserved={already_reserved}). Create a Manufacturing Order to produce the qty.",
        )
    # Partial reserve allowed — lock whatever is available (up to the still-needed qty on this line).
    reserve_qty = min(qty_still_needed, free_stock)
    await db.items.update_one(
        {"id": parent_item_id},
        {"$inc": {"reserved_stock": reserve_qty}},
    )
    new_reserved = prev_line_reserved + reserve_qty
    is_fully_reserved = new_reserved >= qty
    await db.production_orders.update_one(
        {"id": order_id, "lines.line_id": target_line.get("line_id")},
        {"$set": {
            "lines.$.is_reserved": is_fully_reserved,
            "lines.$.reserved_qty": new_reserved,
            "lines.$.reserved_at": datetime.now(timezone.utc),
        }},
    )
    return {
        "message": (
            f"Reserved {reserve_qty} of {qty_still_needed} units of {parent_item.get('part_number')}"
            + (f" — line now fully reserved ({new_reserved}/{qty})" if is_fully_reserved else f" — line partial: {new_reserved}/{qty}. Click Reserve again when more FG arrives, or create an MO for the balance.")
        ),
        "reserved_qty": reserve_qty,
        "line_reserved_total": new_reserved,
        "line_quantity": qty,
        "is_fully_reserved": is_fully_reserved,
        "item_part_number": parent_item.get("part_number"),
        "item_name": parent_item.get("name"),
    }


@production_router.post("/{order_id}/release-line")
async def release_so_line(order_id: str, payload: dict = Body(default={}), request: Request = None):
    """Release a previously-reserved Sales Order line's parent stock."""
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="production", action="edit")
    order = await db.production_orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Sales order not found")
    line_id = payload.get("line_id")
    line_no = payload.get("line_no")
    target_line = None
    for ln in order.get("lines") or []:
        if line_id and ln.get("line_id") == line_id:
            target_line = ln
            break
        if line_no is not None and str(ln.get("line_no")) == str(line_no):
            target_line = ln
            break
    if not target_line:
        raise HTTPException(status_code=404, detail="Line not found")
    reserved_qty = int(target_line.get("reserved_qty", 0) or 0)
    if reserved_qty <= 0:
        raise HTTPException(status_code=400, detail="Line has no reservation to release")
    bom = await db.boms.find_one({"id": target_line.get("bom_id")}, {"_id": 0, "parent_item_id": 1})
    if not bom or not bom.get("parent_item_id"):
        raise HTTPException(status_code=404, detail="BOM (or its parent item) not found for this line")
    parent_item_id = bom["parent_item_id"]
    await db.items.update_one(
        {"id": parent_item_id},
        {"$inc": {"reserved_stock": -reserved_qty}},
    )
    await db.production_orders.update_one(
        {"id": order_id, "lines.line_id": target_line.get("line_id")},
        {"$set": {"lines.$.is_reserved": False, "lines.$.reserved_qty": 0}},
    )
    return {"message": f"Released {reserved_qty} unit(s)", "released_qty": reserved_qty}


@production_router.get("/{order_id}/print-data")
async def get_so_print_data(order_id: str, request: Request):
    """Return Sales Order document hydrated for PDF/Preview rendering."""
    await get_current_user(request)
    order = await db.production_orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Sales order not found")
    # Customer
    customer = None
    if order.get("customer_id"):
        customer = await db.customers.find_one({"id": order["customer_id"]}, {"_id": 0})
    order["customer"] = customer
    # Creator info for signature block
    order["created_by_user"] = await _lookup_creator(order.get("created_by"))
    # Company settings (header) — try both legacy keys for compatibility.
    company = (
        await db.company_settings.find_one({"type": "company"}, {"_id": 0})
        or await db.company_settings.find_one({"_id": "singleton"}, {"_id": 0})
        or {}
    )
    order["company"] = company
    # Hydrate each line with the parent item + BOM (so PDF can show part_number / name / HSN / UOM)
    for ln in (order.get("lines") or []):
        bom_doc = await db.boms.find_one({"id": ln.get("bom_id")}, {"_id": 0})
        if bom_doc:
            ln["bom"] = bom_doc
            item = await db.items.find_one({"id": bom_doc.get("parent_item_id")}, {"_id": 0})
            ln["item"] = item
    # If sourced from a quotation, pull rate/HSN/description per line for the PDF
    if order.get("source_quotation_id"):
        sq = await db.crm_quotations.find_one({"id": order["source_quotation_id"]}, {"_id": 0})
        if sq:
            order["source_quotation"] = {
                "id": sq.get("id"),
                "quotation_no": sq.get("quotation_no"),
                "currency": sq.get("currency", "INR"),
                "valid_until": sq.get("valid_until"),
            }
            for ln in (order.get("lines") or []):
                sqln_no = ln.get("source_quotation_line_no")
                if sqln_no is None:
                    continue
                qln = next(
                    (q_ln for q_ln in (sq.get("lines") or []) if str(q_ln.get("line_no")) == str(sqln_no)),
                    None,
                )
                if qln:
                    ln["rate"] = qln.get("rate", 0)
                    ln["discount_pct"] = qln.get("discount_pct", 0)
                    ln["gst_rate"] = qln.get("gst_rate", 0)
                    ln["hsn_code"] = qln.get("hsn_code") or (ln.get("item") or {}).get("hsn_code", "")
                    ln["description"] = qln.get("description") or ""
    return order


# ================== MRP ROUTES ==================

@mrp_router.get("/demand")
async def calculate_demand(request: Request, production_order_id: Optional[str] = None):
    """Calculate RM demand from open MANUFACTURING ORDERS (MTS + MTO).

    Rationale: SOs only show demand intent — they reserve available stock but
    don't commit to a production plan. Real RM/component demand happens when
    an MO is created (MTS or MTO). MRP therefore explodes:
      1. Reserved MOs → use their pre-computed `reserved_materials` shortfall.
      2. Un-reserved open MOs (pending / in_progress, not yet `materials_reserved`)
         → explode their BOM at MO quantity. Variant-bearing components are
         redirected to the matching variant child SKU when the MO has a
         `variant_selection`.
      3. SOs without any MO yet are NOT counted — create the MO first.

    Net Requirement = (reservation shortfall) + max(0, gross - available)
    where available = on_hand - allocated_for_mo - safety_stock.
    Only Raw Material items are returned."""
    await get_current_user(request)
    
    # Recalculate reservations based on current stock first
    await recalculate_all_reservations()
    
    # Optional filter by source SO (kept for backward compatibility — finds
    # all MOs linked to that SO instead of filtering by SO id directly).
    so_filter_ids: Set[str] = set()
    if production_order_id:
        so_filter_ids.add(production_order_id)
    
    # ---- Step 1: aggregate reserved-MO shortfalls (same as before) ----
    reserved_mos = await db.work_orders.find(
        {"materials_reserved": True, "status": {"$in": ["pending", "in_progress"]}},
        {"_id": 0, "reserved_materials": 1, "consumed_materials": 1, "production_order_id": 1, "quantity": 1}
    ).to_list(5000)
    
    total_allocated = {}
    total_shortfall = {}
    for mo in reserved_mos:
        if so_filter_ids and mo.get("production_order_id") not in so_filter_ids:
            continue
        # Already-consumed quantities are no longer allocated demand —
        # they've already left stores. Subtract them from this MO's
        # reservation contribution so MRP doesn't double-count or
        # re-suggest POs for material that's already been issued.
        consumed_for_mo = {}
        for cm in (mo.get("consumed_materials") or []):
            cid_c = cm.get("item_id")
            if cid_c:
                consumed_for_mo[cid_c] = consumed_for_mo.get(cid_c, 0) + float(cm.get("quantity") or 0)
        for rm in mo.get("reserved_materials", []):
            rid = rm.get("item_id")
            if rid:
                alloc = max(0.0, float(rm.get("allocated_qty", 0) or 0) - float(consumed_for_mo.get(rid, 0)))
                total_allocated[rid] = total_allocated.get(rid, 0) + alloc
                total_shortfall[rid] = total_shortfall.get(rid, 0) + rm.get("shortfall_qty", 0)
    
    # ---- Caches ----
    item_cache = {}
    async def get_item(item_id):
        if item_id not in item_cache:
            item_cache[item_id] = await db.items.find_one({"id": item_id}, {"_id": 0})
        return item_cache[item_id]
    
    bom_cache = {}
    async def get_active_bom(parent_item_id):
        if parent_item_id not in bom_cache:
            bom_cache[parent_item_id] = await db.boms.find_one({"parent_item_id": parent_item_id, "status": "active"}, {"_id": 0})
        return bom_cache[parent_item_id]
    
    rm_demand = {}
    
    async def explode_all_rm(parent_item_id: str, parent_qty: float, order_info: dict, variant_selection: Optional[Dict[str, str]] = None, visited: set = None, consumed_by_item: Optional[Dict[str, float]] = None):
        if visited is None:
            visited = set()
        if parent_item_id in visited:
            return
        visited.add(parent_item_id)
        
        bom = await get_active_bom(parent_item_id)
        if not bom:
            return
        
        for comp in bom.get("components", []):
            if comp.get("is_alternate"):
                continue
            comp_item_id = comp.get("item_id")
            qty_needed = comp.get("quantity", 0) * parent_qty
            
            comp_item = await get_item(comp_item_id)
            if not comp_item:
                continue

            # Variant-aware: redirect demand to the variant child SKU when the
            # MO has a matching variant_selection (same logic as WO /start).
            variant_child = await _resolve_variant_child_item(comp_item, variant_selection)
            if variant_child:
                comp_item_id = variant_child["id"]
                comp_item = variant_child
                # Refresh cache with the variant child so later lookups are consistent.
                item_cache[comp_item_id] = variant_child

            # Net off any qty already consumed by THIS MO so MRP doesn't
            # re-order what's already been issued. The consumed_by_item map
            # is only present at the top-level explode call (per-MO);
            # children inherit an empty map so each consumed entry is
            # subtracted ONCE.
            already_consumed = float((consumed_by_item or {}).get(comp_item_id, 0))
            if already_consumed > 0:
                qty_needed = max(0.0, qty_needed - already_consumed)
                if qty_needed <= 0:
                    # Skip this component entirely — already fully consumed.
                    continue
            
            if comp_item.get("category") == "raw_material":
                if comp_item_id not in rm_demand:
                    rm_demand[comp_item_id] = {
                        "item": comp_item,
                        "gross_requirement": 0,
                        "on_hand": comp_item.get("current_stock", 0),
                        "safety_stock": comp_item.get("safety_stock", 0),
                        "allocated_for_mo": total_allocated.get(comp_item_id, 0),
                        "shortfall_from_mo": total_shortfall.get(comp_item_id, 0),
                        "reserved_for_mo": total_allocated.get(comp_item_id, 0) + total_shortfall.get(comp_item_id, 0),
                        "net_requirement": 0,
                        "orders": []
                    }
                rm_demand[comp_item_id]["gross_requirement"] += qty_needed
                rm_demand[comp_item_id]["orders"].append({
                    "order_id": order_info.get("id"),
                    "order_number": order_info.get("order_number"),
                    "quantity_needed": qty_needed,
                    "due_date": order_info.get("due_date")
                })
            else:
                child_visited = set(visited)
                # Children inherit no consumed_by_item — the parent-level
                # netting was already applied above. Pass {} explicitly to
                # avoid double-subtracting.
                await explode_all_rm(comp_item_id, qty_needed, order_info, variant_selection, child_visited, consumed_by_item={})
    
    # ---- Step 2: explode UNRESERVED open MOs (MTS + MTO) ----
    # Reserved MOs are already accounted for via total_allocated / total_shortfall.
    open_mo_query = {
        "status": {"$in": ["pending", "in_progress"]},
        "materials_reserved": {"$ne": True},
        # Skip child WOs (already exploded through parent) and subcontract.
        "$or": [{"parent_wo_id": None}, {"parent_wo_id": {"$exists": False}}],
    }
    if so_filter_ids:
        open_mo_query["production_order_id"] = {"$in": list(so_filter_ids)}
    open_mos = await db.work_orders.find(open_mo_query, {"_id": 0}).to_list(5000)
    for mo in open_mos:
        item_id = mo.get("item_id")
        if not item_id:
            continue
        qty = mo.get("quantity", 0) or 0
        if qty <= 0:
            continue
        order_info = {
            "id": mo.get("id"),
            "order_number": mo.get("wo_number") or mo.get("mo_number") or mo.get("id"),
            "due_date": mo.get("scheduled_end") or mo.get("scheduled_start"),
        }
        # When an MO has already consumed materials (e.g. operation 1
        # completed and consumed the parts), those quantities must be
        # subtracted from MRP's gross demand — otherwise MRP keeps showing
        # the full BOM requirement and over-purchases. We pass a
        # `consumed_by_item` map down the explosion so the leaf RM
        # calculation can net-off what's already been issued.
        consumed_map = {}
        for cm in (mo.get("consumed_materials") or []):
            cid_c = cm.get("item_id")
            if cid_c:
                consumed_map[cid_c] = consumed_map.get(cid_c, 0) + float(cm.get("quantity") or 0)
        await explode_all_rm(item_id, qty, order_info, mo.get("variant_selection"), consumed_by_item=consumed_map)
    
    # ---- Step 3: Compute net requirement ----
    for item_id, data in rm_demand.items():
        available_for_new = max(0, data["on_hand"] - data["allocated_for_mo"] - data["safety_stock"])
        unreserved_shortfall = max(0, data["gross_requirement"] - available_for_new)
        reservation_shortfall = data["shortfall_from_mo"]
        data["net_requirement"] = reservation_shortfall + unreserved_shortfall
        
        if data["net_requirement"] > 0:
            shortage_orders = []
            running_available = max(available_for_new, 0)
            for order_entry in data["orders"]:
                qty_needed = order_entry.get("quantity_needed", 0)
                if running_available >= qty_needed:
                    running_available -= qty_needed
                else:
                    shortage_orders.append(order_entry)
                    running_available = 0
            data["orders"] = shortage_orders
    
    # ---- Step 3b: pure-reservation shortfalls (no current MO demand) ----
    for item_id, sf_qty in total_shortfall.items():
        if sf_qty > 0 and item_id not in rm_demand:
            item = await get_item(item_id)
            if item and item.get("category") == "raw_material":
                rm_demand[item_id] = {
                    "item": item,
                    "gross_requirement": 0,
                    "on_hand": item.get("current_stock", 0),
                    "safety_stock": item.get("safety_stock", 0),
                    "allocated_for_mo": total_allocated.get(item_id, 0),
                    "shortfall_from_mo": sf_qty,
                    "reserved_for_mo": total_allocated.get(item_id, 0) + sf_qty,
                    "net_requirement": sf_qty,
                    "orders": [{"order_id": "reservation", "order_number": "MO Reservation Shortfall", "quantity_needed": sf_qty}]
                }

    # Step 4: Filter and enrich with PO status
    result = []
    for d in rm_demand.values():
        item = d.get("item", {})
        if d.get("net_requirement", 0) <= 0:
            continue
        
        item_id = item.get("id")
        po_qty = 0
        if item_id:
            pos = await db.purchase_orders.find(
                {"status": {"$in": ["draft", "approved", "sent", "confirmed", "partial"]},
                 "$or": [{"lines.item_id": item_id}, {"items.item_id": item_id}]},
                {"_id": 0, "lines": 1, "items": 1, "po_number": 1}
            ).to_list(100)
            for po in pos:
                counted = False
                for pi in po.get("lines", []):
                    if pi.get("item_id") == item_id:
                        po_qty += max(0, (pi.get("quantity", 0) or 0) - (pi.get("received_quantity", 0) or 0))
                        counted = True
                if not counted:
                    for pi in po.get("items", []):
                        if pi.get("item_id") == item_id:
                            po_qty += max(0, (pi.get("quantity", 0) or 0) - (pi.get("received_quantity", 0) or 0))
        
        d["po_ordered_qty"] = int(po_qty)
        d["po_status"] = "po_sent" if po_qty >= d["net_requirement"] else ("partial_po" if po_qty > 0 else "pending")
        d["remaining_to_order"] = max(0, d["net_requirement"] - po_qty)
        
        result.append(d)
    
    return result

@mrp_router.get("/suggestions")
async def get_purchase_suggestions(request: Request):
    """Get purchase order suggestions based on reorder points and MRP"""
    await get_current_user(request)
    
    suggestions = []
    
    # Check items below reorder point
    items = await db.items.find({"$expr": {"$lte": ["$current_stock", "$reorder_point"]}}, {"_id": 0}).to_list(1000)
    
    for item in items:
        if item.get("category") != "raw_material":
            continue
        if item.get("current_stock", 0) <= item.get("reorder_point", 0):
            suggested_qty = item.get("safety_stock", 0) * 2 - item.get("current_stock", 0)
            suggestions.append({
                "item": item,
                "reason": "below_reorder_point",
                "current_stock": item.get("current_stock", 0),
                "reorder_point": item.get("reorder_point", 0),
                "safety_stock": item.get("safety_stock", 0),
                "suggested_quantity": max(suggested_qty, 1),
                "lead_time_days": item.get("lead_time_days", 0),
                "estimated_cost": max(suggested_qty, 1) * item.get("unit_cost", 0)
            })
    
    # Check MRP demand
    demand = await calculate_demand(request)
    for d in demand:
        if d.get("net_requirement", 0) > 0:
            item = d.get("item")
            if not item:
                continue
            mrp_qty = d.get("net_requirement", 0)
            # Check if already in suggestions (from reorder point)
            existing = next((s for s in suggestions if s.get("item", {}).get("id") == item.get("id")), None)
            if existing:
                # Use the higher of reorder-based qty and MRP demand qty
                if mrp_qty > existing.get("suggested_quantity", 0):
                    existing["suggested_quantity"] = mrp_qty
                    existing["net_requirement"] = mrp_qty
                    existing["reason"] = "mrp_requirement"
                    existing["estimated_cost"] = mrp_qty * item.get("unit_cost", 0)
            else:
                suggestions.append({
                    "item": item,
                    "reason": "mrp_requirement",
                    "current_stock": item.get("current_stock", 0),
                    "net_requirement": mrp_qty,
                    "suggested_quantity": mrp_qty,
                    "lead_time_days": item.get("lead_time_days", 0),
                    "estimated_cost": mrp_qty * item.get("unit_cost", 0)
                })
    
    # Add PO status to each suggestion
    for s in suggestions:
        s_item_id = s.get("item", {}).get("id")
        s_po_qty = 0
        if s_item_id:
            s_pos = await db.purchase_orders.find(
                {"status": {"$in": ["draft", "approved", "sent", "confirmed"]},
                 "$or": [{"lines.item_id": s_item_id}, {"items.item_id": s_item_id}]},
                {"_id": 0, "lines": 1, "items": 1}
            ).to_list(100)
            for po in s_pos:
                counted = False
                for pi in po.get("lines", []):
                    if pi.get("item_id") == s_item_id:
                        s_po_qty += max(0, (pi.get("quantity", 0) or 0) - (pi.get("received_quantity", 0) or 0))
                        counted = True
                if not counted:
                    for pi in po.get("items", []):
                        if pi.get("item_id") == s_item_id:
                            s_po_qty += max(0, (pi.get("quantity", 0) or 0) - (pi.get("received_quantity", 0) or 0))
        s["po_ordered_qty"] = int(s_po_qty)
        suggested = s.get("suggested_quantity", 0)
        s["po_status"] = "po_sent" if s_po_qty >= suggested else ("partial_po" if s_po_qty > 0 else "pending")
    
    return suggestions

# ================== QUALITY ROUTES ==================

@quality_router.get("/templates")
async def get_inspection_templates(request: Request, category: Optional[str] = None):
    await get_current_user(request)
    query = {}
    if category:
        query["category"] = category
    templates = await db.inspection_templates.find(query, {"_id": 0}).to_list(1000)
    return templates

@quality_router.get("/templates/{template_id}")
async def get_inspection_template(template_id: str, request: Request):
    await get_current_user(request)
    template = await db.inspection_templates.find_one({"id": template_id}, {"_id": 0})
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    return template

@quality_router.post("/templates")
async def create_inspection_template(template_data: InspectionTemplateCreate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin", "quality_inspector"], module="quality", action="create")
    template_doc = {
        "id": str(uuid.uuid4()),
        **template_data.model_dump(),
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"]
    }
    await db.inspection_templates.insert_one(template_doc)
    template_doc.pop("_id", None)
    return template_doc

@quality_router.get("/inspections")
async def get_inspections(request: Request, result: Optional[str] = None, item_id: Optional[str] = None):
    await get_current_user(request)
    query = {}
    if result:
        query["overall_result"] = result
    if item_id:
        query["item_id"] = item_id
    inspections = await db.inspections.find(query, {"_id": 0}).to_list(1000)
    
    for insp in inspections:
        item = await db.items.find_one({"id": insp.get("item_id")}, {"_id": 0})
        insp["item"] = item
        template = await db.inspection_templates.find_one({"id": insp.get("template_id")}, {"_id": 0})
        insp["template"] = template
    return inspections

@quality_router.post("/inspections")
async def create_inspection(inspection_data: InspectionRecordCreate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin", "quality_inspector"], module="quality", action="create")
    # Generate inspection number
    count = await db.inspections.count_documents({})
    inspection_number = f"INS-{str(count + 1).zfill(6)}"
    
    inspection_doc = {
        "id": str(uuid.uuid4()),
        "inspection_number": inspection_number,
        **inspection_data.model_dump(),
        "inspected_by": user["id"],
        "inspected_by_name": user.get("name"),
        "inspection_date": datetime.now(timezone.utc),
        "created_at": datetime.now(timezone.utc)
    }
    await db.inspections.insert_one(inspection_doc)
    inspection_doc.pop("_id", None)
    return inspection_doc

@quality_router.get("/metrics")
async def get_quality_metrics(request: Request, days: int = 30):
    """Get quality metrics for dashboard"""
    await get_current_user(request)
    
    start_date = datetime.now(timezone.utc) - timedelta(days=days)
    
    total = await db.inspections.count_documents({"inspection_date": {"$gte": start_date}})
    passed = await db.inspections.count_documents({"inspection_date": {"$gte": start_date}, "overall_result": "pass"})
    failed = await db.inspections.count_documents({"inspection_date": {"$gte": start_date}, "overall_result": "fail"})
    conditional = await db.inspections.count_documents({"inspection_date": {"$gte": start_date}, "overall_result": "conditional"})
    
    pass_rate = (passed / total * 100) if total > 0 else 0
    
    return {
        "total_inspections": total,
        "passed": passed,
        "failed": failed,
        "conditional": conditional,
        "pass_rate": round(pass_rate, 2),
        "period_days": days
    }

# ================== INVENTORY ROUTES ==================

@inventory_router.get("")
async def get_inventory(request: Request, category: Optional[str] = None, low_stock: bool = False):
    await get_current_user(request)
    query = {}
    if category:
        query["category"] = category
    if low_stock:
        query["$expr"] = {"$lte": ["$current_stock", "$reorder_point"]}
    
    items = await db.items.find(query, {"_id": 0}).to_list(50000)
    return items


@inventory_router.put("/items/{item_id}/stock-fields")
async def update_stock_fields(item_id: str, request: Request, data: dict = Body(default={})):
    """Update stock-relevant + (optionally) master fields of an item.
    Permission tiers:
      - STOCK fields (current_stock, safety_stock, reorder_point, lead_time_days, unit_cost)
        accept items.* OR inventory.* (semantic overlap: thresholds are inventory).
      - MASTER fields (name, group_id, hsn_code, gst_rate, purchase_price, sale_price)
        require items.edit / items.create (or admin). If a non-eligible user sends
        them, those keys are silently dropped — stock fields still apply.
      For raw_material category, purchase_price changes auto-sync unit_cost (mirrors
      the full /items form's behaviour) so BOM rollups stay consistent."""
    user = await get_current_user(request)
    perms = (user.get("permissions") or {})
    is_admin = user.get("role") == "admin"
    items_perms = perms.get("items") or []
    inv_perms = perms.get("inventory") or []
    can_edit_stock = is_admin \
        or "edit" in items_perms or "create" in items_perms \
        or "edit" in inv_perms or "create" in inv_perms
    can_edit_master = is_admin \
        or "edit" in items_perms or "create" in items_perms
    if not can_edit_stock:
        raise HTTPException(status_code=403, detail="Not authorized to edit stock fields")
    
    item = await db.items.find_one({"id": item_id})
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    
    # Stock-tier whitelist (allowed if can_edit_stock)
    stock_fields = ["safety_stock", "reorder_point", "lead_time_days", "unit_cost", "current_stock"]
    # Master-tier whitelist (only applied if can_edit_master)
    master_fields = ["name", "group_id", "hsn_code", "gst_rate", "purchase_price", "sale_price"]
    update = {}
    for f in stock_fields:
        if f in data and data[f] is not None:
            try:
                update[f] = float(data[f]) if f in ("unit_cost", "safety_stock", "reorder_point", "current_stock") else int(data[f])
            except (TypeError, ValueError):
                continue
    if can_edit_master:
        for f in master_fields:
            if f in data and data[f] is not None:
                v = data[f]
                if f in ("gst_rate", "purchase_price", "sale_price"):
                    try: v = float(v)
                    except (TypeError, ValueError): continue
                # Validate name not empty
                if f == "name" and not str(v).strip():
                    continue
                update[f] = v
        # Keep unit_cost in sync with purchase_price for raw materials —
        # mirrors the behaviour in the full Items page form.
        if "purchase_price" in update and item.get("category") == "raw_material":
            update["unit_cost"] = float(update["purchase_price"])
    
    if not update:
        raise HTTPException(status_code=400, detail="No valid stock fields provided")
    
    update["updated_at"] = datetime.now(timezone.utc)
    
    # If current_stock is changed, also log an inventory transaction so the
    # audit trail is preserved.
    new_current = update.get("current_stock")
    if new_current is not None and float(new_current) != float(item.get("current_stock", 0) or 0):
        delta = float(new_current) - float(item.get("current_stock", 0) or 0)
        tx_doc = {
            "id": str(uuid.uuid4()),
            "item_id": item_id,
            "transaction_type": "adjust",
            "quantity": delta,
            "reference_type": "stock_edit",
            "reference_id": item_id,
            "previous_stock": float(item.get("current_stock", 0) or 0),
            "new_stock": float(new_current),
            "notes": f"Stock adjusted via Inventory edit dialog by {user.get('email', user.get('id'))}",
            "created_at": datetime.now(timezone.utc),
            "created_by": user["id"],
        }
        await db.inventory_transactions.insert_one(tx_doc)
    
    await db.items.update_one({"id": item_id}, {"$set": update})
    
    updated = await db.items.find_one({"id": item_id}, {"_id": 0})
    return updated


@inventory_router.get("/transactions")
async def get_inventory_transactions(
    request: Request,
    item_id: Optional[str] = None,
    transaction_type: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    limit: int = 100,
):
    await get_current_user(request)
    query = {}
    if item_id:
        query["item_id"] = item_id
    if transaction_type and transaction_type != "all":
        query["transaction_type"] = transaction_type
    # Date range filter — inclusive on both ends. Accepts YYYY-MM-DD.
    if from_date or to_date:
        date_q = {}
        try:
            if from_date:
                date_q["$gte"] = datetime.strptime(from_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            if to_date:
                # Use end-of-day so the same date appears on both sides of the filter.
                end_dt = datetime.strptime(to_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
                end_dt = end_dt + timedelta(days=1) - timedelta(milliseconds=1)
                date_q["$lte"] = end_dt
        except ValueError:
            # Invalid date string — skip the filter rather than 500.
            date_q = {}
        if date_q:
            query["created_at"] = date_q

    transactions = await db.inventory_transactions.find(query, {"_id": 0}).sort("created_at", -1).to_list(limit)

    for tx in transactions:
        item = await db.items.find_one({"id": tx.get("item_id")}, {"_id": 0})
        tx["item"] = item
    return transactions

@inventory_router.post("/reconcile-reservations")
async def reconcile_reservations(request: Request):
    """Admin tool: scan every item's `reserved_stock` and recompute it from the
    canonical source-of-truth — active SO line reservations + active MO child reservations.
    
    Reasons drift can happen:
      • A cancellation path missed releasing some reservation (e.g. older partial fix)
      • Direct DB edits
      • Bug in a previous code revision
    
    This endpoint is idempotent and safe to run repeatedly.
    """
    user = await get_current_user(request)
    _require_access(user, ["admin"], module="inventory", action="edit")
    
    # Build the canonical map: item_id -> expected reserved_stock.
    expected = {}
    
    # Source 1: SO lines that are reserved (status != cancelled)
    so_cur = db.production_orders.find(
        {"status": {"$nin": ["cancelled"]}},
        {"_id": 0, "lines": 1, "id": 1}
    )
    async for so in so_cur:
        for ln in so.get("lines") or []:
            if not ln.get("is_reserved") and int(ln.get("reserved_qty", 0) or 0) <= 0:
                continue
            bom = await db.boms.find_one({"id": ln.get("bom_id")}, {"_id": 0, "parent_item_id": 1})
            if not bom or not bom.get("parent_item_id"):
                continue
            qty = int(ln.get("reserved_qty", 0) or 0)
            if qty <= 0:
                continue
            expected[bom["parent_item_id"]] = expected.get(bom["parent_item_id"], 0) + qty
    
    # Source 2: MO child_reservations (status not cancelled / completed)
    mo_cur = db.work_orders.find(
        {"status": {"$nin": ["cancelled", "completed"]}, "child_reservations": {"$exists": True, "$ne": []}},
        {"_id": 0, "child_reservations": 1, "id": 1}
    )
    async for mo in mo_cur:
        for resv in (mo.get("child_reservations") or []):
            iid = resv.get("item_id")
            qty = int(resv.get("qty", 0) or 0)
            if iid and qty > 0:
                expected[iid] = expected.get(iid, 0) + qty
    
    # Now compare against every item and reset drift.
    drift_records = []
    items_cur = db.items.find({}, {"_id": 0, "id": 1, "part_number": 1, "reserved_stock": 1, "current_stock": 1})
    async for it in items_cur:
        current = int(it.get("reserved_stock", 0) or 0)
        canonical = expected.get(it["id"], 0)
        if current != canonical:
            await db.items.update_one(
                {"id": it["id"]},
                {"$set": {"reserved_stock": canonical}},
            )
            drift_records.append({
                "item_id": it["id"],
                "part_number": it.get("part_number"),
                "before": current,
                "after": canonical,
                "delta": canonical - current,
            })
    return {
        "message": f"Reconciled {len(drift_records)} item(s)" if drift_records else "All items already in sync — no drift found.",
        "drift_count": len(drift_records),
        "drift": drift_records,
    }


@inventory_router.post("/transactions")
async def create_inventory_transaction(tx_data: InventoryTransactionCreate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin", "inventory_manager", "production_manager"], module="inventory", action="create")
    # Warehouse is mandatory for stock-changing transactions
    if tx_data.transaction_type in ["receive", "issue", "adjust"] and not tx_data.warehouse_id:
        raise HTTPException(status_code=400, detail="Warehouse is required for stock-changing transactions")
    
    item = await db.items.find_one({"id": tx_data.item_id})
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    
    warehouse = None
    if tx_data.warehouse_id:
        warehouse = await db.warehouses.find_one({"id": tx_data.warehouse_id}, {"_id": 0})
        if not warehouse:
            raise HTTPException(status_code=404, detail="Warehouse not found")
    
    # Calculate new stock
    current_stock = item.get("current_stock", 0)
    if tx_data.transaction_type == "receive":
        new_stock = current_stock + tx_data.quantity
    elif tx_data.transaction_type == "issue":
        new_stock = current_stock - tx_data.quantity
        if new_stock < 0:
            raise HTTPException(status_code=400, detail="Insufficient stock")
    elif tx_data.transaction_type == "adjust":
        new_stock = tx_data.quantity  # Direct adjustment
    else:
        new_stock = current_stock
    
    # Create transaction record
    tx_doc = {
        "id": str(uuid.uuid4()),
        **tx_data.model_dump(),
        "previous_stock": current_stock,
        "new_stock": new_stock,
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"]
    }
    await db.inventory_transactions.insert_one(tx_doc)
    
    # Update item stock
    await db.items.update_one({"id": tx_data.item_id}, {"$set": {"current_stock": new_stock}})
    
    # Update warehouse-specific stock
    if tx_data.warehouse_id:
        wh_stock = await db.warehouse_stock.find_one({"warehouse_id": tx_data.warehouse_id, "item_id": tx_data.item_id})
        wh_current = wh_stock.get("quantity", 0) if wh_stock else 0
        if tx_data.transaction_type == "receive":
            wh_new = wh_current + tx_data.quantity
        elif tx_data.transaction_type == "issue":
            wh_new = wh_current - tx_data.quantity
            if wh_new < 0:
                raise HTTPException(status_code=400, detail=f"Insufficient stock in warehouse {warehouse.get('name') if warehouse else ''}")
        elif tx_data.transaction_type == "adjust":
            wh_new = tx_data.quantity
        else:
            wh_new = wh_current
        await db.warehouse_stock.update_one(
            {"warehouse_id": tx_data.warehouse_id, "item_id": tx_data.item_id},
            {"$set": {"quantity": wh_new, "updated_at": datetime.now(timezone.utc)}},
            upsert=True
        )
    
    tx_doc.pop("_id", None)
    return tx_doc

# ================== USERS ROUTES (Admin only) ==================

@users_router.get("")
async def get_users(request: Request):
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    
    users_list = []
    async for u in db.users.find({}, {"password_hash": 0}):
        u["id"] = str(u["_id"])
        del u["_id"]
        if "permissions" not in u or not u["permissions"]:
            u["permissions"] = get_default_permissions(u.get("role", "inventory_manager"))
        users_list.append(u)
    return users_list


@users_router.get("/assignable")
async def get_assignable_users(request: Request):
    """Light user list (id/name/email) usable by ANY authenticated user — meant
    for assignment dropdowns (e.g. Support tickets, Lead owner) so non-admins
    can also delegate work without exposing the full user-management payload."""
    await get_current_user(request)
    users_list = []
    async for u in db.users.find({}, {"_id": 1, "name": 1, "email": 1}):
        users_list.append({
            "id": str(u["_id"]),
            "name": u.get("name", ""),
            "email": u.get("email", ""),
        })
    return users_list

@users_router.post("", status_code=201)
async def create_user(user_data: UserCreate, request: Request):
    admin = await get_current_user(request)
    if admin["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can create users")
    
    email = user_data.email.lower()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")

    # Permissions are now defined ONLY at the role-group level. The user-level
    # `role` field is auto-derived from the group (admin-group → "admin",
    # otherwise we default to "inventory_manager" purely so legacy
    # role-checking code keeps working). Per-user permission overrides are no
    # longer accepted from the client.
    derived_role = "inventory_manager"
    if user_data.role_group_id:
        rg = await db.role_groups.find_one({"id": user_data.role_group_id}, {"_id": 0, "is_admin_group": 1})
        if rg and rg.get("is_admin_group"):
            derived_role = "admin"

    user_doc = {
        "email": email,
        "password_hash": hash_password(user_data.password),
        "name": user_data.name,
        "role": derived_role,
        # Empty per-user permissions — `get_current_user` will overlay the
        # role-group's permissions on every request.
        "permissions": {},
        "role_group_id": user_data.role_group_id,
        "status": "active",
        "created_at": datetime.now(timezone.utc),
        "created_by": admin["id"]
    }
    result = await db.users.insert_one(user_doc)
    user_doc["id"] = str(result.inserted_id)
    user_doc.pop("_id", None)
    user_doc.pop("password_hash", None)
    return user_doc

@users_router.put("/{user_id}")
async def update_user(user_id: str, data: UserUpdate, request: Request):
    admin = await get_current_user(request)
    is_self = (admin.get("id") == user_id)
    if admin["role"] != "admin" and not is_self:
        raise HTTPException(status_code=403, detail="Only admin can update other users")
    
    update_data = {}
    # Non-admin self-update is restricted to signature_url + password + name
    if admin["role"] == "admin":
        if data.email is not None:
            new_email = data.email.lower().strip()
            if new_email:
                # Enforce unique email
                clash = await db.users.find_one({"email": new_email, "id": {"$ne": user_id}})
                if clash:
                    raise HTTPException(status_code=400, detail=f"Email '{new_email}' is already in use by another user")
                update_data["email"] = new_email
        if data.name is not None:
            update_data["name"] = data.name
        # Per-user `role` field is auto-derived from the role group below.
        # Ignore any explicit role sent from the client.
        if data.permissions is not None:
            # Per-user permission overrides are deprecated. Empty out the field
            # so role-group permissions become the single source of truth.
            update_data["permissions"] = {}
        if data.status is not None:
            update_data["status"] = data.status
        if data.role_group_id is not None:
            new_group_id = data.role_group_id or None
            update_data["role_group_id"] = new_group_id
            # Re-derive role from the new group
            derived_role = "inventory_manager"
            if new_group_id:
                rg = await db.role_groups.find_one({"id": new_group_id}, {"_id": 0, "is_admin_group": 1})
                if rg and rg.get("is_admin_group"):
                    derived_role = "admin"
            update_data["role"] = derived_role
        if data.assigned_customer_ids is not None:
            update_data["assigned_customer_ids"] = data.assigned_customer_ids
    else:
        if data.name is not None:
            update_data["name"] = data.name
    if data.password is not None and data.password.strip():
        update_data["password_hash"] = hash_password(data.password)
    if data.signature_url is not None:
        update_data["signature_url"] = data.signature_url
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    # Try matching by ObjectId
    try:
        result = await db.users.update_one({"_id": ObjectId(user_id)}, {"$set": update_data})
    except:
        result = await db.users.update_one({"id": user_id}, {"$set": update_data})
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Return updated user
    try:
        updated = await db.users.find_one({"_id": ObjectId(user_id)}, {"password_hash": 0})
    except:
        updated = await db.users.find_one({"id": user_id}, {"password_hash": 0})
    
    if updated:
        updated["id"] = str(updated["_id"])
        del updated["_id"]
        return updated
    return {"message": "User updated"}

@users_router.delete("/{user_id}")
async def delete_user(user_id: str, request: Request):
    admin = await get_current_user(request)
    if admin["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can delete users")
    
    # Prevent self-deletion
    if admin["id"] == user_id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    try:
        result = await db.users.delete_one({"_id": ObjectId(user_id)})
    except:
        result = await db.users.delete_one({"id": user_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": "User deleted"}

@users_router.get("/modules")
async def get_modules(request: Request):
    """Get list of all modules and actions for permission UI"""
    await get_current_user(request)
    return {
        "modules": ALL_MODULES,
        "actions": ALL_ACTIONS,
        "module_actions": MODULE_ACTIONS,
        "default_permissions": DEFAULT_PERMISSIONS
    }

# ================== ROLE GROUPS ==================
# Admin can create named role groups (e.g. "Production Admin", "Purchase User") with
# custom permission matrices. Users are then mapped to a group and inherit its permissions.
# Only groups flagged `is_admin_group=true` can see BOM rollup costs.

class RoleGroupCreate(BaseModel):
    name: str
    description: Optional[str] = ""
    permissions: dict = {}
    is_admin_group: bool = False
    view_all_parties: bool = False  # When True, members can see ALL customers + suppliers (not just their own/assigned)

class RoleGroupUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    permissions: Optional[dict] = None
    is_admin_group: Optional[bool] = None
    view_all_parties: Optional[bool] = None

@users_router.get("/role-groups")
async def list_role_groups(request: Request):
    await get_current_user(request)
    groups = await db.role_groups.find({}, {"_id": 0}).sort("name", 1).to_list(500)
    return groups

@users_router.post("/role-groups", status_code=201)
async def create_role_group(data: RoleGroupCreate, request: Request):
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can create role groups")
    existing = await db.role_groups.find_one({"name": data.name})
    if existing:
        raise HTTPException(status_code=400, detail=f"Role group '{data.name}' already exists")
    doc = {
        "id": str(uuid.uuid4()),
        "name": data.name,
        "description": data.description or "",
        "permissions": data.permissions or {},
        "is_admin_group": bool(data.is_admin_group),
        "view_all_parties": bool(data.view_all_parties),
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"]
    }
    await db.role_groups.insert_one(doc)
    doc.pop("_id", None)
    return doc

@users_router.put("/role-groups/{group_id}")
async def update_role_group(group_id: str, data: RoleGroupUpdate, request: Request):
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can update role groups")
    update = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update:
        raise HTTPException(status_code=400, detail="No data to update")
    update["updated_at"] = datetime.now(timezone.utc)
    result = await db.role_groups.update_one({"id": group_id}, {"$set": update})
    if not result.matched_count:
        raise HTTPException(status_code=404, detail="Role group not found")
    return await db.role_groups.find_one({"id": group_id}, {"_id": 0})

@users_router.delete("/role-groups/{group_id}")
async def delete_role_group(group_id: str, request: Request):
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can delete role groups")
    # Unassign users from this group before delete
    await db.users.update_many({"role_group_id": group_id}, {"$unset": {"role_group_id": ""}})
    result = await db.role_groups.delete_one({"id": group_id})
    if not result.deleted_count:
        raise HTTPException(status_code=404, detail="Role group not found")
    return {"message": "Role group deleted"}

# ================== DASHBOARD ROUTES ==================

@dashboard_router.get("/stats")
async def get_dashboard_stats(request: Request):
    await get_current_user(request)
    
    # Item counts by category
    items_count = await db.items.count_documents({})
    raw_materials = await db.items.count_documents({"category": "raw_material"})
    components = await db.items.count_documents({"category": "component"})
    finished_goods = await db.items.count_documents({"category": "finished_good"})
    
    # Low stock items
    low_stock = await db.items.count_documents({"$expr": {"$lte": ["$current_stock", "$reorder_point"]}})
    
    # BOM counts
    active_boms = await db.boms.count_documents({"status": "active"})
    
    # Production orders
    pending_orders = await db.production_orders.count_documents({"status": {"$in": ["draft", "confirmed", "planned", "released"]}})
    in_progress = await db.production_orders.count_documents({"status": "in_progress"})
    
    # Quality metrics
    quality_metrics = await get_quality_metrics(request, days=30)
    
    return {
        "inventory": {
            "total_items": items_count,
            "raw_materials": raw_materials,
            "components": components,
            "finished_goods": finished_goods,
            "low_stock_alerts": low_stock
        },
        "bom": {
            "active_boms": active_boms
        },
        "production": {
            "pending_orders": pending_orders,
            "in_progress": in_progress
        },
        "quality": quality_metrics
    }

# ================== SUPPLIER ROUTES ==================

@suppliers_router.get("")
async def get_suppliers(request: Request, status: Optional[str] = None, mine: Optional[bool] = False):
    user = await get_current_user(request)
    query = {}
    if status:
        query["status"] = status
    # Same per-user contact ownership rule as customers:
    #   - Admins / admin-group / view_all_parties group → see ALL suppliers.
    #     `mine=true` narrows to only ones they personally created.
    #   - Others → only suppliers they created OR are listed in
    #     `supplier.assigned_user_ids`.
    is_admin = user.get("role") == "admin"
    can_view_all = False
    if not is_admin and user.get("role_group_id"):
        rg = await db.role_groups.find_one({"id": user["role_group_id"]}, {"_id": 0, "is_admin_group": 1, "view_all_parties": 1})
        if rg and rg.get("is_admin_group"):
            is_admin = True
        if rg and rg.get("view_all_parties"):
            can_view_all = True
    if is_admin or can_view_all:
        if mine:
            query["created_by"] = user["id"]
    else:
        query["$or"] = [
            {"created_by": user["id"]},
            {"assigned_user_ids": user["id"]},
        ]
    suppliers = await db.suppliers.find(query, {"_id": 0}).to_list(2000)
    return suppliers

@suppliers_router.get("/{supplier_id}")
async def get_supplier(supplier_id: str, request: Request):
    await get_current_user(request)
    supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    return supplier

@suppliers_router.post("", status_code=201)
async def create_supplier(supplier_data: SupplierCreate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="suppliers", action="create")
    # Mandatory GST identity fields for correct CGST/SGST/IGST split on POs/Invoices.
    if not (supplier_data.state_code or "").strip():
        raise HTTPException(status_code=400, detail="State code is required (needed for GST CGST/SGST/IGST logic).")
    pin = (supplier_data.pin_code or "").strip()
    if not pin or not pin.isdigit() or len(pin) != 6:
        raise HTTPException(status_code=400, detail="PIN Code is required and must be a 6-digit number.")
    # Auto-generate supplier code from configurable series if not provided
    provided_code = (supplier_data.code or "").strip()
    if not provided_code:
        supplier_code = await get_next_series_number("supplier_code")
    else:
        supplier_code = provided_code
    existing = await db.suppliers.find_one({"code": supplier_code})
    if existing:
        raise HTTPException(status_code=400, detail="Supplier code already exists")

    # GST duplicate check — block creating a second supplier with the same
    # GSTIN. GST numbers are statutory unique identifiers, so accidentally
    # creating "ABC Industries" and "ABC Ind Pvt Ltd" with the same 27ABCDE...
    # produces wrong PI/PO trails and breaks GSTR-2 reconciliation.
    gstin_clean = (supplier_data.gstin or "").strip().upper()
    if gstin_clean:
        gst_dup = await db.suppliers.find_one({"gstin": gstin_clean}, {"name": 1, "code": 1})
        if gst_dup:
            raise HTTPException(
                status_code=400,
                detail=f"A supplier with GSTIN {gstin_clean} already exists ({gst_dup.get('name','?')} · code {gst_dup.get('code','?')}). Please use the existing record."
            )
    
    supplier_doc = {
        "id": str(uuid.uuid4()),
        **supplier_data.model_dump(),
        "code": supplier_code,
        "gstin": gstin_clean or supplier_data.gstin,
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"]
    }
    await db.suppliers.insert_one(supplier_doc)
    supplier_doc.pop("_id", None)
    return supplier_doc

@suppliers_router.put("/{supplier_id}")
async def update_supplier(supplier_id: str, supplier_data: SupplierUpdate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="suppliers", action="edit")
    # If the update payload sets these fields, validate them (None means "don't change").
    if supplier_data.state_code is not None and not supplier_data.state_code.strip():
        raise HTTPException(status_code=400, detail="State code cannot be blank.")
    if supplier_data.pin_code is not None:
        pin = (supplier_data.pin_code or "").strip()
        if not pin or not pin.isdigit() or len(pin) != 6:
            raise HTTPException(status_code=400, detail="PIN Code must be a 6-digit number.")
    update_data = {k: v for k, v in supplier_data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    # Same GST uniqueness rule as create — but exclude the supplier being
    # edited so the user can save a record without changing the GSTIN.
    if "gstin" in update_data:
        gstin_clean = (update_data["gstin"] or "").strip().upper()
        if gstin_clean:
            gst_dup = await db.suppliers.find_one(
                {"gstin": gstin_clean, "id": {"$ne": supplier_id}},
                {"name": 1, "code": 1},
            )
            if gst_dup:
                raise HTTPException(
                    status_code=400,
                    detail=f"A supplier with GSTIN {gstin_clean} already exists ({gst_dup.get('name','?')} · code {gst_dup.get('code','?')}). GSTIN must be unique."
                )
            update_data["gstin"] = gstin_clean
    
    update_data["updated_at"] = datetime.now(timezone.utc)
    result = await db.suppliers.update_one({"id": supplier_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    return supplier

@suppliers_router.delete("/{supplier_id}")
async def delete_supplier(supplier_id: str, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin"], module="suppliers", action="delete")
    result = await db.suppliers.delete_one({"id": supplier_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    return {"message": "Supplier deleted"}


# ---------- GSTIN Lookup (Appyflow) ----------
class GSTINLookupRequest(BaseModel):
    gstin: str


def _get_appyflow_key(company_settings: dict) -> str:
    """Prefer Settings override, fall back to backend .env."""
    if company_settings and company_settings.get("appyflow_api_key"):
        return str(company_settings["appyflow_api_key"]).strip()
    return (os.environ.get("APPYFLOW_API_KEY") or "").strip()


@suppliers_router.post("/lookup-gstin")
async def lookup_gstin(payload: GSTINLookupRequest, request: Request):
    """
    Look up GSTIN details via Appyflow API and return a normalized dict
    the Supplier form can use to pre-fill legal name, trade name, state, PIN,
    principal address and registration status.

    Key resolution order:
      1) `company_settings.appyflow_api_key` (editable in Settings UI)
      2) `APPYFLOW_API_KEY` env var (backend/.env)
    """
    await get_current_user(request)
    gstin = (payload.gstin or "").strip().upper()
    if len(gstin) != 15 or not gstin.isalnum():
        raise HTTPException(status_code=400, detail="GSTIN must be exactly 15 alphanumeric characters")

    company = await db.company_settings.find_one({"type": "company"}, {"_id": 0})
    key = _get_appyflow_key(company or {})
    if not key:
        raise HTTPException(
            status_code=503,
            detail="GSTIN lookup is not configured. Set 'Appyflow API Key' in Settings → Integrations "
                   "or APPYFLOW_API_KEY in backend .env."
        )

    import httpx
    url = "https://appyflow.in/api/verifyGST"
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            resp = await client.get(url, params={"gstNo": gstin, "key_secret": key},
                                    headers={"Accept": "application/json"})
    except httpx.TimeoutException:
        raise HTTPException(status_code=504, detail="GSTIN lookup timed out")
    except httpx.RequestError as e:
        raise HTTPException(status_code=502, detail=f"Unable to reach Appyflow: {e}")

    if resp.status_code != 200:
        raise HTTPException(status_code=502, detail=f"Appyflow error: HTTP {resp.status_code}")

    try:
        data = resp.json()
    except Exception:
        raise HTTPException(status_code=502, detail="Invalid JSON from Appyflow")

    if data.get("error") is True:
        msg = str(data.get("message", "GSTIN lookup failed"))
        # Map common errors to precise status codes.
        m = msg.lower()
        if "invalid" in m:
            raise HTTPException(status_code=400, detail=f"Invalid GSTIN: {msg}")
        if "not found" in m or "not exist" in m:
            raise HTTPException(status_code=404, detail="GSTIN not found in GST database")
        raise HTTPException(status_code=422, detail=msg)

    tp = data.get("taxpayerInfo") or {}
    if not tp:
        raise HTTPException(status_code=404, detail="GSTIN not found (empty response)")

    # Detect Appyflow "free tier / sandbox" responses. On the free plan, Appyflow
    # ALWAYS returns the same demo record (DISHANT MAHAJAN / AppyFlow Technologies)
    # and attaches a notice in `message`. Surface this to the caller so the UI
    # can show a clear banner instead of silently trusting the stub data.
    appyflow_notice = str(data.get("message") or "")
    sandbox_mode = False
    if appyflow_notice:
        m = appyflow_notice.lower()
        if "free credits" in m or "paid credits" in m or "sandbox" in m:
            sandbox_mode = True

    pradr = tp.get("pradr") or {}
    addr = pradr.get("addr") or {}
    status_raw = str(tp.get("sts") or "").upper()
    normalized = {
        "gstin": gstin,
        "legal_name": tp.get("lgnm") or "",
        "trade_name": tp.get("tradeNam") or tp.get("tradeName") or "",
        "status": "active" if status_raw not in {"CANCELLED", "INACTIVE", "SUSPENDED"} else status_raw.lower(),
        "registration_date": tp.get("rgdt") or "",
        "taxpayer_type": tp.get("dty") or tp.get("dpty") or "",
        "constitution": tp.get("ctb") or "",
        "state_jurisdiction": tp.get("stj") or "",
        "principal_address": {
            "building": addr.get("bno") or "",
            "street": addr.get("st") or "",
            "locality": addr.get("loc") or "",
            "city": addr.get("dst") or addr.get("city") or "",
            "state_name": addr.get("stcd") or "",  # "Maharashtra" etc — caller should map to code
            "pin_code": addr.get("pncd") or "",
            "full": pradr.get("adr") or "",
        },
        # State code from GSTIN (first 2 digits) — authoritative for CGST/SGST logic.
        "state_code_from_gstin": gstin[:2],
        # Provider diagnostics — let the UI warn the user.
        "provider": "appyflow",
        "provider_message": appyflow_notice,
        "sandbox_mode": sandbox_mode,
    }
    return normalized

# ================== PURCHASE ORDER ROUTES ==================

@purchase_orders_router.get("")
async def get_purchase_orders(request: Request, status: Optional[str] = None, supplier_id: Optional[str] = None):
    await get_current_user(request)
    query = {}
    if status:
        query["status"] = status
    if supplier_id:
        query["supplier_id"] = supplier_id
    
    orders = await db.purchase_orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    for order in orders:
        supplier = await db.suppliers.find_one({"id": order.get("supplier_id")}, {"_id": 0})
        order["supplier"] = supplier
        # Enrich lines with item details and compute subtotal
        subtotal = 0
        for line in order.get("lines", []):
            item = await db.items.find_one({"id": line.get("item_id")}, {"_id": 0})
            line["item"] = item
            subtotal += line.get("total_price", 0) or (line.get("quantity", 0) * line.get("unit_price", 0))
        if not order.get("subtotal"):
            order["subtotal"] = subtotal
    return orders

@purchase_orders_router.get("/{po_id}")
async def get_purchase_order(po_id: str, request: Request):
    await get_current_user(request)
    order = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Purchase order not found")
    
    supplier = await db.suppliers.find_one({"id": order.get("supplier_id")}, {"_id": 0})
    order["supplier"] = supplier
    for line in order.get("lines", []):
        item = await db.items.find_one({"id": line.get("item_id")}, {"_id": 0})
        line["item"] = item
    return order

@purchase_orders_router.get("/{po_id}/print-data")
async def get_po_print_data(po_id: str, request: Request):
    """Get PO data with company settings for printing."""
    await get_current_user(request)
    order = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Purchase order not found")
    supplier = await db.suppliers.find_one({"id": order.get("supplier_id")}, {"_id": 0})
    order["supplier"] = supplier
    for line in order.get("lines", []):
        item = await db.items.find_one({"id": line.get("item_id")}, {"_id": 0})
        line["item"] = item
    company = await db.company_settings.find_one({"type": "company"}, {"_id": 0})
    order["company"] = company
    # Attach creator's name + signature for print (not current logged-in user).
    order["created_by_user"] = await _lookup_creator(order.get("created_by"))
    # Get delivery warehouse details
    if order.get("delivery_warehouse_id"):
        wh = await db.warehouses.find_one({"id": order["delivery_warehouse_id"]}, {"_id": 0})
        order["delivery_warehouse"] = wh

    # Recompute is_inter_state + tax split from CURRENT supplier/company state
    # (stored totals can go stale if state_code is later corrected on either party).
    company_state = (company or {}).get("state_code", "")
    supplier_state = (supplier or {}).get("state_code", "")
    is_inter_state = bool(company_state and supplier_state and company_state != supplier_state)
    order["is_inter_state"] = is_inter_state

    # Rebuild tax totals from line items for a consistent print.
    subtotal = 0.0
    total_cgst = 0.0
    total_sgst = 0.0
    total_igst = 0.0
    for ln in order.get("lines", []):
        qty = float(ln.get("quantity", 0) or 0)
        price = float(ln.get("unit_price", 0) or 0)
        gross = qty * price
        disc_amt = float(ln.get("discount_amount", 0) or 0)
        if not disc_amt and ln.get("discount_value"):
            dv = float(ln.get("discount_value", 0) or 0)
            disc_amt = gross * dv / 100 if ln.get("discount_type") == "percentage" else dv
        net = gross - disc_amt
        gst_rate = float(ln.get("gst_rate", 0) or 0)
        tax = round(net * gst_rate / 100, 2)
        if is_inter_state:
            ln["igst_amount"] = tax
            ln["cgst_amount"] = 0
            ln["sgst_amount"] = 0
            total_igst += tax
        else:
            half = round(tax / 2, 2)
            ln["igst_amount"] = 0
            ln["cgst_amount"] = half
            ln["sgst_amount"] = tax - half  # absorb rounding into SGST
            total_cgst += half
            total_sgst += (tax - half)
        subtotal += net
    order["subtotal"] = round(subtotal, 2)
    order["total_cgst"] = round(total_cgst, 2)
    order["total_sgst"] = round(total_sgst, 2)
    order["total_igst"] = round(total_igst, 2)
    order["total_tax"] = round(total_cgst + total_sgst + total_igst, 2)
    # Respect any additional charges already on the order (don't recompute here).
    charges_total = sum(float(c.get("amount", 0) or 0) + float(c.get("tax_amount", 0) or 0)
                        for c in order.get("additional_charges", []) or [])
    order["total_amount"] = round(subtotal + total_cgst + total_sgst + total_igst + charges_total, 2)

    # Inject Default PO Terms & Conditions (from Inventory → Configuration) so
    # every printed PO carries the terms unless the PO itself already has custom terms.
    if not order.get("terms_conditions") and company and company.get("po_terms_conditions"):
        order["terms_conditions"] = company.get("po_terms_conditions")

    return order

@purchase_orders_router.post("", status_code=201)
async def create_purchase_order(po_data: PurchaseOrderCreate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager", "inventory_manager"], module="purchase_orders", action="create")
    supplier = await db.suppliers.find_one({"id": po_data.supplier_id})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    # Generate PO number from configurable series
    po_number = await get_next_series_number("po_number")
    
    # Currency: INR (default) keeps GST. Non-INR (export/import) → no GST.
    po_currency = (po_data.currency or "INR").upper()
    is_export = po_currency != "INR"

    # Get company settings for GST calculation
    company_settings = await db.company_settings.find_one({"type": "company"}, {"_id": 0})
    company_state = company_settings.get("state_code", "") if company_settings else ""
    supplier_state = supplier.get("state_code", "")
    is_inter_state = (not is_export) and bool(company_state) and bool(supplier_state) and company_state != supplier_state
    
    # Calculate line totals with discount and GST
    lines_with_tax = []
    subtotal = 0
    total_cgst = 0
    total_sgst = 0
    total_igst = 0
    
    for line in po_data.lines:
        line_data = line.model_dump()
        gross_amount = line.quantity * line.unit_price
        
        # Calculate discount
        discount_amount = 0
        if line.discount_value and line.discount_value > 0:
            if line.discount_type == "percentage":
                discount_amount = round(gross_amount * line.discount_value / 100, 2)
            else:
                discount_amount = round(line.discount_value, 2)
        line_data["discount_amount"] = discount_amount
        line_amount = gross_amount - discount_amount
        
        # Auto-fill UOM/HSN from item if not provided
        item_doc = await db.items.find_one({"id": line.item_id}, {"_id": 0})
        if item_doc:
            if not line_data.get("hsn_code"):
                line_data["hsn_code"] = item_doc.get("hsn_code", "")
            if not line_data.get("uom") or line_data["uom"] == "pcs":
                line_data["uom"] = item_doc.get("unit_of_measure", "pcs")
            if not line_data.get("description"):
                line_data["description"] = item_doc.get("description", "")
        
        gst_rate = 0 if is_export else (line.gst_rate or 0)
        tax_amount = round(line_amount * gst_rate / 100, 2)
        
        if is_inter_state:
            line_data["igst_amount"] = tax_amount
            line_data["cgst_amount"] = 0
            line_data["sgst_amount"] = 0
            total_igst += tax_amount
        else:
            half_tax = round(tax_amount / 2, 2)
            line_data["cgst_amount"] = half_tax
            line_data["sgst_amount"] = half_tax
            line_data["igst_amount"] = 0
            total_cgst += half_tax
            total_sgst += half_tax
        
        line_data["gross_amount"] = round(gross_amount, 2)
        line_data["line_amount"] = round(line_amount, 2)
        line_data["tax_amount"] = tax_amount
        subtotal += line_amount
        lines_with_tax.append(line_data)
    
    # Process additional charges with GST
    charges_with_tax = []
    charges_subtotal = 0
    for charge in (po_data.additional_charges or []):
        c_data = charge.model_dump()
        c_amount = charge.amount
        c_gst_rate = 0 if is_export else (charge.gst_rate or 0)
        c_tax = round(c_amount * c_gst_rate / 100, 2)
        
        if is_inter_state:
            c_data["igst_amount"] = c_tax
            c_data["cgst_amount"] = 0
            c_data["sgst_amount"] = 0
            total_igst += c_tax
        else:
            c_half = round(c_tax / 2, 2)
            c_data["cgst_amount"] = c_half
            c_data["sgst_amount"] = c_half
            c_data["igst_amount"] = 0
            total_cgst += c_half
            total_sgst += c_half
        
        c_data["tax_amount"] = c_tax
        charges_subtotal += c_amount
        charges_with_tax.append(c_data)
    
    total_tax = total_cgst + total_sgst + total_igst
    total_amount = subtotal + charges_subtotal + total_tax
    
    # Get delivery address from warehouse
    delivery_address = ""
    if po_data.delivery_warehouse_id:
        wh = await db.warehouses.find_one({"id": po_data.delivery_warehouse_id}, {"_id": 0})
        if wh:
            delivery_address = wh.get("address", "") or wh.get("location", "")
    
    po_doc = {
        "id": str(uuid.uuid4()),
        "po_number": po_number,
        "revision": 0,
        "revision_history": [],
        "supplier_id": po_data.supplier_id,
        "expected_date": po_data.expected_date,
        "delivery_warehouse_id": po_data.delivery_warehouse_id or "",
        "delivery_address": delivery_address,
        "quotation_ref": po_data.quotation_ref or "",
        "quotation_date": po_data.quotation_date,
        "lines": lines_with_tax,
        "additional_charges": charges_with_tax,
        "subtotal": round(subtotal, 2),
        "charges_subtotal": round(charges_subtotal, 2),
        "total_cgst": round(total_cgst, 2),
        "total_sgst": round(total_sgst, 2),
        "total_igst": round(total_igst, 2),
        "total_tax": round(total_tax, 2),
        "total_amount": round(total_amount, 2),
        "is_inter_state": is_inter_state,
        "currency": po_currency,
        "status": "draft",
        "notes": po_data.notes,
        "terms_conditions": po_data.terms_conditions if po_data.terms_conditions is not None else None,
        "revision_label": (po_data.revision_label or "").strip() or None,
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"]
    }
    await db.purchase_orders.insert_one(po_doc)
    po_doc.pop("_id", None)
    # Auto-update item purchase_price & unit_cost to the latest PO line rate
    for line in lines_with_tax:
        if line.get("item_id") and line.get("unit_price"):
            await db.items.update_one(
                {"id": line["item_id"]},
                {"$set": {"purchase_price": line["unit_price"], "unit_cost": line["unit_price"]}}
            )
    return po_doc

@purchase_orders_router.post("/from-mrp", status_code=201)
async def create_po_from_mrp(data: MRPCreatePORequest, request: Request):
    """Create PO from MRP suggestions — blocks if items already covered by existing POs"""
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager", "inventory_manager"], module="purchase_orders", action="create")
    supplier = await db.suppliers.find_one({"id": data.supplier_id})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    lines = []
    skipped_items = []
    for entry in data.items:
        item_id = entry.get("item_id") if isinstance(entry, dict) else entry
        item = await db.items.find_one({"id": item_id}, {"_id": 0})
        if not item:
            continue
        
        if isinstance(entry, dict):
            qty = entry.get("quantity") or entry.get("suggested_quantity") or max(item.get("safety_stock", 0) * 2 - item.get("current_stock", 0), 1)
            price = entry.get("unit_price") or item.get("unit_cost", 0)
        else:
            qty = max(item.get("safety_stock", 0) * 2 - item.get("current_stock", 0), 1)
            price = item.get("unit_cost", 0)
        
        # Check if existing non-cancelled POs already cover this item
        existing_po_qty = 0
        existing_pos = await db.purchase_orders.find(
            {"status": {"$nin": ["cancelled", "received", "short_closed"]},
             "$or": [{"lines.item_id": item_id}, {"items.item_id": item_id}]},
            {"_id": 0, "lines": 1, "items": 1}
        ).to_list(100)
        for epo in existing_pos:
            found_in_lines = False
            for pi in epo.get("lines", []):
                if pi.get("item_id") == item_id:
                    existing_po_qty += max(0, (pi.get("quantity", 0) or 0) - (pi.get("received_quantity", 0) or 0))
                    found_in_lines = True
            if not found_in_lines:
                for pi in epo.get("items", []):
                    if pi.get("item_id") == item_id:
                        existing_po_qty += max(0, (pi.get("quantity", 0) or 0) - (pi.get("received_quantity", 0) or 0))
        
        if existing_po_qty >= int(qty):
            skipped_items.append(f"{item.get('part_number')} (PO already covers {int(existing_po_qty)} >= {int(qty)})")
            continue
        
        # Use the full user-requested qty (user already sees suggested amount and confirmed it)
        final_qty = max(int(qty), 1)
        
        # Resolve description: client-provided override > item.description > item name
        client_desc = entry.get("description") if isinstance(entry, dict) else None
        line_desc = (client_desc if client_desc not in (None, "") else (item.get("description") or item.get("name") or ""))
        
        lines.append({
            "item_id": item_id,
            "description": line_desc,
            "quantity": final_qty,
            "unit_price": price,
            "hsn_code": item.get("hsn_code", ""),
            "gst_rate": item.get("gst_rate", 18),
            "uom": item.get("unit_of_measure", "pcs"),
            "notes": ""
        })
    
    if not lines:
        skip_msg = "\n".join(skipped_items) if skipped_items else "No valid items."
        raise HTTPException(status_code=400, detail=f"All items already have POs covering the required quantity.\n{skip_msg}")
    
    # GST calculation
    company_settings = await db.company_settings.find_one({"type": "company"}, {"_id": 0})
    company_state = company_settings.get("state_code", "") if company_settings else ""
    supplier_state = supplier.get("state_code", "")
    is_inter_state = company_state and supplier_state and company_state != supplier_state
    
    subtotal = 0
    total_cgst = 0
    total_sgst = 0
    total_igst = 0
    
    for line in lines:
        line_amount = line["quantity"] * line["unit_price"]
        gst_rate = line.get("gst_rate", 0)
        tax_amount = round(line_amount * gst_rate / 100, 2)
        line["line_amount"] = line_amount
        line["tax_amount"] = tax_amount
        if is_inter_state:
            line["igst_amount"] = tax_amount
            line["cgst_amount"] = 0
            line["sgst_amount"] = 0
            total_igst += tax_amount
        else:
            half_tax = round(tax_amount / 2, 2)
            line["cgst_amount"] = half_tax
            line["sgst_amount"] = half_tax
            line["igst_amount"] = 0
            total_cgst += half_tax
            total_sgst += half_tax
        subtotal += line_amount
    
    total_tax = total_cgst + total_sgst + total_igst
    total_amount = subtotal + total_tax
    
    count = await db.purchase_orders.count_documents({})
    po_number = await get_next_series_number("po_number")
    
    po_doc = {
        "id": str(uuid.uuid4()),
        "po_number": po_number,
        "revision": 0,
        "revision_history": [],
        "supplier_id": data.supplier_id,
        "expected_date": datetime.now(timezone.utc) + timedelta(days=supplier.get("lead_time_days", 7)),
        "delivery_warehouse_id": "",
        "delivery_address": "",
        "quotation_ref": "",
        "quotation_date": None,
        "lines": lines,
        "additional_charges": [],
        "subtotal": round(subtotal, 2),
        "charges_subtotal": 0,
        "total_cgst": round(total_cgst, 2),
        "total_sgst": round(total_sgst, 2),
        "total_igst": round(total_igst, 2),
        "total_tax": round(total_tax, 2),
        "total_amount": round(total_amount, 2),
        "is_inter_state": is_inter_state,
        "status": "draft",
        "notes": "Created from MRP suggestions",
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"]
    }
    await db.purchase_orders.insert_one(po_doc)
    po_doc.pop("_id", None)
    # Auto-update item purchase_price & unit_cost to the latest PO line rate
    for line in lines:
        if line.get("item_id") and line.get("unit_price"):
            await db.items.update_one(
                {"id": line["item_id"]},
                {"$set": {"purchase_price": line["unit_price"], "unit_cost": line["unit_price"]}}
            )
    return po_doc

@purchase_orders_router.put("/{po_id}")
async def update_purchase_order(po_id: str, po_data: PurchaseOrderUpdate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager", "inventory_manager"], module="purchase_orders", action="edit")
    existing_po = await db.purchase_orders.find_one({"id": po_id})
    if not existing_po:
        raise HTTPException(status_code=404, detail="Purchase order not found")
    
    if existing_po.get("status") in ["received", "cancelled"]:
        raise HTTPException(status_code=400, detail="Cannot edit a received or cancelled PO")
    
    is_draft = existing_po.get("status") == "draft"
    
    # If PO is already submitted (sent/partial), create a revision
    if not is_draft and po_data.lines is not None:
        # Save current state to revision history
        snapshot = {
            "revision": existing_po.get("revision", 0),
            "lines": existing_po.get("lines", []),
            "additional_charges": existing_po.get("additional_charges", []),
            "subtotal": existing_po.get("subtotal", 0),
            "total_amount": existing_po.get("total_amount", 0),
            "revised_at": datetime.now(timezone.utc),
            "revised_by": user["id"]
        }
        await db.purchase_orders.update_one(
            {"id": po_id},
            {"$push": {"revision_history": snapshot},
             "$inc": {"revision": 1}}
        )
    
    # Build update data - recalculate if lines or charges changed
    if po_data.lines is not None:
        supplier_id = po_data.supplier_id or existing_po.get("supplier_id")
        supplier = await db.suppliers.find_one({"id": supplier_id})
        company_settings = await db.company_settings.find_one({"type": "company"}, {"_id": 0})
        company_state = company_settings.get("state_code", "") if company_settings else ""
        supplier_state = supplier.get("state_code", "") if supplier else ""
        # Currency: explicit update value > existing > default INR. Non-INR ⇒ no GST.
        po_currency = (po_data.currency or existing_po.get("currency") or "INR").upper()
        is_export = po_currency != "INR"
        is_inter_state = (not is_export) and bool(company_state) and bool(supplier_state) and company_state != supplier_state
        
        lines_with_tax = []
        subtotal = 0
        total_cgst = 0
        total_sgst = 0
        total_igst = 0
        
        for line in po_data.lines:
            line_data = line.model_dump()
            gross_amount = line.quantity * line.unit_price
            discount_amount = 0
            if line.discount_value and line.discount_value > 0:
                if line.discount_type == "percentage":
                    discount_amount = round(gross_amount * line.discount_value / 100, 2)
                else:
                    discount_amount = round(line.discount_value, 2)
            line_data["discount_amount"] = discount_amount
            line_amount = gross_amount - discount_amount
            
            item_doc = await db.items.find_one({"id": line.item_id}, {"_id": 0})
            if item_doc:
                if not line_data.get("hsn_code"):
                    line_data["hsn_code"] = item_doc.get("hsn_code", "")
                if not line_data.get("uom") or line_data["uom"] == "pcs":
                    line_data["uom"] = item_doc.get("unit_of_measure", "pcs")
                if not line_data.get("description"):
                    line_data["description"] = item_doc.get("description", "")
            
            gst_rate = 0 if is_export else (line.gst_rate or 0)
            tax_amount = round(line_amount * gst_rate / 100, 2)
            if is_inter_state:
                line_data["igst_amount"] = tax_amount
                line_data["cgst_amount"] = 0
                line_data["sgst_amount"] = 0
                total_igst += tax_amount
            else:
                half_tax = round(tax_amount / 2, 2)
                line_data["cgst_amount"] = half_tax
                line_data["sgst_amount"] = half_tax
                line_data["igst_amount"] = 0
                total_cgst += half_tax
                total_sgst += half_tax
            
            line_data["gross_amount"] = round(gross_amount, 2)
            line_data["line_amount"] = round(line_amount, 2)
            line_data["tax_amount"] = tax_amount
            subtotal += line_amount
            lines_with_tax.append(line_data)
        
        charges_with_tax = []
        charges_subtotal = 0
        # Empty list "[]" MUST clear charges (was incorrectly falling back to existing
        # charges via `or existing_po.get(...)`). Use explicit "is not None" check.
        charges_source = (
            po_data.additional_charges
            if po_data.additional_charges is not None
            else existing_po.get("additional_charges", [])
        )
        for charge in charges_source:
            c_data = charge.model_dump() if hasattr(charge, 'model_dump') else dict(charge)
            c_amount = c_data.get("amount", 0)
            c_gst_rate = 0 if is_export else c_data.get("gst_rate", 0)
            c_tax = round(c_amount * c_gst_rate / 100, 2)
            if is_inter_state:
                c_data["igst_amount"] = c_tax
                c_data["cgst_amount"] = 0
                c_data["sgst_amount"] = 0
                total_igst += c_tax
            else:
                c_half = round(c_tax / 2, 2)
                c_data["cgst_amount"] = c_half
                c_data["sgst_amount"] = c_half
                c_data["igst_amount"] = 0
                total_cgst += c_half
                total_sgst += c_half
            c_data["tax_amount"] = c_tax
            charges_subtotal += c_amount
            charges_with_tax.append(c_data)
        
        total_tax = total_cgst + total_sgst + total_igst
        total_amount = subtotal + charges_subtotal + total_tax
        
        update_data = {
            "lines": lines_with_tax,
            "additional_charges": charges_with_tax,
            "subtotal": round(subtotal, 2),
            "charges_subtotal": round(charges_subtotal, 2),
            "total_cgst": round(total_cgst, 2),
            "total_sgst": round(total_sgst, 2),
            "total_igst": round(total_igst, 2),
            "total_tax": round(total_tax, 2),
            "total_amount": round(total_amount, 2),
            "is_inter_state": is_inter_state,
            "currency": po_currency,
        }
        if po_data.supplier_id:
            update_data["supplier_id"] = po_data.supplier_id
    else:
        update_data = {}
    
    # Set simple fields
    if po_data.expected_date is not None:
        update_data["expected_date"] = po_data.expected_date
    if po_data.status is not None:
        update_data["status"] = po_data.status
    if po_data.notes is not None:
        update_data["notes"] = po_data.notes
    if po_data.delivery_warehouse_id is not None:
        update_data["delivery_warehouse_id"] = po_data.delivery_warehouse_id
        wh = await db.warehouses.find_one({"id": po_data.delivery_warehouse_id}, {"_id": 0})
        update_data["delivery_address"] = (wh.get("address", "") or wh.get("location", "")) if wh else ""
    if po_data.quotation_ref is not None:
        update_data["quotation_ref"] = po_data.quotation_ref
    if po_data.quotation_date is not None:
        update_data["quotation_date"] = po_data.quotation_date
    # Persist PO-specific Terms & Conditions (overrides the default from Inventory → Configuration).
    # Use hasattr-style explicit include so an empty string ("") is saved as "clear override".
    if po_data.terms_conditions is not None:
        update_data["terms_conditions"] = po_data.terms_conditions
    # Manual revision label override (BOM-style). Blank string clears it, falling back to numeric revision.
    if po_data.revision_label is not None:
        update_data["revision_label"] = po_data.revision_label.strip() or None
    # Currency-only update (no line edit) — persist as-is. (When `lines` was provided
    # above, currency was already written into update_data via the recompute branch.)
    if po_data.currency is not None and "currency" not in update_data:
        update_data["currency"] = (po_data.currency or "INR").upper()
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    update_data["updated_at"] = datetime.now(timezone.utc)
    await db.purchase_orders.update_one({"id": po_id}, {"$set": update_data})
    
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    return po


@purchase_orders_router.post("/{po_id}/cancel")
async def cancel_purchase_order(po_id: str, request: Request):
    """Cancel a PO. Only draft/approved/sent POs can be cancelled."""
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="purchase_orders", action="create")
    po = await db.purchase_orders.find_one({"id": po_id})
    if not po:
        raise HTTPException(status_code=404, detail="Purchase order not found")
    
    if po.get("status") in ["received", "cancelled"]:
        raise HTTPException(status_code=400, detail=f"Cannot cancel a {po['status']} PO")
    
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"status": "cancelled", "cancelled_at": datetime.now(timezone.utc), "cancelled_by": user["id"]}}
    )
    
    return {"message": f"Purchase Order {po.get('po_number')} cancelled successfully"}


@purchase_orders_router.post("/{po_id}/short-close")
async def short_close_purchase_order(po_id: str, request: Request, data: dict = Body(default={})):
    """Manually short-close a PO. Used when supplier denies further supply.
    Sets status='short_closed' so the un-received qty no longer counts toward
    'PO ordered qty' in MRP — letting users place fresh POs for the shortage.
    Allowed only on draft/approved/sent/partial POs."""
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="purchase_orders", action="edit")
    po = await db.purchase_orders.find_one({"id": po_id})
    if not po:
        raise HTTPException(status_code=404, detail="Purchase order not found")
    
    if po.get("status") in ["received", "cancelled", "short_closed"]:
        raise HTTPException(status_code=400, detail=f"Cannot short-close a {po['status']} PO")
    
    reason = (data.get("reason") or "").strip() if isinstance(data, dict) else ""
    
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {
            "status": "short_closed",
            "short_closed_at": datetime.now(timezone.utc),
            "short_closed_by": user["id"],
            "short_close_reason": reason,
        }}
    )
    
    return {"message": f"Purchase Order {po.get('po_number')} short-closed successfully"}


@purchase_orders_router.post("/{po_id}/receive")
async def receive_purchase_order(po_id: str, request: Request):
    """Receive PO and update inventory"""
    user = await get_current_user(request)
    _require_access(user, ["admin", "inventory_manager"], module="purchase_orders", action="create")
    po = await db.purchase_orders.find_one({"id": po_id})
    if not po:
        raise HTTPException(status_code=404, detail="Purchase order not found")
    
    if po.get("status") == "received":
        raise HTTPException(status_code=400, detail="GRN already completed for this PO")
    
    # Create inventory transactions for each line
    for line in po.get("lines", []):
        item = await db.items.find_one({"id": line.get("item_id")})
        if item:
            current_stock = item.get("current_stock", 0)
            new_stock = current_stock + line.get("quantity", 0)
            
            tx_doc = {
                "id": str(uuid.uuid4()),
                "item_id": line.get("item_id"),
                "transaction_type": "receive",
                "quantity": line.get("quantity", 0),
                "reference_type": "purchase_order",
                "reference_id": po_id,
                "previous_stock": current_stock,
                "new_stock": new_stock,
                "notes": f"Received from PO {po.get('po_number')}",
                "created_at": datetime.now(timezone.utc),
                "created_by": user["id"]
            }
            await db.inventory_transactions.insert_one(tx_doc)
            await db.items.update_one({"id": line.get("item_id")}, {"$set": {"current_stock": new_stock}})
    
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"status": "received", "received_at": datetime.now(timezone.utc), "received_by": user["id"]}}
    )
    
    # If PO was created from SC order (without material), update SC order received qty + complete it
    sc_order_id = po.get("reference_sc_order_id")
    if sc_order_id:
        sc_order = await db.subcontract_orders.find_one({"id": sc_order_id})
        if sc_order:
            # Update received quantities on SC order lines
            sc_lines = sc_order.get("lines", [])
            for po_line in po.get("lines", []) + po.get("items", []):
                for sc_line in sc_lines:
                    if sc_line["item_id"] == po_line.get("item_id"):
                        sc_line["received_quantity"] = sc_line.get("received_quantity", 0) + po_line.get("quantity", 0)
            
            # Check if all received
            all_received = all(sl.get("received_quantity", 0) >= sl.get("quantity", 0) for sl in sc_lines)
            sc_new_status = "completed" if all_received else "in_progress"
            
            await db.subcontract_orders.update_one(
                {"id": sc_order_id},
                {"$set": {"lines": sc_lines, "status": sc_new_status, "updated_at": datetime.now(timezone.utc)}}
            )
            
            # If SC completed, auto-complete linked MOs
            if sc_new_status == "completed":
                ref_wo_ids = sc_order.get("reference_wo_ids", [])
                if not ref_wo_ids and sc_order.get("reference_wo_id"):
                    ref_wo_ids = [sc_order["reference_wo_id"]]
                
                for ref_wo_id in ref_wo_ids:
                    ref_wo = await db.work_orders.find_one({"id": ref_wo_id})
                    if ref_wo and ref_wo.get("is_subcontract") and ref_wo.get("status") != "completed":
                        ops = ref_wo.get("operations_status", [])
                        for op in ops:
                            if op.get("status") != "completed":
                                op["status"] = "completed"
                                op["actual_end"] = datetime.now(timezone.utc)
                                op["quantity_completed"] = ref_wo.get("quantity", 0)
                                op["quantity_accepted"] = ref_wo.get("quantity", 0)
                        
                        await db.work_orders.update_one(
                            {"id": ref_wo_id},
                            {"$set": {
                                "operations_status": ops,
                                "status": "completed",
                                "quantity_completed": ref_wo.get("quantity", 0),
                                "actual_end": datetime.now(timezone.utc),
                                "updated_at": datetime.now(timezone.utc)
                            }}
                        )
    
    return {"message": "GRN completed successfully"}

# ================== WAREHOUSE ROUTES ==================

@warehouses_router.get("")
async def get_warehouses(request: Request, status: Optional[str] = None):
    await get_current_user(request)
    query = {}
    if status:
        query["status"] = status
    warehouses = await db.warehouses.find(query, {"_id": 0}).to_list(100)
    return warehouses

@warehouses_router.get("/{warehouse_id}")
async def get_warehouse(warehouse_id: str, request: Request):
    await get_current_user(request)
    warehouse = await db.warehouses.find_one({"id": warehouse_id}, {"_id": 0})
    if not warehouse:
        raise HTTPException(status_code=404, detail="Warehouse not found")
    return warehouse

@warehouses_router.get("/stock/by-item")
async def get_stock_by_item(request: Request):
    """Aggregate warehouse stock grouped by item. Returns {item_id: [{warehouse_id, warehouse_name, warehouse_code, quantity}]}"""
    await get_current_user(request)
    all_stock = await db.warehouse_stock.find({}, {"_id": 0}).to_list(10000)
    warehouses = {w["id"]: w for w in await db.warehouses.find({}, {"_id": 0}).to_list(1000)}
    grouped = {}
    for s in all_stock:
        iid = s.get("item_id")
        if not iid:
            continue
        wh = warehouses.get(s.get("warehouse_id"), {})
        grouped.setdefault(iid, []).append({
            "warehouse_id": s.get("warehouse_id"),
            "warehouse_name": wh.get("name", ""),
            "warehouse_code": wh.get("code", ""),
            "quantity": s.get("quantity", 0),
        })
    return grouped

@warehouses_router.get("/{warehouse_id}/stock")
async def get_warehouse_stock(warehouse_id: str, request: Request):
    """Get stock levels for a specific warehouse"""
    await get_current_user(request)
    
    warehouse = await db.warehouses.find_one({"id": warehouse_id}, {"_id": 0})
    if not warehouse:
        raise HTTPException(status_code=404, detail="Warehouse not found")
    
    stock = await db.warehouse_stock.find({"warehouse_id": warehouse_id}, {"_id": 0}).to_list(1000)
    
    for s in stock:
        item = await db.items.find_one({"id": s.get("item_id")}, {"_id": 0})
        s["item"] = item
    
    return {"warehouse": warehouse, "stock": stock}

@warehouses_router.post("", status_code=201)
async def create_warehouse(warehouse_data: WarehouseCreate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin", "inventory_manager"], module="stores", action="create")
    existing = await db.warehouses.find_one({"code": warehouse_data.code})
    if existing:
        raise HTTPException(status_code=400, detail="Warehouse code already exists")
    
    # If this is default, unset other defaults
    if warehouse_data.is_default:
        await db.warehouses.update_many({}, {"$set": {"is_default": False}})
    
    warehouse_doc = {
        "id": str(uuid.uuid4()),
        **warehouse_data.model_dump(),
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"]
    }
    await db.warehouses.insert_one(warehouse_doc)
    warehouse_doc.pop("_id", None)
    return warehouse_doc

@warehouses_router.put("/{warehouse_id}")
async def update_warehouse(warehouse_id: str, warehouse_data: WarehouseUpdate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin", "inventory_manager"], module="stores", action="edit")
    update_data = {k: v for k, v in warehouse_data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    # If setting as default, unset other defaults
    if update_data.get("is_default"):
        await db.warehouses.update_many({"id": {"$ne": warehouse_id}}, {"$set": {"is_default": False}})
    
    update_data["updated_at"] = datetime.now(timezone.utc)
    result = await db.warehouses.update_one({"id": warehouse_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Warehouse not found")
    
    warehouse = await db.warehouses.find_one({"id": warehouse_id}, {"_id": 0})
    return warehouse

@warehouses_router.post("/transfer", status_code=201)
async def create_stock_transfer(transfer_data: StockTransferCreate, request: Request):
    """Transfer stock between warehouses"""
    user = await get_current_user(request)
    _require_access(user, ["admin", "inventory_manager"], module="stores", action="create")
    # Verify warehouses exist
    from_wh = await db.warehouses.find_one({"id": transfer_data.from_warehouse_id})
    to_wh = await db.warehouses.find_one({"id": transfer_data.to_warehouse_id})
    if not from_wh or not to_wh:
        raise HTTPException(status_code=404, detail="Warehouse not found")
    
    # Check stock in source warehouse
    from_stock = await db.warehouse_stock.find_one({
        "warehouse_id": transfer_data.from_warehouse_id,
        "item_id": transfer_data.item_id
    })
    
    if not from_stock or from_stock.get("quantity", 0) < transfer_data.quantity:
        raise HTTPException(status_code=400, detail="Insufficient stock in source warehouse")
    
    # Update source warehouse stock
    new_from_qty = from_stock.get("quantity", 0) - transfer_data.quantity
    await db.warehouse_stock.update_one(
        {"warehouse_id": transfer_data.from_warehouse_id, "item_id": transfer_data.item_id},
        {"$set": {"quantity": new_from_qty, "updated_at": datetime.now(timezone.utc)}}
    )
    
    # Update destination warehouse stock
    to_stock = await db.warehouse_stock.find_one({
        "warehouse_id": transfer_data.to_warehouse_id,
        "item_id": transfer_data.item_id
    })
    
    if to_stock:
        new_to_qty = to_stock.get("quantity", 0) + transfer_data.quantity
        await db.warehouse_stock.update_one(
            {"warehouse_id": transfer_data.to_warehouse_id, "item_id": transfer_data.item_id},
            {"$set": {"quantity": new_to_qty, "updated_at": datetime.now(timezone.utc)}}
        )
    else:
        await db.warehouse_stock.insert_one({
            "id": str(uuid.uuid4()),
            "warehouse_id": transfer_data.to_warehouse_id,
            "item_id": transfer_data.item_id,
            "quantity": transfer_data.quantity,
            "created_at": datetime.now(timezone.utc)
        })
    
    # Create transfer record
    transfer_doc = {
        "id": str(uuid.uuid4()),
        **transfer_data.model_dump(),
        "status": "completed",
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"]
    }
    await db.stock_transfers.insert_one(transfer_doc)
    transfer_doc.pop("_id", None)
    return transfer_doc

@warehouses_router.get("/transfers/history")
async def get_transfer_history(request: Request, limit: int = 50):
    await get_current_user(request)
    transfers = await db.stock_transfers.find({}, {"_id": 0}).sort("created_at", -1).to_list(limit)
    
    for t in transfers:
        t["from_warehouse"] = await db.warehouses.find_one({"id": t.get("from_warehouse_id")}, {"_id": 0})
        t["to_warehouse"] = await db.warehouses.find_one({"id": t.get("to_warehouse_id")}, {"_id": 0})
        t["item"] = await db.items.find_one({"id": t.get("item_id")}, {"_id": 0})
    
    return transfers

# ================== GRN (GOODS RECEIPT NOTE) ROUTES ==================

@grn_router.get("")
async def get_grn_list(request: Request):
    """Get all GRN records with supplier + line items enriched (batch fetched)"""
    await get_current_user(request)
    grns = await db.grn.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Batch-fetch POs
    po_ids = {g.get("po_id") for g in grns if g.get("po_id")}
    pos_map = {}
    if po_ids:
        async for po in db.purchase_orders.find({"id": {"$in": list(po_ids)}}, {"_id": 0}):
            pos_map[po["id"]] = po
    # Batch-fetch JW orders for JW GRNs
    jw_ids = {g.get("sc_order_id") or g.get("jw_order_id") for g in grns if (g.get("sc_order_id") or g.get("jw_order_id"))}
    jws_map = {}
    if jw_ids:
        async for jw in db.subcontract_orders.find({"id": {"$in": list(jw_ids)}}, {"_id": 0}):
            jws_map[jw["id"]] = jw
    # Collect supplier ids
    supplier_ids = {po.get("supplier_id") for po in pos_map.values() if po.get("supplier_id")}
    for jw in jws_map.values():
        if jw.get("supplier_id"):
            supplier_ids.add(jw["supplier_id"])
    suppliers_map = {}
    if supplier_ids:
        async for s in db.suppliers.find({"id": {"$in": list(supplier_ids)}}, {"_id": 0}):
            suppliers_map[s["id"]] = s
    # Collect item ids
    item_ids = set()
    for g in grns:
        for line in g.get("lines", []):
            if line.get("item_id"):
                item_ids.add(line["item_id"])
    items_map = {}
    if item_ids:
        async for it in db.items.find({"id": {"$in": list(item_ids)}}, {"_id": 0}):
            items_map[it["id"]] = it
    
    # Batch-fetch DCs for JW GRNs
    dc_ids = {g.get("dc_id") for g in grns if g.get("dc_id")}
    dcs_map = {}
    if dc_ids:
        async for dc in db.delivery_challans.find({"id": {"$in": list(dc_ids)}}, {"_id": 0}):
            dcs_map[dc["id"]] = dc
    
    for grn in grns:
        po = pos_map.get(grn.get("po_id"))
        grn["po"] = po
        jw = jws_map.get(grn.get("sc_order_id") or grn.get("jw_order_id"))
        grn["jw_order"] = jw
        if jw and not grn.get("jw_order_number"):
            grn["jw_order_number"] = jw.get("order_number")
        # Enrich DC number for JW GRNs (prefer explicit dc_id; otherwise, fall back to latest sent DC for this SC)
        dc = dcs_map.get(grn.get("dc_id")) if grn.get("dc_id") else None
        if not dc and jw:
            dc = await db.delivery_challans.find_one(
                {"subcontract_order_id": jw["id"], "status": "sent"},
                {"_id": 0}, sort=[("created_at", -1)]
            )
        grn["dc"] = dc
        if dc and not grn.get("dc_number"):
            grn["dc_number"] = dc.get("dc_number")
        supplier_id = (po and po.get("supplier_id")) or (jw and jw.get("supplier_id"))
        grn["supplier"] = suppliers_map.get(supplier_id) if supplier_id else None
        for line in grn.get("lines", []):
            line["item"] = items_map.get(line.get("item_id"))
    return grns

@grn_router.get("/pending-pos")
async def get_pending_grn_pos(request: Request):
    """Get POs that are approved/sent/partial and ready for GRN"""
    await get_current_user(request)
    pos = await db.purchase_orders.find(
        {"status": {"$in": ["approved", "sent", "partial"]}}, {"_id": 0}
    ).sort("created_at", -1).to_list(1000)
    
    # Batch fetch suppliers + items (avoid N+1)
    supplier_ids = {po.get("supplier_id") for po in pos if po.get("supplier_id")}
    item_ids = set()
    for po in pos:
        for line in po.get("lines", []):
            if line.get("item_id"): item_ids.add(line["item_id"])
    suppliers_map = {}
    if supplier_ids:
        async for s in db.suppliers.find({"id": {"$in": list(supplier_ids)}}, {"_id": 0}):
            suppliers_map[s["id"]] = s
    items_map = {}
    if item_ids:
        async for it in db.items.find({"id": {"$in": list(item_ids)}}, {"_id": 0}):
            items_map[it["id"]] = it
    for po in pos:
        po["supplier"] = suppliers_map.get(po.get("supplier_id"))
        for line in po.get("lines", []):
            line["item"] = items_map.get(line.get("item_id"))
    return pos

@grn_router.post("", status_code=201)
async def create_grn(grn_data: GRNCreate, request: Request):
    """Create GRN - verify material, price, update inventory.
    Supports partial receipts: cumulatively updates each PO line's
    received_quantity. PO status becomes 'partial' until every line is fully
    received, then flips to 'received'.

    NEW: When `status='draft'`, the GRN is saved as a draft (editable) and
    NO inventory/PO/SC updates happen until the user explicitly approves it
    via POST /grn/{grn_id}/approve. This lets the user double-check the
    received qty/price before stock is committed."""
    user = await get_current_user(request)
    _require_access(user, ["admin", "inventory_manager"], module="stores", action="create")
    po = await db.purchase_orders.find_one({"id": grn_data.po_id})
    if not po:
        raise HTTPException(status_code=404, detail="Purchase order not found")
    if po.get("status") in ["received", "short_closed", "cancelled"]:
        raise HTTPException(status_code=400, detail=f"Cannot create GRN for a {po.get('status')} PO")

    # Generate GRN number
    count = await db.grn.count_documents({})
    grn_number = f"GRN-{str(count + 1).zfill(6)}"

    save_as_draft = (grn_data.status or "posted").lower() == "draft"

    if save_as_draft:
        # Build raw GRN lines from incoming data without touching inventory/PO.
        draft_lines = []
        for grn_line in grn_data.lines:
            if (grn_line.received_quantity or 0) <= 0:
                continue
            po_line = next((l for l in po.get("lines", []) if l.get("item_id") == grn_line.item_id), {})
            draft_lines.append({
                "item_id": grn_line.item_id,
                "po_quantity": po_line.get("quantity", 0),
                "received_quantity": grn_line.received_quantity,
                "po_price": po_line.get("unit_price", 0),
                "verified_price": grn_line.verified_price,
                "uom": po_line.get("uom", "pcs"),
                "hsn_code": po_line.get("hsn_code", ""),
            })
        if not draft_lines:
            raise HTTPException(status_code=400, detail="No lines with received quantity > 0")
        draft_doc = {
            "id": str(uuid.uuid4()),
            "grn_number": grn_number,
            "po_id": grn_data.po_id,
            "po_number": po.get("po_number", ""),
            "supplier_id": po.get("supplier_id", ""),
            "supplier_invoice_no": grn_data.supplier_invoice_no,
            "supplier_invoice_date": grn_data.supplier_invoice_date,
            "warehouse_id": grn_data.warehouse_id or po.get("delivery_warehouse_id", ""),
            "lines": draft_lines,
            "notes": grn_data.notes,
            "status": "draft",
            "created_at": datetime.now(timezone.utc),
            "created_by": user["id"],
        }
        await db.grn.insert_one(draft_doc)
        draft_doc.pop("_id", None)
        return draft_doc

    # Posted path — original commit-now behaviour. Delegated to helper so
    # /approve can reuse exactly the same logic.
    return await _post_grn_to_inventory(
        user=user,
        po=po,
        grn_number=grn_number,
        supplier_invoice_no=grn_data.supplier_invoice_no,
        supplier_invoice_date=grn_data.supplier_invoice_date,
        warehouse_id=grn_data.warehouse_id,
        notes=grn_data.notes,
        line_items=[(ln.item_id, float(ln.received_quantity or 0), float(ln.verified_price or 0)) for ln in grn_data.lines],
    )


async def _post_grn_to_inventory(
    *,
    user: dict,
    po: dict,
    grn_number: str,
    supplier_invoice_no: str,
    supplier_invoice_date,
    warehouse_id: Optional[str],
    notes: Optional[str],
    line_items: List[Tuple[str, float, float]],
    existing_grn_id: Optional[str] = None,
):
    """Shared posting logic — used by direct GRN creation AND by the
    /grn/{id}/approve endpoint that promotes a draft to a committed GRN.

    `line_items` is a list of (item_id, received_quantity, verified_price).
    `existing_grn_id` — when re-using a draft GRN's UUID instead of inserting
    a new doc; the helper updates the existing draft to status=posted.
    """
    # Build a quick lookup of incoming receive qty per item
    incoming = {}
    for item_id, recv_qty, _verified_price in line_items:
        if recv_qty > 0:
            incoming[item_id] = incoming.get(item_id, 0.0) + recv_qty

    # Prepare updated PO lines (cumulative received_quantity per line)
    updated_po_lines = []
    for po_line in po.get("lines", []):
        prev_recv = float(po_line.get("received_quantity", 0) or 0)
        ord_qty = float(po_line.get("quantity", 0) or 0)
        add = float(incoming.get(po_line.get("item_id"), 0))
        new_recv = prev_recv + add
        new_recv_capped = min(new_recv, ord_qty) if ord_qty > 0 else new_recv
        updated_po_lines.append({**po_line, "received_quantity": new_recv_capped})

    # Process each line - update inventory with verified quantities
    grn_lines = []
    for item_id, recv_qty, verified_price in line_items:
        if recv_qty <= 0:
            continue
        item = await db.items.find_one({"id": item_id})
        if not item:
            continue

        current_stock = item.get("current_stock", 0)
        new_stock = current_stock + recv_qty

        # Find matching PO line for reference
        po_line = next((l for l in po.get("lines", []) if l.get("item_id") == item_id), {})

        grn_lines.append({
            "item_id": item_id,
            "po_quantity": po_line.get("quantity", 0),
            "received_quantity": recv_qty,
            "po_price": po_line.get("unit_price", 0),
            "verified_price": verified_price,
            "uom": po_line.get("uom", "pcs"),
            "hsn_code": po_line.get("hsn_code", ""),
        })

        # Create inventory transaction
        tx_doc = {
            "id": str(uuid.uuid4()),
            "item_id": item_id,
            "transaction_type": "receive",
            "quantity": recv_qty,
            "reference_type": "grn",
            "reference_id": grn_number,
            "previous_stock": current_stock,
            "new_stock": new_stock,
            "notes": f"GRN {grn_number} from PO {po.get('po_number')}",
            "created_at": datetime.now(timezone.utc),
            "created_by": user["id"],
        }
        await db.inventory_transactions.insert_one(tx_doc)
        await db.items.update_one({"id": item_id}, {"$set": {"current_stock": new_stock}})

    if not grn_lines:
        raise HTTPException(status_code=400, detail="No lines with received quantity > 0")

    grn_doc = {
        "id": existing_grn_id or str(uuid.uuid4()),
        "grn_number": grn_number,
        "po_id": po.get("id"),
        "po_number": po.get("po_number", ""),
        "supplier_id": po.get("supplier_id", ""),
        "supplier_invoice_no": supplier_invoice_no,
        "supplier_invoice_date": supplier_invoice_date,
        "warehouse_id": warehouse_id or po.get("delivery_warehouse_id", ""),
        "lines": grn_lines,
        "notes": notes,
        "status": "posted",
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"],
    }
    if existing_grn_id:
        # Promoting a draft → posted. Preserve original created_at if present.
        existing = await db.grn.find_one({"id": existing_grn_id}, {"_id": 0})
        if existing and existing.get("created_at"):
            grn_doc["created_at"] = existing["created_at"]
        grn_doc["approved_at"] = datetime.now(timezone.utc)
        grn_doc["approved_by"] = user["id"]
        await db.grn.update_one({"id": existing_grn_id}, {"$set": {k: v for k, v in grn_doc.items() if k != "id"}})
    else:
        await db.grn.insert_one(grn_doc)
        grn_doc.pop("_id", None)

    # Decide new PO status: 'received' iff every line is fully received, else 'partial'
    fully_received = all(
        float(pl.get("received_quantity", 0) or 0) >= float(pl.get("quantity", 0) or 0)
        for pl in updated_po_lines
    ) and len(updated_po_lines) > 0
    new_status = "received" if fully_received else "partial"

    grn_numbers_history = list(po.get("grn_numbers", []) or [])
    if grn_number not in grn_numbers_history:
        grn_numbers_history.append(grn_number)

    po_update = {
        "status": new_status,
        "lines": updated_po_lines,
        "grn_numbers": grn_numbers_history,
        "grn_number": grn_number,
    }
    if fully_received:
        po_update["received_at"] = datetime.now(timezone.utc)
        po_update["received_by"] = user["id"]

    await db.purchase_orders.update_one({"id": po.get("id")}, {"$set": po_update})

    # SC linkage logic only fires when the PO is fully received.
    sc_order_ids = po.get("reference_sc_order_ids", []) if fully_received else []
    if not sc_order_ids and po.get("reference_sc_order_id") and fully_received:
        sc_order_ids = [po["reference_sc_order_id"]]

    for sc_id in sc_order_ids:
        sc_order = await db.subcontract_orders.find_one({"id": sc_id})
        if not sc_order or sc_order.get("status") == "completed":
            continue

        for p in sc_order.get("job_work_parts", []):
            for gl in grn_lines:
                if gl["item_id"] == p.get("item_id"):
                    p["received_quantity"] = p.get("received_quantity", 0) + gl["received_quantity"]

        await db.subcontract_orders.update_one({"id": sc_id}, {"$set": {
            "status": "completed",
            "job_work_parts": sc_order.get("job_work_parts", []),
            "last_receipt_date": datetime.now(timezone.utc).isoformat(),
            "completed_at": datetime.now(timezone.utc).isoformat(),
        }})

        all_wo_ids = list(set(filter(None, [
            sc_order.get("reference_wo_id"),
            *(sc_order.get("reference_wo_ids", []))
        ])))

        for ref_wo_id in all_wo_ids:
            ref_wo = await db.work_orders.find_one({"id": ref_wo_id})
            if not ref_wo or ref_wo.get("status") == "completed":
                continue

            ops = ref_wo.get("operations_status", [])
            for op in ops:
                if op.get("status") != "completed":
                    op["status"] = "completed"
                    op["actual_end"] = datetime.now(timezone.utc)
                    op["quantity_completed"] = ref_wo.get("quantity", 0)
                    op["quantity_accepted"] = ref_wo.get("quantity", 0)

            mo_qty = ref_wo.get("quantity", 0)

            await db.work_orders.update_one({"id": ref_wo_id}, {"$set": {
                "operations_status": ops, "status": "completed",
                "quantity_completed": mo_qty, "actual_end": datetime.now(timezone.utc),
            }})

    return grn_doc


@grn_router.put("/{grn_id}")
async def update_draft_grn(grn_id: str, data: GRNUpdate, request: Request):
    """Edit a draft GRN. Posted GRNs are immutable."""
    user = await get_current_user(request)
    _require_access(user, ["admin", "inventory_manager"], module="stores", action="edit")
    grn = await db.grn.find_one({"id": grn_id})
    if not grn:
        raise HTTPException(status_code=404, detail="GRN not found")
    if (grn.get("status") or "posted") != "draft":
        raise HTTPException(status_code=400, detail="Only draft GRNs can be edited. This GRN has already been approved.")

    update_data = {}
    if data.supplier_invoice_no is not None:
        update_data["supplier_invoice_no"] = data.supplier_invoice_no
    if data.supplier_invoice_date is not None:
        update_data["supplier_invoice_date"] = data.supplier_invoice_date
    if data.warehouse_id is not None:
        update_data["warehouse_id"] = data.warehouse_id
    if data.notes is not None:
        update_data["notes"] = data.notes
    if data.lines is not None:
        # Rebuild lines: pull po_quantity / po_price / uom / hsn from the PO
        # so the user never has to send those (and can't tamper with them).
        po = await db.purchase_orders.find_one({"id": grn.get("po_id")})
        new_lines = []
        for ln in data.lines:
            if (ln.received_quantity or 0) <= 0:
                continue
            po_line = next((l for l in (po or {}).get("lines", []) if l.get("item_id") == ln.item_id), {})
            new_lines.append({
                "item_id": ln.item_id,
                "po_quantity": po_line.get("quantity", 0),
                "received_quantity": ln.received_quantity,
                "po_price": po_line.get("unit_price", 0),
                "verified_price": ln.verified_price,
                "uom": po_line.get("uom", "pcs"),
                "hsn_code": po_line.get("hsn_code", ""),
            })
        if not new_lines:
            raise HTTPException(status_code=400, detail="Draft GRN must have at least one line with received qty > 0")
        update_data["lines"] = new_lines

    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc)
        update_data["updated_by"] = user["id"]
        await db.grn.update_one({"id": grn_id}, {"$set": update_data})
    return await db.grn.find_one({"id": grn_id}, {"_id": 0})


@grn_router.post("/{grn_id}/approve")
async def approve_draft_grn(grn_id: str, request: Request):
    """Promote a draft GRN to posted — runs the full inventory + PO + SC
    cascade that a direct-post GRN would have run."""
    user = await get_current_user(request)
    _require_access(user, ["admin", "inventory_manager"], module="stores", action="edit")
    grn = await db.grn.find_one({"id": grn_id}, {"_id": 0})
    if not grn:
        raise HTTPException(status_code=404, detail="GRN not found")
    if (grn.get("status") or "posted") != "draft":
        raise HTTPException(status_code=400, detail="GRN is not in draft status")
    po = await db.purchase_orders.find_one({"id": grn.get("po_id")})
    if not po:
        raise HTTPException(status_code=404, detail="Source PO no longer exists")
    if po.get("status") in ("received", "short_closed", "cancelled"):
        raise HTTPException(status_code=400, detail=f"Cannot approve GRN: source PO is {po.get('status')}")
    line_items = [
        (ln.get("item_id"), float(ln.get("received_quantity") or 0), float(ln.get("verified_price") or 0))
        for ln in (grn.get("lines") or [])
    ]
    return await _post_grn_to_inventory(
        user=user,
        po=po,
        grn_number=grn.get("grn_number"),
        supplier_invoice_no=grn.get("supplier_invoice_no") or "",
        supplier_invoice_date=grn.get("supplier_invoice_date"),
        warehouse_id=grn.get("warehouse_id"),
        notes=grn.get("notes") or "",
        line_items=line_items,
        existing_grn_id=grn_id,
    )


@grn_router.post("/manual", status_code=201)
async def create_manual_grn(data: ManualGRNCreate, request: Request):
    """Create a GRN WITHOUT a preceding Purchase Order.
    Used for: direct receipts, cash purchases, emergency deliveries, returns-in etc.
    Treats user-entered supplier + lines as authoritative, updates inventory + transactions.
    """
    user = await get_current_user(request)
    _require_access(user, ["admin", "inventory_manager"], module="stores", action="create")
    supplier = await db.suppliers.find_one({"id": data.supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")

    count = await db.grn.count_documents({})
    grn_number = f"GRN-{str(count + 1).zfill(6)}"

    grn_lines = []
    for ln in data.lines:
        item = await db.items.find_one({"id": ln.item_id})
        if not item:
            raise HTTPException(status_code=404, detail=f"Item {ln.item_id} not found")
        new_stock = (item.get("current_stock") or 0) + ln.received_quantity
        grn_lines.append({
            "item_id": ln.item_id,
            "po_quantity": 0,
            "received_quantity": ln.received_quantity,
            "po_price": 0,
            "verified_price": ln.verified_price,
            "uom": ln.uom or item.get("unit_of_measure") or "pcs",
            "hsn_code": ln.hsn_code or item.get("hsn_code", ""),
        })
        # Inventory transaction
        await db.inventory_transactions.insert_one({
            "id": str(uuid.uuid4()),
            "item_id": ln.item_id,
            "warehouse_id": data.warehouse_id or None,
            "transaction_type": "goods_receipt",
            "quantity": ln.received_quantity,
            "unit_cost": ln.verified_price,
            "reference_id": grn_number,
            "reference_type": "manual_grn",
            "notes": f"Manual GRN {grn_number} (supplier: {supplier.get('name')})",
            "created_at": datetime.now(timezone.utc),
            "created_by": user["id"],
        })
        await db.items.update_one(
            {"id": ln.item_id},
            {"$set": {"current_stock": new_stock, "purchase_price": ln.verified_price, "unit_cost": ln.verified_price}},
        )

    grn_doc = {
        "id": str(uuid.uuid4()),
        "grn_number": grn_number,
        "po_id": None,
        "manual": True,
        "supplier_id": data.supplier_id,
        "supplier_invoice_no": data.supplier_invoice_no,
        "supplier_invoice_date": data.supplier_invoice_date,
        "warehouse_id": data.warehouse_id,
        "notes": data.notes,
        "lines": grn_lines,
        "total_received_quantity": sum(l["received_quantity"] for l in grn_lines),
        "total_cost": sum(l["received_quantity"] * l["verified_price"] for l in grn_lines),
        "qty_mismatches": [],
        "price_mismatches": [],
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"],
    }
    await db.grn.insert_one(grn_doc)
    grn_doc.pop("_id", None)
    # Bump matching Manual-DC line received_qty so the DC picker filters
    # out fully-received DCs and partials stay open with the balance. We
    # match by item_id; if multiple lines have the same item, qty is
    # distributed FIFO into the lines with remaining balance.
    if data.manual_dc_id:
        dc = await db.delivery_challans.find_one({"id": data.manual_dc_id, "is_manual": True})
        if dc:
            new_lines = list(dc.get("lines") or [])
            for grn_line in grn_lines:
                remaining = float(grn_line["received_quantity"])
                if remaining <= 0:
                    continue
                for dl in new_lines:
                    if remaining <= 0:
                        break
                    if dl.get("item_id") != grn_line["item_id"]:
                        continue
                    open_qty = float(dl.get("quantity") or 0) - float(dl.get("received_qty") or 0)
                    if open_qty <= 0:
                        continue
                    take = min(open_qty, remaining)
                    dl["received_qty"] = float(dl.get("received_qty") or 0) + take
                    remaining -= take
            # If every line is now fully received, flip DC status to received.
            all_done = all(float(l.get("quantity") or 0) <= float(l.get("received_qty") or 0) for l in new_lines)
            update_set = {"lines": new_lines, "updated_at": datetime.now(timezone.utc)}
            if all_done:
                update_set["status"] = "received"
                update_set["received_at"] = datetime.now(timezone.utc)
            await db.delivery_challans.update_one({"id": data.manual_dc_id}, {"$set": update_set})
    return grn_doc


@grn_router.get("/{grn_id}/print-data")
async def get_grn_print_data(grn_id: str, request: Request):
    """Get GRN data for printing. Supports both PO-based GRNs and JW (subcontract) GRNs."""
    await get_current_user(request)
    grn = await db.grn.find_one({"id": grn_id}, {"_id": 0})
    if not grn:
        raise HTTPException(status_code=404, detail="GRN not found")
    
    is_jw = bool(grn.get("jw_order_id") or grn.get("sc_order_id"))
    grn["is_jw"] = is_jw
    
    # PO reference (PO GRNs only)
    po = await db.purchase_orders.find_one({"id": grn.get("po_id")}, {"_id": 0}) if grn.get("po_id") else None
    grn["po"] = po
    
    # JW reference + DC — fetch JW order details
    jw_order = None
    dc = None
    if is_jw:
        jw_id = grn.get("jw_order_id") or grn.get("sc_order_id")
        jw_order = await db.subcontract_orders.find_one({"id": jw_id}, {"_id": 0}) if jw_id else None
        grn["jw_order"] = jw_order
        # Find the DC referenced by this GRN (or the latest sent DC for this SC)
        if grn.get("dc_id"):
            dc = await db.delivery_challans.find_one({"id": grn["dc_id"]}, {"_id": 0})
        elif jw_id:
            dc = await db.delivery_challans.find_one(
                {"subcontract_order_id": jw_id, "status": "sent"},
                {"_id": 0}, sort=[("created_at", -1)]
            )
        grn["dc"] = dc
    
    # Supplier — resolve from GRN, fall back to JW order
    supplier_id = grn.get("supplier_id") or (jw_order.get("supplier_id") if jw_order else None)
    grn["supplier"] = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0}) if supplier_id else None
    
    # Enrich lines with item info; for JW GRNs also surface the stored process_charges as
    # verified_price so the print template (which is written for PO GRNs) renders correctly.
    for line in grn.get("lines", []):
        item = await db.items.find_one({"id": line.get("item_id")}, {"_id": 0})
        line["item"] = item
        if is_jw:
            # JW GRN stores `process_charges` (the rate user entered at receiving) — expose it as
            # `verified_price` so the print template can compare it to the expected jw_rate.
            if "verified_price" not in line or not line.get("verified_price"):
                line["verified_price"] = line.get("process_charges", 0) or 0
            # Pull sent qty and rate from matching DC/SC line for this item.
            # Three strategies:
            #   1) Match by DC line item_id (works for processing-only JW where the same Part
            #      was sent and received)
            #   2) Match by SC-order lines item_id
            #   3) Match by SC job_work_parts (works for JW WITH RM where the line being received
            #      is the FG — FG is not in DC/SC.lines but IS in job_work_parts). Sent qty here
            #      means the qty we expected back; rate means the processing charges per unit.
            _sent_qty = 0
            _rate = 0
            if dc:
                for dl in (dc.get("lines") or []):
                    if dl.get("item_id") == line.get("item_id"):
                        _sent_qty = dl.get("quantity") or dl.get("sent_quantity") or 0
                        # Prefer processing_charges (per-unit vendor charge) over rate.
                        # rate on Job OS DC lines = RM cost per unit (BOM rollup) which is
                        # our internal accounting, not what we pay the vendor. GRN "Rate/Unit"
                        # should be the processing charges.
                        _rate = dl.get("processing_charges") or dl.get("rate") or 0
                        break
            if (_sent_qty == 0 or _rate == 0) and jw_order:
                for sl in (jw_order.get("lines") or []):
                    if sl.get("item_id") == line.get("item_id"):
                        if _sent_qty == 0:
                            _sent_qty = sl.get("sent_quantity") or sl.get("quantity") or 0
                        if _rate == 0:
                            _rate = sl.get("rate") or 0
                        break
            # Strategy 3 — JW with RM: FG item we receive is in job_work_parts
            if (_sent_qty == 0 or _rate == 0) and jw_order:
                for jwp in (jw_order.get("job_work_parts") or []):
                    if jwp.get("item_id") == line.get("item_id"):
                        if _sent_qty == 0:
                            _sent_qty = jwp.get("quantity") or 0
                        if _rate == 0:
                            _rate = jwp.get("charges") or 0
                        break
            line["jw_sent_quantity"] = _sent_qty
            line["jw_rate"] = _rate
    
    company = await db.company_settings.find_one({"type": "company"}, {"_id": 0})
    grn["company"] = company
    if grn.get("warehouse_id"):
        wh = await db.warehouses.find_one({"id": grn["warehouse_id"]}, {"_id": 0})
        grn["warehouse"] = wh
    return grn

# ================== WORK CENTER ROUTES ==================

@work_centers_router.get("")
async def get_work_centers(request: Request, status: Optional[str] = None):
    await get_current_user(request)
    query = {}
    if status:
        query["status"] = status
    work_centers = await db.work_centers.find(query, {"_id": 0}).to_list(100)
    return work_centers

@work_centers_router.get("/{wc_id}")
async def get_work_center(wc_id: str, request: Request):
    await get_current_user(request)
    wc = await db.work_centers.find_one({"id": wc_id}, {"_id": 0})
    if not wc:
        raise HTTPException(status_code=404, detail="Work center not found")
    return wc

@work_centers_router.post("", status_code=201)
async def create_work_center(wc_data: WorkCenterCreate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="manufacturing", action="create")
    existing = await db.work_centers.find_one({"code": wc_data.code})
    if existing:
        raise HTTPException(status_code=400, detail="Work center code already exists")
    
    wc_doc = {
        "id": str(uuid.uuid4()),
        **wc_data.model_dump(),
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"]
    }
    await db.work_centers.insert_one(wc_doc)
    wc_doc.pop("_id", None)
    return wc_doc

@work_centers_router.put("/{wc_id}")
async def update_work_center(wc_id: str, wc_data: WorkCenterUpdate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="manufacturing", action="edit")
    update_data = {k: v for k, v in wc_data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    update_data["updated_at"] = datetime.now(timezone.utc)
    result = await db.work_centers.update_one({"id": wc_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Work center not found")
    
    wc = await db.work_centers.find_one({"id": wc_id}, {"_id": 0})
    return wc

# ================== ROUTING ROUTES ==================

@routings_router.get("")
async def get_routings(request: Request, status: Optional[str] = None, item_id: Optional[str] = None):
    await get_current_user(request)
    query = {}
    if status:
        query["status"] = status
    if item_id:
        query["item_id"] = item_id
    
    routings = await db.routings.find(query, {"_id": 0}).to_list(1000)
    return routings

@routings_router.get("/{routing_id}")
async def get_routing(routing_id: str, request: Request):
    await get_current_user(request)
    routing = await db.routings.find_one({"id": routing_id}, {"_id": 0})
    if not routing:
        raise HTTPException(status_code=404, detail="Routing not found")
    return routing

@routings_router.post("", status_code=201)
async def create_routing(routing_data: RoutingCreate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="routings", action="create")
    routing_doc = {
        "id": str(uuid.uuid4()),
        "name": routing_data.name,
        "description": routing_data.description,
        "status": routing_data.status,
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"]
    }
    await db.routings.insert_one(routing_doc)
    routing_doc.pop("_id", None)
    return routing_doc

@routings_router.put("/{routing_id}")
async def update_routing(routing_id: str, routing_data: RoutingUpdate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="routings", action="edit")
    update_data = {}
    for k, v in routing_data.model_dump().items():
        if v is not None:
            update_data[k] = v
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    update_data["updated_at"] = datetime.now(timezone.utc)
    result = await db.routings.update_one({"id": routing_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Routing not found")
    
    routing = await db.routings.find_one({"id": routing_id}, {"_id": 0})
    return routing

# ================== WORK ORDER ROUTES ==================

async def recalculate_all_reservations():
    """Recalculate allocated_qty and shortfall_qty for ALL reserved MOs based on current stock.
    Uses FIFO (reserved_at order) to allocate stock fairly."""
    reserved_mos = await db.work_orders.find(
        {"materials_reserved": True, "status": {"$in": ["pending", "in_progress"]}},
        {"_id": 0}
    ).sort("reserved_at", 1).to_list(5000)  # FIFO by reservation time
    
    if not reserved_mos:
        return
    
    # Get current stock for all RM items across all reservations
    all_rm_ids = set()
    for mo in reserved_mos:
        for rm in mo.get("reserved_materials", []):
            all_rm_ids.add(rm.get("item_id"))
    
    stock_map = {}
    if all_rm_ids:
        async for item in db.items.find({"id": {"$in": list(all_rm_ids)}}, {"_id": 0, "id": 1, "current_stock": 1}):
            stock_map[item["id"]] = item.get("current_stock", 0)
    
    # Remaining pool per item (starts at current stock, decremented as we allocate FIFO)
    remaining = dict(stock_map)
    
    for mo in reserved_mos:
        updated_materials = []
        total_shortfall = 0
        changed = False
        for rm in mo.get("reserved_materials", []):
            rid = rm.get("item_id")
            needed = rm.get("quantity", 0)
            avail = max(0, remaining.get(rid, 0))
            new_alloc = min(avail, needed)
            new_shortfall = max(0, needed - new_alloc)
            remaining[rid] = remaining.get(rid, 0) - new_alloc
            
            if rm.get("allocated_qty") != new_alloc or rm.get("shortfall_qty") != new_shortfall:
                changed = True
            
            rm_copy = dict(rm)
            rm_copy["allocated_qty"] = new_alloc
            rm_copy["shortfall_qty"] = new_shortfall
            updated_materials.append(rm_copy)
            total_shortfall += new_shortfall
        
        if changed:
            await db.work_orders.update_one({"id": mo["id"]}, {"$set": {
                "reserved_materials": updated_materials,
                "reservation_shortfall": total_shortfall
            }})


@work_orders_router.get("")
async def get_work_orders(request: Request, status: Optional[str] = None, production_order_id: Optional[str] = None):
    await get_current_user(request)
    # Recalculate reservations based on current stock before returning
    await recalculate_all_reservations()
    query = {}
    if status:
        query["status"] = status
    if production_order_id:
        query["production_order_id"] = production_order_id
    
    work_orders = await db.work_orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Batch-fetch related docs to eliminate N+1 queries
    routing_ids = {wo.get("routing_id") for wo in work_orders if wo.get("routing_id")}
    item_ids = {wo.get("item_id") for wo in work_orders if wo.get("item_id")}
    po_ids = {wo.get("production_order_id") for wo in work_orders if wo.get("production_order_id")}
    
    routings_map = {}
    if routing_ids:
        async for r in db.routings.find({"id": {"$in": list(routing_ids)}}, {"_id": 0}):
            routings_map[r["id"]] = r
            if r.get("item_id"):
                item_ids.add(r["item_id"])
    
    items_map = {}
    if item_ids:
        async for it in db.items.find({"id": {"$in": list(item_ids)}}, {"_id": 0}):
            items_map[it["id"]] = it
    
    pos_map = {}
    if po_ids:
        async for po in db.production_orders.find({"id": {"$in": list(po_ids)}}, {"_id": 0}):
            pos_map[po["id"]] = po
    
    for wo in work_orders:
        routing = routings_map.get(wo.get("routing_id"))
        wo["routing"] = routing
        resolved_item_id = wo.get("item_id") or (routing.get("item_id") if routing else None)
        if resolved_item_id:
            wo["item"] = items_map.get(resolved_item_id)
        wo["production_order"] = pos_map.get(wo.get("production_order_id"))

    # Backfill `outsourced_quantity` for legacy OS ops where the field was
    # never persisted at SC-creation time (e.g., SCs created via older code
    # paths or admin tooling). Without this, the Job Card UI fails to show
    # the "Outsourced qty: x/y" hint below the vendor name. We do ONE batch
    # lookup of all referenced SC orders, then fill in qty from the matching
    # job_work_part. This keeps the response shape backwards-compatible while
    # guaranteeing the field is populated for the frontend.
    # Collect ALL live OS ops (not just those missing outsourced_quantity)
    # so we can also flag `dc_sent` on each OS run — used by the frontend
    # to hide the Revoke button once a DC has been dispatched (which would
    # otherwise leave the vendor expecting material but the SC erased).
    needs_sc_lookup = []
    all_os_ops = []
    for wo in work_orders:
        for op in (wo.get("operations_status") or []):
            if op.get("is_job_work") and op.get("outsource_sc_order_id"):
                all_os_ops.append((wo.get("id"), wo.get("item_id"), op))
                if not op.get("outsourced_quantity"):
                    needs_sc_lookup.append((wo.get("id"), wo.get("item_id"), op))
    if all_os_ops:
        sc_ids = list({op["outsource_sc_order_id"] for _wo_id, _item_id, op in all_os_ops})
        sc_map = {}
        async for sc in db.subcontract_orders.find({"id": {"$in": sc_ids}}, {"_id": 0, "id": 1, "job_work_parts": 1, "reference_wo_ids": 1, "reference_wo_id": 1, "reference_operation_seqs": 1, "lines": 1, "dc_created": 1}):
            sc_map[sc["id"]] = sc
        # Pre-compute which SCs have an active DC. A DC is "active" when
        # its status is sent/approved (i.e., the material physically left
        # the premises). Pending/draft DCs don't count — those can still
        # be cancelled and the revoke is safe.
        dc_active_scs = set()
        async for dc in db.delivery_challans.find(
            {"subcontract_order_id": {"$in": sc_ids}, "status": {"$in": ["sent", "approved", "received"]}},
            {"_id": 0, "subcontract_order_id": 1}
        ):
            dc_active_scs.add(dc.get("subcontract_order_id"))
        # Build a quick wo_id → wo.quantity lookup so we can fall back to
        # "entire WO outsourced" when no job_work_parts match.
        wo_qty_lookup = {wo.get("id"): float(wo.get("quantity") or 0) for wo in work_orders}
        # Flag dc_sent on each OS run + on op-level so the UI can hide
        # Revoke globally for that vendor allocation.
        for _wo_id, _item_id, op in all_os_ops:
            sc = sc_map.get(op.get("outsource_sc_order_id"))
            op_dc_sent = (sc and bool(sc.get("dc_created"))) or op.get("outsource_sc_order_id") in dc_active_scs
            op["dc_sent"] = op_dc_sent
            for r in (op.get("runs") or []):
                if not (r.get("operator") or "").startswith("OS: "):
                    continue
                run_sc_id = r.get("outsource_sc_order_id") or op.get("outsource_sc_order_id")
                if run_sc_id and (run_sc_id in dc_active_scs or (sc_map.get(run_sc_id) or {}).get("dc_created")):
                    r["dc_sent"] = True
                else:
                    r["dc_sent"] = False
        for _wo_id, item_id, op in needs_sc_lookup:
            sc = sc_map.get(op.get("outsource_sc_order_id"))
            if not sc:
                # Dangling reference — SC was deleted but the WO op still
                # carries `outsource_sc_order_id`. Fall back to wo.quantity
                # so the Job Card UI still shows the OS hint instead of
                # hiding the maroon "Outsourced qty: x/y" line entirely.
                fallback_qty = wo_qty_lookup.get(_wo_id, 0.0)
                if fallback_qty > 0:
                    op["outsourced_quantity"] = fallback_qty
                continue
            op_name = op.get("operation_name", "") or ""
            matched_qty = 0.0
            # Pass 1: surgical match on item + process + (optional) wo_id.
            # Some legacy SCs use `process_names` (plural list) instead of
            # `process_name` (singular string), so we accept both shapes.
            for jp in (sc.get("job_work_parts") or []):
                if jp.get("item_id") != item_id:
                    continue
                jp_processes = []
                if jp.get("process_name"):
                    jp_processes.append((jp.get("process_name") or "").strip())
                jp_processes.extend([(p or "").strip() for p in (jp.get("process_names") or [])])
                if jp_processes and op_name not in jp_processes:
                    continue
                if jp.get("wo_id") and jp.get("wo_id") != _wo_id:
                    continue
                matched_qty += float(jp.get("quantity") or 0)
            # Pass 2: any job_work_part referencing this WO (regardless of
            # process name — legacy SCs sometimes have a single SA line).
            if matched_qty == 0:
                for jp in (sc.get("job_work_parts") or []):
                    if jp.get("wo_id") == _wo_id:
                        matched_qty += float(jp.get("quantity") or 0)
            # Pass 3: SC's reference_wo_ids includes this WO and the SC has
            # exactly one reference — assume the entire WO qty was outsourced.
            # Also handles the legacy `reference_wo_id` (singular) field used
            # by older SCs that pre-date the multi-WO consolidation feature.
            if matched_qty == 0:
                ref_wos = list(sc.get("reference_wo_ids") or [])
                if sc.get("reference_wo_id"):
                    ref_wos.append(sc.get("reference_wo_id"))
                if _wo_id in ref_wos:
                    matched_qty = wo_qty_lookup.get(_wo_id, 0.0)
            # Pass 4 (last resort): SC is completely orphaned (no jwp, no
            # refs) but the WO operation still has is_job_work=True with
            # this SC linked. Fall back to wo.quantity so the Job Card UI
            # at least shows the OS hint. Better than silently hiding.
            if matched_qty == 0 and not (sc.get("job_work_parts") or []) \
                    and not (sc.get("reference_wo_ids") or []) \
                    and not sc.get("reference_wo_id"):
                matched_qty = wo_qty_lookup.get(_wo_id, 0.0)
            # Pass 5: SC exists with jwp but none match THIS wo_id (data
            # drift — SC was originally for a different WO but the current
            # op still points to it). Use wo.quantity as a sane fallback so
            # the operator at least sees an OS hint instead of nothing.
            if matched_qty == 0:
                matched_qty = wo_qty_lookup.get(_wo_id, 0.0)
            if matched_qty > 0:
                op["outsourced_quantity"] = matched_qty

        # Also backfill per-run SC info for legacy OS runs (created before
        # runs carried `outsource_sc_order_id`). Without this, the
        # per-vendor Revoke / Short Close buttons can't find the right SC
        # for runs created via the old code path.
        for _wo_id, item_id, op in needs_sc_lookup:
            for r in (op.get("runs") or []):
                if not (r.get("operator") or "").startswith("OS: "):
                    continue
                if r.get("outsource_sc_order_id"):
                    continue  # already populated
                r["outsource_sc_order_id"] = op.get("outsource_sc_order_id")
                r["outsource_sc_order_number"] = op.get("outsource_sc_order_number")
                r["outsource_supplier_name"] = r.get("outsource_supplier_name") or op.get("outsource_supplier_name") or (r.get("operator") or "").replace("OS: ", "", 1)
                r["outsource_supplier_id"] = r.get("outsource_supplier_id") or op.get("job_work_supplier_id")
                # Also backfill quantity_planned if missing (= op.outsourced_quantity)
                if not r.get("quantity_planned"):
                    r["quantity_planned"] = op.get("outsourced_quantity") or 0

    return work_orders

@work_orders_router.get("/{wo_id}")
async def get_work_order(wo_id: str, request: Request):
    await get_current_user(request)
    wo = await db.work_orders.find_one({"id": wo_id}, {"_id": 0})
    if not wo:
        raise HTTPException(status_code=404, detail="Work order not found")
    
    routing = await db.routings.find_one({"id": wo.get("routing_id")}, {"_id": 0})
    wo["routing"] = routing
    if routing:
        item = await db.items.find_one({"id": routing.get("item_id")}, {"_id": 0})
        wo["item"] = item
        # Enrich operations
        for op in routing.get("operations", []):
            wc = await db.work_centers.find_one({"id": op.get("work_center_id")}, {"_id": 0})
            op["work_center"] = wc
    
    return wo

@work_orders_router.post("", status_code=201)
async def create_work_order(wo_data: WorkOrderCreate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="manufacturing", action="create")

    # Resolve order type: MTS (build to stock — no SO) or MTO (against an SO line).
    order_type = (wo_data.order_type or "mto").lower()
    if order_type not in ("mts", "mto"):
        raise HTTPException(status_code=400, detail="order_type must be 'mts' or 'mto'")

    prod_order = None
    so_line = None
    bom = None
    if order_type == "mto":
        # MTO requires a Sales Order link.
        if not wo_data.production_order_id:
            raise HTTPException(status_code=400, detail="MTO Manufacturing Order requires a Sales Order")
        prod_order = await db.production_orders.find_one({"id": wo_data.production_order_id})
        if not prod_order:
            raise HTTPException(status_code=404, detail="Sales Order not found")
        # When the SO has multiple lines, the caller must pick which line this MO covers.
        so_lines = prod_order.get("lines") or []
        if wo_data.source_so_line_id:
            so_line = next((l for l in so_lines if l.get("line_id") == wo_data.source_so_line_id), None)
            if not so_line:
                raise HTTPException(status_code=404, detail="Selected SO line not found on the Sales Order")
            bom = await db.boms.find_one({"id": so_line.get("bom_id")})
        elif len(so_lines) == 1:
            so_line = so_lines[0]
            bom = await db.boms.find_one({"id": so_line.get("bom_id")})
        else:
            # Multi-line SO without a line selection — fall back to legacy top-level bom_id.
            bom = await db.boms.find_one({"id": prod_order.get("bom_id")})
        if not bom:
            raise HTTPException(status_code=404, detail="BOM not found for the selected Sales Order line")

        # Safeguard: an MTO MO cannot exceed the line's still-available quantity
        # (line.quantity - reserved_qty - already-created non-cancelled MO qty).
        if so_line:
            existing_mos = await db.work_orders.find(
                {
                    "production_order_id": wo_data.production_order_id,
                    "source_so_line_id": so_line.get("line_id", ""),
                    "parent_wo_id": None,
                    "status": {"$ne": "cancelled"},
                },
                {"quantity": 1, "_id": 0},
            ).to_list(1000)
            already_mo_qty = sum(int(m.get("quantity", 0) or 0) for m in existing_mos)
            ln_qty = int(so_line.get("quantity", 0) or 0)
            ln_resv = int(so_line.get("reserved_qty", 0) or 0)
            balance = ln_qty - ln_resv - already_mo_qty
            if int(wo_data.quantity or 0) > max(0, balance):
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"Cannot create MO for {wo_data.quantity} units — SO line balance is only {max(0, balance)} "
                        f"(line qty {ln_qty}, reserved {ln_resv}, MOs already created {already_mo_qty})."
                    ),
                )
    else:
        # MTS — pick the item directly. The item must have an active BOM (otherwise nothing
        # to manufacture). Production_order_id is left blank so MRP/SO flows ignore this MO.
        if not wo_data.item_id:
            raise HTTPException(status_code=400, detail="MTS Manufacturing Order requires an item")
        bom = await db.boms.find_one({"parent_item_id": wo_data.item_id, "status": "active"})
        if not bom:
            # Fall back to any BOM for the item (e.g. legacy 'draft' status).
            bom = await db.boms.find_one({"parent_item_id": wo_data.item_id})
        if not bom:
            raise HTTPException(status_code=404, detail="No BOM found for the selected item. Create a BOM before manufacturing.")

    routing = None
    if wo_data.routing_id:
        routing = await db.routings.find_one({"id": wo_data.routing_id})

    # Get the item from BOM's parent_item_id (the item to manufacture)
    item = await db.items.find_one({"id": bom.get("parent_item_id")})
    if not item:
        raise HTTPException(status_code=404, detail="Item not found for BOM")

    # ===== Phase 2: Resolve variant_selection for this MO =====
    # Priority: explicit payload > inherited from SO line (MTO only) > None.
    variant_sel_for_mo = wo_data.variant_selection or None
    if not variant_sel_for_mo and so_line:
        inherited = so_line.get("variant_selection")
        if inherited:
            variant_sel_for_mo = inherited
    # Validate every attribute in the selection is defined on the parent item.
    # Uses EFFECTIVE variants (own for CP/RM, inherited from BOM components for FG/SG)
    # so MOs for FG/SG with variants flowing up from BOM components are accepted.
    if variant_sel_for_mo:
        effective_attrs = await _get_effective_variants(item)
        defined_attrs = {a["name"]: set(v["value"] for v in (a.get("values") or [])) for a in effective_attrs if a.get("name")}
        for attr_name, attr_val in variant_sel_for_mo.items():
            if attr_name not in defined_attrs:
                raise HTTPException(status_code=400, detail=f"Variant attribute '{attr_name}' is not defined or inherited on this item")
            if defined_attrs[attr_name] and attr_val not in defined_attrs[attr_name]:
                raise HTTPException(status_code=400, detail=f"'{attr_val}' is not a valid value for '{attr_name}' (allowed: {sorted(defined_attrs[attr_name])})")
    
    created_work_orders = []
    
    # Helper function to create work order for an item
    async def create_wo_for_item(item_id: str, qty: int, parent_wo_id: str = None, is_main: bool = False, use_routing_id: str = None):
        # For main MO, use the routing passed in the request
        # For child MOs, try to find a routing by item_id (backward compatible with old routings)
        if use_routing_id:
            item_routing = await db.routings.find_one({"id": use_routing_id, "status": "active"})
        else:
            item_routing = await db.routings.find_one({"item_id": item_id, "status": "active"})
        
        if not item_routing:
            # Routing now lives inside BOM.parent_routings — allow MO creation when
            # the item's BOM has at least one routing entry.
            item_bom_check = await db.boms.find_one(
                {"parent_item_id": item_id, "status": "active"},
                {"_id": 0, "parent_routings": 1}
            )
            if item_bom_check and (item_bom_check.get("parent_routings") or []):
                item_routing = {"id": None, "name": "BOM Routing"}
            elif is_main:
                # Main MO requires a routing source — neither legacy collection nor BOM has one
                return None
            else:
                # For child MOs, create without routing (operations come from BOM)
                item_routing = {"id": None, "name": "No Routing"}
        
        item_doc = await db.items.find_one({"id": item_id})
        if not item_doc:
            return None
        
        # Main WO always gets created with full requested quantity (user explicitly chose to manufacture)
        # Child WOs also always get created for full BOM quantity (needed for this production run)
        qty_to_manufacture = qty
        
        # Generate WO number
        count = await db.work_orders.count_documents({})
        wo_number = f"MO-{str(count + 1).zfill(6)}"
        
        # Create operation statuses from BOM routings
        operations_status = []
        # Find routings for this item:
        # 1) If this is the main/parent item, use bom.parent_routings
        # 2) If this is a child, first check the CHILD's OWN BOM.parent_routings (iter 74+ source of truth),
        #    fall back to the parent BOM's component-level routings for legacy BOMs.
        item_routings_list = []
        if is_main:
            item_bom = await db.boms.find_one({"parent_item_id": item_id, "status": "active"}, {"_id": 0})
            item_routings_list = item_bom.get("parent_routings", []) if item_bom else []
        else:
            # Child MO — try the child's OWN BOM parent_routings first (unified source of truth)
            own_bom = await db.boms.find_one({"parent_item_id": item_id, "status": "active"}, {"_id": 0})
            if own_bom and own_bom.get("parent_routings"):
                item_routings_list = own_bom.get("parent_routings", [])
            if not item_routings_list:
                # Legacy fallback — parent BOM's component entry routings
                parent_bom = await db.boms.find_one({
                    "components.item_id": item_id,
                    "status": "active"
                }, {"_id": 0})
                if parent_bom:
                    for comp in parent_bom.get("components", []):
                        if comp.get("item_id") == item_id:
                            item_routings_list = comp.get("routings", [])
                            break
        
        for seq, r_entry in enumerate(normalize_routings(item_routings_list), 1):
            operations_status.append({
                "sequence": seq * 10,
                "operation_name": r_entry.get("name", ""),
                "process_cost_per_unit": float(r_entry.get("cost", 0) or 0),
                "work_center_id": "",  # Work centre decided at Job Card runtime
                "work_center_name": "",
                "is_job_work": False,
                "job_work_supplier_id": "",
                "status": "pending",
                "quantity_completed": 0
            })
        
        wo_doc = {
            "id": str(uuid.uuid4()),
            "wo_number": wo_number,
            "production_order_id": wo_data.production_order_id or "",
            "source_so_line_id": wo_data.source_so_line_id or "",
            "order_type": order_type if is_main else "mts",  # children of MTO parents are still MTS-style (build sub-items)
            "routing_id": item_routing.get("id"),
            "item_id": item_id,
            "quantity": qty_to_manufacture,
            "quantity_completed": 0,
            "due_date": wo_data.due_date if is_main else None,
            "scheduled_start": wo_data.scheduled_start,
            "scheduled_end": wo_data.scheduled_end,
            "status": "pending",
            "operations_status": operations_status,
            "parent_wo_id": parent_wo_id,
            "is_subcontract": wo_data.is_subcontract if is_main else False,
            "subcontract_supplier_id": wo_data.subcontract_supplier_id if is_main else "",
            "subcontract_type": wo_data.subcontract_type if is_main else "with_material",
            # Filter parent's variant_selection to ONLY the axes that this
            # WO's BOM tree actually contains. A child SG with no
            # variant-bearing leaves gets variant_selection=None (runs plain);
            # a child SG that has a variant-bearing leaf gets only the
            # matching axes. This implements the user's contextual variant
            # propagation rule: variants follow the BOM tree, not the parent MO.
            "variant_selection": (
                (await _filter_variant_selection_for_item(item_id, variant_sel_for_mo))
                if variant_sel_for_mo else None
            ),
            "variant_sku": (_build_variant_sku(item_doc.get("part_number") or "", variant_sel_for_mo) if (is_main and variant_sel_for_mo) else None),
            "notes": wo_data.notes if is_main else f"Auto-created for {item_doc.get('category', 'child item')}",
            "materials_consumed": False,
            "created_at": datetime.now(timezone.utc),
            "created_by": user["id"]
        }
        await db.work_orders.insert_one(wo_doc)
        wo_doc.pop("_id", None)
        
        logger.info(f"Created WO {wo_number} for {item_doc.get('part_number')} (qty={qty_to_manufacture}, parent={parent_wo_id})")
        return wo_doc
    
    # Helper function to recursively create work orders for child items
    async def create_child_work_orders(parent_item_id: str, parent_qty: int, parent_wo_id: str):
        # Find BOM for this item
        bom = await db.boms.find_one({"parent_item_id": parent_item_id, "status": "active"})
        if not bom:
            return

        # Phase 2 — filter components by variant_selection on the top-level MO.
        # variant_selection is captured below at wo_doc time; here we use the helper.
        applicable_components = _filter_components_by_variant(bom.get("components", []) or [], variant_sel_for_mo)

        for component in applicable_components:
            if component.get("is_alternate"):
                continue  # Skip alternate components
            
            child_item_id = component.get("item_id")
            child_qty = int(component.get("quantity", 1) * parent_qty)
            
            child_item = await db.items.find_one({"id": child_item_id})
            if not child_item:
                continue
            
            # Calculate WIP stock: stock produced by child MOs of active (non-cancelled) parent MOs
            # This stock is reserved for those parents and should NOT be available for new MOs
            wip_qty = 0
            active_child_mos = await db.work_orders.find({
                "item_id": child_item_id,
                "status": "completed",
                "parent_wo_id": {"$exists": True, "$ne": None}
            }, {"_id": 0, "parent_wo_id": 1, "quantity": 1, "quantity_completed": 1}).to_list(5000)
            for child_mo in active_child_mos:
                # Check if parent MO is still active (not completed, not cancelled)
                parent_mo = await db.work_orders.find_one(
                    {"id": child_mo["parent_wo_id"]},
                    {"_id": 0, "status": 1}
                )
                if parent_mo and parent_mo.get("status") not in ["completed", "cancelled"]:
                    wip_qty += child_mo.get("quantity_completed", child_mo.get("quantity", 0))
            
            current_stock = child_item.get("current_stock", 0)
            already_reserved = int(child_item.get("reserved_stock", 0) or 0)
            # free_stock = physical stock minus WIP committed to other active
            # parent MOs minus already-reserved-for-other-pending-MOs. This
            # is the qty we can safely consider available for the CURRENT parent MO.
            free_stock = max(0, current_stock - wip_qty - already_reserved)
            
            if free_stock >= child_qty:
                logger.info(f"Skipping child MO for {child_item.get('part_number')} — free stock {free_stock} (total {current_stock}, WIP {wip_qty}, reserved {already_reserved}) >= required {child_qty}")
                # Auto-reserve the in-stock child qty against the parent MO (released on parent cancel).
                await db.items.update_one(
                    {"id": child_item_id},
                    {"$inc": {"reserved_stock": child_qty}},
                )
                await db.work_orders.update_one(
                    {"id": parent_wo_id},
                    {"$push": {"child_reservations": {"item_id": child_item_id, "qty": child_qty}}},
                )
                continue
            
            # Create MO only for shortage qty (required - free stock).
            # Auto-reserve the in-stock portion (free_stock) so it's locked for this MO.
            shortage_qty = child_qty - int(free_stock)
            if free_stock > 0:
                await db.items.update_one(
                    {"id": child_item_id},
                    {"$inc": {"reserved_stock": int(free_stock)}},
                )
                await db.work_orders.update_one(
                    {"id": parent_wo_id},
                    {"$push": {"child_reservations": {"item_id": child_item_id, "qty": int(free_stock)}}},
                )
            
            # Create work orders for any manufacturable child — i.e., any child
            # that has its own BOM. Previously we additionally required a
            # routing source (legacy `routings` doc OR BOM `parent_routings`),
            # which silently dropped parts/SGs that didn't yet have a routing.
            # Per user request, MOs are now created for parts/SGs regardless
            # of routing; routing-less MOs go straight from Start → consume
            # RM → show the Complete button (no Job Card / operations).
            child_routing = await db.routings.find_one({"item_id": child_item_id, "status": "active"})
            child_own_bom = await db.boms.find_one(
                {"parent_item_id": child_item_id, "status": "active"},
                {"_id": 0, "parent_routings": 1, "components": 1}
            )
            is_manufacturable = bool(child_own_bom)
            has_routing_source = bool(child_routing) or bool(child_own_bom and child_own_bom.get("parent_routings"))
            if is_manufacturable or has_routing_source:
                child_wo = await create_wo_for_item(child_item_id, shortage_qty, parent_wo_id)
                if child_wo:
                    created_work_orders.append(child_wo)
                    # Recursively create work orders for this child's children —
                    # pass `shortage_qty`, NOT the full `child_qty`. Reason: the
                    # in-stock portion of THIS SG already physically exists, so
                    # we only need to manufacture sub-parts/sub-SGs to cover the
                    # shortage. Previously passing `child_qty` over-produced
                    # downstream MOs (the user-reported "MO for child SG & Parts
                    # even though parent SG is in stock" symptom — the parent
                    # was partially in stock, so the parent's own MO was
                    # correctly sized to the shortage, but its sub-children
                    # were being exploded against full required qty).
                    await create_child_work_orders(child_item_id, shortage_qty, child_wo["id"])
    
    # For all cases (including subcontract): create MO first, SC order created at "Start SC"
    # Create main work order - use item from BOM (not routing)
    main_item_id = bom.get("parent_item_id")
    main_wo = await create_wo_for_item(main_item_id, wo_data.quantity, None, is_main=True, use_routing_id=wo_data.routing_id)
    if main_wo:
        created_work_orders.insert(0, main_wo)
        # Create work orders for child items
        await create_child_work_orders(main_item_id, wo_data.quantity, main_wo["id"])
    
    # SC order is NOT auto-created here — user clicks "Start SC" / "Create SC" button which calls /create-sc endpoint
    return {"message": f"Created {len(created_work_orders)} work order(s)", "work_orders": created_work_orders}


@work_orders_router.post("/bulk-subcontract")
async def bulk_subcontract(request: Request, data: dict = Body(...)):
    """Mark multiple MOs as SC and create a single consolidated SC Order + DC"""
    user = await get_current_user(request)
    wo_ids = data.get("wo_ids", [])
    supplier_id = data.get("supplier_id")
    sc_type = data.get("subcontract_type", "with_material")
    
    if not wo_ids or not supplier_id:
        raise HTTPException(status_code=400, detail="wo_ids and supplier_id required")
    
    supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    # Mark all MOs as subcontract
    all_sc_lines = []
    all_job_work_parts = []
    mo_numbers = []
    for wo_id in wo_ids:
        wo = await db.work_orders.find_one({"id": wo_id})
        if not wo:
            continue
        
        await db.work_orders.update_one({"id": wo_id}, {"$set": {
            "is_subcontract": True,
            "subcontract_supplier_id": supplier_id,
            "subcontract_type": sc_type,
            "updated_at": datetime.now(timezone.utc)
        }})
        mo_numbers.append(wo.get("wo_number", wo_id))
        
        # Collect materials for SC order
        routing = await db.routings.find_one({"id": wo.get("routing_id")})
        item_id = routing.get("item_id") if routing else wo.get("item_id")
        
        # Add to job work parts (the FG/SA/Part being processed). Charges/pc = item's own BOM parent routing cost.
        _fg_charge = 0
        try:
            _fg_charge = await compute_bom_fg_process_only(item_id)
        except Exception:
            _fg_charge = 0
        all_job_work_parts.append({"item_id": item_id, "quantity": wo.get("quantity", 1), "charges": _fg_charge, "wo_id": wo_id})
        
        if sc_type == "without_material":
            wo_item = await db.items.find_one({"id": item_id}, {"_id": 0})
            if wo_item:
                # SC without RM: Part/SA rate = BOM material cost (RM only), process cost lives on job_work_parts.charges
                _no_rm_costs = await compute_bom_costs(item_id)
                _rate_no_rm = _no_rm_costs.get("rm_cost", 0) or wo_item.get("unit_cost", 0)
                all_sc_lines.append({"item_id": item_id, "quantity": wo.get("quantity", 1), "sent_quantity": 0, "received_quantity": 0, "rate": round(_rate_no_rm, 2), "wo_id": wo_id})
        else:
            consumed = wo.get("consumed_materials", [])
            if consumed:
                for m in consumed:
                    all_sc_lines.append({"item_id": m["item_id"], "quantity": m["quantity"], "sent_quantity": m["quantity"], "received_quantity": 0, "rate": m.get("unit_cost", 0), "wo_id": wo_id})
            else:
                # Smart resolve: walk up to root MO, collect completed items from entire tree
                root_id = wo_id
                cur = wo
                while cur.get("parent_wo_id"):
                    p = await db.work_orders.find_one({"id": cur["parent_wo_id"]})
                    if not p:
                        break
                    root_id = p["id"]
                    cur = p
                
                async def collect_bulk_tree(pid):
                    mos = []
                    ch = await db.work_orders.find({"parent_wo_id": pid}, {"_id": 0}).to_list(500)
                    for c in ch:
                        mos.append(c)
                        mos.extend(await collect_bulk_tree(c["id"]))
                    return mos
                
                tree_mos = await collect_bulk_tree(root_id)
                completed_items = {cm["item_id"] for cm in tree_mos if cm.get("status") == "completed"}
                
                async def bulk_smart_resolve(pid, mult, visited=None):
                    if visited is None:
                        visited = set()
                    if pid in visited:
                        return []
                    visited.add(pid)
                    res = []
                    b = await db.boms.find_one({"parent_item_id": pid, "status": "active"})
                    if not b:
                        return res
                    for c in b.get("components", []):
                        if c.get("is_alternate"):
                            continue
                        ci = await db.items.find_one({"id": c["item_id"]}, {"_id": 0})
                        if not ci:
                            continue
                        cq = c.get("quantity", 1) * mult
                        cat = ci.get("category", "")
                        if cat == "raw_material":
                            # RM rate = item.unit_cost (material rate per unit)
                            res.append({"item_id": c["item_id"], "quantity": int(cq), "sent_quantity": int(cq), "received_quantity": 0, "rate": ci.get("unit_cost", 0), "wo_id": wo_id})
                        elif cat in ["component", "sub_assembly"]:
                            if c["item_id"] in completed_items:
                                # Completed Part/SA — send as BOM Total/Unit (material + all process rolled up)
                                _rate = await compute_bom_total_unit_cost(c["item_id"]) or ci.get("unit_cost", 0)
                                res.append({"item_id": c["item_id"], "quantity": int(cq), "sent_quantity": int(cq), "received_quantity": 0, "rate": round(_rate, 2), "wo_id": wo_id})
                            else:
                                child_rm = await bulk_smart_resolve(c["item_id"], cq, visited)
                                if child_rm:
                                    res.extend(child_rm)
                                else:
                                    # Not yet produced and no BOM to explode — send as BOM Total/Unit too
                                    _rate = await compute_bom_total_unit_cost(c["item_id"]) or ci.get("unit_cost", 0)
                                    res.append({"item_id": c["item_id"], "quantity": int(cq), "sent_quantity": int(cq), "received_quantity": 0, "rate": round(_rate, 2), "wo_id": wo_id})
                    return res
                
                resolved = await bulk_smart_resolve(item_id, wo.get("quantity", 1))
                if resolved:
                    all_sc_lines.extend(resolved)
                else:
                    wo_item = await db.items.find_one({"id": item_id}, {"_id": 0})
                    if wo_item:
                        all_sc_lines.append({"item_id": item_id, "quantity": wo.get("quantity", 1), "sent_quantity": wo.get("quantity", 1), "received_quantity": 0, "rate": wo_item.get("unit_cost", 0), "wo_id": wo_id})
    
    if not all_sc_lines:
        raise HTTPException(status_code=400, detail="No materials to create SC order")
    
    # Consolidate lines (merge same item_id)
    consolidated = {}
    for line in all_sc_lines:
        iid = line["item_id"]
        if iid in consolidated:
            consolidated[iid]["quantity"] += line["quantity"]
            consolidated[iid]["sent_quantity"] += line["sent_quantity"]
        else:
            consolidated[iid] = {k: v for k, v in line.items() if k != "wo_id"}
    
    sc_lines = list(consolidated.values())
    
    # Get FG item name from first MO
    first_wo = await db.work_orders.find_one({"id": wo_ids[0]})
    first_routing = await db.routings.find_one({"id": first_wo.get("routing_id")}) if first_wo else None
    fg_item_id = first_routing.get("item_id") if first_routing else (first_wo.get("item_id") if first_wo else "")
    fg_item = await db.items.find_one({"id": fg_item_id}, {"_id": 0}) if fg_item_id else None
    
    # Create single SC order
    sc_count = await db.subcontract_orders.count_documents({})
    sc_doc = {
        "id": str(uuid.uuid4()),
        "order_number": f"JW-{str(sc_count + 1).zfill(6)}",
        "supplier_id": supplier_id,
        "reference_wo_ids": wo_ids,
        "reference_wo_id": wo_ids[0],
        "subcontract_type": sc_type,
        "fg_item_id": fg_item_id,
        "fg_item_name": f"{fg_item.get('part_number', '')} - {fg_item.get('name', '')}" if fg_item else "",
        "fg_quantity": sum(line["quantity"] for line in sc_lines),
        "job_work_parts": [{k: v for k, v in p.items() if k != "wo_id"} for p in all_job_work_parts],
        "lines": sc_lines,
        "status": "in_progress",
        "notes": f"Bulk SC for MOs: {', '.join(mo_numbers)} ({sc_type.replace('_', ' ')})",
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"]
    }
    await db.subcontract_orders.insert_one(sc_doc)
    sc_doc.pop("_id", None)
    
    # Create DC if with_material
    dc_doc = None
    if sc_type == "with_material" and sc_lines:
        dc_lines = [{"item_id": l["item_id"], "quantity": l["quantity"], "rate": l["rate"]} for l in sc_lines]
        dc_doc = {
            "id": str(uuid.uuid4()),
            "dc_number": await get_next_series_number("delivery_challan"),
            "subcontract_order_id": sc_doc["id"],
            "lines": dc_lines,
            "status": "draft",
            "notes": f"Bulk DC for MOs: {', '.join(mo_numbers)}",
            "created_at": datetime.now(timezone.utc),
            "created_by": user["id"]
        }
        await db.delivery_challans.insert_one(dc_doc)
        dc_doc.pop("_id", None)
    
    return {
        "message": f"Created SC Order {sc_doc['order_number']}" + (f" + DC {dc_doc['dc_number']}" if dc_doc else "") + f" for {len(wo_ids)} MOs",
        "sc_order": sc_doc,
        "dc": dc_doc
    }


@work_orders_router.post("/{wo_id}/release")
async def release_work_order(wo_id: str, request: Request):
    """Release an MO — transitions status from 'pending' to 'released' and
    increments `reserved_stock` on every required child component by the
    full required quantity (one BOM-explosion level).

    This is the explicit user action that commits child stock to this MO.
    Subsequent /start (consume materials) decrements both current_stock AND
    the reservation booked here.

    On MO cancel, the recorded `child_reservations` are released.
    """
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="manufacturing", action="edit")
    wo = await db.work_orders.find_one({"id": wo_id})
    if not wo:
        raise HTTPException(status_code=404, detail="Manufacturing order not found")
    if wo.get("status") not in ("pending",):
        raise HTTPException(status_code=400, detail=f"Can only release MOs in 'pending' status (current: {wo.get('status')})")
    if wo.get("child_reservations"):
        raise HTTPException(status_code=400, detail="MO already has child reservations — already released")

    item_id = wo.get("item_id")
    qty = int(wo.get("quantity", 0) or 0)
    if not item_id or qty <= 0:
        raise HTTPException(status_code=400, detail="MO has no item or zero qty — cannot release")

    bom = await db.boms.find_one({"parent_item_id": item_id, "status": "active"}, {"_id": 0})
    if not bom:
        bom = await db.boms.find_one({"parent_item_id": item_id}, {"_id": 0})
    if not bom:
        raise HTTPException(status_code=404, detail="No BOM found for this MO's item")

    reservations = []
    for component in bom.get("components", []) or []:
        if component.get("is_alternate"):
            continue
        child_item_id = component.get("item_id")
        if not child_item_id:
            continue
        per_unit = float(component.get("quantity", 0) or 0)
        if per_unit <= 0:
            continue
        required_qty = int(per_unit * qty)
        if required_qty <= 0:
            continue
        # Increment reserved_stock on the child item.
        await db.items.update_one(
            {"id": child_item_id},
            {"$inc": {"reserved_stock": required_qty}},
        )
        reservations.append({"item_id": child_item_id, "qty": required_qty})

    await db.work_orders.update_one(
        {"id": wo_id},
        {"$set": {
            "status": "released",
            "released_at": datetime.now(timezone.utc),
            "released_by": user["id"],
            "child_reservations": reservations,
        }},
    )
    return {
        "message": f"Released MO {wo.get('wo_number')} — reserved {len(reservations)} child component(s).",
        "reservations": reservations,
    }


@work_orders_router.get("/{wo_id}/material-requirements")
async def get_wo_material_requirements(wo_id: str, request: Request):
    """Read-only material requirement list for a single MO.

    Returns the IMMEDIATE components from the MO item's own active BOM
    (single level, no recursion). For an FG MO, that's the FG's direct
    components (typically SAs + parts). For an SG/Part MO, that's the
    SG/Part's direct components (typically raw materials). Mirrors the
    `consumed_materials` shape so the frontend can reuse the same renderer
    and PDF template.

    Notes:
      - Excludes alternate components.
      - Deduplicates same item that may appear twice in the same BOM
        (sums quantity).
    """
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager", "inventory_manager", "quality_inspector"], module="manufacturing", action="view")
    wo = await db.work_orders.find_one({"id": wo_id}, {"_id": 0})
    if not wo:
        raise HTTPException(status_code=404, detail="Work order not found")
    routing = await db.routings.find_one({"id": wo.get("routing_id")}, {"_id": 0}) if wo.get("routing_id") else None
    fg_item_id = (routing or {}).get("item_id") or wo.get("item_id")
    if not fg_item_id:
        return {"wo_number": wo.get("wo_number"), "materials": []}
    wo_qty = float(wo.get("quantity") or 1)

    # Top-level only — fetch the ACTIVE BOM for this MO's item and emit one
    # row per non-alternate component. No recursion into child BOMs.
    bom = await db.boms.find_one({"parent_item_id": fg_item_id, "status": "active"}, {"_id": 0})
    collected = []
    # Pre-load UOM master so we can attach `decimal_places` to each row.
    # Frontend uses this to round/format the `Required` column per the UOM
    # master setting (e.g., kgs = 2 dp, pcs = 0 dp). Falls back to 2 dp.
    uoms_list = await db.uoms.find({}, {"_id": 0}).to_list(500)
    uom_decimal_map = {}
    for u in uoms_list:
        code = (u.get("code") or "").strip()
        if code:
            try:
                uom_decimal_map[code.lower()] = int(u.get("decimal_places") if u.get("decimal_places") is not None else 2)
            except (ValueError, TypeError):
                uom_decimal_map[code.lower()] = 2
    if bom:
        # Pre-compute allocations across all open MOs so we can show
        # planners the FREE stock (current_stock − allocated_by_others).
        # Read once for the whole set of components.
        open_mos = await db.work_orders.find(
            {"materials_reserved": True, "status": {"$in": ["pending", "in_progress"]}, "id": {"$ne": wo_id}},
            {"_id": 0, "reserved_materials": 1},
        ).to_list(5000)
        allocated_by_others: dict = {}
        for mo in open_mos:
            for rm in (mo.get("reserved_materials") or []):
                rid = rm.get("item_id")
                if rid:
                    allocated_by_others[rid] = allocated_by_others.get(rid, 0) + (rm.get("allocated_qty") or 0)
        # Materials this MO has ALREADY consumed (from WO /start path).
        # Without this, partially-completed MOs kept showing the full BOM
        # requirement even after their first operation had consumed the
        # parts — confusing planners and double-counting in MRP.
        consumed_by_item: dict = {}
        for cm in (wo.get("consumed_materials") or []):
            cid_c = cm.get("item_id")
            if cid_c:
                consumed_by_item[cid_c] = consumed_by_item.get(cid_c, 0) + float(cm.get("quantity") or 0)
        for comp in bom.get("components", []) or []:
            if comp.get("is_alternate"):
                continue
            cid = comp.get("item_id")
            if not cid:
                continue
            comp_qty = float(comp.get("quantity") or 0) * wo_qty
            consumed_qty = float(consumed_by_item.get(cid, 0))
            outstanding = max(0.0, comp_qty - consumed_qty)
            citem = await db.items.find_one({"id": cid}, {"_id": 0})
            if not citem:
                continue
            on_hand = float(citem.get("current_stock") or 0)
            allocated = float(allocated_by_others.get(cid, 0))
            available = max(0.0, on_hand - allocated)
            # Shortage now reflects only what's STILL OUTSTANDING — what
            # we've already consumed is no longer "required".
            shortage = max(0.0, outstanding - available)
            uom_code = (citem.get("unit_of_measure") or citem.get("uom") or "pcs")
            collected.append({
                "item_id": cid,
                "item": citem.get("part_number", ""),
                "name": citem.get("name", ""),
                "category": citem.get("category", ""),
                "quantity": comp_qty,
                "consumed_qty": consumed_qty,
                "outstanding_qty": outstanding,
                "uom": uom_code,
                "uom_decimal_places": uom_decimal_map.get((uom_code or "").strip().lower(), 2),
                "unit_cost": float(citem.get("purchase_price") or 0),
                "available_stock": available,
                "shortage": shortage,
            })

    # Consolidate duplicates (same item listed twice in the BOM) while
    # preserving original order for a stable PDF.
    seen_order = []
    bucket = {}
    for r in collected:
        key = r["item_id"]
        if key not in bucket:
            bucket[key] = dict(r)
            seen_order.append(key)
        else:
            bucket[key]["quantity"] += r["quantity"]
            # consumed/outstanding columns must also be summed when the same
            # component appears twice in the same BOM (legacy multi-line).
            bucket[key]["consumed_qty"] = bucket[key].get("consumed_qty", 0) + r.get("consumed_qty", 0)
            bucket[key]["outstanding_qty"] = bucket[key].get("outstanding_qty", 0) + r.get("outstanding_qty", 0)
            bucket[key]["shortage"] = max(0.0, bucket[key]["outstanding_qty"] - bucket[key]["available_stock"])
    materials = [bucket[k] for k in seen_order]
    return {
        "wo_number": wo.get("wo_number"),
        "item": {
            "part_number": (await db.items.find_one({"id": fg_item_id}, {"_id": 0, "part_number": 1, "name": 1})) or {},
        },
        "quantity": wo_qty,
        "materials": materials,
    }


@work_orders_router.post("/{wo_id}/reconcile-consumption")
async def reconcile_wo_consumption(wo_id: str, request: Request):
    """Heal the decimal-truncation bug on already-started MOs.

    Compares each BOM component's expected consumption (`bom.qty × wo.qty`)
    against what's actually stored in `consumed_materials`. For every
    component whose stored qty is short:
      • Issues the delta from stock (creates a stock_movement audit row).
      • Updates `consumed_materials` so the new total matches expected.
      • Decrements the item's `current_stock`.

    Safe-by-design:
      • Only runs against the FG/SG MO's *direct* BOM (no recursion).
      • Skips components that aren't short.
      • If stock is insufficient for the delta, that component is reported
        in the response but NOT touched (caller can issue the missing qty
        via Inventory → Adjustment first, then re-run reconciliation).
      • Idempotent — calling twice on a healed MO is a no-op.
    """
    user = await get_current_user(request)
    role = user.get("role")
    perms = ((user.get("permissions") or {}).get("manufacturing") or [])
    if role != "admin" and "edit" not in perms:
        raise HTTPException(status_code=403, detail="You don't have permission to reconcile consumption")

    wo = await db.work_orders.find_one({"id": wo_id}, {"_id": 0})
    if not wo:
        raise HTTPException(status_code=404, detail="Work order not found")

    fg_item_id = wo.get("item_id")
    if not fg_item_id:
        raise HTTPException(status_code=400, detail="MO has no item_id")

    bom = await db.boms.find_one({"parent_item_id": fg_item_id, "status": "active"}, {"_id": 0})
    if not bom:
        raise HTTPException(status_code=400, detail="No active BOM for this MO's item")

    wo_qty = float(wo.get("quantity") or 0)
    consumed_now: dict = {}
    for cm in (wo.get("consumed_materials") or []):
        cid_c = cm.get("item_id")
        if cid_c:
            consumed_now[cid_c] = consumed_now.get(cid_c, 0) + float(cm.get("quantity") or 0)

    healed = []           # components we successfully topped-up
    skipped_no_stock = []  # components we couldn't heal due to insufficient stock
    now = datetime.now(timezone.utc)
    new_consumed_entries = []

    for comp in bom.get("components", []) or []:
        if comp.get("is_alternate"):
            continue
        cid = comp.get("item_id")
        if not cid:
            continue
        expected = float(comp.get("quantity") or 0) * wo_qty
        already = float(consumed_now.get(cid, 0))
        delta = round(expected - already, 6)
        if delta <= 0:
            continue  # nothing to heal

        citem = await db.items.find_one({"id": cid})
        if not citem:
            continue
        on_hand = float(citem.get("current_stock") or 0)
        if on_hand < delta:
            skipped_no_stock.append({
                "item_id": cid,
                "part_number": citem.get("part_number", ""),
                "name": citem.get("name", ""),
                "delta_required": delta,
                "available": on_hand,
            })
            continue

        # Decrement stock + create audit movement.
        new_stock = on_hand - delta
        await db.items.update_one({"id": cid}, {"$set": {"current_stock": new_stock}})
        await db.stock_movements.insert_one({
            "id": str(uuid.uuid4()),
            "item_id": cid,
            "transaction_type": "issue",
            "quantity": delta,
            "reference_type": "work_order",
            "reference_id": wo_id,
            "reference_number": wo.get("wo_number", ""),
            "notes": f"Reconciliation top-up for {wo.get('wo_number', '')} (decimal truncation healing)",
            "performed_by": user.get("id"),
            "created_at": now,
        })
        new_consumed_entries.append({
            "item_id": cid,
            "item_code": citem.get("part_number", ""),
            "item_name": citem.get("name", ""),
            "quantity": delta,
            "uom": citem.get("unit_of_measure") or citem.get("uom") or "pcs",
            "unit_cost": float(citem.get("purchase_price") or 0),
            "consumed_at": now,
            "reconciliation": True,
        })
        healed.append({
            "item_id": cid,
            "part_number": citem.get("part_number", ""),
            "name": citem.get("name", ""),
            "delta_consumed": delta,
            "new_total_consumed": already + delta,
        })

    if new_consumed_entries:
        # Append the top-up entries instead of overwriting the existing
        # consumed_materials list — keeps the original issue history intact
        # while the rolled-up total now matches the BOM expectation.
        await db.work_orders.update_one(
            {"id": wo_id},
            {"$push": {"consumed_materials": {"$each": new_consumed_entries}}},
        )

    return {
        "ok": True,
        "wo_id": wo_id,
        "wo_number": wo.get("wo_number"),
        "healed_components": healed,
        "skipped_due_to_stock": skipped_no_stock,
        "healed_count": len(healed),
    }


class RevokeOpPayload(BaseModel):
    run_number: Optional[int] = None  # When provided, revoke ONLY that vendor's run/SC line
    reason: Optional[str] = ""



async def _recompute_wo_status_after_op_change(wo_id: str) -> Optional[str]:
    """Re-evaluate a Work Order's overall status after one of its operations
    has been mutated. Used by every endpoint that flips an op status (short-
    close-no-grn op-level, per-vendor short-close, etc.) so a completed last
    operation always promotes the WO to 'completed' instead of leaving it
    stuck on 'in_progress'.

    Rules (mirrors the inline logic in PUT /work-orders/{wo}/operations/{seq}):
      - All ops completed AND no blockers (subcontract MO not yet GRN-ed, or
        any op still 'sent' to a job-worker) → wo.status = 'completed'.
      - At least one op in progress/completed/stopped → 'in_progress'.
      - Otherwise → leave status as-is.

    Returns the new status (or None if no change). Also flips
    `actual_end` and triggers FG stock + cost roll-up when the WO transitions
    to completed for the first time.
    """
    wo = await db.work_orders.find_one({"id": wo_id})
    if not wo:
        return None
    operations = wo.get("operations_status") or []
    if not operations:
        return None
    current = wo.get("status")

    # Auto-heal stuck outsourced ops — if a Job Work part covering THIS
    # WO has already been fully received via GRN (received_quantity >=
    # quantity) BUT the corresponding op is still showing `outsource_status
    # == 'sent'` and status != completed, flip it now. This is the main
    # safety net for legacy MOs that got stuck because the GRN posting
    # path didn't propagate completion (consolidated SCs, double-matching
    # parts etc.). Mutated in-memory so the rest of this function sees the
    # healed state; persisted at the end if anything changed.
    op_mutated = False
    for op in operations:
        if not op.get("is_job_work"):
            continue
        if op.get("status") == "completed" or op.get("short_closed"):
            continue
        sc_id = op.get("outsource_sc_order_id")
        if not sc_id:
            continue
        sc = await db.subcontract_orders.find_one({"id": sc_id}, {"_id": 0, "job_work_parts": 1})
        if not sc:
            continue
        # Total received qty for THIS WO from matching parts. Prefer parts
        # tagged with `wo_id == wo_id`; fall back to item-id match for
        # legacy SCs that never recorded wo_id on the part.
        mo_qty = float(wo.get("quantity") or 0)
        received_for_wo = 0.0
        any_wo_tagged = any(p.get("wo_id") == wo_id for p in sc.get("job_work_parts", []))
        for p in sc.get("job_work_parts", []):
            if any_wo_tagged:
                if p.get("wo_id") == wo_id:
                    received_for_wo += float(p.get("received_quantity") or 0)
            elif p.get("item_id") == wo.get("item_id"):
                received_for_wo += float(p.get("received_quantity") or 0)
        if mo_qty > 0 and received_for_wo >= mo_qty:
            op["status"] = "completed"
            op["outsource_status"] = "received"
            op["actual_end"] = datetime.now(timezone.utc)
            op["quantity_completed"] = max(float(op.get("quantity_completed") or 0), mo_qty)
            op["quantity_accepted"] = max(float(op.get("quantity_accepted") or 0), mo_qty)
            op_mutated = True

    # An op is "effectively completed" if its status is `completed` OR it
    # was short-closed (short_closed=True). Legacy short-close paths used
    # to leave the op status as 'in_progress' while flipping short_closed,
    # so we treat both as terminal here when deciding the MO's overall
    # status.
    def _op_done(op):
        return op.get("status") == "completed" or op.get("short_closed") is True
    all_completed = all(_op_done(op) for op in operations)
    any_active = any(
        op.get("status") in ("in_progress", "completed", "stopped") or op.get("short_closed") is True
        for op in operations
    )

    new_status = current
    if all_completed:
        can_complete = True
        if wo.get("is_subcontract"):
            sc_order = await db.subcontract_orders.find_one({"reference_wo_id": wo_id})
            if sc_order and sc_order.get("status") != "completed":
                can_complete = False
        for op in operations:
            # Short-closed ops are effectively settled — they should not
            # block MO completion even if outsource_status is still 'sent'
            # (legacy short-close paths didn't clear that field).
            if op.get("short_closed"):
                continue
            if op.get("is_job_work") and op.get("outsource_status") == "sent":
                can_complete = False
                break
        if can_complete:
            new_status = "completed"
    elif any_active and current == "pending":
        new_status = "in_progress"

    if new_status == current and not op_mutated:
        return None

    update_payload = {}
    if op_mutated:
        update_payload["operations_status"] = operations
    if new_status != current:
        update_payload["status"] = new_status
        if new_status == "completed":
            update_payload["actual_end"] = datetime.now(timezone.utc)
            update_payload["quantity_completed"] = wo.get("quantity", 0)
            # Heal legacy short-closed ops whose status was never flipped to
            # 'completed' by the old short-close paths — so the UI shows them
            # consistently as Done.
            healed_ops = []
            any_healed = False
            for op in operations:
                if op.get("short_closed") and op.get("status") != "completed":
                    op = {**op, "status": "completed"}
                    any_healed = True
                healed_ops.append(op)
            if any_healed:
                update_payload["operations_status"] = healed_ops
    if update_payload:
        await db.work_orders.update_one({"id": wo_id}, {"$set": update_payload})
    return new_status if new_status != current else None


@work_orders_router.post("/{wo_id}/sync-status")
async def sync_wo_status(wo_id: str, request: Request):
    """Manually re-evaluate a Work Order's status based on its current
    operations. Useful for fixing MOs that got "stuck" in `in_progress`
    after some legacy short-close paths didn't propagate the status flip.

    Returns the resulting status (or null if nothing changed).
    """
    user = await get_current_user(request)
    role = user.get("role")
    perms = ((user.get("permissions") or {}).get("manufacturing") or [])
    if role != "admin" and "edit" not in perms:
        raise HTTPException(status_code=403, detail="You don't have permission to sync MO status")
    wo = await db.work_orders.find_one({"id": wo_id})
    if not wo:
        raise HTTPException(status_code=404, detail="Work order not found")
    new_status = await _recompute_wo_status_after_op_change(wo_id)
    return {"ok": True, "wo_id": wo_id, "previous_status": wo.get("status"), "new_status": new_status or wo.get("status"), "changed": new_status is not None}



@work_orders_router.post("/{wo_id}/operations/{sequence}/short-close")
async def short_close_wo_operation(wo_id: str, sequence: int, request: Request, payload: Optional[RevokeOpPayload] = None):
    """Revoke an OS allocation on a WO operation.

    Two modes:
      - **Op-level** (no `run_number` in body): Revoke ALL OS allocations on
        this op (legacy behaviour). Removes every OS run, clears OS fields
        on the op, drops every matching JWP line from the linked SC, and
        hard-deletes the SC if no other refs remain.
      - **Per-vendor** (with `run_number`): Revoke ONLY that specific run's
        vendor allocation. Reads the run's `outsource_sc_order_id` (stored
        at run-creation time) to find the right SC line. Removes only that
        line + that run. If the op still has other OS runs, op stays
        in 'pending' (partial OS); outsource_sc_order_id is repointed at
        a remaining OS run's SC for compatibility. If this was the LAST
        OS run, the op is fully reverted to pending (no OS metadata).

    Permission: admin OR any user with manufacturing 'edit' access — the
    latter covers production managers, supervisors etc. who need to manage
    in-flight outsource allocations without admin role escalation.
    """
    user = await get_current_user(request)
    # Allow admin OR users with manufacturing 'edit' permission. The
    # previous policy was admin-only, which blocked production managers
    # from managing their own job-card outsource flows.
    role = user.get("role")
    perms = ((user.get("permissions") or {}).get("manufacturing") or [])
    if role != "admin" and "edit" not in perms and "create" not in perms:
        raise HTTPException(status_code=403, detail="You don't have permission to revoke operations")
    run_number_filter = payload.run_number if payload else None
    wo = await db.work_orders.find_one({"id": wo_id})
    if not wo:
        raise HTTPException(status_code=404, detail="Work order not found")
    ops = wo.get("operations_status") or []
    target = None
    for op in ops:
        if op.get("sequence") == sequence:
            target = op
            break
    if not target:
        raise HTTPException(status_code=404, detail=f"Operation sequence {sequence} not found")
    if not target.get("is_job_work") or not target.get("outsource_sc_order_id"):
        raise HTTPException(status_code=400, detail="Operation must be an outsourced (OS) operation to short close")
    if target.get("status") == "completed":
        raise HTTPException(status_code=400, detail="Operation is already completed — cannot short close")

    # ---------- PER-VENDOR (per-run) revoke branch -----------------------
    # When `run_number` is supplied, only that vendor's allocation is
    # revoked. We use the run's stored `outsource_sc_order_id` (added at
    # run-creation time) to find the right SC. Legacy runs that pre-date
    # that field fall back to the op-level outsource_sc_order_id.
    if run_number_filter is not None:
        runs_all = target.get("runs") or []
        target_run = next((r for r in runs_all if r.get("run_number") == run_number_filter and (r.get("operator") or "").startswith("OS: ")), None)
        if not target_run:
            raise HTTPException(status_code=404, detail=f"OS run #{run_number_filter} not found on this operation")
        run_sc_id = target_run.get("outsource_sc_order_id") or target.get("outsource_sc_order_id")
        run_sc_number = target_run.get("outsource_sc_order_number") or target.get("outsource_sc_order_number")
        run_qty = float(target_run.get("quantity_planned") or 0)
        run_supplier_id = target_run.get("outsource_supplier_id") or target.get("job_work_supplier_id")
        op_name_val = target.get("operation_name") or ""
        if isinstance(op_name_val, dict):
            op_name_val = op_name_val.get("name", "")
        op_name = (op_name_val or "").strip()
        # GRN safety check on the SC line matching this run's supplier
        if run_sc_id:
            sc = await db.subcontract_orders.find_one({"id": run_sc_id})
            if sc:
                for line in (sc.get("job_work_parts") or []):
                    same_op = (line.get("process_name") or "").strip() == op_name or op_name in (line.get("process_names") or [])
                    if line.get("wo_id") == wo_id and same_op and (line.get("received_quantity") or 0) > 0:
                        raise HTTPException(status_code=400, detail=f"Cannot revoke this vendor allocation — GRN already received against {run_sc_number}. Reverse the GRN first.")
        # Remove just this run; recompute outsourced_quantity from remaining OS runs
        target["runs"] = [r for r in runs_all if r.get("run_number") != run_number_filter]
        remaining_os_runs = [r for r in target["runs"] if (r.get("operator") or "").startswith("OS: ")]
        remaining_os_qty = sum(float(r.get("quantity_planned") or 0) for r in remaining_os_runs)
        target["outsourced_quantity"] = remaining_os_qty
        if not remaining_os_runs:
            # No OS runs left → fully revert op to pending
            for f in ("is_job_work", "job_work_supplier_id", "outsource_status",
                      "outsource_supplier_name", "outsource_charges",
                      "outsource_sc_order_id", "outsource_sc_order_number",
                      "actual_start", "operator"):
                target.pop(f, None)
            target["status"] = "pending"
        else:
            # Re-point the op's outsource_sc_order_id at one of the remaining runs
            # so the rest of the workflow keeps working.
            first_remaining = remaining_os_runs[0]
            target["outsource_sc_order_id"] = first_remaining.get("outsource_sc_order_id") or target.get("outsource_sc_order_id")
            target["outsource_sc_order_number"] = first_remaining.get("outsource_sc_order_number") or target.get("outsource_sc_order_number")
            target["outsource_supplier_name"] = first_remaining.get("outsource_supplier_name") or target.get("outsource_supplier_name")
            target["job_work_supplier_id"] = first_remaining.get("outsource_supplier_id") or target.get("job_work_supplier_id")
            target["operator"] = first_remaining.get("operator")
        await db.work_orders.update_one({"id": wo_id}, {"$set": {"operations_status": ops}})

        # Update the SC: remove ONLY the JWP line that matches this wo +
        # process + supplier. If the SC ends up empty, hard-delete it.
        sc_deleted_pv = False
        sc_updated_pv = False
        if run_sc_id:
            sc = await db.subcontract_orders.find_one({"id": run_sc_id})
            if sc:
                new_jwp = []
                dropped_any = False
                for line in (sc.get("job_work_parts") or []):
                    same_op = (line.get("process_name") or "").strip() == op_name or op_name in (line.get("process_names") or [])
                    # Match by wo+op; if the SC was vendor-specific (single
                    # supplier_id matches run_supplier_id at top level), the
                    # match is already unambiguous.
                    if line.get("wo_id") == wo_id and same_op and not dropped_any:
                        dropped_any = True
                        continue
                    new_jwp.append(line)
                remaining_wo_ids = {(p.get("wo_id") or "") for p in new_jwp if p.get("wo_id")}
                ref_wo_ids = [w for w in (sc.get("reference_wo_ids") or []) if w in remaining_wo_ids]
                already_sent_dc = bool(sc.get("dc_created")) or await db.delivery_challans.find_one({
                    "subcontract_order_id": run_sc_id, "status": {"$in": ["sent", "approved"]}
                })
                has_received = any(float(p.get("received_quantity") or 0) > 0 for p in new_jwp)
                if not new_jwp and not ref_wo_ids and not already_sent_dc and not has_received:
                    await db.subcontract_orders.delete_one({"id": run_sc_id})
                    sc_deleted_pv = True
                else:
                    update_payload = {
                        "job_work_parts": new_jwp,
                        "reference_wo_ids": ref_wo_ids,
                        "processing_charges": sum(float(p.get("charges") or 0) * float(p.get("quantity") or 0) for p in new_jwp),
                        "updated_at": datetime.now(timezone.utc),
                    }
                    if not new_jwp:
                        update_payload["status"] = "short_closed"
                        update_payload["short_closed_at"] = datetime.now(timezone.utc)
                        update_payload["short_closed_by"] = user["id"]
                    await db.subcontract_orders.update_one({"id": run_sc_id}, {"$set": update_payload})
                sc_updated_pv = True

        return {
            "ok": True,
            "released": True,
            "per_vendor": True,
            "run_number": run_number_filter,
            "sc_order_id": run_sc_id,
            "sc_order_number": run_sc_number,
            "sc_updated": sc_updated_pv,
            "sc_deleted": sc_deleted_pv,
            "supplier_id": run_supplier_id,
        }
    # ---------- end per-vendor branch — fall through to op-level revoke --
    sc_order_id = target.get("outsource_sc_order_id")
    sc_order_number = target.get("outsource_sc_order_number")
    op_name_val = target.get("operation_name") or ""
    if isinstance(op_name_val, dict):
        op_name_val = op_name_val.get("name", "")
    op_name = (op_name_val or "").strip()
    # GRN safety — check the SC line for received_quantity > 0
    if sc_order_id:
        sc = await db.subcontract_orders.find_one({"id": sc_order_id})
        if sc:
            for line in (sc.get("job_work_parts") or []):
                if line.get("wo_id") == wo_id and (line.get("process_name") or "").strip() == op_name:
                    if (line.get("received_quantity") or 0) > 0:
                        raise HTTPException(status_code=400, detail="Cannot short-close: this operation has received quantity > 0. Reverse the GRN first.")
                    break

    # Clear OS fields on the WO operation
    for f in (
        "is_job_work", "job_work_supplier_id", "outsource_status",
        "outsource_supplier_name", "outsource_charges",
        "outsource_sc_order_id", "outsource_sc_order_number",
        "actual_start", "operator",
    ):
        target.pop(f, None)
    target["status"] = "pending"
    target["outsourced_quantity"] = 0
    runs = target.get("runs") or []
    target["runs"] = [r for r in runs if not (r.get("operator") or "").startswith("OS: ")]
    await db.work_orders.update_one({"id": wo_id}, {"$set": {"operations_status": ops}})

    # Trim the SC's job_work_parts to remove this WO+process pair.
    # Behaviour change (per user request): when the revoke removes the
    # LAST line from the SC AND no other WOs reference it, DELETE the SC
    # entirely so the Job Work list stays clean. If there are still
    # references (consolidated SC across multiple MOs) we just trim the
    # parts and update reference_wo_ids; the SC remains visible for the
    # other MOs.
    sc_updated = False
    sc_deleted = False
    if sc_order_id:
        sc = await db.subcontract_orders.find_one({"id": sc_order_id})
        if sc:
            new_jwp = []
            for line in (sc.get("job_work_parts") or []):
                if line.get("wo_id") == wo_id and (line.get("process_name") or "").strip() == op_name:
                    continue  # drop this line
                new_jwp.append(line)
            # Prune reference_wo_ids if this WO no longer appears.
            remaining_wo_ids = {(p.get("wo_id") or "") for p in new_jwp if p.get("wo_id")}
            ref_wo_ids = [w for w in (sc.get("reference_wo_ids") or []) if w in remaining_wo_ids]
            # If NO parts remain AND no other WOs reference it AND no DC/GRN
            # was ever sent, hard-delete the SC. Otherwise keep it open with
            # whatever lines/refs are left.
            already_sent_dc = bool(sc.get("dc_created")) or await db.delivery_challans.find_one({
                "subcontract_order_id": sc_order_id, "status": {"$in": ["sent", "approved"]}
            })
            has_received = any(float(p.get("received_quantity") or 0) > 0 for p in new_jwp)
            if not new_jwp and not ref_wo_ids and not already_sent_dc and not has_received:
                await db.subcontract_orders.delete_one({"id": sc_order_id})
                sc_deleted = True
                sc_updated = True
            else:
                update_payload = {"job_work_parts": new_jwp, "reference_wo_ids": ref_wo_ids, "updated_at": datetime.now(timezone.utc)}
                # Recompute processing_charges from remaining parts.
                update_payload["processing_charges"] = sum(
                    float(p.get("charges") or 0) * float(p.get("quantity") or 0) for p in new_jwp
                )
                # If empty but can't delete (DC sent / received), mark short_closed.
                if not new_jwp:
                    update_payload["status"] = "short_closed"
                    update_payload["short_closed_at"] = datetime.now(timezone.utc)
                    update_payload["short_closed_by"] = user["id"]
                await db.subcontract_orders.update_one({"id": sc_order_id}, {"$set": update_payload})
                sc_updated = True

    return {
        "ok": True,
        "released": True,
        "sc_order_id": sc_order_id,
        "sc_order_number": sc_order_number,
        "sc_updated": sc_updated,
        "sc_deleted": sc_deleted,
    }


class ShortCloseNoGRNPayload(BaseModel):
    reason: Optional[str] = ""
    run_number: Optional[int] = None  # When provided, short-close ONLY that vendor's run/SC line


@work_orders_router.post("/{wo_id}/operations/{sequence}/short-close-no-grn")
async def short_close_wo_operation_no_grn(wo_id: str, sequence: int, payload: ShortCloseNoGRNPayload, request: Request):
    """Hard-close an in-progress OS operation/run as COMPLETED without GRN.

    Two modes (mirrors `/short-close`):
      - **Op-level** (no `run_number`): all OS runs short-closed.
      - **Per-vendor** (`run_number`): only that vendor's run is short-closed.
        The op stays in-progress with the other vendor's run still active.

    Permission: admin OR users with manufacturing 'edit' permission.
    """
    user = await get_current_user(request)
    role = user.get("role")
    perms = ((user.get("permissions") or {}).get("manufacturing") or [])
    if role != "admin" and "edit" not in perms and "create" not in perms:
        raise HTTPException(status_code=403, detail="You don't have permission to short-close (no GRN) operations")
    wo = await db.work_orders.find_one({"id": wo_id})
    if not wo:
        raise HTTPException(status_code=404, detail="Work order not found")
    ops = wo.get("operations_status") or []
    target = None
    for op in ops:
        if op.get("sequence") == sequence:
            target = op
            break
    if not target:
        raise HTTPException(status_code=404, detail=f"Operation sequence {sequence} not found")
    if not target.get("is_job_work") or not target.get("outsource_sc_order_id"):
        raise HTTPException(status_code=400, detail="Operation must be an outsourced (OS) operation")
    if target.get("status") == "completed":
        raise HTTPException(status_code=400, detail="Operation is already completed")
    run_number_filter = payload.run_number if payload else None

    # ---------- PER-VENDOR (per-run) short-close branch -----------------
    if run_number_filter is not None:
        now = datetime.now(timezone.utc)
        runs_all = target.get("runs") or []
        target_run = next((r for r in runs_all if r.get("run_number") == run_number_filter and (r.get("operator") or "").startswith("OS: ")), None)
        if not target_run:
            raise HTTPException(status_code=404, detail=f"OS run #{run_number_filter} not found")
        run_sc_id = target_run.get("outsource_sc_order_id") or target.get("outsource_sc_order_id")
        run_sc_number = target_run.get("outsource_sc_order_number") or target.get("outsource_sc_order_number")
        run_qty = float(target_run.get("quantity_planned") or 0)
        op_name_val = target.get("operation_name") or ""
        if isinstance(op_name_val, dict):
            op_name_val = op_name_val.get("name", "")
        op_name = (op_name_val or "").strip()
        # Mark just this run as short-closed/completed; sum into op's completed/short_closed buckets.
        target_run["ended_at"] = now
        target_run["short_closed"] = True
        target_run["short_close_reason"] = payload.reason or ""
        target_run["quantity_completed"] = run_qty
        # Recompute op-level totals
        target["quantity_completed"] = (target.get("quantity_completed") or 0) + run_qty
        # If all OS runs are now short_closed/completed AND in-house completed
        # matches plan, mark the whole op completed.
        os_runs = [r for r in runs_all if (r.get("operator") or "").startswith("OS: ")]
        if all(r.get("short_closed") or r.get("ended_at") for r in os_runs):
            # Only flip to completed if there's no remaining in-house qty pending.
            total_done = float(target.get("quantity_completed") or 0)
            if total_done >= float(wo.get("quantity") or 0):
                target["status"] = "completed"
                target["short_closed"] = True
                target["short_close_reason"] = payload.reason or ""
                target["actual_end"] = now
        await db.work_orders.update_one({"id": wo_id}, {"$set": {"operations_status": ops}})
        # Re-evaluate WO status: if this run-close pushed the op to completed
        # and it was the LAST pending op, the MO should auto-finish.
        await _recompute_wo_status_after_op_change(wo_id)
        # Update the matching SC JWP line — set short_closed + charges=0.
        sc_updated_pv = False
        if run_sc_id:
            sc = await db.subcontract_orders.find_one({"id": run_sc_id})
            if sc:
                new_jwp = list(sc.get("job_work_parts") or [])
                dropped_any = False
                for line in new_jwp:
                    same_op = (line.get("process_name") or "").strip() == op_name or op_name in (line.get("process_names") or [])
                    if line.get("wo_id") == wo_id and same_op and not dropped_any:
                        dropped_any = True
                        line["short_closed"] = True
                        line["short_close_reason"] = payload.reason or ""
                        line["short_closed_at"] = now
                        line["charges"] = 0  # zero out so the JW list shows ₹0 for written-off work
                still_open = [p for p in new_jwp if not p.get("short_closed") and float(p.get("received_quantity") or 0) < float(p.get("quantity") or 0)]
                update_payload = {
                    "job_work_parts": new_jwp,
                    "processing_charges": sum(float(p.get("charges") or 0) * float(p.get("quantity") or 0) for p in new_jwp),
                    "updated_at": now,
                }
                if not still_open:
                    update_payload["status"] = "short_closed"
                    update_payload["short_closed_at"] = now
                    update_payload["short_closed_by"] = user["id"]
                await db.subcontract_orders.update_one({"id": run_sc_id}, {"$set": update_payload})
                sc_updated_pv = True
        return {
            "ok": True,
            "per_vendor": True,
            "run_number": run_number_filter,
            "sc_order_id": run_sc_id,
            "sc_order_number": run_sc_number,
            "sc_updated": sc_updated_pv,
        }
    # ---------- end per-vendor — fall through to op-level short-close --
    sc_order_id = target.get("outsource_sc_order_id")
    sc_order_number = target.get("outsource_sc_order_number")
    op_name_val = target.get("operation_name") or ""
    if isinstance(op_name_val, dict):
        op_name_val = op_name_val.get("name", "")
    op_name = (op_name_val or "").strip()

    # Mark the op completed + short_closed. Keep the OS metadata so the
    # audit trail of who/where the material went remains visible.
    now = datetime.now(timezone.utc)
    target["status"] = "completed"
    target["short_closed"] = True
    target["short_close_reason"] = payload.reason or ""
    target["short_closed_at"] = now
    target["short_closed_by"] = user["id"]
    target["actual_end"] = now
    # Mark the OS leg as settled so the WO-status recompute (and the rest
    # of the system) no longer thinks materials are out at the vendor.
    if target.get("outsource_status") == "sent":
        target["outsource_status"] = "short_closed"
    # Treat the entire op qty as accounted for so downstream MO logic
    # (e.g. quantity_completed roll-up) doesn't think work is still pending.
    op_qty = float(target.get("allocated_qty") or wo.get("quantity") or 0)
    target["quantity_completed"] = op_qty
    target["quantity_accepted"] = op_qty
    await db.work_orders.update_one({"id": wo_id}, {"$set": {"operations_status": ops}})
    # Re-evaluate WO status: a short-closed op is functionally "completed",
    # so if this was the last pending op the MO should now finish.
    await _recompute_wo_status_after_op_change(wo_id)

    # Mark the SC's matching line short_closed (don't delete — keep audit).
    # Per user request: when short-closing a line, ZERO its charges so the
    # JW list's "CHARGES" column shows ₹0 for short-closed lines (the
    # vendor isn't paid for work that was written off). The original
    # quantity stays for the audit trail.
    sc_updated = False
    if sc_order_id:
        sc = await db.subcontract_orders.find_one({"id": sc_order_id})
        if sc:
            new_jwp = list(sc.get("job_work_parts") or [])
            for line in new_jwp:
                if line.get("wo_id") == wo_id and (line.get("process_name") or "").strip() == op_name:
                    line["short_closed"] = True
                    line["short_close_reason"] = payload.reason or ""
                    line["short_closed_at"] = now
                    # Zero out charges for the short-closed line so the
                    # SC list/print doesn't show fees for written-off work.
                    line["charges"] = 0
            # If all remaining lines are short_closed OR received in full, mark the SC short_closed
            still_open = [
                p for p in new_jwp
                if not p.get("short_closed")
                and float(p.get("received_quantity") or 0) < float(p.get("quantity") or 0)
            ]
            # Recompute processing_charges from the remaining (non-zero) lines.
            processing_charges = sum(
                float(p.get("charges") or 0) * float(p.get("quantity") or 0) for p in new_jwp
            )
            update_payload = {
                "job_work_parts": new_jwp,
                "processing_charges": processing_charges,
                "updated_at": now,
            }
            if not still_open:
                update_payload["status"] = "short_closed"
                update_payload["short_closed_at"] = now
                update_payload["short_closed_by"] = user["id"]
            await db.subcontract_orders.update_one({"id": sc_order_id}, {"$set": update_payload})
            sc_updated = True

    return {
        "ok": True,
        "sc_order_id": sc_order_id,
        "sc_order_number": sc_order_number,
        "sc_updated": sc_updated,
        "completed_qty": op_qty,
    }


@work_orders_router.post("/{wo_id}/reserve")
async def reserve_materials_for_wo(wo_id: str, request: Request):
    """Reserve materials for an MO by computing its full BOM requirement recursively.
    Stores reserved_materials on the MO. MRP uses this to calculate net RM demand."""
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="manufacturing", action="create")
    wo = await db.work_orders.find_one({"id": wo_id})
    if not wo:
        raise HTTPException(status_code=404, detail="Work order not found")
    
    if wo.get("status") not in ["pending", "in_progress"]:
        raise HTTPException(status_code=400, detail="Can only reserve for pending or in-progress MOs")
    
    if wo.get("materials_reserved"):
        raise HTTPException(status_code=400, detail="Materials already reserved for this MO")
    
    routing = await db.routings.find_one({"id": wo.get("routing_id")})
    if not routing:
        raise HTTPException(status_code=404, detail="Routing not found")
    
    item_id = routing.get("item_id")
    wo_qty = wo.get("quantity", 1)
    
    # Recursively explode BOM and collect ALL material requirements (RM, SA, Parts)
    reserved = []
    
    async def collect_bom_materials(parent_item_id: str, parent_qty: float, visited: set = None):
        if visited is None:
            visited = set()
        if parent_item_id in visited:
            return
        visited.add(parent_item_id)
        
        bom = await db.boms.find_one({"parent_item_id": parent_item_id, "status": "active"}, {"_id": 0})
        if not bom:
            return
        for comp in bom.get("components", []):
            if comp.get("is_alternate"):
                continue
            comp_item_id = comp.get("item_id")
            comp_qty = comp.get("quantity", 0) * parent_qty
            comp_item = await db.items.find_one({"id": comp_item_id}, {"_id": 0})
            if not comp_item:
                continue
            reserved.append({
                "item_id": comp_item_id,
                "part_number": comp_item.get("part_number", ""),
                "name": comp_item.get("name", ""),
                "category": comp_item.get("category", ""),
                "quantity": comp_qty,
                "uom": comp_item.get("unit_of_measure", "pcs")
            })
            # Recurse into child BOMs to find more RM
            child_visited = set(visited)
            await collect_bom_materials(comp_item_id, comp_qty, child_visited)
    
    await collect_bom_materials(item_id, wo_qty)
    
    # Only keep raw_material items
    rm_needed = [r for r in reserved if r.get("category") == "raw_material"]
    
    # Consolidate duplicates (same RM may appear from different BOM paths)
    consolidated = {}
    for r in rm_needed:
        rid = r["item_id"]
        if rid in consolidated:
            consolidated[rid]["quantity"] += r["quantity"]
        else:
            consolidated[rid] = dict(r)
    rm_needed = list(consolidated.values())
    
    # Calculate already-allocated stock from OTHER reserved MOs (FIFO priority)
    other_reserved = await db.work_orders.find(
        {"materials_reserved": True, "id": {"$ne": wo_id}, "status": {"$in": ["pending", "in_progress"]}},
        {"_id": 0, "reserved_materials": 1}
    ).to_list(5000)
    already_allocated = {}  # item_id -> qty already locked by other MOs
    for other_mo in other_reserved:
        for orm in other_mo.get("reserved_materials", []):
            orid = orm.get("item_id")
            if orid:
                already_allocated[orid] = already_allocated.get(orid, 0) + orm.get("allocated_qty", 0)
    
    # Allocate from available stock (stock - already_allocated_by_others)
    rm_reserved = []
    shortfall_items = []
    for r in rm_needed:
        rid = r["item_id"]
        item = await db.items.find_one({"id": rid}, {"_id": 0})
        current_stock = item.get("current_stock", 0) if item else 0
        other_alloc = already_allocated.get(rid, 0)
        available = max(0, current_stock - other_alloc)
        needed = r["quantity"]
        allocated = min(available, needed)
        shortfall = max(0, needed - allocated)
        
        entry = {
            "item_id": rid,
            "part_number": r.get("part_number", ""),
            "name": r.get("name", ""),
            "category": "raw_material",
            "quantity": needed,
            "allocated_qty": allocated,
            "shortfall_qty": shortfall,
            "uom": r.get("uom", "pcs")
        }
        rm_reserved.append(entry)
        if shortfall > 0:
            shortfall_items.append(f"{r.get('part_number','')}: need {needed}, allocated {allocated}, shortfall {shortfall}")
    
    total_shortfall = sum(r["shortfall_qty"] for r in rm_reserved)
    
    await db.work_orders.update_one({"id": wo_id}, {"$set": {
        "materials_reserved": True,
        "reserved_materials": rm_reserved,
        "reservation_shortfall": total_shortfall,
        "reserved_at": datetime.now(timezone.utc),
        "reserved_by": user["id"]
    }})
    
    msg = f"Materials reserved for {wo.get('wo_number')} ({len(rm_reserved)} RM items)"
    if total_shortfall > 0:
        msg += f"\n\nShortfall (purchase needed):\n" + "\n".join(shortfall_items)
    
    return {
        "success": True,
        "message": msg,
        "reserved_materials": rm_reserved,
        "total_shortfall": total_shortfall
    }


@work_orders_router.post("/{wo_id}/create-sc")
async def create_sc_for_wo(wo_id: str, request: Request):
    """Dedicated endpoint: create SC order for a subcontract MO. Simple and direct."""
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="manufacturing", action="create")
    wo = await db.work_orders.find_one({"id": wo_id})
    if not wo:
        raise HTTPException(status_code=404, detail="Work order not found")
    
    if not wo.get("is_subcontract"):
        raise HTTPException(status_code=400, detail="This MO is not marked as subcontract")
    
    if not wo.get("subcontract_supplier_id"):
        raise HTTPException(status_code=400, detail="No supplier set for this SC MO")
    
    # Check if SC already exists for THIS specific MO
    existing = await db.subcontract_orders.find_one({
        "$or": [
            {"reference_wo_id": wo_id},
            {"reference_wo_ids": wo_id}
        ]
    })
    if existing:
        # If SC exists but no DC sent yet, recalculate lines with smart resolution
        sent_dc = await db.delivery_challans.find_one({"subcontract_order_id": existing["id"], "status": "sent"})
        if not sent_dc and existing.get("status") in ["draft", "in_progress"]:
            # Recalculate lines below and update existing SC
            pass  # Fall through to line building logic
        else:
            existing.pop("_id", None)
            return {"success": True, "message": f"SC order already exists: {existing.get('order_number')}", "sc_order": existing}
    
    sc_type = wo.get("subcontract_type", "with_material")
    item_id = wo.get("item_id")
    item = await db.items.find_one({"id": item_id}, {"_id": 0}) if item_id else None
    qty = wo.get("quantity", 1)
    
    # Build lines — smart resolution based on MO completion status across the work order tree
    sc_lines = []
    if sc_type == "without_material":
        # SC without RM: FG/SA/Part rate = BOM RM cost (material only). No process cost embedded here —
        # process cost lives on job_work_parts.charges (FG parent routing).
        _no_rm_costs = await compute_bom_costs(item_id)
        _no_rm_rate = _no_rm_costs.get("rm_cost", 0) or (item.get("unit_cost", 0) if item else 0)
        sc_lines = [{"item_id": item_id, "quantity": qty, "sent_quantity": 0, "received_quantity": 0, "rate": round(_no_rm_rate, 2)}]
    else:
        # Walk up to find root parent MO, then collect ALL MOs in the tree
        root_wo_id = wo_id
        current_wo = wo
        while current_wo.get("parent_wo_id"):
            parent_wo = await db.work_orders.find_one({"id": current_wo["parent_wo_id"]})
            if not parent_wo:
                break
            root_wo_id = parent_wo["id"]
            current_wo = parent_wo
        
        # Collect all MOs in the tree (recursive)
        async def collect_tree_mos(pid):
            mos = []
            children = await db.work_orders.find({"parent_wo_id": pid}, {"_id": 0}).to_list(500)
            for c in children:
                mos.append(c)
                mos.extend(await collect_tree_mos(c["id"]))
            return mos
        
        all_tree_mos = await collect_tree_mos(root_wo_id)
        completed_item_ids = {m["item_id"] for m in all_tree_mos if m.get("status") == "completed"}
        
        # Helper to calculate BOM Total/Unit (material + all process rollup) for a completed Part/SA
        # being sent in an SC-with-material flow. Matches what the BOM viewer shows in the Total/Unit column.
        async def calc_bom_rollup(item_id_calc):
            try:
                val = await compute_bom_total_unit_cost(item_id_calc)
                if val and val > 0:
                    return val
            except Exception:
                pass
            # Fallback to item.unit_cost
            ci_fallback = await db.items.find_one({"id": item_id_calc}, {"_id": 0})
            return ci_fallback.get("unit_cost", 0) if ci_fallback else 0
        
        # Smart resolve: completed parts sent as-is, unprocessed parts resolved to RM
        async def smart_resolve(parent_item_id, multiplier, visited=None):
            if visited is None:
                visited = set()
            if parent_item_id in visited:
                return []
            visited.add(parent_item_id)
            result = []
            bom = await db.boms.find_one({"parent_item_id": parent_item_id, "status": "active"}, {"_id": 0})
            if not bom:
                return result
            for comp in bom.get("components", []):
                if comp.get("is_alternate"):
                    continue
                ci = await db.items.find_one({"id": comp["item_id"]}, {"_id": 0})
                if not ci:
                    continue
                comp_qty = comp.get("quantity", 1) * multiplier
                cat = ci.get("category", "")
                if cat == "raw_material":
                    # Always add RM directly
                    result.append({"item_id": comp["item_id"], "quantity": int(comp_qty), "sent_quantity": 0, "received_quantity": 0, "rate": ci.get("unit_cost", 0)})
                elif cat in ["component", "sub_assembly"]:
                    if comp["item_id"] in completed_item_ids:
                        # Part already completed — send the finished part with BOM rollup cost
                        rollup_cost = await calc_bom_rollup(comp["item_id"])
                        result.append({"item_id": comp["item_id"], "quantity": int(comp_qty), "sent_quantity": 0, "received_quantity": 0, "rate": round(rollup_cost, 2)})
                    else:
                        # Part NOT completed — resolve to its RM via BOM
                        child_rm = await smart_resolve(comp["item_id"], comp_qty, visited)
                        if child_rm:
                            result.extend(child_rm)
                        else:
                            # No BOM found, use rollup cost
                            rollup_cost = await calc_bom_rollup(comp["item_id"])
                            result.append({"item_id": comp["item_id"], "quantity": int(comp_qty), "sent_quantity": 0, "received_quantity": 0, "rate": round(rollup_cost, 2)})
            return result
        
        rm_items = await smart_resolve(item_id, qty)
        # Deduplicate by item_id (sum quantities)
        rm_map = {}
        for rm in rm_items:
            if rm["item_id"] in rm_map:
                rm_map[rm["item_id"]]["quantity"] += rm["quantity"]
            else:
                rm_map[rm["item_id"]] = rm
        sc_lines = list(rm_map.values())
        
        if not sc_lines:
            sc_lines = [{"item_id": item_id, "quantity": qty, "sent_quantity": 0, "received_quantity": 0, "rate": item.get("unit_cost", 0) if item else 0}]
    
    # If existing SC was found (no DC sent), update its lines with recalculated values
    if existing and not await db.delivery_challans.find_one({"subcontract_order_id": existing["id"], "status": "sent"}):
        await db.subcontract_orders.update_one({"id": existing["id"]}, {"$set": {
            "lines": sc_lines,
            "updated_at": datetime.now(timezone.utc)
        }})
        # Also update draft DCs linked to this SC
        draft_dcs = await db.delivery_challans.find({"subcontract_order_id": existing["id"], "status": "draft"}).to_list(100)
        for ddc in draft_dcs:
            dc_lines = [{"item_id": l["item_id"], "quantity": l["quantity"], "rate": l["rate"]} for l in sc_lines]
            await db.delivery_challans.update_one({"id": ddc["id"]}, {"$set": {"lines": dc_lines}})
        existing.pop("_id", None)
        existing["lines"] = sc_lines
        return {"success": True, "message": f"SC order {existing.get('order_number')} lines recalculated", "sc_order": existing}
    
    # Try to consolidate into existing SC for same supplier (no sent DC, no PO created)
    consolidate_sc = await db.subcontract_orders.find_one({
        "supplier_id": wo["subcontract_supplier_id"],
        "subcontract_type": sc_type,
        "status": {"$in": ["draft", "in_progress"]},
    }, sort=[("created_at", -1)])
    
    # Skip if it already has PO or sent DC
    if consolidate_sc:
        if consolidate_sc.get("po_created"):
            consolidate_sc = None
        else:
            sent_dc = await db.delivery_challans.find_one({"subcontract_order_id": consolidate_sc["id"], "status": "sent"})
            if sent_dc:
                consolidate_sc = None
    
    if consolidate_sc:
        # Merge parts and lines into existing SC
        parts = consolidate_sc.get("job_work_parts", [])
        part_found = False
        # Compute FG process charge once (used if we need to append a new part entry)
        _fg_charge = 0
        _bom_rollup = 0
        _process_names = []
        try:
            _bc = await compute_bom_costs(item_id)
            _fg_charge = _bc.get("fg_process_cost", 0) or 0
            _bom_rollup = (_bc.get("rm_cost", 0) or 0) + (_bc.get("process_cost", 0) or 0)
            _process_names = _bc.get("process_names", []) or []
        except Exception:
            pass
        for ep in parts:
            if ep.get("item_id") == item_id:
                ep["quantity"] = ep.get("quantity", 0) + qty
                # Populate charges from BOM if currently 0 (migration safe)
                if not ep.get("charges"):
                    ep["charges"] = _fg_charge
                if not ep.get("bom_rollup_cost"):
                    ep["bom_rollup_cost"] = round(_bom_rollup, 2)
                if not ep.get("process_names") and _process_names:
                    ep["process_names"] = _process_names
                part_found = True
                break
        if not part_found:
            parts.append({
                "item_id": item_id, "quantity": qty,
                "charges": _fg_charge,
                "bom_rollup_cost": round(_bom_rollup, 2),
                "process_names": _process_names,
                "received_quantity": 0
            })
        
        lines = consolidate_sc.get("lines", [])
        for nl in sc_lines:
            found = False
            for el in lines:
                if el["item_id"] == nl["item_id"]:
                    el["quantity"] += nl["quantity"]
                    found = True
                    break
            if not found:
                lines.append(nl)
        
        ref_ids = consolidate_sc.get("reference_wo_ids", [])
        ref_ids.append(wo_id)
        
        await db.subcontract_orders.update_one({"id": consolidate_sc["id"]}, {"$set": {
            "job_work_parts": parts, "lines": lines, "reference_wo_ids": ref_ids,
            "updated_at": datetime.now(timezone.utc)
        }})
        
        if wo.get("status") == "pending":
            await db.work_orders.update_one({"id": wo_id}, {"$set": {"status": "in_progress", "actual_start": datetime.now(timezone.utc)}})
        
        # Mark child MOs as outsourced
        async def mark_children_outsourced_c(parent_id):
            children = await db.work_orders.find({"parent_wo_id": parent_id}, {"_id": 0}).to_list(500)
            for child in children:
                if child.get("status") not in ["completed", "cancelled"]:
                    await db.work_orders.update_one({"id": child["id"]}, {"$set": {
                        "outsourced_by_parent": True,
                        "outsourced_sc_order": consolidate_sc.get("order_number", ""),
                        "status": "outsourced"
                    }})
                    await mark_children_outsourced_c(child["id"])
        await mark_children_outsourced_c(wo_id)
        
        consolidate_sc.pop("_id", None)
        return {"success": True, "message": f"Consolidated into {consolidate_sc.get('order_number')}", "sc_order": consolidate_sc}
    
    # Create SC order
    sc_count = await db.subcontract_orders.count_documents({})
    # Pull BOM-based costs for the FG/SA item
    bom_costs = await compute_bom_costs(item_id)
    # For SC (with material) — job_work_parts.charges/pc = FG PARENT ROUTING cost only ("FG Process")
    # For SC (without material) — same: charges/pc = FG parent routing cost of the item's BOM
    fg_process_only = bom_costs.get("fg_process_cost", 0) or 0
    bom_rollup_cost_val = bom_costs["rm_cost"] + bom_costs["process_cost"]  # Total/Unit for reference
    default_process_charges = fg_process_only
    
    # Pull previous process charges for this item from latest SC order (override only if BOM has no FG routing)
    prev_charges = 0
    if not default_process_charges:
        prev_sc = await db.subcontract_orders.find_one(
            {"job_work_parts.item_id": item_id, "status": {"$in": ["in_progress", "completed"]}},
            {"_id": 0, "job_work_parts": 1},
            sort=[("created_at", -1)]
        )
        if prev_sc:
            for pjwp in prev_sc.get("job_work_parts", []):
                if pjwp.get("item_id") == item_id and pjwp.get("charges"):
                    prev_charges = pjwp["charges"]
                    break
    effective_charges = default_process_charges or prev_charges
    
    sc_doc = {
        "id": str(uuid.uuid4()),
        "order_number": f"JW-{str(sc_count + 1).zfill(6)}",
        "supplier_id": wo["subcontract_supplier_id"],
        "reference_wo_id": wo_id,
        "reference_wo_ids": [wo_id],
        "subcontract_type": sc_type,
        "fg_item_id": item_id,
        "fg_item_name": f"{item.get('part_number', '')} - {item.get('name', '')}" if item else "",
        "fg_quantity": qty,
        "job_work_parts": [{"item_id": item_id, "quantity": qty, "charges": effective_charges, "received_quantity": 0, "bom_rollup_cost": round(bom_rollup_cost_val, 2), "process_names": bom_costs.get("process_names", [])}],
        "lines": sc_lines,
        "status": "in_progress",
        "notes": f"SC for MO {wo.get('wo_number')} ({sc_type.replace('_', ' ')})",
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"]
    }
    await db.subcontract_orders.insert_one(sc_doc)
    sc_doc.pop("_id", None)
    
    # Update MO status to in_progress if still pending
    if wo.get("status") == "pending":
        await db.work_orders.update_one({"id": wo_id}, {"$set": {"status": "in_progress", "actual_start": datetime.now(timezone.utc)}})
    
    # Mark ALL child MOs as outsourced (recursively) — they're covered by parent's SC
    async def mark_children_outsourced(parent_id):
        children = await db.work_orders.find({"parent_wo_id": parent_id}, {"_id": 0}).to_list(500)
        for child in children:
            if child.get("status") not in ["completed", "cancelled"]:
                await db.work_orders.update_one({"id": child["id"]}, {"$set": {
                    "outsourced_by_parent": True,
                    "outsourced_sc_order": sc_doc.get("order_number", ""),
                    "status": "outsourced"
                }})
                await mark_children_outsourced(child["id"])
    
    await mark_children_outsourced(wo_id)
    
    return {"success": True, "message": f"SC Order {sc_doc['order_number']} created", "sc_order": sc_doc}


@work_orders_router.post("/{wo_id}/unreserve")
async def unreserve_materials_for_wo(wo_id: str, request: Request):
    """Remove material reservation from an MO"""
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="manufacturing", action="create")
    wo = await db.work_orders.find_one({"id": wo_id})
    if not wo:
        raise HTTPException(status_code=404, detail="Work order not found")
    
    if not wo.get("materials_reserved"):
        raise HTTPException(status_code=400, detail="No reservation exists for this MO")
    
    await db.work_orders.update_one({"id": wo_id}, {"$set": {
        "materials_reserved": False,
        "reserved_materials": [],
    }, "$unset": {"reserved_at": "", "reserved_by": ""}})
    
    return {"success": True, "message": f"Reservation removed for {wo.get('wo_number')}"}


@work_orders_router.post("/{wo_id}/start")
async def start_work_order(wo_id: str, request: Request, preview: bool = False):
    """Start a work order — consumes required materials from inventory.

    Pass `?preview=true` to compute what WOULD be consumed (and surface any
    insufficient-stock / reserved-by-other-MO conflicts) without actually
    deducting stock or marking the MO as started. The frontend uses this to
    show a confirmation dialog before committing.
    """
    user = await get_current_user(request)
    # Preview-only needs VIEW/EDIT access; only the actual start (which
    # consumes stock) requires create/edit rights. Splitting these means
    # operators with edit-only permission can still SEE the material
    # consumption details before requesting an admin to confirm the start.
    _require_access(user, ["admin", "production_manager", "production_operator"], module="manufacturing", action="view" if preview else "edit")
    wo = await db.work_orders.find_one({"id": wo_id})
    if not wo:
        raise HTTPException(status_code=404, detail="Work order not found")
    
    if wo.get("status") not in ["pending", "in_progress"]:
        raise HTTPException(status_code=400, detail="Work order is not in pending/in-progress status")
    
    already_started = wo.get("status") == "in_progress"
    
    if wo.get("materials_consumed") and not wo.get("is_subcontract"):
        raise HTTPException(status_code=400, detail="Materials already consumed for this work order")
    
    # Get the routing and item
    routing = await db.routings.find_one({"id": wo.get("routing_id")})
    if not routing:
        if wo.get("is_subcontract"):
            # For SC MOs without routing, use item_id directly
            item_id = wo.get("item_id")
            bom = await db.boms.find_one({"parent_item_id": item_id, "status": "active"}) if item_id else None
        elif wo.get("parent_wo_id"):
            # Child MO (SG/Part) — no routing; use its own BOM (if any) for material consumption
            item_id = wo.get("item_id")
            bom = await db.boms.find_one({"parent_item_id": item_id, "status": "active"}) if item_id else None
        elif wo.get("operations_status"):
            # Main MO whose routing lives inside BOM.parent_routings (no legacy
            # `routings` row). The WO was created with operations already embedded
            # from those parent_routings, so we just need the BOM for material
            # consumption.
            item_id = wo.get("item_id")
            bom = await db.boms.find_one({"parent_item_id": item_id, "status": "active"}) if item_id else None
        else:
            raise HTTPException(status_code=404, detail="Routing not found")
    else:
        item_id = routing.get("item_id")
        bom = await db.boms.find_one({"parent_item_id": item_id, "status": "active"})
    
    consumed_materials = []
    insufficient_materials = []
    reserved_conflicts = []
    
    # Skip material consumption for ALL subcontract MOs:
    # - "without_material": vendor sources own RM
    # - "with_material": RM sent via DC (stock deducted at DC Send, not MO Start)
    skip_material_consumption = bool(wo.get("is_subcontract"))
    
    # Check if materials are reserved by OTHER MOs (only for non-SC inhouse MOs)
    if not wo.get("materials_reserved") and not skip_material_consumption:
        # Walk up the parent chain to find all ancestor MO IDs
        ancestor_ids = set()
        current = wo
        while current.get("parent_wo_id"):
            ancestor_ids.add(current["parent_wo_id"])
            current = await db.work_orders.find_one({"id": current["parent_wo_id"]}, {"_id": 0, "id": 1, "parent_wo_id": 1})
            if not current:
                break
        
        # Exclude self AND all ancestors from the "other reserved" check
        exclude_ids = {wo_id} | ancestor_ids
        other_reserved_mos = await db.work_orders.find(
            {"materials_reserved": True, "id": {"$nin": list(exclude_ids)}, "status": {"$in": ["pending", "in_progress"]}},
            {"_id": 0, "wo_number": 1, "reserved_materials": 1}
        ).to_list(5000)
        
        # Build map: item_id -> [{mo_number, allocated_qty}]
        reserved_by_others = {}
        for other_mo in other_reserved_mos:
            for rm in other_mo.get("reserved_materials", []):
                rid = rm.get("item_id")
                alloc = rm.get("allocated_qty", 0)
                if rid and alloc > 0:
                    if rid not in reserved_by_others:
                        reserved_by_others[rid] = {"total": 0, "mos": []}
                    reserved_by_others[rid]["total"] += alloc
                    reserved_by_others[rid]["mos"].append({"mo": other_mo.get("wo_number"), "qty": alloc})
        
        # Check each BOM component against reserved stock
        if bom and not skip_material_consumption:
            wo_qty_check = wo.get("quantity", 1)
            for component in bom.get("components", []):
                if component.get("is_alternate"):
                    continue
                comp_item_id = component.get("item_id")
                comp_item = await db.items.find_one({"id": comp_item_id}, {"_id": 0})
                if not comp_item or comp_item.get("category") not in ["raw_material", "component", "sub_assembly"]:
                    continue
                # Use float so fractional BOM quantities (e.g. 1.68 kgs of
                # HR sheet for a 24-piece batch) survive the reserved-stock
                # check. The `int(...)` cast would round 1.68 → 1, silently
                # consuming less than the BOM specifies.
                required_qty = float(component.get("quantity", 1) * wo_qty_check)
                current_stock = comp_item.get("current_stock", 0)
                reserved_info = reserved_by_others.get(comp_item_id)
                if reserved_info:
                    free_stock = max(0, current_stock - reserved_info["total"])
                    if free_stock < required_qty:
                        mo_details = ", ".join([f"{m['mo']}({m['qty']})" for m in reserved_info["mos"]])
                        reserved_conflicts.append({
                            "item": comp_item.get("part_number"),
                            "name": comp_item.get("name"),
                            "required": required_qty,
                            "total_stock": current_stock,
                            "reserved_by": reserved_info["total"],
                            "free_stock": free_stock,
                            "reserved_mos": mo_details
                        })
        
        if reserved_conflicts:
            return {
                "success": False,
                "message": "Cannot start — materials are reserved by other MOs. Reserve this MO first or wait for stock.",
                "reserved_conflicts": reserved_conflicts
            }
    
    if bom and not skip_material_consumption:
        wo_qty = wo.get("quantity", 1)
        mo_variant_selection = wo.get("variant_selection") or None
        
        for component in bom.get("components", []):
            if component.get("is_alternate"):
                continue
            
            comp_item_id = component.get("item_id")
            comp_item = await db.items.find_one({"id": comp_item_id})
            if not comp_item:
                continue

            # Variant-aware consumption: if this component carries its own
            # variant_attributes AND the MO's variant_selection picks one of
            # those axes, redirect consumption to the matching variant child
            # SKU (e.g. CRW0E8000091-30GT) instead of the parent. The variant
            # child must already exist (user generates them via the Item dialog).
            variant_child = await _resolve_variant_child_item(comp_item, mo_variant_selection)
            if variant_child:
                comp_item_id = variant_child["id"]
                comp_item = variant_child
            
            # Consume ALL BOM components (RM, components, sub-assemblies) from stock
            if comp_item.get("category") in ["raw_material", "component", "sub_assembly"]:
                # Use float so fractional BOM quantities (e.g. 1.68 kgs of HR
                # sheet) are consumed exactly. Previously this was `int(...)`
                # which silently truncated 1.68 → 1, leaving 0.68 outstanding
                # forever and skewing MRP.
                required_qty = float(component.get("quantity", 1) * wo_qty)
                current_stock = float(comp_item.get("current_stock") or 0)
                
                if current_stock < required_qty:
                    insufficient_materials.append({
                        "item": comp_item.get("part_number"),
                        "name": comp_item.get("name"),
                        "required": required_qty,
                        "available": current_stock
                    })
                else:
                    # Compute what we'd consume. Skip the DB writes entirely when
                    # the request is just a preview — we still build
                    # `consumed_materials` so the dialog shows the same numbers
                    # the user will see after confirming.
                    new_stock = current_stock - required_qty
                    if not preview:
                        tx_doc = {
                            "id": str(uuid.uuid4()),
                            "item_id": comp_item_id,
                            "transaction_type": "issue",
                            "quantity": required_qty,
                            "reference_type": "work_order",
                            "reference_id": wo_id,
                            "previous_stock": current_stock,
                            "new_stock": new_stock,
                            "notes": f"Consumed for WO {wo.get('wo_number')}",
                            "created_at": datetime.now(timezone.utc),
                            "created_by": user["id"]
                        }
                        await db.inventory_transactions.insert_one(tx_doc)
                        await db.items.update_one(
                            {"id": comp_item_id},
                            {"$set": {"current_stock": new_stock}}
                        )

                    consumed_materials.append({
                        "item_id": comp_item_id,
                        "item": comp_item.get("part_number"),
                        "name": comp_item.get("name"),
                        "quantity": required_qty,
                        "uom": comp_item.get("unit_of_measure", "pcs"),
                        "unit_cost": comp_item.get("unit_cost", 0)
                    })
                    # Release any reservation tied to THIS MO for this component.
                    # The stock is now physically consumed (decremented above),
                    # so the booking is no longer needed. Without this, the
                    # reserved_stock counter would over-count after material
                    # consumption (`current_stock - reserved_stock` would go
                    # negative for future MOs of the same component).
                    if not preview:
                        for resv in (wo.get("child_reservations") or []):
                            if resv.get("item_id") == comp_item_id:
                                resv_qty = int(resv.get("qty", 0) or 0)
                                if resv_qty > 0:
                                    await db.items.update_one(
                                        {"id": comp_item_id},
                                        {"$inc": {"reserved_stock": -resv_qty}},
                                    )
                                break
    
    if insufficient_materials:
        return {
            "success": False,
            "message": "Insufficient materials to start work order",
            "insufficient_materials": insufficient_materials
        }
    
    # PREVIEW mode: return what WOULD happen without persisting state.
    # The frontend uses this to show a confirmation dialog; user can close it
    # to abort without any side-effects.
    if preview:
        return {
            "success": True,
            "preview": True,
            "message": "Preview only — no materials consumed yet. Confirm to start.",
            "consumed_materials": consumed_materials,
            "wo_number": wo.get("wo_number"),
        }

    # Update work order status to in_progress (skip if already started)
    if not already_started:
        await db.work_orders.update_one(
            {"id": wo_id},
            {"$set": {
                "status": "in_progress",
                "actual_start": datetime.now(timezone.utc),
                "materials_consumed": True,
                "consumed_materials": consumed_materials,
                "updated_at": datetime.now(timezone.utc)
            }}
        )
    
    # SC order is NOT created here — only via /create-sc endpoint
    return {
        "success": True,
        "message": "Work order started" + (", materials consumed" if consumed_materials else ""),
        "consumed_materials": consumed_materials,
        "wo_number": wo.get("wo_number")
    }


@work_orders_router.get("/{wo_id}/tree")
async def get_work_order_tree(wo_id: str, request: Request):
    """Get the MO tree: Finished Good → Semi-Finished → Parts"""
    await get_current_user(request)
    
    async def build_tree(wid):
        wo = await db.work_orders.find_one({"id": wid}, {"_id": 0})
        if not wo:
            return None
        item = await db.items.find_one({"id": wo.get("item_id")}, {"_id": 0})
        wo["item"] = item
        children = await db.work_orders.find({"parent_wo_id": wid}, {"_id": 0}).to_list(100)
        wo["children"] = []
        for child in children:
            child_tree = await build_tree(child["id"])
            if child_tree:
                wo["children"].append(child_tree)
        return wo
    
    # Find root MO (may be this one or its parent)
    wo = await db.work_orders.find_one({"id": wo_id}, {"_id": 0})
    if not wo:
        raise HTTPException(status_code=404, detail="Work order not found")
    
    root_id = wo_id
    if wo.get("parent_wo_id"):
        # Walk up to root
        current = wo
        while current.get("parent_wo_id"):
            parent = await db.work_orders.find_one({"id": current["parent_wo_id"]}, {"_id": 0})
            if not parent:
                break
            current = parent
            root_id = current["id"]
    
    return await build_tree(root_id)

@work_orders_router.put("/{wo_id}")
async def update_work_order(wo_id: str, wo_data: WorkOrderUpdate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="manufacturing", action="edit")
    wo = await db.work_orders.find_one({"id": wo_id})
    if not wo:
        raise HTTPException(status_code=404, detail="Work order not found")
    
    update_data = {k: v for k, v in wo_data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")

    # Cancellation is destructive — require delete permission.
    if update_data.get("status") == "cancelled" and wo.get("status") != "cancelled":
        _require_access(user, ["admin", "production_manager"], module="manufacturing", action="delete")
    
    # If trying to change status to in_progress, redirect to start endpoint
    if update_data.get("status") == "in_progress" and wo.get("status") == "pending":
        raise HTTPException(
            status_code=400, 
            detail="Use POST /api/work-orders/{wo_id}/start to start a work order (this will consume materials)"
        )

    # If cancelling, release any child stock reservations recorded against this MO,
    # and cascade-cancel all uncompleted child WOs (and their reservations).
    if update_data.get("status") == "cancelled" and wo.get("status") != "cancelled":
        for resv in (wo.get("child_reservations") or []):
            resv_item = resv.get("item_id")
            resv_qty = int(resv.get("qty", 0) or 0)
            if resv_item and resv_qty > 0:
                await db.items.update_one(
                    {"id": resv_item},
                    {"$inc": {"reserved_stock": -resv_qty}},
                )
        if wo.get("child_reservations"):
            update_data["child_reservations"] = []
            update_data["reservation_released_at"] = datetime.now(timezone.utc)

        # Fix: cascade-cancel every uncompleted descendant WO. BFS over the
        # parent_wo_id graph. Skips already-completed / already-cancelled WOs
        # (those represent real produced output and must NOT be undone).
        now_ts = datetime.now(timezone.utc)
        cancelled_children: List[str] = []
        frontier = [wo_id]
        while frontier:
            cursor = await db.work_orders.find(
                {"parent_wo_id": {"$in": frontier}, "status": {"$nin": ["completed", "cancelled"]}},
                {"_id": 0, "id": 1, "child_reservations": 1}
            ).to_list(5000)
            if not cursor:
                break
            child_ids = [c["id"] for c in cursor]
            # Release each child WO's own reservations before cancelling it.
            for child in cursor:
                for resv in (child.get("child_reservations") or []):
                    resv_item = resv.get("item_id")
                    resv_qty = int(resv.get("qty", 0) or 0)
                    if resv_item and resv_qty > 0:
                        await db.items.update_one(
                            {"id": resv_item},
                            {"$inc": {"reserved_stock": -resv_qty}},
                        )
            await db.work_orders.update_many(
                {"id": {"$in": child_ids}},
                {"$set": {
                    "status": "cancelled",
                    "child_reservations": [],
                    "reservation_released_at": now_ts,
                    "cancelled_by_parent": wo_id,
                    "updated_at": now_ts,
                }},
            )
            cancelled_children.extend(child_ids)
            frontier = child_ids
        if cancelled_children:
            update_data["cascade_cancelled_children"] = cancelled_children
    
    # If completing the work order, update finished goods stock
    if update_data.get("status") == "completed" and wo.get("status") == "in_progress":
        operations = wo.get("operations_status", [])
        mo_qty = wo.get("quantity", 0)
        
        # Block completion if ANY operation is not completed (enforced for ALL MOs — parent and child)
        for op in operations:
            if op.get("status") != "completed":
                raise HTTPException(status_code=400, detail=f"Cannot complete: Operation '{op.get('operation_name')}' (Seq {op.get('sequence')}) is not completed yet. Complete all operations via Job Card first.")
        
        # Block completion if subcontracted and materials not received
        if wo.get("is_subcontract"):
            sc_order = await db.subcontract_orders.find_one({"reference_wo_id": wo_id})
            if sc_order and sc_order.get("status") != "completed":
                raise HTTPException(status_code=400, detail="Cannot complete: Subcontracted materials have not been fully received. Please receive materials from subcontractor first.")
        
        # Block completion if any operation is outsourced and SC order not received
        for op in operations:
            if op.get("is_job_work") and op.get("outsource_status") == "sent":
                raise HTTPException(status_code=400, detail=f"Cannot complete: Outsourced operation '{op.get('operation_name')}' materials not received back. Receive from vendor first.")
        
        # Block completion if last operation produced less than MO quantity
        if operations:
            last_op = operations[-1]
            last_op_qty = last_op.get("quantity_completed", 0)
            if last_op_qty < mo_qty:
                raise HTTPException(status_code=400, detail=f"Cannot complete: Last operation produced {last_op_qty}/{mo_qty} units. Full quantity must be produced before completing the MO.")
        
        routing = await db.routings.find_one({"id": wo.get("routing_id")})
        item_id = routing.get("item_id") if routing else wo.get("item_id")
        if item_id:
            item = await db.items.find_one({"id": item_id})
            if item:
                current_stock = item.get("current_stock", 0)
                # Use actual produced qty from last operation
                last_op = operations[-1] if operations else {}
                produced_qty = last_op.get("quantity_accepted", last_op.get("quantity_completed", mo_qty))
                new_stock = current_stock + produced_qty
                
                # Create inventory transaction for produced items
                tx_doc = {
                    "id": str(uuid.uuid4()),
                    "item_id": item_id,
                    "transaction_type": "receive",
                    "quantity": produced_qty,
                    "reference_type": "work_order",
                    "reference_id": wo_id,
                    "previous_stock": current_stock,
                    "new_stock": new_stock,
                    "notes": f"Produced from WO {wo.get('wo_number')}",
                    "created_at": datetime.now(timezone.utc),
                    "created_by": user["id"]
                }
                await db.inventory_transactions.insert_one(tx_doc)
                
                # Update item stock
                await db.items.update_one(
                    {"id": item_id},
                    {"$set": {"current_stock": new_stock}}
                )
        
        update_data["actual_end"] = datetime.now(timezone.utc)
        update_data["quantity_completed"] = wo.get("quantity", 0)
    
    update_data["updated_at"] = datetime.now(timezone.utc)
    await db.work_orders.update_one({"id": wo_id}, {"$set": update_data})
    
    # If marking an already-started MO as subcontract, create SC Order + DC now
    if update_data.get("is_subcontract") and wo.get("status") == "in_progress":
        updated_wo_fresh = await db.work_orders.find_one({"id": wo_id})
        sc_supplier = updated_wo_fresh.get("subcontract_supplier_id") or update_data.get("subcontract_supplier_id")
        sc_type = updated_wo_fresh.get("subcontract_type", "with_material")
        if sc_supplier:
            existing_sc = await db.subcontract_orders.find_one({"reference_wo_id": wo_id})
            if not existing_sc:
                # Build lines based on subcontract type
                if sc_type == "without_material":
                    routing = await db.routings.find_one({"id": updated_wo_fresh.get("routing_id")})
                    wo_item_id = routing.get("item_id") if routing else updated_wo_fresh.get("item_id")
                    wo_item = await db.items.find_one({"id": wo_item_id}, {"_id": 0})
                    sc_lines_data = [{"item_id": wo_item_id, "quantity": updated_wo_fresh.get("quantity", 1), "unit_cost": wo_item.get("unit_cost", 0)}] if wo_item else []
                else:
                    consumed = updated_wo_fresh.get("consumed_materials", [])
                    sc_lines_data = consumed
                    if not sc_lines_data:
                        routing = await db.routings.find_one({"id": updated_wo_fresh.get("routing_id")})
                        wo_item_id = routing.get("item_id") if routing else updated_wo_fresh.get("item_id")
                        wo_item = await db.items.find_one({"id": wo_item_id}, {"_id": 0})
                        if wo_item:
                            sc_lines_data = [{"item_id": wo_item_id, "quantity": updated_wo_fresh.get("quantity", 1), "unit_cost": wo_item.get("unit_cost", 0)}]
                
                if sc_lines_data:
                    sc_count = await db.subcontract_orders.count_documents({})
                    sc_sent_qty = lambda m: m["quantity"] if sc_type == "with_material" else 0
                    # Get FG item name
                    upd_routing = await db.routings.find_one({"id": updated_wo_fresh.get("routing_id")})
                    upd_fg_item_id = upd_routing.get("item_id") if upd_routing else updated_wo_fresh.get("item_id")
                    upd_fg_item = await db.items.find_one({"id": upd_fg_item_id}, {"_id": 0})
                    sc_doc = {
                        "id": str(uuid.uuid4()),
                        "order_number": f"JW-{str(sc_count + 1).zfill(6)}",
                        "supplier_id": sc_supplier,
                        "reference_wo_id": wo_id,
                        "subcontract_type": sc_type,
                        "fg_item_id": upd_fg_item_id,
                        "fg_item_name": f"{upd_fg_item.get('part_number', '')} - {upd_fg_item.get('name', '')}" if upd_fg_item else "",
                        "fg_quantity": updated_wo_fresh.get("quantity", 0),
                        "lines": [{"item_id": m["item_id"], "quantity": m["quantity"], "sent_quantity": sc_sent_qty(m), "received_quantity": 0, "rate": m.get("unit_cost", 0)} for m in sc_lines_data],
                        "status": "in_progress",
                        "notes": f"Auto-created from sub-contract MO {updated_wo_fresh.get('wo_number')} ({sc_type.replace('_', ' ')})",
                        "created_at": datetime.now(timezone.utc),
                        "created_by": user["id"]
                    }
                    await db.subcontract_orders.insert_one(sc_doc)
                    sc_doc.pop("_id", None)
                    
                    # Only create DC for "with_material" type
                    if sc_type == "with_material":
                        dc_lines = [{"item_id": m["item_id"], "quantity": m["quantity"], "rate": m.get("unit_cost", 0)} for m in sc_lines_data]
                        if dc_lines:
                            dc_doc = {
                                "id": str(uuid.uuid4()),
                                "dc_number": await get_next_series_number("delivery_challan"),
                                "subcontract_order_id": sc_doc["id"],
                                "reference_wo_id": wo_id,
                                "lines": dc_lines,
                                "status": "draft",
                                "notes": f"Auto-DC for sub-contract MO {updated_wo_fresh.get('wo_number')}",
                                "created_at": datetime.now(timezone.utc),
                                "created_by": user["id"]
                            }
                            await db.delivery_challans.insert_one(dc_doc)
    
    updated_wo = await db.work_orders.find_one({"id": wo_id}, {"_id": 0})
    return updated_wo

@work_orders_router.put("/{wo_id}/operations/{sequence}")
async def update_work_order_operation(wo_id: str, sequence: int, op_data: WorkOrderOperationUpdate, request: Request):
    """Update a specific operation status within a work order (Job Card)"""
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="manufacturing", action="edit")
    wo = await db.work_orders.find_one({"id": wo_id})
    if not wo:
        raise HTTPException(status_code=404, detail="Work order not found")
    
    if wo.get("status") == "pending":
        raise HTTPException(status_code=400, detail="Cannot start Job Card operation: the Manufacturing Order is still pending. Please click 'Inhouse Start' on the MO first (this will consume RM from inventory), then re-open the Job Card.")
    if wo.get("status") not in ["in_progress"]:
        raise HTTPException(status_code=400, detail=f"Cannot update operations on {wo.get('status')} manufacturing order")
    
    operations = wo.get("operations_status", [])
    target_op = None
    target_idx = None
    
    for idx, op in enumerate(operations):
        if op.get("sequence") == sequence:
            target_op = op
            target_idx = idx
            break
    
    if target_op is None:
        raise HTTPException(status_code=404, detail="Operation not found")
    
    mo_qty = wo.get("quantity", 0)
    
    # START operation
    if op_data.status == "in_progress":
        for prev_op in operations[:target_idx]:
            if prev_op.get("status") not in ["completed", "stopped"]:
                raise HTTPException(status_code=400, detail=f"Previous operation '{prev_op.get('operation_name')}' must be completed first")
        
        is_outsource = op_data.is_outsource or False
        
        if is_outsource:
            if not op_data.outsource_supplier_id:
                raise HTTPException(status_code=400, detail="Supplier is required for outsourced operations")
            
            supplier = await db.suppliers.find_one({"id": op_data.outsource_supplier_id}, {"_id": 0})
            supplier_name = supplier.get("name", "Outsourced") if supplier else "Outsourced"

            # Partial OS qty handling: if the user entered a smaller qty than
            # the full MO, only that subset is shipped to the vendor — the
            # remainder remains on the operation as pending so the user can
            # still Start it in-house (or outsource it again to a different
            # vendor later). When omitted or >= MO qty, behave as before
            # (full operation outsourced).
            os_qty = float(op_data.outsource_quantity or 0)
            if os_qty <= 0 or os_qty > mo_qty:
                os_qty = float(mo_qty)
            is_partial_os = os_qty < float(mo_qty)
            
            # For operation outsourcing: SC lines = Part/SA item only (NOT RM)
            wo_item = await db.items.find_one({"id": wo.get("item_id")}, {"_id": 0})
            op_name = target_op.get("operation_name", "")
            
            # Pull RM cost from BOM (material only, no process cost)
            bom_costs = await compute_bom_costs(wo.get("item_id"))
            bom_rollup_cost = bom_costs["rm_cost"] or (wo_item.get("unit_cost", 0) if wo_item else 0)
            
            # Default charges: look up matching operation in ANY BOM (item as parent or component)
            outsource_charges = op_data.outsource_charges or 0
            if not outsource_charges and op_name:
                outsource_charges = await find_routing_cost(wo.get("item_id"), op_name)
            
            # Fallback: pull from previous SC
            if not outsource_charges:
                prev_sc_op = await db.subcontract_orders.find_one(
                    {"job_work_parts.item_id": wo.get("item_id"), "status": {"$in": ["in_progress", "completed"]}},
                    {"_id": 0, "job_work_parts": 1},
                    sort=[("created_at", -1)]
                )
                if prev_sc_op:
                    for pjwp in prev_sc_op.get("job_work_parts", []):
                        if pjwp.get("item_id") == wo.get("item_id") and pjwp.get("charges"):
                            outsource_charges = pjwp["charges"]
                            break
            
            sc_part = {"item_id": wo.get("item_id"), "quantity": os_qty, "charges": outsource_charges, "received_quantity": 0, "bom_rollup_cost": round(bom_rollup_cost, 2), "process_name": op_name, "wo_id": wo_id}
            sc_lines = []  # No RM lines for operation outsourcing — only the part goes and comes back
            
            # Check for existing SC order for same supplier (consolidate across all MOs).
            # Consolidate ONLY if:
            #   (a) SC is still open (draft/in_progress),
            #   (b) It's a Job Card OS SC (has reference_operation_seqs — so we don't merge
            #       into a plain MO→SC), AND
            #   (c) NO delivery challan has been sent for it yet (dc_created must be falsy AND
            #       no sent DC document linked). Once the vendor has received materials, the
            #       next outsource for the same vendor MUST go into a fresh SC so that the
            #       next DC/GRN cycle is independent.
            candidate_scs = db.subcontract_orders.find({
                "supplier_id": op_data.outsource_supplier_id,
                "status": {"$in": ["draft", "in_progress"]},
                "subcontract_type": "without_material",
                "reference_operation_seqs": {"$exists": True, "$ne": []},
                "$or": [
                    {"dc_created": {"$exists": False}},
                    {"dc_created": None},
                    {"dc_created": False}
                ]
            }, {"_id": 0}).sort("created_at", -1)
            existing_sc = None
            async for _cand in candidate_scs:
                # Double-check: no DC (draft OR sent) already linked to this SC
                linked_dc = await db.delivery_challans.find_one({
                    "subcontract_order_id": _cand["id"],
                    "status": {"$in": ["sent", "draft"]}
                })
                if not linked_dc:
                    existing_sc = _cand
                    break
            
            if existing_sc:
                # Consolidate into existing SC — add this part to job_work_parts
                jwp = existing_sc.get("job_work_parts", [])
                # Check if this item + process already exists in job_work_parts (for consolidation)
                found_jwp = False
                for jp in jwp:
                    if jp.get("item_id") == wo.get("item_id") and jp.get("process_name", "") == target_op.get("operation_name", ""):
                        jp["quantity"] += os_qty
                        # charges are per-unit — update only if not already set
                        if not jp.get("charges") and outsource_charges:
                            jp["charges"] = outsource_charges
                        found_jwp = True
                        break
                if not found_jwp:
                    jwp.append(sc_part)
                
                ref_ops = existing_sc.get("reference_operation_seqs", [existing_sc.get("reference_operation_seq")])
                if sequence not in ref_ops:
                    ref_ops.append(sequence)
                
                # Track all reference WO IDs
                ref_wo_ids = existing_sc.get("reference_wo_ids", [existing_sc.get("reference_wo_id")])
                if wo_id not in ref_wo_ids:
                    ref_wo_ids.append(wo_id)
                
                total_charges = sum(p.get("charges", 0) * p.get("quantity", 1) for p in jwp)
                
                await db.subcontract_orders.update_one({"id": existing_sc["id"]}, {"$set": {
                    "job_work_parts": jwp,
                    "reference_wo_ids": ref_wo_ids,
                    "processing_charges": total_charges,
                    "reference_operation_seqs": ref_ops,
                    "status": "in_progress",
                    "notes": f"Consolidated Job OS - {len(jwp)} parts, {len(ref_wo_ids)} MOs",
                    "updated_at": datetime.now(timezone.utc)
                }})
                
                sc_order_doc = existing_sc
            else:
                # Create new SC Order — Part/SA only, no RM
                sc_count = await db.subcontract_orders.count_documents({})
                sc_order_number = f"JW-{str(sc_count + 1).zfill(6)}"
                
                sc_order_doc = {
                    "id": str(uuid.uuid4()),
                    "order_number": sc_order_number,
                    "supplier_id": op_data.outsource_supplier_id,
                    "subcontract_type": "without_material",
                    "reference_wo_id": wo_id,
                    "reference_wo_ids": [wo_id],
                    "reference_operation_seq": sequence,
                    "reference_operation_seqs": [sequence],
                    "job_work_parts": [sc_part],
                    "lines": [],  # No RM lines
                    "processing_charges": outsource_charges,
                    "dc_created": False,
                    "status": "in_progress",
                    "notes": f"Operation outsource: {target_op.get('operation_name')} on MO {wo.get('wo_number')}",
                    "created_at": datetime.now(timezone.utc),
                    "created_by": user["id"]
                }
                await db.subcontract_orders.insert_one(sc_order_doc)
            
            target_op["is_job_work"] = True
            target_op["job_work_supplier_id"] = op_data.outsource_supplier_id
            target_op["outsource_status"] = "sent"
            target_op["outsource_supplier_name"] = supplier_name
            target_op["outsource_charges"] = op_data.outsource_charges or 0
            target_op["outsource_sc_order_id"] = sc_order_doc["id"]
            target_op["outsource_sc_order_number"] = sc_order_doc.get("order_number", "")
            target_op["operator"] = f"OS: {supplier_name}"
            target_op["actual_start"] = target_op.get("actual_start") or datetime.now(timezone.utc)
            # For PARTIAL OS we want the user to still be able to Start the
            # remaining qty in-house. So we leave the operation in `pending`
            # state when only part of the qty has gone out; only flip to
            # `in_progress` once the full qty is outsourced.
            target_op["status"] = "pending" if is_partial_os else "in_progress"
            target_op["outsourced_quantity"] = (target_op.get("outsourced_quantity") or 0) + os_qty

            runs = target_op.get("runs", [])
            runs.append({
                "run_number": len(runs) + 1,
                "operator": f"OS: {supplier_name}",
                "quantity_planned": os_qty,
                "quantity_completed": 0,
                "started_at": datetime.now(timezone.utc),
                "ended_at": None,
                "quality_result": None,
                "reject_qty": 0,
                "rework_qty": 0,
                "notes": f"Outsourced to {supplier_name}",
                # Per-run SC info — required so per-vendor Revoke / Short Close
                # can find the right SC line (different vendors create different
                # SCs OR different lines on a shared SC). Without this, multi-
                # vendor partial OS can't be individually managed.
                "outsource_sc_order_id": sc_order_doc["id"],
                "outsource_sc_order_number": sc_order_doc.get("order_number", ""),
                "outsource_supplier_id": op_data.outsource_supplier_id,
                "outsource_supplier_name": supplier_name,
            })
            target_op["runs"] = runs
        else:
            if not op_data.operator or not op_data.operator.strip():
                raise HTTPException(status_code=400, detail="Operator name is required to start an operation")
            
            # Initialize or append to runs list
            runs = target_op.get("runs", [])
            # Prevent over-allocation across parallel operators
            already_allocated = sum((r.get("quantity_completed") or 0) if r.get("ended_at") else (r.get("quantity_planned") or r.get("quantity_completed") or 0) for r in runs)
            remaining_to_allocate = max(0, mo_qty - already_allocated)
            req_qty = op_data.quantity_completed or remaining_to_allocate or mo_qty
            if remaining_to_allocate <= 0:
                raise HTTPException(status_code=400, detail=f"Cannot start: all {mo_qty} units are already allocated across operators. Stop/complete existing runs first.")
            if req_qty > remaining_to_allocate:
                raise HTTPException(status_code=400, detail=f"Requested quantity {req_qty} exceeds remaining unallocated quantity {remaining_to_allocate}.")
            planned_qty = req_qty
            run_entry = {
                "run_number": len(runs) + 1,
                "operator": op_data.operator.strip(),
                "quantity_planned": planned_qty,
                "quantity_completed": 0,
                "started_at": datetime.now(timezone.utc),
                "ended_at": None,
                "quality_result": None,
                "reject_qty": 0,
                "rework_qty": 0,
                "notes": op_data.notes or ""
            }
            runs.append(run_entry)
            target_op["runs"] = runs
            target_op["actual_start"] = target_op.get("actual_start") or datetime.now(timezone.utc)
            target_op["operator"] = op_data.operator.strip()
            target_op["status"] = "in_progress"
            # Save work center if provided
            if op_data.work_center_id:
                wc = await db.work_centers.find_one({"id": op_data.work_center_id}, {"_id": 0})
                target_op["work_center_id"] = op_data.work_center_id
                target_op["work_center_name"] = wc.get("name", "") if wc else ""
            if op_data.process_cost_per_unit is not None:
                target_op["process_cost_per_unit"] = op_data.process_cost_per_unit
    
    # STOP operation (per-run)
    elif op_data.status == "stopped":
        runs = target_op.get("runs", [])
        produced_qty = min(op_data.quantity_completed or 0, mo_qty)
        # Pick which run to close: explicit run_number > open run for operator > last open run
        target_run_idx = None
        if op_data.run_number is not None:
            for _i, _r in enumerate(runs):
                if _r.get("run_number") == op_data.run_number and _r.get("ended_at") is None:
                    target_run_idx = _i
                    break
        if target_run_idx is None and op_data.operator:
            for _i in range(len(runs) - 1, -1, -1):
                if runs[_i].get("ended_at") is None and runs[_i].get("operator") == op_data.operator:
                    target_run_idx = _i
                    break
        if target_run_idx is None:
            for _i in range(len(runs) - 1, -1, -1):
                if runs[_i].get("ended_at") is None:
                    target_run_idx = _i
                    break
        if target_run_idx is None:
            raise HTTPException(status_code=400, detail="No open run to stop for this operation")
        runs[target_run_idx]["ended_at"] = datetime.now(timezone.utc)
        runs[target_run_idx]["quantity_completed"] = produced_qty
        runs[target_run_idx]["quality_result"] = op_data.quality_result or "accept"
        runs[target_run_idx]["reject_qty"] = min(op_data.reject_qty or 0, produced_qty)
        runs[target_run_idx]["rework_qty"] = min(op_data.rework_qty or 0, produced_qty)
        runs[target_run_idx]["notes"] = op_data.notes or runs[target_run_idx].get("notes", "")

        total_accepted = sum(r.get("quantity_completed", 0) - r.get("reject_qty", 0) - r.get("rework_qty", 0) for r in runs)
        total_completed = sum(r.get("quantity_completed", 0) for r in runs)
        target_op["quantity_completed"] = total_completed
        target_op["quantity_accepted"] = total_accepted
        target_op["quantity_rejected"] = sum(r.get("reject_qty", 0) for r in runs)
        target_op["quantity_rework"] = sum(r.get("rework_qty", 0) for r in runs)
        target_op["runs"] = runs
        # Op status derived from runs: in_progress if ANY run still open, else stopped
        target_op["status"] = "in_progress" if any(r.get("ended_at") is None for r in runs) else "stopped"
    
    # COMPLETE operation (per-run)
    elif op_data.status == "completed":
        # Block outsourced operation completion if SC order not received
        if target_op.get("is_job_work"):
            sc_order_id = target_op.get("outsource_sc_order_id")
            if sc_order_id:
                sc_order = await db.subcontract_orders.find_one({"id": sc_order_id})
                if sc_order and sc_order.get("status") != "completed":
                    raise HTTPException(status_code=400, detail=f"Cannot complete outsourced operation: Materials not received back from subcontractor. Receive items via Job Work page first.")
            elif target_op.get("outsource_status") == "sent":
                raise HTTPException(status_code=400, detail=f"Cannot complete outsourced operation: Materials not received back. Receive items via Job Work page first.")
        
        runs = target_op.get("runs", [])
        produced_qty = min(op_data.quantity_completed or mo_qty, mo_qty)
        # Pick run to close (same precedence as stop)
        target_run_idx = None
        if op_data.run_number is not None:
            for _i, _r in enumerate(runs):
                if _r.get("run_number") == op_data.run_number and _r.get("ended_at") is None:
                    target_run_idx = _i
                    break
        if target_run_idx is None and op_data.operator:
            for _i in range(len(runs) - 1, -1, -1):
                if runs[_i].get("ended_at") is None and runs[_i].get("operator") == op_data.operator:
                    target_run_idx = _i
                    break
        if target_run_idx is None:
            for _i in range(len(runs) - 1, -1, -1):
                if runs[_i].get("ended_at") is None:
                    target_run_idx = _i
                    break
        # Close the target run (if any open)
        if target_run_idx is not None:
            runs[target_run_idx]["ended_at"] = datetime.now(timezone.utc)
            runs[target_run_idx]["quantity_completed"] = produced_qty
            runs[target_run_idx]["quality_result"] = op_data.quality_result or "accept"
            runs[target_run_idx]["reject_qty"] = min(op_data.reject_qty or 0, produced_qty)
            runs[target_run_idx]["rework_qty"] = min(op_data.rework_qty or 0, produced_qty)
        
        total_completed = sum(r.get("quantity_completed", 0) for r in runs)
        total_accepted = sum(r.get("quantity_completed", 0) - r.get("reject_qty", 0) - r.get("rework_qty", 0) for r in runs)
        
        target_op["quantity_completed"] = total_completed
        target_op["quantity_accepted"] = total_accepted
        target_op["quantity_rejected"] = sum(r.get("reject_qty", 0) for r in runs)
        target_op["quantity_rework"] = sum(r.get("rework_qty", 0) for r in runs)
        target_op["runs"] = runs
        # Status derivation: any open run → in_progress; else, completed ONLY if target_op qty >= mo_qty
        if any(r.get("ended_at") is None for r in runs):
            target_op["status"] = "in_progress"
        elif total_completed >= mo_qty:
            target_op["status"] = "completed"
            target_op["actual_end"] = datetime.now(timezone.utc)
        else:
            # All runs closed but partial qty — user should allocate remaining first
            target_op["status"] = "stopped"
        
        # Calculate actual time only when truly completed
        if target_op["status"] == "completed" and target_op.get("actual_start"):
            actual_start = target_op["actual_start"]
            if isinstance(actual_start, str):
                actual_start = datetime.fromisoformat(actual_start.replace('Z', '+00:00'))
            actual_end = target_op["actual_end"]
            if actual_start.tzinfo is None:
                actual_start = actual_start.replace(tzinfo=timezone.utc)
            if actual_end.tzinfo is None:
                actual_end = actual_end.replace(tzinfo=timezone.utc)
            delta = (actual_end - actual_start).total_seconds() / 60
            target_op["actual_time_min"] = round(delta, 1)
    
    if op_data.notes:
        target_op["notes"] = op_data.notes
    
    # Save process cost per unit if provided (explicit override)
    if op_data.process_cost_per_unit is not None and op_data.process_cost_per_unit > 0:
        target_op["process_cost_per_unit"] = op_data.process_cost_per_unit
    elif op_data.status in ("stopped", "completed") and not target_op.get("is_job_work"):
        # Auto-compute from Work Center hourly_rate × total run duration ÷ total qty
        try:
            wc_id = target_op.get("work_center_id")
            wc_doc = await db.work_centers.find_one({"id": wc_id}, {"_id": 0}) if wc_id else None
            hourly_rate = float((wc_doc or {}).get("hourly_rate", 0) or 0)
            total_minutes = 0.0
            for _r in (target_op.get("runs") or []):
                _s = _r.get("started_at")
                _e = _r.get("ended_at")
                if _s and _e:
                    _ds = _s if isinstance(_s, datetime) else datetime.fromisoformat(str(_s).replace('Z', '+00:00'))
                    _de = _e if isinstance(_e, datetime) else datetime.fromisoformat(str(_e).replace('Z', '+00:00'))
                    if _ds.tzinfo is None: _ds = _ds.replace(tzinfo=timezone.utc)
                    if _de.tzinfo is None: _de = _de.replace(tzinfo=timezone.utc)
                    total_minutes += max(0.0, (_de - _ds).total_seconds() / 60)
            total_qty = sum((_r.get("quantity_completed") or 0) for _r in (target_op.get("runs") or []))
            total_labor_cost = (total_minutes / 60.0) * hourly_rate
            if total_qty > 0 and hourly_rate > 0:
                target_op["process_cost_per_unit"] = round(total_labor_cost / total_qty, 2)
                target_op["total_labor_cost"] = round(total_labor_cost, 2)
        except Exception:
            pass
    
    # Auto-determine WO status. Treat short-closed ops as effectively
    # completed (legacy short-close paths may not have flipped status).
    def _op_done(op):
        return op.get("status") == "completed" or op.get("short_closed") is True
    all_completed = all(_op_done(op) for op in operations)
    any_in_progress = any(
        op.get("status") in ["in_progress", "completed", "stopped"] or op.get("short_closed") is True
        for op in operations
    )
    
    wo_status = wo.get("status")
    if all_completed:
        # Check subcontract receipt before auto-completing
        can_complete = True
        if wo.get("is_subcontract"):
            sc_order = await db.subcontract_orders.find_one({"reference_wo_id": wo_id})
            if sc_order and sc_order.get("status") != "completed":
                can_complete = False
        # Check any outsourced operation has pending receipt — but skip
        # short-closed ops; those are settled regardless of outsource_status.
        for op in operations:
            if op.get("short_closed"):
                continue
            if op.get("is_job_work") and op.get("outsource_status") == "sent":
                can_complete = False
        
        if can_complete:
            wo_status = "completed"
    elif any_in_progress:
        wo_status = "in_progress"
    
    update_fields = {
        "operations_status": operations,
        "status": wo_status,
        "updated_at": datetime.now(timezone.utc)
    }
    
    # If all operations completed, update stock
    if all_completed and wo_status == "completed":
        # Calculate final accepted qty from last operation
        last_op = operations[-1] if operations else {}
        final_accepted = last_op.get("quantity_accepted", mo_qty)
        
        item = await db.items.find_one({"id": wo.get("item_id")})
        if item:
            # If the WO has a variant_selection, credit stock to the FG/SG variant
            # child SKU instead of the parent. Auto-create the child SKU if it
            # doesn't exist yet (e.g. user added a new BOM-component variant after
            # the parent was first set up).
            variant_sel = wo.get("variant_selection") or None
            target_item_id = item["id"]
            target_sku = item.get("part_number") or ""
            if variant_sel:
                eff = await _get_effective_variants(item)
                # Build the {attr -> {value, short_code}} combo from the selection.
                combo: Dict[str, Dict[str, str]] = {}
                for attr in eff:
                    v = variant_sel.get(attr["name"])
                    if v is None:
                        continue
                    match = next((x for x in attr["values"] if x["value"] == v), None)
                    if match:
                        combo[attr["name"]] = {"value": match["value"], "short_code": match["short_code"]}
                if combo:
                    variant_sku = _build_variant_sku_from_short_codes(item.get("part_number") or "", combo)
                    existing = await db.items.find_one({"part_number": variant_sku}, {"_id": 0})
                    if not existing:
                        # Auto-create the variant child item (mirrors generate_item_variants logic).
                        short_codes = {k: v["short_code"] for k, v in combo.items()}
                        values_map = {k: v["value"] for k, v in combo.items()}
                        child = {k: v for k, v in item.items() if k not in ("_id", "id", "variant_attributes", "auto_suffix_variant_sku")}
                        child["id"] = str(uuid.uuid4())
                        child["part_number"] = variant_sku
                        child["name"] = (item.get("name", "") or "") + " · " + ", ".join(f"{k}: {v}" for k, v in values_map.items())
                        child["parent_item_id"] = item["id"]
                        child["is_variant"] = True
                        child["variant_short_codes"] = short_codes
                        child["variant_values"] = values_map
                        child["variant_attributes"] = None
                        child["is_active"] = True
                        child["current_stock"] = 0
                        child["reserved_stock"] = 0
                        child["created_at"] = datetime.now(timezone.utc)
                        await db.items.insert_one(child)
                        target_item_id = child["id"]
                    else:
                        target_item_id = existing["id"]
                    target_sku = variant_sku
            await db.items.update_one(
                {"id": target_item_id},
                {"$inc": {"current_stock": final_accepted}}
            )
            update_fields["fg_credited_item_id"] = target_item_id
            update_fields["fg_credited_sku"] = target_sku
        update_fields["actual_end"] = datetime.now(timezone.utc)
        update_fields["quantity_completed"] = final_accepted
    
    await db.work_orders.update_one({"id": wo_id}, {"$set": update_fields})
    
    return await db.work_orders.find_one({"id": wo_id}, {"_id": 0})

@work_orders_router.get("/{wo_id}/print-data")
async def get_work_order_print_data(wo_id: str, request: Request):
    """Get full work order data for printing (WO sheet + Job Card)"""
    await get_current_user(request)
    
    wo = await db.work_orders.find_one({"id": wo_id}, {"_id": 0})
    if not wo:
        raise HTTPException(status_code=404, detail="Work order not found")
    
    # Get item details
    item = await db.items.find_one({"id": wo.get("item_id")}, {"_id": 0})
    wo["item"] = item
    
    # Get routing details
    routing = await db.routings.find_one({"id": wo.get("routing_id")}, {"_id": 0})
    wo["routing"] = routing
    
    # Get work center names (+ hourly rate) for operations
    for op in wo.get("operations_status", []):
        wc = await db.work_centers.find_one({"id": op.get("work_center_id")}, {"_id": 0})
        op["work_center_name"] = wc.get("name", "") if wc else ""
        op["work_center"] = wc  # Full WC object (contains hourly_rate) used by Job Card print
    
    # Get consumed materials (stored on WO doc or from inventory transactions)
    consumed = wo.get("consumed_materials", [])
    if not consumed and wo.get("materials_consumed"):
        # Fallback: fetch from inventory transactions
        txns = await db.inventory_transactions.find(
            {"reference_id": wo_id, "transaction_type": "issue"}, {"_id": 0}
        ).to_list(100)
        for tx in txns:
            tx_item = await db.items.find_one({"id": tx.get("item_id")}, {"_id": 0})
            consumed.append({
                "item_id": tx.get("item_id"),
                "item": tx_item.get("part_number", "") if tx_item else "",
                "name": tx_item.get("name", "") if tx_item else "",
                "quantity": tx.get("quantity", 0),
                "uom": tx_item.get("unit_of_measure", "pcs") if tx_item else "pcs",
                "unit_cost": tx_item.get("unit_cost", 0) if tx_item else 0
            })
    wo["consumed_materials"] = consumed
    
    # Get child manufacturing orders
    child_mos = await db.work_orders.find({"parent_wo_id": wo_id}, {"_id": 0}).to_list(100)
    for child in child_mos:
        child_item = await db.items.find_one({"id": child.get("item_id")}, {"_id": 0})
        child["item"] = child_item
    wo["child_mos"] = child_mos
    
    # Get company settings for header
    company = await db.company_settings.find_one({"type": "company"}, {"_id": 0})
    wo["company"] = company
    
    return wo

# ================== EXPORT / IMPORT ROUTES ==================

@items_router.get("/export/excel")
async def export_items_excel(request: Request, category: Optional[str] = None, group_id: Optional[str] = None):
    """Export items to Excel. Optional `category` filter: raw_material | component | sub_assembly | finished_good | all.
    Optional `group_id` filter restricts the export to a single Item Group (e.g. only fasteners)."""
    await get_current_user(request)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    try:
        query = {}
        valid_cats = {"raw_material", "component", "sub_assembly", "finished_good"}
        if category and category != "all" and category in valid_cats:
            query["category"] = category
        # Item-Group filter — when provided, scope export to that group only.
        group_label = None
        if group_id:
            grp = await db.item_groups.find_one({"id": group_id}, {"_id": 0})
            if not grp:
                raise HTTPException(status_code=404, detail=f"Item Group '{group_id}' not found")
            query["group_id"] = group_id
            group_label = grp.get("name") or grp.get("code") or "Group"
        items = await db.items.find(query, {"_id": 0}).to_list(10000)

        wb = Workbook()
        ws = wb.active
        cat_label_map = {"raw_material": "Raw Materials", "component": "Parts", "sub_assembly": "Sub-Assemblies", "finished_good": "Finished Goods"}
        # Compose worksheet title from group + category labels (Excel sheet names cap at 31 chars).
        title_parts = []
        if group_label:
            title_parts.append(group_label)
        if category and category != "all":
            title_parts.append(cat_label_map.get(category, category))
        ws.title = (" - ".join(title_parts) or "Items Master")[:31]

        headers = ["Part Number", "Name", "Description", "Category", "Group", "UOM", "Purchase Cost", "Sales Price", "Lead Time (Days)", "Safety Stock", "Current Stock", "Reorder Point", "HSN Code", "GST Rate (%)"]
        header_fill = PatternFill(start_color="1D3557", end_color="1D3557", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # Safe value coercion — openpyxl rejects dicts/lists/None in some versions
        def _safe_str(v):
            if v is None:
                return ""
            if isinstance(v, (dict, list)):
                return str(v)
            return v

        def _safe_num(v, default=0):
            if v is None:
                return default
            try:
                return float(v) if isinstance(v, (int, float, str)) and str(v).strip() != "" else default
            except (ValueError, TypeError):
                return default

        # Build group lookup for export (only once)
        group_map = {g["id"]: g.get("name", "") async for g in db.item_groups.find({}, {"_id": 0, "id": 1, "name": 1})}

        for row, item in enumerate(items, 2):
            data = [
                _safe_str(item.get("part_number", "")),
                _safe_str(item.get("name", "")),
                _safe_str(item.get("description", "")),
                _safe_str(item.get("category", "")),
                _safe_str(group_map.get(item.get("group_id"), "")),
                _safe_str(item.get("unit_of_measure", "")),
                _safe_num(item.get("purchase_price") or item.get("unit_cost"), 0),
                _safe_num(item.get("sale_price") or item.get("sales_price") or item.get("standard_price"), 0),
                _safe_num(item.get("lead_time_days"), 0),
                _safe_num(item.get("safety_stock"), 0),
                _safe_num(item.get("current_stock"), 0),
                _safe_num(item.get("reorder_point"), 0),
                _safe_str(item.get("hsn_code", "")),
                _safe_num(item.get("gst_rate"), 18),
            ]
            for col, value in enumerate(data, 1):
                try:
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = thin_border
                except Exception as cell_err:
                    # If one cell fails, substitute with stringified fallback so the whole export doesn't 500
                    logger.warning(f"[export_items] row={row} col={col} value={value!r} err={cell_err}; coerced to str")
                    ws.cell(row=row, column=col, value=str(value)).border = thin_border

        for col in range(1, len(headers) + 1):
            # Safe column letter (handles up to 52 cols via AA-AZ path)
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(col)].width = 18

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename_map = {"raw_material": "items_raw_materials.xlsx", "component": "items_parts.xlsx",
                        "sub_assembly": "items_sub_assemblies.xlsx", "finished_good": "items_finished_goods.xlsx"}
        out_name = filename_map.get(category, "items_master.xlsx") if category and category != "all" else "items_master.xlsx"

        # Read bytes once — avoids streaming-generator exceptions that would strip CORS headers
        data_bytes = output.getvalue()
        return Response(
            content=data_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={out_name}",
                "Content-Length": str(len(data_bytes)),
            },
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[export_items_excel] Unhandled error (category={category}): {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Export failed: {type(e).__name__}: {e}")

@items_router.post("/import/excel")
async def import_items_excel(request: Request, file: UploadFile = File(...)):
    """Import items from Excel file"""
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager", "inventory_manager"], module="items", action="create")
    from openpyxl import load_workbook
    
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    results = {"created": 0, "updated": 0, "errors": []}
    
    # Map header names to field names
    # NOTE: "Purchase Cost" maps to `purchase_price` (the field used in PO creation).
    # "Unit Cost" is retained as a fallback for backward compatibility with older templates.
    field_map = {
        "Part Number": "part_number", "Name": "name", "Description": "description",
        "Category": "category", "Group": "group_name", "UOM": "unit_of_measure",
        "Purchase Cost": "purchase_price", "Unit Cost": "unit_cost",
        "Sales Price": "sale_price",
        "Lead Time (Days)": "lead_time_days", "Safety Stock": "safety_stock",
        "Current Stock": "current_stock", "Reorder Point": "reorder_point",
        "HSN Code": "hsn_code", "GST Rate (%)": "gst_rate"
    }
    
    col_indices = {}
    for idx, header in enumerate(headers):
        if header in field_map:
            col_indices[field_map[header]] = idx
    
    # Pre-load existing groups for fast matching (name -> id)
    existing_groups = {g["name"].lower(): g["id"] async for g in db.item_groups.find({}, {"_id": 0, "id": 1, "name": 1})}

    # PERF: previously each row issued its own find_one + insert_one/update_one
    # to MongoDB, which made imports of >300 rows take many seconds (round-trip
    # per row x 2). Now we:
    #   1. Pre-fetch ALL existing items keyed by part_number (single query).
    #   2. Buffer inserts and updates locally.
    #   3. Flush via insert_many() + a single bulk_write() for updates.
    existing_items_by_pn = {
        it["part_number"]: it
        async for it in db.items.find(
            {"part_number": {"$ne": None}},
            {"_id": 0, "id": 1, "part_number": 1},
        )
        if it.get("part_number")
    }
    inserts: list = []
    updates: list = []  # list of (part_number, item_data)
    new_groups: list = []  # rows from cache that need to be flushed

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        try:
            if not row or not row[0]:
                continue

            item_data = {}
            group_name_raw = None
            for field, col_idx in col_indices.items():
                if col_idx < len(row) and row[col_idx] is not None:
                    val = row[col_idx]
                    if field == "group_name":
                        group_name_raw = str(val).strip() if val else None
                        continue
                    if field in ["unit_cost", "purchase_price", "sale_price", "gst_rate"]:
                        val = float(val) if val else 0
                    elif field in ["lead_time_days", "safety_stock", "current_stock", "reorder_point"]:
                        val = int(val) if val else 0
                    else:
                        val = str(val).strip()
                    item_data[field] = val

            # If "Group" column provided, find-or-create the group LOCALLY
            # against the in-memory cache. Newly-created groups are queued
            # for a single insert_many() flush at the end.
            if group_name_raw:
                parent_cat = item_data.get("category")
                gid = existing_groups.get(group_name_raw.lower())
                if not gid:
                    gid = str(uuid.uuid4())
                    new_groups.append({
                        "id": gid,
                        "name": group_name_raw,
                        "parent_category": parent_cat,
                        "default_hsn_code": None,
                        "default_gst_rate": None,
                        "description": "",
                        "created_at": datetime.now(timezone.utc),
                        "created_by": user["id"],
                    })
                    existing_groups[group_name_raw.lower()] = gid
                item_data["group_id"] = gid

            if not item_data.get("part_number"):
                results["errors"].append(f"Row {row_num}: Missing part number")
                continue

            # Validate category
            valid_cats = ["raw_material", "component", "sub_assembly", "finished_good"]
            if item_data.get("category") and item_data["category"] not in valid_cats:
                results["errors"].append(f"Row {row_num}: Invalid category '{item_data['category']}'")
                continue

            # Mirror purchase_price → unit_cost (if unit_cost missing) so downstream BOM/valuation logic still works
            if item_data.get("purchase_price") and not item_data.get("unit_cost"):
                item_data["unit_cost"] = item_data["purchase_price"]

            pn = item_data["part_number"]
            if pn in existing_items_by_pn:
                updates.append((pn, item_data))
                results["updated"] += 1
            else:
                item_data["id"] = str(uuid.uuid4())
                item_data.setdefault("unit_of_measure", "pcs")
                item_data.setdefault("category", "raw_material")
                item_data.setdefault("gst_rate", 18)
                item_data["created_at"] = datetime.now(timezone.utc)
                inserts.append(item_data)
                # Mark in cache so a duplicate row in the same file won't
                # be inserted twice — second occurrence becomes an update.
                existing_items_by_pn[pn] = {"id": item_data["id"], "part_number": pn}
                results["created"] += 1
        except Exception as e:  # noqa: BLE001 — surface per-row failures, keep going.
            results["errors"].append(f"Row {row_num}: {str(e)}")

    # ---- Flush DB writes in bulk ------------------------------------------------
    if new_groups:
        try:
            await db.item_groups.insert_many(new_groups, ordered=False)
        except Exception as e:  # noqa: BLE001
            results["errors"].append(f"item_groups bulk insert: {str(e)}")
    if inserts:
        try:
            await db.items.insert_many(inserts, ordered=False)
        except Exception as e:  # noqa: BLE001
            results["errors"].append(f"items bulk insert: {str(e)}")
    if updates:
        # Use bulk_write with one UpdateOne per row — still one DB round-trip.
        from pymongo import UpdateOne
        ops = [UpdateOne({"part_number": pn}, {"$set": data}) for pn, data in updates]
        try:
            await db.items.bulk_write(ops, ordered=False)
        except Exception as e:  # noqa: BLE001
            results["errors"].append(f"items bulk update: {str(e)}")

    return results

@bom_router.get("/export/parts-only/excel")
async def export_bom_parts_only_excel(request: Request, bom_id: str):
    """Export ONLY the Component → Raw-Material BOMs of the root BOM tree.

    Behaviour:
        - Skip FG and SG parent rows entirely.
        - Walk the full BOM tree starting at the root BOM. For every item
          encountered whose category == 'component' (and that has its own
          active BOM), emit one row per Raw-Material child in that
          component's BOM.
        - Each component's routing operations are output as dynamic columns
          (one per distinct operation name across the whole tree) with the
          per-unit cost as the value, plus a Total Routing Cost column.
        - Routing is shown only on the FIRST RM row of each component to
          avoid visual duplication; subsequent rows for the same component
          leave routing cells blank.
        - Each unique component is exported only once even if used in
          multiple parent assemblies.
    """
    await get_current_user(request)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    root_bom = await db.boms.find_one({"id": bom_id}, {"_id": 0})
    if not root_bom:
        raise HTTPException(status_code=404, detail="BOM not found")

    # Pre-load items + active BOMs for in-memory recursion.
    items_by_id: Dict[str, Dict[str, Any]] = {}
    async for it in db.items.find({}, {"_id": 0}):
        items_by_id[it["id"]] = it
    active_bom_by_parent: Dict[str, Dict[str, Any]] = {}
    async for b in db.boms.find({"status": "active"}, {"_id": 0}):
        active_bom_by_parent[b["parent_item_id"]] = b

    # Step 1: walk the tree from the root BOM and collect every component-category
    # item that has its own active BOM. Dedupe by item_id.
    component_boms: List[Dict[str, Any]] = []
    seen_component_ids: set = set()

    def walk_tree(bom):
        parent_item = items_by_id.get(bom["parent_item_id"]) or {}
        parent_cat = (parent_item.get("category") or "").lower()
        if parent_cat == "component" and bom["parent_item_id"] not in seen_component_ids:
            seen_component_ids.add(bom["parent_item_id"])
            component_boms.append(bom)
        for comp in bom.get("components", []) or []:
            cid = comp.get("item_id")
            child = active_bom_by_parent.get(cid)
            if child:
                walk_tree(child)

    walk_tree(root_bom)

    # Step 2: keep only component BOMs that have at least one Raw-Material
    # child — those are the rows we will actually emit. We collect the union
    # of routing operation names from THESE components only, so columns
    # represent real data (no empty operation columns).
    emit_list: List[Dict[str, Any]] = []
    routing_names: List[str] = []
    seen_names: set = set()
    for cb in component_boms:
        rm_children = [
            c for c in (cb.get("components") or [])
            if (items_by_id.get(c.get("item_id"), {}).get("category") or "").lower() == "raw_material"
        ]
        if not rm_children:
            continue
        emit_list.append({"bom": cb, "rm_children": rm_children})
        for r in normalize_routings(cb.get("parent_routings") or []):
            nm = (r.get("name") or "").strip()
            if nm and nm not in seen_names:
                seen_names.add(nm)
                routing_names.append(nm)

    def routing_cost(routings, name):
        for r in normalize_routings(routings or []):
            if (r.get("name") or "").strip() == name:
                return r.get("cost", 0) or 0
        return ""

    # Step 3: build the row list. Only RM children are emitted.
    rows: List[Dict[str, Any]] = []
    for entry in emit_list:
        cb = entry["bom"]
        rm_children = entry["rm_children"]
        comp_item = items_by_id.get(cb["parent_item_id"]) or {}
        comp_routings = cb.get("parent_routings") or []
        total_route_cost = round(routings_total_cost(comp_routings), 2)
        for idx, comp in enumerate(rm_children):
            citem = items_by_id.get(comp.get("item_id")) or {}
            rows.append({
                "component_part_number": comp_item.get("part_number", ""),
                "component_name": comp_item.get("name", ""),
                "revision": cb.get("revision", ""),
                "rm_part_number": citem.get("part_number", ""),
                "rm_name": citem.get("name", ""),
                "quantity": comp.get("quantity", 0),
                "uom": comp.get("unit_of_measure") or citem.get("unit_of_measure") or "",
                "is_alternate": "Yes" if comp.get("is_alternate") else "No",
                "effectivity_date": (comp.get("effectivity_date") or cb.get("effectivity_date") or ""),
                "routings": comp_routings if idx == 0 else [],
                "total_route_cost": total_route_cost if idx == 0 else "",
            })

    wb = Workbook()
    ws = wb.active
    ws.title = "Component BOM"

    base_headers = [
        "Component Part Number", "Component Name", "Revision",
        "RM Part Number", "RM Name", "Quantity", "UOM",
        "Is Alternate", "Effectivity Date",
    ]
    headers = base_headers + routing_names + ["Total Routing Cost"]

    header_fill = PatternFill(start_color="1D3557", end_color="1D3557", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    row_num = 2
    for r in rows:
        # Effectivity date — format ISO datetime → YYYY-MM-DD for readability.
        eff = r["effectivity_date"]
        if hasattr(eff, "strftime"):
            eff = eff.strftime("%Y-%m-%d")
        elif isinstance(eff, str) and "T" in eff:
            eff = eff.split("T", 1)[0]
        base_values = [
            r["component_part_number"],
            r["component_name"],
            r["revision"],
            r["rm_part_number"],
            r["rm_name"],
            r["quantity"],
            r["uom"],
            r["is_alternate"],
            eff,
        ]
        routing_values = [routing_cost(r["routings"], nm) for nm in routing_names]
        all_values = base_values + routing_values + [r["total_route_cost"]]
        for col, v in enumerate(all_values, 1):
            cell = ws.cell(row=row_num, column=col, value=v)
            cell.border = thin_border
            if col == 6:  # Quantity
                cell.alignment = Alignment(horizontal='right')
            if col > len(base_headers):  # routing + total cost columns
                cell.alignment = Alignment(horizontal='right')
        row_num += 1

    if not rows:
        note_cell = ws.cell(
            row=2, column=1,
            value="No component-level BOMs found in this assembly. Ensure your components "
                  "(category = 'component') have their own active BOMs authored.",
        )
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        note_cell.font = Font(italic=True, color="723B13", size=10)
        note_cell.alignment = Alignment(horizontal='left', wrap_text=True)

    # Sensible column widths.
    base_widths = [22, 32, 9, 22, 32, 9, 8, 10, 13]
    for i, w in enumerate(base_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    for i in range(len(base_widths) + 1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 14

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    root_item = items_by_id.get(root_bom["parent_item_id"]) or {}
    fname = f"{root_item.get('part_number','BOM')} Component BOM.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@bom_router.get("/export/excel")
async def export_boms_excel(request: Request, bom_id: Optional[str] = None, top_level_only: bool = True):
    """
    Export BOMs to Excel.

    Args:
      bom_id: optional single BOM to export (used by the "Export this BOM"
              button in the UI). If omitted, every BOM in the database is
              dumped (legacy bulk export).
      top_level_only: when True (default), the parent's tree stops at its
              direct children — sub-assemblies are NOT exploded into the
              parent's sheet. SG users export each sub-assembly via its
              own per-row Excel button (which calls this same endpoint with
              the SG's bom_id). When False, the legacy multi-level
              recursion behavior is preserved.

    Routings layout (per latest user spec):
      One column per master Routing (sourced from db.routings, status=active).
      The parent FG/SA row of each BOM holds its routing cost in those columns;
      component rows leave them blank. Empty cell = routing not used.

    Columns (in order):
      Level | Parent PN | Parent Name | Revision | Status |
      Component PN | Component Name | Quantity | Is Alternate | Effectivity Date |
      <Routing 1> | <Routing 2> | … | <Routing N>

    A second "Routings Summary" sheet aggregates each routing's total cost
    contribution across the exported tree.
    """
    await get_current_user(request)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if bom_id:
        boms = await db.boms.find({"id": bom_id}, {"_id": 0}).to_list(1)
    else:
        boms = await db.boms.find({}, {"_id": 0}).to_list(10000)

    items_map = {}
    async for item in db.items.find({}, {"_id": 0}):
        items_map[item["id"]] = item

    # Index BOMs by parent_item_id -> active revision preferred for recursion
    boms_by_parent = {}
    for b in await db.boms.find({}, {"_id": 0}).to_list(10000):
        pid = b.get("parent_item_id")
        if not pid:
            continue
        existing = boms_by_parent.get(pid)
        if existing is None or (b.get("status") == "active" and existing.get("status") != "active"):
            boms_by_parent[pid] = b

    # Master routings list (active) — these become column headers.
    master_routings = await db.routings.find(
        {"status": "active"}, {"_id": 0, "name": 1}
    ).sort("name", 1).to_list(500)
    routing_names = []
    seen_lower = set()
    for r in master_routings:
        nm = (r.get("name") or "").strip()
        if nm and nm.lower() not in seen_lower:
            seen_lower.add(nm.lower())
            routing_names.append(nm)

    wb = Workbook()
    ws = wb.active
    ws.title = "BOM Tree"

    base_headers = [
        "Level", "Parent Part Number", "Parent Name", "Revision", "Status",
        "Component Part Number", "Component Name", "Component Type", "Quantity",
        "Is Alternate", "Effectivity Date",
    ]
    headers = base_headers + routing_names
    header_fill = PatternFill(start_color="1D3557", end_color="1D3557", fill_type="solid")
    routing_fill = PatternFill(start_color="723B13", end_color="723B13", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    sub_fill = PatternFill(start_color="E1EFFE", end_color="E1EFFE", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = routing_fill if col > len(base_headers) else header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    # Aggregator for the second sheet — across all parent routings written.
    routings_summary = {nm: {"occurrences": 0, "total_cost": 0.0} for nm in routing_names}

    # Short Material-Type badge — same short codes the UI shows (FG / SG / Part / RM).
    # Anything unmapped falls through to "Part" which matches the existing UI convention.
    def _type_badge(item):
        cat = (item or {}).get("category") or ""
        return {
            "finished_good": "FG",
            "sub_assembly": "SG",
            "raw_material": "RM",
            "component": "Part",
            "purchased_part": "Part",
        }.get(cat, "Part" if cat else "")

    def _routing_cost_map(rs):
        """Convert a parent_routings list ([str | {name,cost}]) to {name_lower: cost}.
        Multiple entries with the same name accumulate their cost."""
        out = {}
        for r in (rs or []):
            if isinstance(r, str):
                key = r.strip()
                if key:
                    out[key.lower()] = out.get(key.lower(), 0.0)  # cost-less
            elif isinstance(r, dict):
                nm = str(r.get("name", "")).strip()
                if not nm:
                    continue
                try:
                    cost = float(r.get("cost", 0) or 0)
                except (TypeError, ValueError):
                    cost = 0.0
                out[nm.lower()] = out.get(nm.lower(), 0.0) + cost
        return out

    row_num = 2
    visited_parents = set()

    def _walk(parent_bom, level):
        nonlocal row_num
        if not parent_bom:
            return
        parent_id = parent_bom.get("parent_item_id")
        if parent_id in visited_parents:
            return
        visited_parents.add(parent_id)

        parent_item = items_map.get(parent_id, {})
        parent_routing_costs = _routing_cost_map(parent_bom.get("parent_routings", []))
        # Record into summary sheet
        for nm in routing_names:
            cost = parent_routing_costs.get(nm.lower(), 0.0)
            if cost > 0:
                routings_summary[nm]["occurrences"] += 1
                routings_summary[nm]["total_cost"] = round(routings_summary[nm]["total_cost"] + cost, 2)
        eff_date = str(parent_bom.get("effectivity_date", ""))[:10] if parent_bom.get("effectivity_date") else ""

        components = parent_bom.get("components", []) or []

        def _emit_row(comp, is_first):
            nonlocal row_num
            comp_item = items_map.get((comp or {}).get("item_id"), {}) if comp else {}
            indent = "  " * level
            base = [
                level,
                (indent + (parent_item.get("part_number", "") or "")),
                parent_item.get("name", ""),
                parent_bom.get("revision", ""), parent_bom.get("status", ""),
                comp_item.get("part_number", "") if comp else "",
                comp_item.get("name", "") if comp else "",
                _type_badge(comp_item) if comp else "",
                comp.get("quantity", 0) if comp else "",
                ("Yes" if (comp and comp.get("is_alternate")) else ("No" if comp else "")),
                eff_date,
            ]
            # Routing cells: only the FIRST row of the parent group carries the
            # cost. Subsequent rows leave them blank — matches the parent-only
            # spec and keeps the sheet visually clean.
            routing_cells = []
            for nm in routing_names:
                if is_first:
                    cost = parent_routing_costs.get(nm.lower(), 0.0)
                    routing_cells.append(round(cost, 2) if cost else "")
                else:
                    routing_cells.append("")
            for col, value in enumerate(base + routing_cells, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                if level > 0:
                    cell.fill = sub_fill
            row_num += 1

        if not components:
            _emit_row(None, True)
        else:
            for idx, comp in enumerate(components):
                _emit_row(comp, idx == 0)
                # Recurse only when the caller explicitly asked for a full
                # multi-level dump. With `top_level_only=True` (the default
                # for per-BOM exports), each sub-assembly stays as a one-line
                # entry — the user downloads the SG's own Excel separately.
                if not top_level_only:
                    child_bom = boms_by_parent.get(comp.get("item_id"))
                    if child_bom:
                        _walk(child_bom, level + 1)

        visited_parents.discard(parent_id)

    if bom_id:
        for b in boms:
            _walk(b, 0)
    else:
        used_as_component = set()
        for b in boms:
            for c in b.get("components", []) or []:
                if c.get("item_id"):
                    used_as_component.add(c["item_id"])
        top_level = [b for b in boms if b.get("parent_item_id") not in used_as_component]
        if not top_level:
            top_level = boms
        for b in top_level:
            visited_parents = set()
            _walk(b, 0)

    # Auto-size columns. Routing columns get a sensible default width.
    base_widths = [7, 28, 28, 10, 10, 28, 28, 10, 10, 14]
    for idx, w in enumerate(base_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = w
    for idx in range(len(base_widths) + 1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = 14
    ws.freeze_panes = "F2"  # freeze through Quantity so routing columns scroll with data

    # ---------- Routings Summary Sheet ----------
    ws2 = wb.create_sheet("Routings Summary")
    summary_headers = ["Routing Name", "Occurrences", "Total Cost"]
    for col, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    r = 2
    summary_entries = sorted(
        ((nm, vals) for nm, vals in routings_summary.items() if vals["occurrences"] > 0),
        key=lambda kv: -kv[1]["total_cost"],
    )
    for name, vals in summary_entries:
        for col, value in enumerate([name, vals["occurrences"], vals["total_cost"]], 1):
            cell = ws2.cell(row=r, column=col, value=value)
            cell.border = thin_border
        r += 1
    for idx, w in enumerate([32, 14, 16], 1):
        ws2.column_dimensions[ws2.cell(row=1, column=idx).column_letter].width = w
    ws2.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"bom_{boms[0].get('revision','')}.xlsx" if bom_id and boms else "bom_full_tree.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@bom_router.post("/import/excel")
async def import_bom_excel(request: Request, file: UploadFile = File(...)):
    """
    Import BOMs from Excel — groups by parent part number.
    Uses HEADER-NAME lookup (not column position) so any column order works.

    Required headers:
      - "Parent Part Number"
      - "Component Part Number"
      - "Quantity"

    Optional headers:
      - "Revision"               (default "A")
      - "Status"                 (default "active")
      - "Is Alternate"           ("Yes"/"No")
      - "Effectivity Date"       (ISO yyyy-mm-dd; default today)

    Routings layout (parent-only):
      Each master Routing in db.routings becomes its own column. To set the
      parent BOM's routing cost, put the cost in the corresponding column on
      ANY row of that parent (typically the first row). All values across the
      parent's rows for the same routing column are SUMMED — so it's safe to
      enter the cost on whichever row is most readable.

      Any column whose header doesn't match Parent/Component/Qty/Revision/
      Status/Is Alternate/Effectivity is treated as a routing column. If the
      header doesn't match an existing master Routing (case-insensitive),
      a NEW master Routing is auto-created so future exports include it.
    """
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="bom", action="create")
    from openpyxl import load_workbook

    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active

    results = {"created": 0, "updated": 0, "errors": [], "imported_bom_ids": [], "imported_part_numbers": []}

    items_by_pn = {}
    # Build BOTH a case-sensitive lookup AND a case-insensitive / trimmed
    # fallback. The Excel that the user fills in commonly has subtle whitespace
    # or case drift (e.g. `cgf0g0000093` vs the master's `CGF0G0000093`) that
    # used to fail silently with "Parent X not found in items" — even when the
    # item clearly exists in the catalogue.
    items_by_pn_ci = {}
    async for item in db.items.find({}, {"_id": 0}):
        pn = item.get("part_number", "") or ""
        items_by_pn[pn] = item
        items_by_pn_ci[pn.strip().lower()] = item

    def _lookup_item(pn: str):
        """Find an item by part number, tolerant of whitespace + case."""
        if pn is None:
            return None
        s = str(pn).strip()
        if s in items_by_pn:
            return items_by_pn[s]
        return items_by_pn_ci.get(s.lower())

    # Read header row → map of normalized name → column index
    headers = {}
    raw_headers = []
    for c_idx, cell in enumerate(ws[1], 1):
        if cell.value is not None and str(cell.value).strip() != "":
            raw = str(cell.value).strip()
            raw_headers.append((c_idx, raw))
            headers[raw.lower()] = c_idx

    def col(row, *names):
        for name in names:
            key = name.lower()
            if key in headers:
                val = row[headers[key] - 1]
                if val is not None and str(val).strip() != "":
                    return val
        return None

    required = ["parent part number", "component part number", "quantity"]
    missing = [h for h in required if h not in headers]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required column(s): {', '.join(missing)}. "
                   f"Tip: download a fresh template via BOM → Export to see the expected headers."
        )

    # Determine routing columns: every header that isn't one of the known core
    # columns becomes a routing column. We also map them to existing master
    # Routings (case-insensitive) so we keep canonical naming.
    core_keys = {
        "level", "parent part number", "parent name", "revision", "status",
        "component part number", "component name", "component type", "quantity",
        "is alternate", "effectivity date",
        # "parent type" — tolerated but ignored if an older export includes it
        "parent type",
        # Legacy columns that should NOT be misread as routing columns
        "parent routing count", "parent routing cost total",
        "parent routings (name:cost)", "parent routings",
    }
    master_routings_list = await db.routings.find(
        {}, {"_id": 0, "id": 1, "name": 1, "status": 1}
    ).to_list(500)
    master_by_lower = {(r.get("name") or "").strip().lower(): r for r in master_routings_list}

    routing_cols = []  # list of (col_idx, canonical_name)
    new_routings_to_create = []  # canonical names that don't yet exist
    for c_idx, raw in raw_headers:
        if raw.lower() in core_keys:
            continue
        existing = master_by_lower.get(raw.lower())
        canonical = (existing or {}).get("name") or raw
        routing_cols.append((c_idx, canonical))
        if not existing and raw.lower() not in {n.lower() for n in new_routings_to_create}:
            new_routings_to_create.append(raw)

    # Auto-create any unknown master routings so a fresh export reflects them.
    for new_nm in new_routings_to_create:
        try:
            await db.routings.insert_one({
                "id": str(uuid.uuid4()),
                "name": new_nm,
                "description": f"Auto-created during BOM import on {datetime.now(timezone.utc).isoformat()}",
                "status": "active",
                "created_at": datetime.now(timezone.utc),
                "created_by": user["id"],
            })
        except Exception:
            # If concurrent inserts race or the name has a unique index conflict,
            # ignore — what matters is that a record exists.
            pass

    bom_groups = {}
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row:
            continue
        parent_pn = col(row, "Parent Part Number")
        if not parent_pn:
            continue
        parent_pn = str(parent_pn).strip()

        if parent_pn not in bom_groups:
            bom_groups[parent_pn] = {
                "revision": str(col(row, "Revision") or "A").strip(),
                "status": str(col(row, "Status") or "active").strip(),
                # routings_acc accumulates {name_lower: {"name": canonical, "cost": float}}
                "routings_acc": {},
                "components": [],
            }

        # Accumulate routing costs from THIS row (parent-only model — every
        # row of a parent group can contribute; we sum so the user can place
        # the cost on whichever row is convenient).
        for c_idx, canonical in routing_cols:
            raw_val = row[c_idx - 1]
            if raw_val is None or str(raw_val).strip() == "":
                continue
            try:
                cost = float(raw_val)
            except (TypeError, ValueError):
                # Non-numeric routing cell: treat as zero-cost named routing
                cost = 0.0
            key = canonical.lower()
            entry = bom_groups[parent_pn]["routings_acc"].setdefault(
                key, {"name": canonical, "cost": 0.0}
            )
            entry["cost"] = round(entry["cost"] + cost, 2)

        comp_pn = col(row, "Component Part Number")
        if not comp_pn:
            # Allow header-only rows (e.g. parent with no components yet) to
            # still register parent + routing values without an error.
            continue
        comp_pn = str(comp_pn).strip()

        try:
            qty = float(col(row, "Quantity") or 1)
        except (TypeError, ValueError):
            qty = 1.0

        is_alt_raw = col(row, "Is Alternate") or ""
        is_alt = str(is_alt_raw).strip().lower() in ("yes", "true", "1")

        parent_item = _lookup_item(parent_pn)
        comp_item = _lookup_item(comp_pn)
        if not parent_item:
            results["errors"].append(f"Row {row_num}: Parent '{parent_pn}' not found in items")
            continue
        if not comp_item:
            results["errors"].append(f"Row {row_num}: Component '{comp_pn}' not found in items")
            continue

        # RM items are leaf materials — they cannot have their own BOM or
        # routing. Block them at import (manual BOM creation already blocks
        # RM parents in the UI; this mirrors that server-side for imports).
        if (parent_item.get("category") or "").lower() == "raw_material":
            results["errors"].append(
                f"Row {row_num}: Parent '{parent_pn}' is a Raw Material (RM) — RM items cannot have a BOM. Skipped."
            )
            continue

        bom_groups[parent_pn]["parent_item_id"] = parent_item["id"]
        bom_groups[parent_pn]["components"].append({
            "item_id": comp_item["id"],
            "quantity": qty,
            "is_alternate": is_alt,
        })

    # Create or update BOMs
    for parent_pn, bom_data in bom_groups.items():
        if not bom_data.get("parent_item_id"):
            continue

        # Normalize status to canonical lowercase value so the BOM list filter
        # ("active" / "draft" / "obsolete") sees consistent data regardless of
        # how the user typed it in Excel ("Active", "ACTIVE", " active " etc.).
        # Anything outside the known set falls back to "active".
        raw_status = (bom_data.get("status") or "active").strip().lower()
        if raw_status not in ("draft", "active", "obsolete"):
            raw_status = "active"
        bom_data["status"] = raw_status

        # Convert accumulated routing dict back to a list, dropping zero-cost
        # entries that the user didn't actually fill (avoid noise in the doc).
        parent_routings = [
            {"name": v["name"], "cost": v["cost"]}
            for v in bom_data["routings_acc"].values()
            if v["cost"] > 0
        ]

        existing = await db.boms.find_one({"parent_item_id": bom_data["parent_item_id"], "revision": bom_data["revision"]})
        if existing:
            await db.boms.update_one(
                {"id": existing["id"]},
                {"$set": {
                    "components": bom_data["components"],
                    "parent_routings": parent_routings,
                    "status": bom_data["status"],
                    "updated_at": datetime.now(timezone.utc)
                }}
            )
            # When the imported BOM is being marked active, demote any OTHER
            # revisions of the same parent that were previously active so the
            # business rule "only one active revision per parent" stays true.
            if bom_data["status"] == "active":
                await db.boms.update_many(
                    {"parent_item_id": bom_data["parent_item_id"], "id": {"$ne": existing["id"]}, "status": "active"},
                    {"$set": {"status": "obsolete", "updated_at": datetime.now(timezone.utc)}},
                )
            results["updated"] += 1
            results["imported_bom_ids"].append(existing["id"])
            results["imported_part_numbers"].append(parent_pn)
        else:
            bom_doc = {
                "id": str(uuid.uuid4()),
                "parent_item_id": bom_data["parent_item_id"],
                "name": f"BOM - {parent_pn}",
                "revision": bom_data["revision"],
                "status": bom_data["status"],
                "parent_routings": parent_routings,
                "components": bom_data["components"],
                "effectivity_date": datetime.now(timezone.utc),
                "created_at": datetime.now(timezone.utc),
                "created_by": user["id"]
            }
            await db.boms.insert_one(bom_doc)
            # Same single-active-revision rule applies on insert.
            if bom_data["status"] == "active":
                await db.boms.update_many(
                    {"parent_item_id": bom_data["parent_item_id"], "id": {"$ne": bom_doc["id"]}, "status": "active"},
                    {"$set": {"status": "obsolete", "updated_at": datetime.now(timezone.utc)}},
                )
            results["created"] += 1
            results["imported_bom_ids"].append(bom_doc["id"])
            results["imported_part_numbers"].append(parent_pn)

    return results

@routings_router.get("/export/excel")
async def export_routings_excel(request: Request):
    """Export all routings to Excel"""
    await get_current_user(request)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    routings = await db.routings.find({}, {"_id": 0}).to_list(10000)
    items_map = {}
    async for item in db.items.find({}, {"_id": 0}):
        items_map[item["id"]] = item
    
    wc_map = {}
    async for wc in db.work_centers.find({}, {"_id": 0}):
        wc_map[wc["id"]] = wc
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Routings"
    
    headers = ["Item Part Number", "Item Name", "Routing Name", "Status", "Seq", "Operation", "Work Center", "Setup Time (min)", "Cycle Time (min)", "Description"]
    header_fill = PatternFill(start_color="1D3557", end_color="1D3557", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    row_num = 2
    for routing in routings:
        item = items_map.get(routing.get("item_id"), {})
        for op in routing.get("operations", []):
            wc = wc_map.get(op.get("work_center_id"), {})
            data = [
                item.get("part_number", ""), item.get("name", ""),
                routing.get("name", ""), routing.get("status", ""),
                op.get("sequence", ""), op.get("operation_name", ""),
                wc.get("name", ""), op.get("setup_time", 0), op.get("cycle_time", 0),
                op.get("description", "")
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
            row_num += 1
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col) if col <= 26 else 'A'].width = 18
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=routings.xlsx"}
    )

# ================== COMPANY SETTINGS / GST ROUTES ==================

@settings_router.get("/company")
async def get_company_settings(request: Request):
    await get_current_user(request)
    settings = await db.company_settings.find_one({"type": "company"}, {"_id": 0})
    if not settings:
        settings = {
            "type": "company",
            "company_name": "My Manufacturing Company",
            "gstin": "",
            "state_code": "",
            "address": "",
            "address_line2": "",
            "city": "",
            "state": "",
            "pin_code": "",
            "pan": "",
            "cin": "",
            "phone": "",
            "email": "",
            "logo_data": "",
            "tagline": "",
            "primary_currency": "INR",
            "secondary_currency": "USD"
        }
    return settings

@settings_router.put("/company")
async def update_company_settings(data: CompanySettingsUpdate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin"], module="settings", action="edit")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    update_data["type"] = "company"
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    await db.company_settings.update_one(
        {"type": "company"},
        {"$set": update_data},
        upsert=True
    )
    return await db.company_settings.find_one({"type": "company"}, {"_id": 0})

# ================== NUMBER SERIES (Vendor/Customer/PO/Sales Invoice) ==================
# Each key stores: {prefix, padding, next_number}
DEFAULT_NUMBER_SERIES = {
    "supplier_code":  {"prefix": "SUP",  "padding": 4, "next_number": 1, "label": "Vendor / Supplier Code", "reset_yearly": False, "group": "masters"},
    "customer_code":  {"prefix": "CUST", "padding": 4, "next_number": 1, "label": "Customer Code", "reset_yearly": False, "group": "masters"},
    "po_number":      {"prefix": "PO",   "padding": 6, "next_number": 1, "label": "Purchase Order", "reset_yearly": False, "group": "procurement"},
    "delivery_challan":{"prefix": "DC",  "padding": 6, "next_number": 1, "label": "Delivery Challan", "reset_yearly": False, "group": "stores"},
}
# NOTE: 'sales_invoice' series removed — duplicated the CRM 'tax_invoice' series.
#       The Tax Invoice flow is the single source of truth for invoice numbering now.

# CRM-side series (new style, keyed by `doc_type` on same collection). Merged into the Settings UI.
CRM_DEFAULT_NUMBER_SERIES = {
    "quotation":       {"prefix": "QUO",  "padding": 6, "next_number": 1, "label": "Quotation",       "reset_yearly": False, "group": "crm"},
    "proforma":        {"prefix": "PI",   "padding": 6, "next_number": 1, "label": "Proforma Invoice","reset_yearly": False, "group": "crm"},
    "tax_invoice":     {"prefix": "INV",  "padding": 6, "next_number": 1, "label": "Tax Invoice",     "reset_yearly": True,  "group": "crm"},
    "sales_order":     {"prefix": "SO",   "padding": 6, "next_number": 1, "label": "Sales Order",     "reset_yearly": False, "group": "sales"},
    "purchase_invoice":{"prefix": "PUR",  "padding": 6, "next_number": 1, "label": "Purchase Invoice","reset_yearly": False, "group": "procurement"},
    "packing_list":    {"prefix": "PL",   "padding": 6, "next_number": 1, "label": "Packing List",    "reset_yearly": False, "group": "stores"},
}

async def get_next_series_number(key: str) -> str:
    """Atomically fetch and increment the next number for a series. Returns formatted string.

    Honors the `reset_yearly` flag: when true, prepends the compact FY (e.g. `2627`)
    between prefix and padded number and auto-resets the counter to 1 on FY rollover.
    This is what the Settings → Number Series preview shows, so actual generated
    numbers MUST match that preview.
    """
    default = DEFAULT_NUMBER_SERIES.get(key, {"prefix": "", "padding": 4, "next_number": 1, "reset_yearly": False})
    # Ensure the doc exists with defaults
    existing = await db.number_series.find_one({"key": key})
    if not existing:
        await db.number_series.insert_one({
            "key": key,
            "prefix": default.get("prefix", ""),
            "padding": default.get("padding", 4),
            "next_number": default.get("next_number", 1),
            "reset_yearly": default.get("reset_yearly", False),
        })
        existing = await db.number_series.find_one({"key": key})
    prefix = existing.get("prefix", default.get("prefix", ""))
    padding = int(existing.get("padding", default.get("padding", 4)))
    current = int(existing.get("next_number", default.get("next_number", 1)))
    reset_yearly = bool(existing.get("reset_yearly", default.get("reset_yearly", False)))
    year_part = ""
    if reset_yearly:
        today = datetime.now(timezone.utc)
        fy_start = today.year if today.month >= 4 else today.year - 1
        fy_str = f"{str(fy_start)[-2:]}{str(fy_start + 1)[-2:]}"
        stored_fy = existing.get("current_fy")
        if stored_fy and stored_fy != fy_str:
            # Actual FY rollover — reset counter to 1
            current = 1
            await db.number_series.update_one({"key": key}, {"$set": {"current_fy": fy_str, "next_number": 1}})
        elif not stored_fy:
            # First activation of reset_yearly — just stamp the FY, keep user's next_number
            await db.number_series.update_one({"key": key}, {"$set": {"current_fy": fy_str}})
        year_part = fy_str
    # Increment next_number for the next call
    await db.number_series.update_one({"key": key}, {"$set": {"next_number": current + 1}})
    return f"{prefix}{year_part}{str(current).zfill(padding)}"

class NumberSeriesUpdate(BaseModel):
    prefix: Optional[str] = None
    padding: Optional[int] = None
    next_number: Optional[int] = None
    reset_yearly: Optional[bool] = None

def _current_fy_string() -> str:
    """India FY starts Apr 1. Returns compact 4-digit form 'YYYY' made of last-2 digits of fy_start + fy_end.
    Example: Apr 2026 - Mar 2027 → '2627'."""
    today = datetime.now(timezone.utc).date()
    fy_start = today.year if today.month >= 4 else today.year - 1
    return f"{str(fy_start)[-2:]}{str(fy_start + 1)[-2:]}"

def _normalize_fy(fy: str) -> str:
    """Coerce any legacy FY format (FY26-27, FY2627, 26-27, 26/27, "FY 2026-27") → compact '2627'."""
    if not fy:
        return _current_fy_string()
    import re
    digits = re.sub(r"[^0-9]", "", str(fy))
    if len(digits) == 4:
        return digits  # already compact
    if len(digits) == 8:  # e.g. "20262027"
        return digits[2:4] + digits[6:8]
    if len(digits) == 6:  # e.g. "202627" or "262627"
        # take last 2 of first half and last 2 of second half
        return digits[:2] + digits[-2:]
    return _current_fy_string()

def _format_series_preview(prefix: str, padding: int, next_num: int, reset_yearly: bool, current_fy: str = "") -> str:
    """Build a preview string matching what `_get_next_number` would produce — compact, no dashes/slashes."""
    padded = str(next_num).zfill(padding)
    if reset_yearly:
        fy = current_fy or _current_fy_string()
        return f"{prefix}{fy}{padded}"
    return f"{prefix}{padded}"

@settings_router.get("/number-series")
async def get_number_series(request: Request):
    """Return all configured number series (masters + CRM) with current FY + next-number preview."""
    await get_current_user(request)
    series_list = []
    current_fy = _current_fy_string()
    # Legacy/masters series (keyed by `key`)
    for key, default in DEFAULT_NUMBER_SERIES.items():
        doc = await db.number_series.find_one({"key": key}, {"_id": 0})
        merged = {"key": key, **default}
        if doc:
            merged.update({k: v for k, v in doc.items() if k != "label"})
        merged["current_fy"] = _normalize_fy(merged.get("current_fy") or current_fy)
        merged["preview"] = _format_series_preview(merged.get("prefix", ""), int(merged.get("padding", 4)), int(merged.get("next_number", 1)), bool(merged.get("reset_yearly", False)), merged["current_fy"])
        series_list.append(merged)
    # CRM-side series (keyed by `doc_type`)
    for dt, default in CRM_DEFAULT_NUMBER_SERIES.items():
        doc = await db.number_series.find_one({"doc_type": dt}, {"_id": 0})
        merged = {"key": dt, "doc_type": dt, **default}
        if doc:
            merged.update({k: v for k, v in doc.items() if k != "label"})
        merged["current_fy"] = _normalize_fy(merged.get("current_fy") or current_fy)
        merged["preview"] = _format_series_preview(merged.get("prefix", ""), int(merged.get("padding", 4)), int(merged.get("next_number", 1)), bool(merged.get("reset_yearly", False)), merged["current_fy"])
        series_list.append(merged)
    return series_list

@settings_router.put("/number-series/{key}")
async def update_number_series(key: str, data: NumberSeriesUpdate, request: Request):
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can update number series")
    # Which table: legacy masters (key-based) OR CRM (doc_type-based)?
    if key in DEFAULT_NUMBER_SERIES:
        id_field, default = "key", DEFAULT_NUMBER_SERIES[key]
    elif key in CRM_DEFAULT_NUMBER_SERIES:
        id_field, default = "doc_type", CRM_DEFAULT_NUMBER_SERIES[key]
    else:
        raise HTTPException(status_code=404, detail=f"Unknown series key: {key}")

    update_fields = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update_fields:
        raise HTTPException(status_code=400, detail="No data to update")

    existing = await db.number_series.find_one({id_field: key})
    if not existing:
        seed = {id_field: key, **default}
        seed.pop("label", None); seed.pop("group", None)
        seed.update(update_fields)
        await db.number_series.insert_one(seed)
    else:
        await db.number_series.update_one({id_field: key}, {"$set": update_fields})
    doc = await db.number_series.find_one({id_field: key}, {"_id": 0})
    current_fy = _current_fy_string()
    doc["label"] = default["label"]
    doc["group"] = default.get("group", "misc")
    doc["key"] = key
    doc["current_fy"] = doc.get("current_fy") or current_fy
    doc["preview"] = _format_series_preview(doc.get("prefix", ""), int(doc.get("padding", 4)), int(doc.get("next_number", 1)), bool(doc.get("reset_yearly", False)), doc["current_fy"])
    return doc

@settings_router.get("/states")
async def get_indian_states(request: Request):
    await get_current_user(request)
    return [{"code": k, "name": v} for k, v in INDIAN_STATES.items()]

@settings_router.get("/gst-slabs")
async def get_gst_slabs(request: Request):
    """Return sorted list of configured GST slab percentages."""
    await get_current_user(request)
    rows = await db.tax_slabs.find({}, {"_id": 0}).to_list(200)
    if not rows:
        # Seed defaults on first read
        default_docs = [{"id": str(uuid.uuid4()), "rate": float(r)} for r in GST_SLABS]
        if default_docs:
            await db.tax_slabs.insert_many(default_docs)
        rows = default_docs
    rates = sorted({float(r.get("rate", 0)) for r in rows})
    return rates


@settings_router.post("/gst-slabs")
async def add_gst_slab(payload: dict, request: Request):
    """Add a new GST slab (admin only)."""
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can edit tax slabs")
    try:
        rate = float(payload.get("rate"))
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Rate must be a number")
    if rate < 0 or rate > 100:
        raise HTTPException(status_code=400, detail="Rate must be between 0 and 100")
    existing = await db.tax_slabs.find_one({"rate": rate})
    if existing:
        raise HTTPException(status_code=400, detail=f"Slab {rate}% already exists")
    doc = {"id": str(uuid.uuid4()), "rate": rate, "created_at": datetime.now(timezone.utc)}
    await db.tax_slabs.insert_one(doc)
    doc.pop("_id", None)
    return {"rate": rate}


@settings_router.delete("/gst-slabs/{rate}")
async def delete_gst_slab(rate: float, request: Request):
    """Remove a GST slab (admin only)."""
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can edit tax slabs")
    result = await db.tax_slabs.delete_one({"rate": float(rate)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Slab not found")
    return {"deleted": rate}


# ============================ UNITS OF MEASURE CRUD ============================
@settings_router.get("/uoms")
async def list_uoms(request: Request):
    """List units of measure, seeding defaults on first read."""
    await get_current_user(request)
    rows = await db.uoms.find({}, {"_id": 0}).sort("code", 1).to_list(500)
    if not rows:
        seed = [
            {"id": str(uuid.uuid4()), "code": u["code"], "name": u["name"], "description": "", "decimal_places": 2}
            for u in DEFAULT_UOMS
        ]
        if seed:
            await db.uoms.insert_many(seed)
        rows = seed
    # Backfill decimal_places for legacy rows so the UI never receives undefined.
    for r in rows:
        if "decimal_places" not in r or r.get("decimal_places") is None:
            r["decimal_places"] = 2
    return rows


@settings_router.post("/uoms", status_code=201)
async def create_uom(payload: dict, request: Request):
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can manage units")
    code = str(payload.get("code", "")).strip().lower()
    name = str(payload.get("name", "")).strip()
    if not code or not name:
        raise HTTPException(status_code=400, detail="Both 'code' and 'name' are required")
    exists = await db.uoms.find_one({"code": code})
    if exists:
        raise HTTPException(status_code=400, detail=f"UOM code '{code}' already exists")
    # Decimal places — clamped to [0..6]; defaults to 2 if not provided.
    try:
        decimal_places = int(payload.get("decimal_places") if payload.get("decimal_places") is not None else 2)
    except (TypeError, ValueError):
        decimal_places = 2
    decimal_places = max(0, min(6, decimal_places))
    doc = {
        "id": str(uuid.uuid4()),
        "code": code,
        "name": name,
        "description": str(payload.get("description", "") or ""),
        "decimal_places": decimal_places,
        "created_at": datetime.now(timezone.utc),
    }
    await db.uoms.insert_one(doc)
    doc.pop("_id", None)
    return doc


@settings_router.put("/uoms/{uom_id}")
async def update_uom(uom_id: str, payload: dict, request: Request):
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can manage units")
    current = await db.uoms.find_one({"id": uom_id}, {"_id": 0})
    if not current:
        raise HTTPException(status_code=404, detail="UOM not found")
    old_code = current.get("code", "")
    update = {}
    if payload.get("code") is not None:
        new_code = str(payload["code"]).strip().lower()
        if not new_code:
            raise HTTPException(status_code=400, detail="Code cannot be empty")
        # Ensure code stays unique
        dup = await db.uoms.find_one({"code": new_code, "id": {"$ne": uom_id}})
        if dup:
            raise HTTPException(status_code=400, detail=f"UOM code '{new_code}' already exists")
        update["code"] = new_code
    if payload.get("name") is not None:
        name = str(payload["name"]).strip()
        if not name:
            raise HTTPException(status_code=400, detail="Name cannot be empty")
        update["name"] = name
    if payload.get("description") is not None:
        update["description"] = str(payload.get("description") or "")
    if payload.get("decimal_places") is not None:
        try:
            dp = int(payload.get("decimal_places"))
        except (TypeError, ValueError):
            dp = 2
        update["decimal_places"] = max(0, min(6, dp))
    if not update:
        raise HTTPException(status_code=400, detail="No fields to update")
    update["updated_at"] = datetime.now(timezone.utc)
    result = await db.uoms.update_one({"id": uom_id}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="UOM not found")

    # ---- Cascade code change to dependent documents ----
    cascaded = {"items": 0, "purchase_orders": 0, "purchase_invoices": 0, "sales_orders": 0, "boms": 0}
    new_code_val = update.get("code")
    if new_code_val and new_code_val != old_code and old_code:
        # 1) Items master (unit_of_measure)
        r1 = await db.items.update_many(
            {"unit_of_measure": old_code}, {"$set": {"unit_of_measure": new_code_val}}
        )
        cascaded["items"] = r1.modified_count

        # 2) Purchase Orders — line_items[].uom (arrayFilters to update nested docs)
        r2 = await db.purchase_orders.update_many(
            {"line_items.uom": old_code},
            {"$set": {"line_items.$[elem].uom": new_code_val}},
            array_filters=[{"elem.uom": old_code}],
        )
        cascaded["purchase_orders"] = r2.modified_count

        # 3) Purchase Invoices — lines[].uom
        r3 = await db.purchase_invoices.update_many(
            {"lines.uom": old_code},
            {"$set": {"lines.$[elem].uom": new_code_val}},
            array_filters=[{"elem.uom": old_code}],
        )
        cascaded["purchase_invoices"] = r3.modified_count

        # 4) Sales Orders — lines[].uom (collection may have a different field name; ignore failures)
        try:
            r4 = await db.sales_orders.update_many(
                {"lines.uom": old_code},
                {"$set": {"lines.$[elem].uom": new_code_val}},
                array_filters=[{"elem.uom": old_code}],
            )
            cascaded["sales_orders"] = r4.modified_count
        except Exception:
            pass

        # 5) BOMs — components[].uom
        try:
            r5 = await db.boms.update_many(
                {"components.uom": old_code},
                {"$set": {"components.$[elem].uom": new_code_val}},
                array_filters=[{"elem.uom": old_code}],
            )
            cascaded["boms"] = r5.modified_count
        except Exception:
            pass

    doc = await db.uoms.find_one({"id": uom_id}, {"_id": 0})
    doc["cascaded"] = cascaded
    return doc


@settings_router.delete("/uoms/{uom_id}")
async def delete_uom(uom_id: str, request: Request):
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can manage units")
    uom = await db.uoms.find_one({"id": uom_id}, {"_id": 0})
    if not uom:
        raise HTTPException(status_code=404, detail="UOM not found")
    # Refuse to delete a UOM that is in active use by any item — referential
    # integrity guard. Without this, deleting a UOM silently breaks every item
    # using it and the UI would render a blank UOM column.
    in_use = await db.items.count_documents({"unit_of_measure": uom["code"]})
    if in_use > 0:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete UOM '{uom['code']}' — it is used by {in_use} item(s). Re-assign those items to another UOM first.",
        )
    await db.uoms.delete_one({"id": uom_id})
    return {"deleted": uom_id}

@settings_router.post("/migrate-addresses")
async def migrate_addresses(request: Request):
    """One-time migration: split existing single-line address fields into structured fields"""
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can run migrations")

    migrated = {"suppliers": 0, "customers": 0, "company": 0}

    def split_address(addr_str):
        """Best-effort split of comma-separated address into structured fields"""
        if not addr_str:
            return {}
        parts = [p.strip() for p in addr_str.replace("\n", ",").split(",") if p.strip()]
        result = {}
        # Try to detect pin code (6-digit number at end)
        pin = ""
        for i, p in enumerate(parts):
            import re as _re
            pin_match = _re.search(r'\b(\d{6})\b', p)
            if pin_match:
                pin = pin_match.group(1)
                parts[i] = _re.sub(r'\s*-?\s*\d{6}\b', '', p).strip()
                break
        parts = [p for p in parts if p]
        if pin:
            result["pin_code"] = pin
        if len(parts) >= 4:
            result["address"] = parts[0]
            result["address_line2"] = parts[1]
            result["city"] = parts[2]
            result["state"] = parts[3]
        elif len(parts) == 3:
            result["address"] = parts[0]
            result["city"] = parts[1]
            result["state"] = parts[2]
        elif len(parts) == 2:
            result["address"] = parts[0]
            result["city"] = parts[1]
        elif len(parts) == 1:
            result["address"] = parts[0]
        return result

    # Migrate suppliers
    async for sup in db.suppliers.find({"city": {"$exists": False}, "address": {"$exists": True, "$ne": ""}}):
        fields = split_address(sup.get("address", ""))
        if fields:
            await db.suppliers.update_one({"_id": sup["_id"]}, {"$set": fields})
            migrated["suppliers"] += 1

    # Migrate customers
    async for cust in db.customers.find({"city": {"$exists": False}, "address": {"$exists": True, "$ne": ""}}):
        fields = split_address(cust.get("address", ""))
        if fields:
            await db.customers.update_one({"_id": cust["_id"]}, {"$set": fields})
            migrated["customers"] += 1

    # Migrate company settings
    company = await db.company_settings.find_one({"type": "company"})
    if company and not company.get("city"):
        fields = split_address(company.get("address", ""))
        if fields:
            await db.company_settings.update_one({"_id": company["_id"]}, {"$set": fields})
            migrated["company"] = 1

    return {"message": "Address migration complete", "migrated": migrated}

# ================== PO CHARGE TYPES SETTINGS ==================

@settings_router.get("/po-charges")
async def get_po_charge_types(request: Request):
    await get_current_user(request)
    charges = await db.po_charge_types.find({"is_active": True}, {"_id": 0}).to_list(100)
    return charges

@settings_router.post("/po-charges", status_code=201)
async def create_po_charge_type(data: POChargeTypeCreate, request: Request):
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can manage charge types")
    doc = {
        "id": str(uuid.uuid4()),
        **data.model_dump(),
        "is_active": True,
        "created_at": datetime.now(timezone.utc)
    }
    await db.po_charge_types.insert_one(doc)
    doc.pop("_id", None)
    return doc

@settings_router.put("/po-charges/{charge_id}")
async def update_po_charge_type(charge_id: str, data: POChargeTypeUpdate, request: Request):
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can manage charge types")
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    result = await db.po_charge_types.update_one({"id": charge_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Charge type not found")
    return await db.po_charge_types.find_one({"id": charge_id}, {"_id": 0})

@settings_router.delete("/po-charges/{charge_id}")
async def delete_po_charge_type(charge_id: str, request: Request):
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can manage charge types")
    result = await db.po_charge_types.delete_one({"id": charge_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Charge type not found")
    return {"message": "Charge type deleted"}

@settings_router.post("/clear-transactions")
async def clear_transaction_data(request: Request):
    """Clear all transactional data. Master data (items, BOMs, routings, suppliers, etc.) is preserved."""
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can clear transaction data")
    
    collections = [
        "purchase_orders", "inventory_transactions", "delivery_challans",
        "subcontract_orders", "subcontract_receipts", "work_orders",
        "production_orders", "job_work_orders", "login_attempts",
        "grn_records", "purchase_invoices", "inspections"
    ]
    result = {}
    for col in collections:
        r = await db[col].delete_many({})
        result[col] = r.deleted_count
    
    return {"message": "All transaction data cleared", "deleted": result}

# ================== CUSTOMER ROUTES ==================

@customers_router.post("/lookup-gstin")
async def customer_lookup_gstin(payload: GSTINLookupRequest, request: Request):
    # Customer alias of the supplier GSTIN lookup — same Appyflow logic, same
    # response shape. Lets the Customer form auto-fill name/state/PIN/address
    # from a GSTIN like the Supplier form already does.
    return await lookup_gstin(payload, request)


@customers_router.get("")
async def get_customers(request: Request, status: Optional[str] = None, mine: Optional[bool] = False):
    user = await get_current_user(request)
    query = {}
    if status:
        query["status"] = status
    # Per-user contact ownership (customer-centric assignment, Odoo-style):
    #   - Admins see ALL contacts by default. They can pass `mine=true` to filter to
    #     only contacts they personally created.
    #   - Non-admins see ONLY customers where they are the creator OR where their
    #     user id is listed in `customer.assigned_user_ids`. A non-admin cannot see
    #     any customer until the admin assigns them as a salesperson — no legacy
    #     null-created_by fallback.
    is_admin = user.get("role") == "admin"
    can_view_all = False
    if not is_admin and user.get("role_group_id"):
        rg = await db.role_groups.find_one({"id": user["role_group_id"]}, {"_id": 0, "is_admin_group": 1, "view_all_parties": 1})
        if rg and rg.get("is_admin_group"):
            is_admin = True
        if rg and rg.get("view_all_parties"):
            can_view_all = True
    if is_admin or can_view_all:
        if mine:
            query["created_by"] = user["id"]
    else:
        query["$or"] = [
            {"created_by": user["id"]},
            {"assigned_user_ids": user["id"]},
        ]
    customers = await db.customers.find(query, {"_id": 0}).to_list(2000)
    return customers

@customers_router.get("/{customer_id}")
async def get_customer(customer_id: str, request: Request):
    await get_current_user(request)
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    return customer

@customers_router.post("", status_code=201)
async def create_customer(data: CustomerCreate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="customers", action="create")
    # Auto-generate customer code from configurable series if not provided
    provided_code = (data.code or "").strip()
    if not provided_code:
        customer_code = await get_next_series_number("customer_code")
    else:
        customer_code = provided_code
    existing = await db.customers.find_one({"code": customer_code})
    if existing:
        raise HTTPException(status_code=400, detail="Customer code already exists")

    # GST duplicate check — same rule as suppliers. Two customers with
    # identical GSTIN means the wrong PAN ends up on a Tax Invoice, which
    # the buyer's GSTR-2A reconciliation will then reject.
    gstin_clean = (data.gstin or "").strip().upper()
    if gstin_clean:
        gst_dup = await db.customers.find_one({"gstin": gstin_clean}, {"name": 1, "code": 1})
        if gst_dup:
            raise HTTPException(
                status_code=400,
                detail=f"A customer with GSTIN {gstin_clean} already exists ({gst_dup.get('name','?')} · code {gst_dup.get('code','?')}). Please use the existing record."
            )
    
    customer_doc = {
        "id": str(uuid.uuid4()),
        **data.model_dump(),
        "code": customer_code,
        "gstin": gstin_clean or data.gstin,
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"]
    }
    await db.customers.insert_one(customer_doc)
    customer_doc.pop("_id", None)
    return customer_doc

@customers_router.post("/import")
async def import_customers(payload: dict = Body(...), request: Request = None):
    """Bulk import customers/contacts from rows[]. Each row needs at least `name`."""
    user = await get_current_user(request)
    rows = payload.get("rows") or []
    if not rows:
        raise HTTPException(status_code=400, detail="No rows to import")
    created = 0
    skipped = []
    for idx, r in enumerate(rows, start=1):
        try:
            name = (r.get("name") or "").strip()
            if not name:
                skipped.append({"row": idx, "reason": "missing name"}); continue
            # Match by code or name to avoid duplicates
            if r.get("code"):
                existing = await db.customers.find_one({"code": r["code"]}, {"_id": 0, "id": 1})
                if existing:
                    skipped.append({"row": idx, "reason": f"code {r['code']} exists"}); continue
            count = await db.customers.count_documents({})
            code = r.get("code") or f"CUST-{str(count+1).zfill(6)}"
            doc = {
                "id": str(uuid.uuid4()),
                "code": code,
                "name": name,
                "gstin": r.get("gstin", ""),
                "state_code": r.get("state_code", ""),
                "contact_person": r.get("contact_person", ""),
                "email": r.get("email", ""),
                "phone": r.get("phone", ""),
                "address": r.get("address", ""),
                "address_line2": r.get("address_line2", ""),
                "city": r.get("city", ""),
                "state": r.get("state", ""),
                "pin_code": r.get("pin_code", ""),
                "payment_terms": r.get("payment_terms", "Net 30"),
                "status": r.get("status", "active"),
                "created_at": datetime.now(timezone.utc),
                "created_by": user["id"],
            }
            await db.customers.insert_one(doc)
            created += 1
        except Exception as e:
            skipped.append({"row": idx, "reason": str(e)})
    return {"created": created, "skipped": skipped, "total": len(rows)}

@customers_router.put("/{customer_id}")
async def update_customer(customer_id: str, data: CustomerUpdate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="customers", action="edit")
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    # GST uniqueness — exclude self.
    if "gstin" in update_data:
        gstin_clean = (update_data["gstin"] or "").strip().upper()
        if gstin_clean:
            gst_dup = await db.customers.find_one(
                {"gstin": gstin_clean, "id": {"$ne": customer_id}},
                {"name": 1, "code": 1},
            )
            if gst_dup:
                raise HTTPException(
                    status_code=400,
                    detail=f"A customer with GSTIN {gstin_clean} already exists ({gst_dup.get('name','?')} · code {gst_dup.get('code','?')}). GSTIN must be unique."
                )
            update_data["gstin"] = gstin_clean
    
    update_data["updated_at"] = datetime.now(timezone.utc)
    result = await db.customers.update_one({"id": customer_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    return await db.customers.find_one({"id": customer_id}, {"_id": 0})

@customers_router.delete("/{customer_id}")
async def delete_customer(customer_id: str, request: Request):
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    
    result = await db.customers.delete_one({"id": customer_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    return {"message": "Customer deleted"}

# ================== SEED DATA ==================

async def seed_admin():
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@erp.com").lower()
    admin_password = os.environ.get("ADMIN_PASSWORD", "Admin@123")
    
    existing = await db.users.find_one({"email": admin_email})
    if existing is None:
        hashed = hash_password(admin_password)
        await db.users.insert_one({
            "id": str(uuid.uuid4()),
            "email": admin_email,
            "password_hash": hashed,
            "name": "System Admin",
            "role": "admin",
            "permissions": get_default_permissions("admin"),
            "status": "active",
            "created_at": datetime.now(timezone.utc)
        })
        logger.info(f"Admin user created: {admin_email}")
    elif not verify_password(admin_password, existing.get("password_hash", "")):
        await db.users.update_one(
            {"email": admin_email},
            {"$set": {"password_hash": hash_password(admin_password)}}
        )
        logger.info(f"Admin password updated")
    
    # Ensure admin has permissions
    admin_user = await db.users.find_one({"email": admin_email})
    if admin_user and not admin_user.get("permissions"):
        await db.users.update_one(
            {"email": admin_email},
            {"$set": {"permissions": get_default_permissions("admin"), "status": "active"}}
        )
        logger.info("Admin permissions updated")

async def seed_sample_data():
    """Seed sample data for demonstration"""
    
    # Check if data already exists
    if await db.items.count_documents({}) > 0:
        return
    
    # Sample items
    items = [
        {"id": str(uuid.uuid4()), "part_number": "RM-001", "name": "Steel Sheet 4mm", "description": "Cold rolled steel sheet", "category": "raw_material", "unit_of_measure": "sheet", "unit_cost": 45.00, "lead_time_days": 7, "safety_stock": 50, "current_stock": 120, "reorder_point": 60, "hsn_code": "7208", "gst_rate": 18},
        {"id": str(uuid.uuid4()), "part_number": "RM-002", "name": "Aluminum Bar 25mm", "description": "6061-T6 aluminum bar", "category": "raw_material", "unit_of_measure": "meter", "unit_cost": 28.50, "lead_time_days": 5, "safety_stock": 100, "current_stock": 85, "reorder_point": 80, "hsn_code": "7604", "gst_rate": 18},
        {"id": str(uuid.uuid4()), "part_number": "RM-003", "name": "O-Ring Kit", "description": "NBR O-rings assorted", "category": "raw_material", "unit_of_measure": "kit", "unit_cost": 12.00, "lead_time_days": 3, "safety_stock": 200, "current_stock": 350, "reorder_point": 150, "hsn_code": "4016", "gst_rate": 18},
        {"id": str(uuid.uuid4()), "part_number": "CP-001", "name": "Hydraulic Cylinder", "description": "50mm bore hydraulic cylinder", "category": "component", "unit_of_measure": "pcs", "unit_cost": 280.00, "lead_time_days": 14, "safety_stock": 10, "current_stock": 25, "reorder_point": 15, "hsn_code": "8412", "gst_rate": 18},
        {"id": str(uuid.uuid4()), "part_number": "CP-002", "name": "Control Valve", "description": "Directional control valve 4/3", "category": "component", "unit_of_measure": "pcs", "unit_cost": 165.00, "lead_time_days": 10, "safety_stock": 15, "current_stock": 30, "reorder_point": 20, "hsn_code": "8481", "gst_rate": 18},
        {"id": str(uuid.uuid4()), "part_number": "CP-003", "name": "Electric Motor 5HP", "description": "Three-phase induction motor", "category": "component", "unit_of_measure": "pcs", "unit_cost": 520.00, "lead_time_days": 21, "safety_stock": 5, "current_stock": 8, "reorder_point": 6, "hsn_code": "8501", "gst_rate": 18},
        {"id": str(uuid.uuid4()), "part_number": "SA-001", "name": "Pump Assembly", "description": "Hydraulic pump sub-assembly", "category": "sub_assembly", "unit_of_measure": "pcs", "unit_cost": 850.00, "lead_time_days": 5, "safety_stock": 8, "current_stock": 12, "reorder_point": 10, "hsn_code": "8413", "gst_rate": 18},
        {"id": str(uuid.uuid4()), "part_number": "SA-002", "name": "Control Panel", "description": "PLC control panel assembly", "category": "sub_assembly", "unit_of_measure": "pcs", "unit_cost": 1200.00, "lead_time_days": 7, "safety_stock": 3, "current_stock": 5, "reorder_point": 4, "hsn_code": "8537", "gst_rate": 18},
        {"id": str(uuid.uuid4()), "part_number": "FG-001", "name": "Hydraulic Press 50T", "description": "50-ton hydraulic press machine", "category": "finished_good", "unit_of_measure": "pcs", "unit_cost": 15000.00, "lead_time_days": 30, "safety_stock": 1, "current_stock": 2, "reorder_point": 1, "hsn_code": "8462", "gst_rate": 18},
        {"id": str(uuid.uuid4()), "part_number": "FG-002", "name": "CNC Milling Center", "description": "3-axis CNC milling machine", "category": "finished_good", "unit_of_measure": "pcs", "unit_cost": 45000.00, "lead_time_days": 45, "safety_stock": 1, "current_stock": 1, "reorder_point": 1, "hsn_code": "8459", "gst_rate": 18},
    ]
    
    for item in items:
        item["created_at"] = datetime.now(timezone.utc)
    
    await db.items.insert_many(items)
    logger.info("Sample items seeded")
    
    # Create sample BOMs
    pump_assembly = next(i for i in items if i["part_number"] == "SA-001")
    hydraulic_press = next(i for i in items if i["part_number"] == "FG-001")
    
    boms = [
        {
            "id": str(uuid.uuid4()),
            "parent_item_id": pump_assembly["id"],
            "name": "Pump Assembly BOM",
            "description": "BOM for hydraulic pump sub-assembly",
            "revision": "A",
            "effectivity_date": datetime.now(timezone.utc),
            "status": "active",
            "components": [
                {"item_id": next(i["id"] for i in items if i["part_number"] == "RM-002"), "quantity": 2, "unit_of_measure": "meter"},
                {"item_id": next(i["id"] for i in items if i["part_number"] == "RM-003"), "quantity": 1, "unit_of_measure": "kit"},
                {"item_id": next(i["id"] for i in items if i["part_number"] == "CP-001"), "quantity": 1, "unit_of_measure": "pcs"},
            ],
            "created_at": datetime.now(timezone.utc)
        },
        {
            "id": str(uuid.uuid4()),
            "parent_item_id": hydraulic_press["id"],
            "name": "Hydraulic Press 50T BOM",
            "description": "Complete BOM for 50-ton hydraulic press",
            "revision": "B",
            "effectivity_date": datetime.now(timezone.utc),
            "status": "active",
            "components": [
                {"item_id": next(i["id"] for i in items if i["part_number"] == "RM-001"), "quantity": 10, "unit_of_measure": "sheet"},
                {"item_id": pump_assembly["id"], "quantity": 2, "unit_of_measure": "pcs"},
                {"item_id": next(i["id"] for i in items if i["part_number"] == "CP-002"), "quantity": 4, "unit_of_measure": "pcs"},
                {"item_id": next(i["id"] for i in items if i["part_number"] == "CP-003"), "quantity": 1, "unit_of_measure": "pcs"},
                {"item_id": next(i["id"] for i in items if i["part_number"] == "SA-002"), "quantity": 1, "unit_of_measure": "pcs"},
            ],
            "created_at": datetime.now(timezone.utc)
        }
    ]
    
    await db.boms.insert_many(boms)
    logger.info("Sample BOMs seeded")
    
    # Create inspection templates
    templates = [
        {
            "id": str(uuid.uuid4()),
            "name": "Incoming Material Inspection",
            "description": "Standard inspection for incoming raw materials",
            "category": "incoming",
            "checklist_items": [
                {"id": "1", "name": "Visual Inspection", "description": "Check for visible defects", "required": True},
                {"id": "2", "name": "Dimension Check", "description": "Verify dimensions match spec", "required": True},
                {"id": "3", "name": "Material Certificate", "description": "Verify material certification", "required": True},
                {"id": "4", "name": "Packaging Condition", "description": "Check packaging integrity", "required": False}
            ],
            "created_at": datetime.now(timezone.utc)
        },
        {
            "id": str(uuid.uuid4()),
            "name": "Final Assembly Inspection",
            "description": "QC checklist for finished assemblies",
            "category": "final",
            "checklist_items": [
                {"id": "1", "name": "Functional Test", "description": "Verify all functions work correctly", "required": True},
                {"id": "2", "name": "Pressure Test", "description": "Hydraulic pressure test at 150%", "required": True},
                {"id": "3", "name": "Electrical Safety", "description": "Ground continuity and insulation test", "required": True},
                {"id": "4", "name": "Visual Finish", "description": "Paint and surface finish quality", "required": True},
                {"id": "5", "name": "Documentation", "description": "All documentation complete", "required": True}
            ],
            "created_at": datetime.now(timezone.utc)
        }
    ]
    
    await db.inspection_templates.insert_many(templates)
    logger.info("Sample inspection templates seeded")
    
    # Seed suppliers
    suppliers = [
        {"id": str(uuid.uuid4()), "code": "SUP-001", "name": "Steel Masters Pvt. Ltd.", "contact_person": "Rajesh Kumar", "email": "rajesh@steelmasters.in", "phone": "+91-9876543210", "address": "123 Industrial Area", "address_line2": "Hadapsar", "city": "Pune", "state": "Maharashtra", "pin_code": "411013", "gstin": "27AABCS1234F1Z5", "state_code": "27", "payment_terms": "Net 30", "lead_time_days": 7, "rating": 5, "status": "active", "created_at": datetime.now(timezone.utc)},
        {"id": str(uuid.uuid4()), "code": "SUP-002", "name": "Precision Components Ltd.", "contact_person": "Suresh Patel", "email": "suresh@precisioncomp.in", "phone": "+91-9876543211", "address": "456 GIDC Estate", "address_line2": "Phase II", "city": "Ahmedabad", "state": "Gujarat", "pin_code": "382445", "gstin": "24AABCP5678G1Z3", "state_code": "24", "payment_terms": "Net 45", "lead_time_days": 14, "rating": 4, "status": "active", "created_at": datetime.now(timezone.utc)},
        {"id": str(uuid.uuid4()), "code": "SUP-003", "name": "ElectroPower Systems", "contact_person": "Amit Sharma", "email": "amit@electropower.in", "phone": "+91-9876543212", "address": "789 Electronic City", "address_line2": "Phase I", "city": "Bangalore", "state": "Karnataka", "pin_code": "560100", "gstin": "29AABCE9012H1Z1", "state_code": "29", "payment_terms": "Net 30", "lead_time_days": 21, "rating": 4, "status": "active", "created_at": datetime.now(timezone.utc)},
    ]
    await db.suppliers.insert_many(suppliers)
    logger.info("Sample suppliers seeded")
    
    # Seed company settings
    if await db.company_settings.count_documents({"type": "company"}) == 0:
        await db.company_settings.insert_one({
            "type": "company",
            "company_name": "MachineWorks Manufacturing Pvt. Ltd.",
            "gstin": "27AABCM1234A1Z5",
            "state_code": "27",
            "address": "Plot No. 45, MIDC",
            "address_line2": "Bhosari Industrial Estate",
            "city": "Pune",
            "state": "Maharashtra",
            "pin_code": "411019",
            "pan": "AABCM1234A",
            "cin": "",
            "phone": "+91-20-12345678",
            "email": "info@machineworks.in",
            "logo_data": "",
            "tagline": "Precision Engineering Solutions",
            "primary_currency": "INR",
            "secondary_currency": "USD",
            "created_at": datetime.now(timezone.utc)
        })
        logger.info("Company settings seeded")
    
    # Seed sample customers
    if await db.customers.count_documents({}) == 0:
        customers = [
            {"id": str(uuid.uuid4()), "code": "CUST-001", "name": "Tata Motors Ltd.", "gstin": "27AAACT1234D1Z5", "state_code": "27", "contact_person": "Vikram Singh", "email": "vikram@tatamotors.in", "phone": "+91-9988776655", "address": "Pimpri Plant", "address_line2": "Mumbai-Pune Road", "city": "Pune", "state": "Maharashtra", "pin_code": "411018", "payment_terms": "Net 30", "status": "active", "created_at": datetime.now(timezone.utc)},
            {"id": str(uuid.uuid4()), "code": "CUST-002", "name": "Bharat Heavy Electricals", "gstin": "09AABCB5678E1Z3", "state_code": "09", "contact_person": "Priya Verma", "email": "priya@bhel.in", "phone": "+91-9988776656", "address": "Sector 17", "address_line2": "Industrial Area", "city": "Noida", "state": "Uttar Pradesh", "pin_code": "201301", "payment_terms": "Net 45", "status": "active", "created_at": datetime.now(timezone.utc)},
            {"id": str(uuid.uuid4()), "code": "CUST-003", "name": "Larsen & Toubro", "gstin": "27AABCL9012F1Z1", "state_code": "27", "contact_person": "Anish Mehta", "email": "anish@lnt.in", "phone": "+91-9988776657", "address": "L&T Business Park", "address_line2": "Saki Vihar Road, Powai", "city": "Mumbai", "state": "Maharashtra", "pin_code": "400072", "payment_terms": "Net 30", "status": "active", "created_at": datetime.now(timezone.utc)},
        ]
        await db.customers.insert_many(customers)
        logger.info("Sample customers seeded")
    
    # Seed warehouses
    warehouses = [
        {"id": str(uuid.uuid4()), "code": "WH-MAIN", "name": "Main Warehouse", "location": "Building A, Floor 1", "is_default": True, "status": "active", "created_at": datetime.now(timezone.utc)},
        {"id": str(uuid.uuid4()), "code": "WH-RAW", "name": "Raw Materials Store", "location": "Building B, Floor 1", "is_default": False, "status": "active", "created_at": datetime.now(timezone.utc)},
        {"id": str(uuid.uuid4()), "code": "WH-FG", "name": "Finished Goods Store", "location": "Building C, Floor 1", "is_default": False, "status": "active", "created_at": datetime.now(timezone.utc)},
    ]
    await db.warehouses.insert_many(warehouses)
    logger.info("Sample warehouses seeded")
    
    # Seed warehouse stock (distribute items across warehouses)
    main_wh = warehouses[0]
    raw_wh = warehouses[1]
    fg_wh = warehouses[2]
    
    warehouse_stock = []
    for item in items:
        if item["category"] == "raw_material":
            warehouse_stock.append({"id": str(uuid.uuid4()), "warehouse_id": raw_wh["id"], "item_id": item["id"], "quantity": item["current_stock"], "created_at": datetime.now(timezone.utc)})
        elif item["category"] == "finished_good":
            warehouse_stock.append({"id": str(uuid.uuid4()), "warehouse_id": fg_wh["id"], "item_id": item["id"], "quantity": item["current_stock"], "created_at": datetime.now(timezone.utc)})
        else:
            warehouse_stock.append({"id": str(uuid.uuid4()), "warehouse_id": main_wh["id"], "item_id": item["id"], "quantity": item["current_stock"], "created_at": datetime.now(timezone.utc)})
    await db.warehouse_stock.insert_many(warehouse_stock)
    logger.info("Sample warehouse stock seeded")
    
    # Seed work centers
    work_centers = [
        {"id": str(uuid.uuid4()), "code": "WC-CUT", "name": "Cutting Station", "description": "Laser and plasma cutting machines", "hourly_rate": 75.00, "capacity_per_hour": 10, "status": "active", "created_at": datetime.now(timezone.utc)},
        {"id": str(uuid.uuid4()), "code": "WC-WELD", "name": "Welding Bay", "description": "MIG/TIG welding stations", "hourly_rate": 85.00, "capacity_per_hour": 5, "status": "active", "created_at": datetime.now(timezone.utc)},
        {"id": str(uuid.uuid4()), "code": "WC-MACH", "name": "Machining Center", "description": "CNC milling and turning", "hourly_rate": 120.00, "capacity_per_hour": 3, "status": "active", "created_at": datetime.now(timezone.utc)},
        {"id": str(uuid.uuid4()), "code": "WC-ASSY", "name": "Assembly Line", "description": "Final assembly workstations", "hourly_rate": 65.00, "capacity_per_hour": 8, "status": "active", "created_at": datetime.now(timezone.utc)},
        {"id": str(uuid.uuid4()), "code": "WC-TEST", "name": "Testing Station", "description": "Quality testing and inspection", "hourly_rate": 70.00, "capacity_per_hour": 6, "status": "active", "created_at": datetime.now(timezone.utc)},
    ]
    await db.work_centers.insert_many(work_centers)
    logger.info("Sample work centers seeded")
    
    # Seed routings
    routings = [
        {
            "id": str(uuid.uuid4()),
            "item_id": pump_assembly["id"],
            "name": "Pump Assembly Routing",
            "description": "Manufacturing routing for pump sub-assembly",
            "revision": "A",
            "status": "active",
            "operations": [
                {"sequence": 10, "work_center_id": work_centers[0]["id"], "operation_name": "Cut Aluminum Bar", "description": "Cut to required lengths", "setup_time_minutes": 15, "run_time_minutes": 10},
                {"sequence": 20, "work_center_id": work_centers[2]["id"], "operation_name": "Machine Components", "description": "CNC machining of pump body", "setup_time_minutes": 30, "run_time_minutes": 45},
                {"sequence": 30, "work_center_id": work_centers[3]["id"], "operation_name": "Assemble Pump", "description": "Assemble all components", "setup_time_minutes": 10, "run_time_minutes": 30},
                {"sequence": 40, "work_center_id": work_centers[4]["id"], "operation_name": "Test Pump", "description": "Pressure and leak testing", "setup_time_minutes": 5, "run_time_minutes": 15},
            ],
            "created_at": datetime.now(timezone.utc)
        },
        {
            "id": str(uuid.uuid4()),
            "item_id": hydraulic_press["id"],
            "name": "Hydraulic Press 50T Routing",
            "description": "Manufacturing routing for hydraulic press",
            "revision": "B",
            "status": "active",
            "operations": [
                {"sequence": 10, "work_center_id": work_centers[0]["id"], "operation_name": "Cut Steel Sheets", "description": "Cut frame components", "setup_time_minutes": 20, "run_time_minutes": 60},
                {"sequence": 20, "work_center_id": work_centers[1]["id"], "operation_name": "Weld Frame", "description": "Weld main frame structure", "setup_time_minutes": 45, "run_time_minutes": 120},
                {"sequence": 30, "work_center_id": work_centers[2]["id"], "operation_name": "Machine Mounting Points", "description": "CNC machine mounting surfaces", "setup_time_minutes": 30, "run_time_minutes": 60},
                {"sequence": 40, "work_center_id": work_centers[3]["id"], "operation_name": "Final Assembly", "description": "Install hydraulics and controls", "setup_time_minutes": 30, "run_time_minutes": 180},
                {"sequence": 50, "work_center_id": work_centers[4]["id"], "operation_name": "Final Testing", "description": "Full functional and safety test", "setup_time_minutes": 15, "run_time_minutes": 90},
            ],
            "created_at": datetime.now(timezone.utc)
        }
    ]
    await db.routings.insert_many(routings)
    logger.info("Sample routings seeded")


async def migrate_operations_status():
    """One-time migration: fix work orders with operation_name stored as dict {name, cost}.
    Also correct op.status when open runs still exist but op was previously marked 'stopped'
    due to the old shared-Stop bug that closed only one run but flagged the whole op stopped."""
    try:
        cursor = db.work_orders.find({"operations_status": {"$exists": True, "$ne": []}}, {"_id": 0, "id": 1, "operations_status": 1, "quantity": 1})
        fixed = 0
        async for wo in cursor:
            ops = wo.get("operations_status") or []
            changed = False
            new_ops = []
            mo_qty = wo.get("quantity") or 0
            for op in ops:
                on = op.get("operation_name")
                if isinstance(on, dict):
                    op["operation_name"] = on.get("name", "")
                    if "process_cost_per_unit" not in op:
                        op["process_cost_per_unit"] = float(on.get("cost", 0) or 0)
                    changed = True
                # Re-derive status from runs
                runs = op.get("runs") or []
                if runs:
                    any_open = any(r.get("ended_at") is None for r in runs)
                    total_done = sum((r.get("quantity_completed") or 0) for r in runs)
                    desired_status = (
                        "in_progress" if any_open
                        else ("completed" if total_done >= mo_qty else "stopped")
                    )
                    if op.get("status") != desired_status and op.get("status") != "completed":
                        # Don't override explicit completed state
                        if not (desired_status == "stopped" and op.get("status") == "completed"):
                            op["status"] = desired_status
                            changed = True
                new_ops.append(op)
            if changed:
                await db.work_orders.update_one({"id": wo["id"]}, {"$set": {"operations_status": new_ops}})
                fixed += 1
        if fixed:
            logger.info(f"Migrated operation_name/status on {fixed} work orders")
    except Exception as e:
        logger.exception(f"migrate_operations_status failed: {e}")

async def migrate_sc_jw_charges_from_bom():
    """Refresh job_work_parts.charges & bom_rollup_cost and lines[].rate (for Part/SA lines) on non-completed SC orders.
    Legacy SCs were created with incorrect process cost (either 0 or summed parent+component routings) and lines
    for completed Parts carried unit_cost=0 instead of BOM Total/Unit, which made DC send dialogs show ₹0 rates.
    Per user spec, charges/pc must equal the FG parent routing cost only, and Part/SA line rates must equal BOM Total/Unit."""
    try:
        cursor = db.subcontract_orders.find(
            {"status": {"$in": ["draft", "in_progress"]}},
            {"_id": 0, "id": 1, "job_work_parts": 1, "lines": 1}
        )
        touched = 0
        async for sc in cursor:
            changed = False
            # 1) Refresh job_work_parts
            parts = sc.get("job_work_parts") or []
            new_parts = []
            for p in parts:
                iid = p.get("item_id")
                if not iid:
                    new_parts.append(p); continue
                try:
                    bc = await compute_bom_costs(iid)
                    _fg = round(bc.get("fg_process_cost", 0) or 0, 2)
                    _total_unit = round((bc.get("rm_cost", 0) or 0) + (bc.get("process_cost", 0) or 0), 2)
                    if _fg and p.get("charges") != _fg:
                        p["charges"] = _fg
                        changed = True
                    if _total_unit and p.get("bom_rollup_cost") != _total_unit:
                        p["bom_rollup_cost"] = _total_unit
                        changed = True
                    _names = bc.get("process_names") or []
                    if _names and p.get("process_names") != _names:
                        p["process_names"] = _names
                        changed = True
                except Exception:
                    pass
                new_parts.append(p)
            # 2) Refresh lines[].rate for Part/SA items from BOM Total/Unit
            lines = sc.get("lines") or []
            new_lines = []
            for ln in lines:
                iid = ln.get("item_id")
                if not iid:
                    new_lines.append(ln); continue
                try:
                    it = await db.items.find_one({"id": iid}, {"_id": 0, "category": 1})
                    if it and it.get("category") in ("component", "sub_assembly"):
                        _tu = await compute_bom_total_unit_cost(iid)
                        if _tu > 0 and ln.get("rate") != round(_tu, 2):
                            ln["rate"] = round(_tu, 2)
                            changed = True
                except Exception:
                    pass
                new_lines.append(ln)
            if changed:
                await db.subcontract_orders.update_one({"id": sc["id"]}, {"$set": {"job_work_parts": new_parts, "lines": new_lines}})
                touched += 1
        if touched:
            logger.info(f"Refreshed job_work_parts charges and lines rates from BOM on {touched} SC orders")
    except Exception as e:
        logger.exception(f"migrate_sc_jw_charges_from_bom failed: {e}")

async def migrate_backfill_child_mo_operations_status():
    """Child MOs created BEFORE iter-74 (which moved routings from parent BOM's component
    entries to the child's own BOM parent_routings) may have empty operations_status. This
    migration walks all pending/in_progress inhouse child MOs with no ops and populates
    operations_status from the child item's own BOM.parent_routings."""
    try:
        cursor = db.work_orders.find({
            "parent_wo_id": {"$exists": True, "$ne": None},
            "is_subcontract": {"$ne": True},
            "status": {"$in": ["pending", "in_progress"]}
        }, {"_id": 0})
        fixed = 0
        async for wo in cursor:
            ops = wo.get("operations_status") or []
            if ops:
                continue
            item_id = wo.get("item_id")
            if not item_id:
                continue
            own_bom = await db.boms.find_one({"parent_item_id": item_id, "status": "active"}, {"_id": 0})
            if not own_bom:
                continue
            routings_list = own_bom.get("parent_routings") or []
            if not routings_list:
                continue
            new_ops = []
            for seq, r_entry in enumerate(normalize_routings(routings_list), 1):
                new_ops.append({
                    "sequence": seq * 10,
                    "operation_name": r_entry.get("name", ""),
                    "process_cost_per_unit": float(r_entry.get("cost", 0) or 0),
                    "work_center_id": "",
                    "work_center_name": "",
                    "is_job_work": False,
                    "job_work_supplier_id": "",
                    "status": "pending",
                    "quantity_completed": 0
                })
            if new_ops:
                await db.work_orders.update_one({"id": wo["id"]}, {"$set": {"operations_status": new_ops}})
                fixed += 1
        if fixed:
            logger.info(f"Backfilled operations_status on {fixed} child MOs from their own BOM.parent_routings")
    except Exception as e:
        logger.exception(f"migrate_backfill_child_mo_operations_status failed: {e}")


async def migrate_sync_component_routings_to_child_bom():
    """One-time migration: for every BOM component that has both (a) component-line routings
    AND (b) its own child BOM, copy those routings into the CHILD BOM's parent_routings (if
    the child's parent_routings is empty or has less cost). Then clear the component-line
    routings on the parent. This unifies the "one source of truth" model — PT-1's process
    cost now lives exclusively on PT-1's own BOM parent_routings."""
    try:
        all_boms = await db.boms.find({}, {"_id": 0}).to_list(5000)
        synced = 0
        for bom in all_boms:
            changed = False
            new_components = []
            for comp in (bom.get("components") or []):
                comp_routings = comp.get("routings") or []
                if comp_routings:
                    child_bom = await db.boms.find_one({"parent_item_id": comp.get("item_id")}, {"_id": 0})
                    if child_bom:
                        child_parent_routings = child_bom.get("parent_routings") or []
                        child_total = routings_total_cost(child_parent_routings)
                        comp_total = routings_total_cost(comp_routings)
                        # Only migrate if child has empty OR smaller parent_routings (avoid overwrites)
                        if comp_total > child_total:
                            await db.boms.update_one(
                                {"id": child_bom["id"]},
                                {"$set": {"parent_routings": normalize_routings(comp_routings)}}
                            )
                        # Clear parent's component-line routings — child BOM is now source of truth
                        comp["routings"] = []
                        changed = True
                new_components.append(comp)
            if changed:
                await db.boms.update_one({"id": bom["id"]}, {"$set": {"components": new_components}})
                synced += 1
        if synced:
            logger.info(f"Synced component-line routings → child BOM parent_routings on {synced} BOMs")
    except Exception as e:
        logger.exception(f"migrate_sync_component_routings_to_child_bom failed: {e}")

# ================== APP SETUP ==================

async def migrate_refresh_tax_invoice_qrs():
    """One-shot: rebuild UPI QR on legacy Tax Invoices that used the hardcoded machineworks@upi string."""
    company = await db.company_settings.find_one({"type": "company"}, {"_id": 0}) or {}
    upi_id = (company.get("bank_upi") or "").strip() or "na@upi"
    payee_name = (company.get("company_name") or "Company").strip().replace(" ", "")[:30] or "Company"
    count = 0
    async for ti in db.tax_invoices.find({"qr_code": {"$regex": "machineworks@upi"}}, {"_id": 0, "id": 1, "invoice_no": 1, "grand_total": 1}):
        amt = float(ti.get("grand_total") or 0)
        safe_tn = (ti.get("invoice_no") or "").replace("/", "-")
        new_qr = f"upi://pay?pa={upi_id}&pn={payee_name}&am={amt:.2f}&tn={safe_tn}&cu=INR"
        await db.tax_invoices.update_one({"id": ti["id"]}, {"$set": {"qr_code": new_qr}})
        count += 1
    if count > 0:
        logger.info(f"[migrate] Refreshed UPI QR on {count} legacy Tax Invoices (new payee: {upi_id})")


async def migrate_compact_fy_number_series():
    """Persistent cleanup: normalize any legacy `current_fy` value (e.g. 'FY26-27', '26-27') to the compact 4-digit form.
    Also strips any trailing dashes from stored prefixes that survived the iter-109 migration."""
    count_fy = 0
    count_prefix = 0
    async for doc in db.number_series.find({}, {"_id": 0, "key": 1, "doc_type": 1, "current_fy": 1, "prefix": 1}):
        updates = {}
        fy = doc.get("current_fy")
        if fy:
            norm = _normalize_fy(fy)
            if norm != fy:
                updates["current_fy"] = norm
                count_fy += 1
        prefix = doc.get("prefix")
        if prefix and prefix.endswith(("-", "/", " ")):
            updates["prefix"] = prefix.rstrip("-/ ")
            count_prefix += 1
        if updates:
            q = {"key": doc["key"]} if doc.get("key") else {"doc_type": doc["doc_type"]}
            await db.number_series.update_one(q, {"$set": updates})
    if count_fy or count_prefix:
        logger.info(f"[migrate] Normalized {count_fy} FY field(s) and stripped {count_prefix} prefix trailer(s) in number_series")


async def migrate_purchase_invoices_perm_to_accounts():
    """One-time migration. Previously this folded the legacy `purchase_invoices`
    permission into a unified `accounts` module. The Accounts module has since
    been split back into two granular modules (`purchase_invoices` and
    `tax_invoices`) per user request. So this migration now performs BOTH
    legacy-fold and split logic:

    1. legacy `purchase_invoices` (pre-Accounts) → `purchase_invoices`
       (still a valid key — only needs to be carried forward).
    2. legacy `accounts` (the unified module) → BOTH `purchase_invoices` AND
       `tax_invoices` so users who had Accounts access keep both pages.
    Idempotent — only writes when there's a diff to apply. Never deletes
    legacy keys (safe rollback).
    """
    try:
        # Step 1: backfill `purchase_invoices` from legacy `purchase_invoices`
        # (no-op if already correct; left for clarity if older docs lacked it).
        # Step 2: split `accounts` into both pi + ti.
        cursor = db.role_groups.find(
            {"$or": [
                {"permissions.accounts": {"$exists": True}},
                {"permissions.purchase_invoices": {"$exists": True}},
            ]},
            {"_id": 0, "id": 1, "permissions": 1},
        )
        migrated = 0
        async for rg in cursor:
            perms = rg.get("permissions") or {}
            legacy_pi = perms.get("purchase_invoices") or []
            legacy_acc = perms.get("accounts") or []
            current_pi = perms.get("purchase_invoices") or []
            current_ti = perms.get("tax_invoices") or []
            merged_pi = list({*current_pi, *legacy_pi, *legacy_acc})
            merged_ti = list({*current_ti, *legacy_acc})
            updates = {}
            if set(merged_pi) != set(current_pi):
                updates["permissions.purchase_invoices"] = merged_pi
            if set(merged_ti) != set(current_ti):
                updates["permissions.tax_invoices"] = merged_ti
            if updates:
                await db.role_groups.update_one({"id": rg["id"]}, {"$set": updates})
                migrated += 1
        if migrated:
            logger.info(f"[migrate] Split/backfilled purchase_invoices + tax_invoices on {migrated} role group(s)")
    except Exception as e:
        logger.exception(f"migrate_purchase_invoices_perm_to_accounts failed: {e}")


@app.on_event("startup")
async def startup():
    # Create indexes
    await db.users.create_index("email", unique=True)
    await db.items.create_index("part_number", unique=True)
    await db.items.create_index("category")
    await db.boms.create_index("parent_item_id")
    await db.boms.create_index("status")
    await db.production_orders.create_index("status")
    await db.login_attempts.create_index("identifier")
    await db.suppliers.create_index("code", unique=True)
    await db.purchase_orders.create_index("status")
    await db.warehouses.create_index("code", unique=True)
    await db.work_centers.create_index("code", unique=True)
    await db.routings.create_index("item_id")
    await db.work_orders.create_index("status")
    await db.customers.create_index("code", unique=True)
    await db.company_settings.create_index("type", unique=True)
    
    # Seed data
    await seed_admin()
    await seed_sample_data()
    await migrate_operations_status()
    await migrate_sync_component_routings_to_child_bom()
    await migrate_backfill_child_mo_operations_status()
    await migrate_sc_jw_charges_from_bom()
    await migrate_refresh_tax_invoice_qrs()
    await migrate_compact_fy_number_series()
    await migrate_purchase_invoices_perm_to_accounts()
    
    # Write credentials file (dev environment only, non-fatal on Windows/other OS)
    try:
        Path("/app/memory").mkdir(exist_ok=True)
        with open("/app/memory/test_credentials.md", "w") as f:
            f.write("# Test Credentials\n\n")
            f.write("## Admin Account\n")
            f.write(f"- Email: {os.environ.get('ADMIN_EMAIL', 'admin@erp.com')}\n")
            f.write(f"- Password: {os.environ.get('ADMIN_PASSWORD', 'Admin@123')}\n")
            f.write("- Role: admin\n\n")
            f.write("## Auth Endpoints\n")
            f.write("- POST /api/auth/login\n")
            f.write("- POST /api/auth/register\n")
            f.write("- POST /api/auth/logout\n")
            f.write("- GET /api/auth/me\n")
            f.write("- POST /api/auth/refresh\n")
    except (OSError, PermissionError):
        pass

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

# ================== PURCHASE INVOICE ROUTES ==================

@purchase_invoices_router.get("")
async def get_purchase_invoices(
    request: Request,
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    user = await get_current_user(request)
    query = {}
    if status:
        query["status"] = status
    # Optional date range filter on invoice_date (YYYY-MM-DD inclusive).
    if date_from or date_to:
        rng = {}
        if date_from:
            try:
                rng["$gte"] = datetime.fromisoformat(date_from.replace("Z", "+00:00"))
            except Exception:
                pass
        if date_to:
            try:
                end = datetime.fromisoformat(date_to.replace("Z", "+00:00"))
                # Make `date_to` inclusive of the entire day when only a date is sent.
                if end.hour == 0 and end.minute == 0 and end.second == 0:
                    end = end.replace(hour=23, minute=59, second=59)
                rng["$lte"] = end
            except Exception:
                pass
        if rng:
            query["invoice_date"] = rng
    # Sort by invoice_date desc (newest first) with created_at as tiebreaker.
    invoices = await db.purchase_invoices.find(query, {"_id": 0}).sort([("invoice_date", -1), ("created_at", -1)]).to_list(2000)
    for inv in invoices:
        # JW-based PIs may not carry supplier_id on the invoice doc itself —
        # the supplier lives on the linked subcontract order. Walk that chain
        # so the list always shows a supplier name even for JW invoices.
        supplier_id = inv.get("supplier_id")
        jw_order = None
        jw_id = inv.get("jw_order_id") or inv.get("sc_order_id")
        if not supplier_id and inv.get("grn_id"):
            grn = await db.grn.find_one({"id": inv["grn_id"]}, {"_id": 0, "supplier_id": 1, "jw_order_id": 1, "sc_order_id": 1, "po_id": 1})
            if grn:
                supplier_id = grn.get("supplier_id") or supplier_id
                jw_id = jw_id or grn.get("jw_order_id") or grn.get("sc_order_id")
                if not inv.get("po_id") and grn.get("po_id"):
                    inv["po_id"] = grn["po_id"]
        if jw_id and not jw_order:
            jw_order = await db.subcontract_orders.find_one({"id": jw_id}, {"_id": 0})
            if jw_order:
                inv["jw_order"] = jw_order
                supplier_id = supplier_id or jw_order.get("supplier_id")
        supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0}) if supplier_id else None
        inv["supplier"] = supplier
        if inv.get("po_id"):
            po = await db.purchase_orders.find_one({"id": inv["po_id"]}, {"_id": 0})
            inv["po"] = po
        if inv.get("grn_id"):
            grn = await db.grn.find_one({"id": inv["grn_id"]}, {"_id": 0})
            inv["grn"] = grn
        for line in inv.get("lines", []):
            item = await db.items.find_one({"id": line.get("item_id")}, {"_id": 0})
            line["item"] = item
    return invoices

@purchase_invoices_router.get("/pending-grns")
async def get_grns_pending_invoice(request: Request):
    """Get GRNs (PO-based and JW-based) that don't have a purchase invoice yet.
    Excludes draft GRNs (these have not been approved yet, so no invoice can
    be raised against them). Treats both legacy `status='completed'` and new
    `status='posted'` as valid posted states."""
    user = await get_current_user(request)
    # Get all GRN IDs that already have invoices (single grn_id OR multi grn_ids)
    invoiced_grn_ids = set()
    async for inv in db.purchase_invoices.find(
        {"$or": [
            {"grn_id": {"$exists": True, "$ne": ""}},
            {"grn_ids": {"$exists": True, "$ne": []}},
        ]},
        {"grn_id": 1, "grn_ids": 1},
    ):
        if inv.get("grn_id"):
            invoiced_grn_ids.add(inv["grn_id"])
        for gid in (inv.get("grn_ids") or []):
            invoiced_grn_ids.add(gid)

    # Get all posted GRNs (legacy "completed" + new "posted"). Drafts excluded.
    grns = await db.grn.find(
        {"status": {"$in": ["completed", "posted"]}},
        {"_id": 0},
    ).sort("created_at", -1).to_list(1000)
    
    pending = []
    for grn in grns:
        if grn.get("id") in invoiced_grn_ids:
            continue
        
        # Determine GRN type
        is_jw = bool(grn.get("jw_order_id") or grn.get("sc_order_id"))
        grn["is_jw"] = is_jw
        
        # Fetch supplier from JW order or PO
        supplier_id = grn.get("supplier_id")
        if is_jw and not supplier_id:
            jw_id = grn.get("jw_order_id") or grn.get("sc_order_id")
            jw_order = await db.subcontract_orders.find_one({"id": jw_id}, {"_id": 0})
            if jw_order:
                supplier_id = jw_order.get("supplier_id")
                grn["jw_order"] = jw_order
        elif is_jw:
            jw_id = grn.get("jw_order_id") or grn.get("sc_order_id")
            grn["jw_order"] = await db.subcontract_orders.find_one({"id": jw_id}, {"_id": 0})
        
        if not supplier_id:
            po = await db.purchase_orders.find_one({"id": grn.get("po_id")}, {"_id": 0})
            grn["po"] = po
            supplier_id = po.get("supplier_id") if po else None
        else:
            if grn.get("po_id"):
                grn["po"] = await db.purchase_orders.find_one({"id": grn.get("po_id")}, {"_id": 0})
        
        grn["supplier"] = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0}) if supplier_id else None
        
        for line in grn.get("lines", []):
            item = await db.items.find_one({"id": line.get("item_id")}, {"_id": 0})
            line["item"] = item
        pending.append(grn)
    
    return pending

@purchase_invoices_router.post("", status_code=201)
async def create_purchase_invoice(data: PurchaseInvoiceCreate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="purchase_invoices", action="create")
    count = await db.purchase_invoices.count_documents({})
    inv_number = f"PI-{str(count + 1).zfill(6)}"

    supplier = await db.suppliers.find_one({"id": data.supplier_id})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")

    # ------------------------------------------------------------------
    # Multi-GRN validation: when `grn_ids` carries multiple entries every
    # referenced GRN MUST belong to the same supplier as the invoice. This
    # prevents an operator from accidentally mixing two suppliers' GRNs
    # into a single invoice (which would break tax breakdown + tally exports).
    # We also normalise — if the caller supplied a single-element grn_ids
    # list, we copy it into grn_id for backward compat with the legacy field.
    # ------------------------------------------------------------------
    grn_ids: List[str] = list(data.grn_ids or [])
    if data.grn_id and data.grn_id not in grn_ids:
        grn_ids.insert(0, data.grn_id)
    if grn_ids:
        for gid in grn_ids:
            grn = await db.grn.find_one({"id": gid}, {"_id": 0, "supplier_id": 1, "po_id": 1, "status": 1})
            if not grn:
                raise HTTPException(status_code=404, detail=f"GRN {gid} not found")
            if (grn.get("status") or "posted") not in ("posted", "completed"):
                raise HTTPException(status_code=400, detail=f"GRN {gid} is not posted yet — approve it before raising an invoice.")
            grn_supplier = grn.get("supplier_id")
            if not grn_supplier and grn.get("po_id"):
                po_doc = await db.purchase_orders.find_one({"id": grn["po_id"]}, {"_id": 0, "supplier_id": 1})
                grn_supplier = (po_doc or {}).get("supplier_id")
            if grn_supplier and grn_supplier != data.supplier_id:
                raise HTTPException(status_code=400, detail="All selected GRNs must belong to the same supplier as the invoice.")

    company = await db.company_settings.find_one({"type": "company"}, {"_id": 0})
    is_inter_state = supplier.get("state_code", "") != (company or {}).get("state_code", "")
    
    lines = []
    subtotal = 0
    total_cgst = 0
    total_sgst = 0
    total_igst = 0
    
    for line in data.lines:
        item = await db.items.find_one({"id": line.item_id}, {"_id": 0})
        if not item:
            continue
        line_total = line.quantity * line.unit_price
        discount_amount = line.discount or 0
        line_after_discount = line_total - discount_amount
        gst_rate = line.gst_rate if line.gst_rate is not None else item.get("gst_rate", 18)
        gst_amount = line_after_discount * gst_rate / 100
        
        if is_inter_state:
            total_igst += gst_amount
        else:
            total_cgst += gst_amount / 2
            total_sgst += gst_amount / 2
        
        subtotal += line_after_discount
        lines.append({
            "item_id": line.item_id,
            "quantity": line.quantity,
            "unit_price": line.unit_price,
            "discount": discount_amount,
            "hsn_code": line.hsn_code or item.get("hsn_code", ""),
            "gst_rate": gst_rate,
            "line_total": line_after_discount,
            "gst_amount": gst_amount,
            "is_process_charge": bool(getattr(line, "is_process_charge", False)),
            "description": line.description or ""
        })
    
    total_tax = total_cgst + total_sgst + total_igst

    # Additional charges (freight / packaging / insurance) — mirror PO behaviour:
    # taxed individually, summed into totals. Stored on the invoice doc so the
    # print template can render the breakdown.
    charges_subtotal = 0.0
    charges_cgst_total = 0.0
    charges_sgst_total = 0.0
    charges_igst_total = 0.0
    charges_with_tax = []
    for charge in (data.additional_charges or []):
        c_amount = float(charge.amount or 0)
        c_gst_rate = float(charge.gst_rate or 0)
        c_tax = round(c_amount * c_gst_rate / 100, 2)
        if is_inter_state:
            charges_igst_total += c_tax
        else:
            charges_cgst_total += c_tax / 2
            charges_sgst_total += c_tax / 2
        charges_subtotal += c_amount
        charges_with_tax.append({
            "name": charge.name,
            "amount": c_amount,
            "gst_rate": c_gst_rate,
            "tax_amount": c_tax,
            "total_with_tax": round(c_amount + c_tax, 2),
        })
    total_cgst += charges_cgst_total
    total_sgst += charges_sgst_total
    total_igst += charges_igst_total
    total_tax = total_cgst + total_sgst + total_igst
    total_amount = subtotal + charges_subtotal + total_tax
    
    invoice_doc = {
        "id": str(uuid.uuid4()),
        "invoice_number": inv_number,
        "supplier_id": data.supplier_id,
        "po_id": data.po_id or "",
        "grn_id": (grn_ids[0] if grn_ids else (data.grn_id or "")),
        "grn_ids": grn_ids,  # Multi-GRN linkage (always populated when GRN-based)
        "is_manual": bool(data.is_manual),
        "invoice_no": data.invoice_no,
        "invoice_date": data.invoice_date,
        "due_date": data.due_date,
        "lines": lines,
        "additional_charges": charges_with_tax,
        "subtotal": round(subtotal, 2),
        "charges_subtotal": round(charges_subtotal, 2),
        "total_cgst": round(total_cgst, 2),
        "total_sgst": round(total_sgst, 2),
        "total_igst": round(total_igst, 2),
        "total_tax": round(total_tax, 2),
        "total_amount": round(total_amount, 2),
        "is_inter_state": is_inter_state,
        "status": "draft",
        "notes": data.notes or "",
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"]
    }
    
    await db.purchase_invoices.insert_one(invoice_doc)
    del invoice_doc["_id"]
    return invoice_doc

@purchase_invoices_router.put("/{invoice_id}")
async def update_purchase_invoice(invoice_id: str, data: PurchaseInvoiceUpdate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="purchase_invoices", action="edit")
    invoice = await db.purchase_invoices.find_one({"id": invoice_id})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if update_data:
        # Edits to an APPROVED invoice automatically reset it to "draft" so it
        # has to be re-approved. Mirrors how PO revisions clear the approval.
        # Only triggers if the user actually modified content (lines / charges
        # / amount / dates), not when toggling status itself.
        content_changed = any(
            k in update_data for k in ("lines", "additional_charges", "invoice_no", "invoice_date", "due_date", "notes")
        )
        was_approved = invoice.get("status") == "approved"
        if content_changed and was_approved and "status" not in update_data:
            update_data["status"] = "draft"
            update_data["approved_at"] = None
            update_data["approved_by"] = None
        update_data["updated_at"] = datetime.now(timezone.utc)
        # Recompute totals if lines or additional_charges changed.
        if "lines" in update_data or "additional_charges" in update_data:
            supplier = await db.suppliers.find_one({"id": invoice.get("supplier_id")})
            company_settings = await db.company_settings.find_one({"type": "company"}, {"_id": 0})
            company_state = (company_settings or {}).get("state_code", "")
            supplier_state = (supplier or {}).get("state_code", "")
            is_inter_state = bool(company_state) and bool(supplier_state) and company_state != supplier_state
            new_lines = update_data.get("lines") if "lines" in update_data else invoice.get("lines") or []
            new_charges_input = update_data.get("additional_charges") if "additional_charges" in update_data else invoice.get("additional_charges") or []
            subtotal = 0.0
            total_cgst = 0.0
            total_sgst = 0.0
            total_igst = 0.0
            for ln in new_lines:
                amt = float(ln.get("quantity", 0)) * float(ln.get("unit_price", 0)) - float(ln.get("discount", 0) or 0)
                tax = amt * float(ln.get("gst_rate", 0) or 0) / 100
                if is_inter_state:
                    total_igst += tax
                else:
                    total_cgst += tax / 2
                    total_sgst += tax / 2
                subtotal += amt
            charges_subtotal = 0.0
            charges_with_tax = []
            for c in new_charges_input:
                c_amt = float(c.get("amount", 0))
                c_rate = float(c.get("gst_rate", 0) or 0)
                c_tax = round(c_amt * c_rate / 100, 2)
                if is_inter_state:
                    total_igst += c_tax
                else:
                    total_cgst += c_tax / 2
                    total_sgst += c_tax / 2
                charges_subtotal += c_amt
                charges_with_tax.append({
                    "name": c.get("name", ""),
                    "amount": c_amt,
                    "gst_rate": c_rate,
                    "tax_amount": c_tax,
                    "total_with_tax": round(c_amt + c_tax, 2),
                })
            update_data["additional_charges"] = charges_with_tax
            update_data["subtotal"] = round(subtotal, 2)
            update_data["charges_subtotal"] = round(charges_subtotal, 2)
            update_data["total_cgst"] = round(total_cgst, 2)
            update_data["total_sgst"] = round(total_sgst, 2)
            update_data["total_igst"] = round(total_igst, 2)
            total_tax = total_cgst + total_sgst + total_igst
            update_data["total_tax"] = round(total_tax, 2)
            update_data["total_amount"] = round(subtotal + charges_subtotal + total_tax, 2)
        await db.purchase_invoices.update_one({"id": invoice_id}, {"$set": update_data})
    
    return await db.purchase_invoices.find_one({"id": invoice_id}, {"_id": 0})

@purchase_invoices_router.post("/{invoice_id}/approve")
async def approve_purchase_invoice(invoice_id: str, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin"], module="purchase_invoices", action="edit")
    invoice = await db.purchase_invoices.find_one({"id": invoice_id})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    if invoice.get("status") != "draft":
        raise HTTPException(status_code=400, detail="Only draft invoices can be approved")
    await db.purchase_invoices.update_one({"id": invoice_id}, {"$set": {"status": "approved", "approved_at": datetime.now(timezone.utc), "approved_by": user["id"]}})
    return await db.purchase_invoices.find_one({"id": invoice_id}, {"_id": 0})

@purchase_invoices_router.post("/{invoice_id}/mark-paid")
async def mark_invoice_paid(invoice_id: str, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin"], module="purchase_invoices", action="edit")
    invoice = await db.purchase_invoices.find_one({"id": invoice_id})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    if invoice.get("status") != "approved":
        raise HTTPException(status_code=400, detail="Only approved invoices can be marked as paid")
    await db.purchase_invoices.update_one({"id": invoice_id}, {"$set": {"status": "paid", "paid_at": datetime.now(timezone.utc), "paid_by": user["id"]}})
    return await db.purchase_invoices.find_one({"id": invoice_id}, {"_id": 0})

# ================== TALLY XML EXPORT ==================
# Generates Tally-compatible XML for Purchase Invoice import. The file can be imported
# via Tally Gateway → Import Data → Vouchers. No Tally HTTP daemon required.
# Reference voucher structure: https://tallyhelp.tallysolutions.com/docs/te9rel64/index.htm

def _xml_escape(v):
    if v is None:
        return ""
    s = str(v)
    return (s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
             .replace('"', "&quot;").replace("'", "&apos;"))

def _tally_date(dt):
    """Tally wants YYYYMMDD format."""
    if isinstance(dt, str):
        try:
            dt = datetime.fromisoformat(dt.replace("Z", "+00:00"))
        except Exception:
            return ""
    if isinstance(dt, datetime):
        return dt.strftime("%Y%m%d")
    return ""


# Indian GST state code → state name. Tally Prime stores the full state name
# (e.g. "Maharashtra"), but most ERP supplier records carry only the 2-digit
# state_code (the first two characters of the GSTIN). Without this lookup, the
# Tally Ledger Alteration screen shows "•Not Applicable" for State / Country
# even when the GSTIN is valid.
_GST_STATE_CODE_TO_NAME = {
    "01": "Jammu and Kashmir", "02": "Himachal Pradesh", "03": "Punjab",
    "04": "Chandigarh", "05": "Uttarakhand", "06": "Haryana", "07": "Delhi",
    "08": "Rajasthan", "09": "Uttar Pradesh", "10": "Bihar", "11": "Sikkim",
    "12": "Arunachal Pradesh", "13": "Nagaland", "14": "Manipur",
    "15": "Mizoram", "16": "Tripura", "17": "Meghalaya", "18": "Assam",
    "19": "West Bengal", "20": "Jharkhand", "21": "Odisha",
    "22": "Chhattisgarh", "23": "Madhya Pradesh", "24": "Gujarat",
    "25": "Daman and Diu", "26": "Dadra and Nagar Haveli",
    "27": "Maharashtra", "28": "Andhra Pradesh", "29": "Karnataka",
    "30": "Goa", "31": "Lakshadweep", "32": "Kerala", "33": "Tamil Nadu",
    "34": "Puducherry", "35": "Andaman and Nicobar Islands",
    "36": "Telangana", "37": "Andhra Pradesh", "38": "Ladakh",
    "97": "Other Territory", "99": "Centre Jurisdiction",
}


def _resolve_state_name(party):
    """Return the best human-readable state name for a supplier / customer.

    Preference:
      1) party['state']  (if non-empty after strip)
      2) GST-code-to-name lookup on party['state_code']
      3) First 2 chars of GSTIN if state_code is missing
      4) "" (lets caller emit blank rather than wrong data)
    """
    if not party:
        return ""
    state = (party.get("state") or "").strip()
    if state:
        return state
    code = (party.get("state_code") or "").strip()
    if not code:
        gstin = (party.get("gstin") or "").strip()
        if len(gstin) >= 2 and gstin[:2].isdigit():
            code = gstin[:2]
    if code and len(code) == 1:
        code = "0" + code  # pad e.g. "9" → "09"
    return _GST_STATE_CODE_TO_NAME.get(code, "")


def _build_tally_party_ledger_xml(party, parent_group, ref_date):
    """Build the LEDGER master <TALLYMESSAGE> block for a supplier/customer.

    Used for both supplier (parent="Sundry Creditors") and customer
    (parent="Sundry Debtors") so the Tally Prime Ledger Alteration screen
    shows the full Address / State / Country / GSTIN block instead of
    "•Not Applicable".

    `ref_date` is the YYYYMMDD applicable-from date used inside
    LEDMAILINGDETAILS.LIST and LEDGSTREGDETAILS.LIST.
    """
    if not party or not party.get("name"):
        return ""
    party_name = _xml_escape(party.get("name", ""))
    gstin = _xml_escape(party.get("gstin") or "")
    pan = _xml_escape((party.get("pan") or (party.get("gstin", "")[2:12] if party.get("gstin") else "")))
    addr_lines = []
    for k in ("address", "address_line2", "city"):
        v = party.get(k)
        if v:
            addr_lines.append(_xml_escape(str(v)))
    state = _xml_escape(_resolve_state_name(party))
    pincode = _xml_escape(party.get("pin_code") or party.get("pincode") or "")
    country = _xml_escape(party.get("country") or "India")
    addr_xml = ("<ADDRESS.LIST>" + "".join(f"<ADDRESS>{a}</ADDRESS>" for a in addr_lines) + "</ADDRESS.LIST>") if addr_lines else ""
    mail_addr_xml = "".join(f"<ADDRESS>{_xml_escape(str(party.get(k, '') or ''))}</ADDRESS>" for k in ("address", "address_line2", "city") if party.get(k))
    reg_type = 'Regular' if gstin else 'Unregistered/Consumer'
    parent_x = _xml_escape(parent_group)
    return f"""
    <TALLYMESSAGE xmlns:UDF="TallyUDF">
      <LEDGER NAME="{party_name}" RESERVEDNAME="" ACTION="Alter">
        <MAILINGNAME.LIST TYPE="String">
          <MAILINGNAME>{party_name}</MAILINGNAME>
        </MAILINGNAME.LIST>
        <NAME.LIST>
          <NAME>{party_name}</NAME>
        </NAME.LIST>
        <PARENT>{parent_x}</PARENT>
        <ISBILLWISEON>Yes</ISBILLWISEON>
        <ISCOSTCENTRESON>No</ISCOSTCENTRESON>
        {addr_xml}
        <COUNTRYNAME>{country}</COUNTRYNAME>
        <LEDSTATENAME>{state}</LEDSTATENAME>
        <STATENAME>{state}</STATENAME>
        <PINCODE>{pincode}</PINCODE>
        <PARTYGSTIN>{gstin}</PARTYGSTIN>
        <GSTREGISTRATIONTYPE>{reg_type}</GSTREGISTRATIONTYPE>
        <INCOMETAXNUMBER>{pan}</INCOMETAXNUMBER>
        <LEDMAILINGDETAILS.LIST>
          <APPLICABLEFROM>{ref_date}</APPLICABLEFROM>
          <MAILINGNAME>{party_name}</MAILINGNAME>
          <ADDRESS.LIST TYPE="String">{mail_addr_xml}</ADDRESS.LIST>
          <STATE>{state}</STATE>
          <COUNTRY>{country}</COUNTRY>
          <PINCODE>{pincode}</PINCODE>
        </LEDMAILINGDETAILS.LIST>
        <LEDGSTREGDETAILS.LIST>
          <APPLICABLEFROM>{ref_date}</APPLICABLEFROM>
          <GSTREGISTRATIONTYPE>{reg_type}</GSTREGISTRATIONTYPE>
          <STATE>{state}</STATE>
          <GSTIN>{gstin}</GSTIN>
          <PARTYTYPE>Not Applicable</PARTYTYPE>
        </LEDGSTREGDETAILS.LIST>
      </LEDGER>
    </TALLYMESSAGE>"""


def _build_tally_master_messages(invoice, supplier, lines, additional_charges):
    """Emit the stock-item + ledger master <TALLYMESSAGE> blocks the voucher
    references. Without these, Tally treats unknown stock items as text-only
    (qty/rate blank) and unknown ledgers as plain-amount entries. Each
    master uses `ACTION="Alter"` with `CANDELETE="No"` so re-importing the
    same XML is idempotent.

    Two categories of master are pushed here:
      1) Stock items for every PI line — gives Tally a valid stock master
         so BILLEDQTY + RATE actually show up in the voucher view.
      2) The supplier ledger (Sundry Creditor) with full address + GSTIN —
         fixes the "Party A/c GSTIN and address not showing" bug.
    """
    blocks: List[str] = []

    # ---------------- Supplier ledger master with GSTIN + address ----------
    inv_date_xml = _tally_date(invoice.get('invoice_date')) or '20170701'
    sup_block = _build_tally_party_ledger_xml(supplier, "Sundry Creditors", inv_date_xml)
    if sup_block:
        blocks.append(sup_block)

    # ---------------- Stock item masters --------------------------------
    blocks.append(_build_tally_inventory_masters(invoice, lines, additional_charges))
    return "".join(blocks)


def _stock_item_name_for_line(ln):
    """Resolve the Tally STOCKITEMNAME for an invoice line.

    For PI lines the item doc lives under ln['item']; for TI / SO lines we
    only have description + hsn_code. The fallback chain is:
      1) part_number + name (PI shape)
      2) description (TI / SO shape)
      3) HSN-XXXXX (last resort so Tally still sees a master)
    """
    it = ln.get("item") or {}
    pn = (it.get("part_number") or "").strip()
    nm = (it.get("name") or "").strip()
    composite = f"{pn} - {nm}".strip(" -")
    if composite:
        return composite
    desc = (ln.get("description") or "").strip().splitlines()[0].strip() if ln.get("description") else ""
    if desc:
        return desc[:80]
    hsn = (ln.get("hsn_code") or it.get("hsn_code") or "").strip()
    if hsn:
        return f"HSN-{hsn}"
    return "Misc Item"


def _build_tally_inventory_masters(invoice, lines, additional_charges):
    """Emit STOCKITEM + UNIT + additional-charge ledger masters.

    Without these masters Tally treats STOCKITEMNAME as plain text → quantity
    + rate disappear in the voucher view. Reused by both Purchase (Sundry
    Creditor flow) and Sales (Sundry Debtor flow) exports.
    """
    blocks: List[str] = []
    ref_date = _tally_date(invoice.get("invoice_date")) or "20170701"
    # Stock items
    seen_stock = set()
    for ln in lines:
        si_name = _stock_item_name_for_line(ln)
        if not si_name or si_name in seen_stock:
            continue
        seen_stock.add(si_name)
        it = ln.get("item") or {}
        uom = _xml_escape(it.get("unit_of_measure") or ln.get("uom") or "Nos")
        si_name_x = _xml_escape(si_name)
        hsn = _xml_escape(ln.get("hsn_code") or it.get("hsn_code") or "")
        gst_rate = float(ln.get("gst_rate") or it.get("gst_rate") or 0)
        blocks.append(f"""
    <TALLYMESSAGE xmlns:UDF="TallyUDF">
      <STOCKITEM NAME="{si_name_x}" RESERVEDNAME="" ACTION="Alter">
        <NAME.LIST>
          <NAME>{si_name_x}</NAME>
        </NAME.LIST>
        <BASEUNITS>{uom}</BASEUNITS>
        <GSTAPPLICABLE>&#4; Applicable</GSTAPPLICABLE>
        <GSTTYPEOFSUPPLY>Goods</GSTTYPEOFSUPPLY>
        <HSNCODE>{hsn}</HSNCODE>
        <GSTDETAILS.LIST>
          <APPLICABLEFROM>{ref_date}</APPLICABLEFROM>
          <HSNCODE>{hsn}</HSNCODE>
          <STATEWISEDETAILS.LIST>
            <STATENAME>&#4; Any</STATENAME>
            <RATEDETAILS.LIST>
              <GSTRATEDUTYHEAD>CGST</GSTRATEDUTYHEAD>
              <GSTRATE>{gst_rate/2:.2f}</GSTRATE>
            </RATEDETAILS.LIST>
            <RATEDETAILS.LIST>
              <GSTRATEDUTYHEAD>SGST/UTGST</GSTRATEDUTYHEAD>
              <GSTRATE>{gst_rate/2:.2f}</GSTRATE>
            </RATEDETAILS.LIST>
            <RATEDETAILS.LIST>
              <GSTRATEDUTYHEAD>IGST</GSTRATEDUTYHEAD>
              <GSTRATE>{gst_rate:.2f}</GSTRATE>
            </RATEDETAILS.LIST>
          </STATEWISEDETAILS.LIST>
        </GSTDETAILS.LIST>
      </STOCKITEM>
    </TALLYMESSAGE>""")

    # UOM masters
    seen_uoms = set()
    for ln in lines:
        it = ln.get("item") or {}
        uom = (it.get("unit_of_measure") or ln.get("uom") or "Nos").strip()
        if not uom or uom in seen_uoms:
            continue
        seen_uoms.add(uom)
        uom_x = _xml_escape(uom)
        blocks.append(f"""
    <TALLYMESSAGE xmlns:UDF="TallyUDF">
      <UNIT NAME="{uom_x}" RESERVEDNAME="" ACTION="Alter">
        <NAME.LIST><NAME>{uom_x}</NAME></NAME.LIST>
        <ISSIMPLEUNIT>Yes</ISSIMPLEUNIT>
      </UNIT>
    </TALLYMESSAGE>""")

    # Additional-charge ledger masters
    for charge in (additional_charges or []):
        ch_name = (charge.get("name") or "").strip()
        if not ch_name:
            continue
        ch_name_x = _xml_escape(ch_name)
        blocks.append(f"""
    <TALLYMESSAGE xmlns:UDF="TallyUDF">
      <LEDGER NAME="{ch_name_x}" RESERVEDNAME="" ACTION="Alter">
        <NAME.LIST><NAME>{ch_name_x}</NAME></NAME.LIST>
        <PARENT>Indirect Expenses</PARENT>
        <ISBILLWISEON>No</ISBILLWISEON>
      </LEDGER>
    </TALLYMESSAGE>""")
    return "".join(blocks)


def _build_tally_purchase_voucher_xml(invoice, supplier, company, lines, is_inter_state):
    """Build a single <TALLYMESSAGE> block for a purchase voucher.

    Each line item is emitted as an ALLINVENTORYENTRIES.LIST block containing
    BILLEDQTY (quantity) + RATE (unit price) + AMOUNT (extended), plus per-
    line CGST/SGST/IGST classifications so Tally can re-derive the tax slabs.

    Additional charges (Freight, Packaging, Insurance, etc.) are emitted as
    SEPARATE ledger entries with their own GST classifications — keeps the
    purchase voucher's lines clean and lets Tally allocate the charge to its
    own ledger account.
    """
    inv_no = _xml_escape(invoice.get("invoice_no") or invoice.get("invoice_number") or "")
    inv_date = _tally_date(invoice.get("invoice_date"))
    party_name = _xml_escape(supplier.get("name", ""))
    narration = _xml_escape(invoice.get("notes", "") or f"Purchase against invoice {inv_no}")
    state_name = _xml_escape(_resolve_state_name(supplier))

    subtotal = float(invoice.get("subtotal", 0) or 0)
    charges_subtotal = float(invoice.get("charges_subtotal", 0) or 0)
    total_cgst = float(invoice.get("total_cgst", 0) or 0)
    total_sgst = float(invoice.get("total_sgst", 0) or 0)
    total_igst = float(invoice.get("total_igst", 0) or 0)
    total_amount = float(invoice.get("total_amount", 0) or 0)
    additional_charges = invoice.get("additional_charges", []) or []

    # Inventory entries per line — Tally's preferred shape for purchase
    # invoice imports. Includes qty + rate + per-line tax classifications.
    inv_entries = []
    for ln in lines:
        it = ln.get("item") or {}
        name = _xml_escape(f"{it.get('part_number','')} - {it.get('name','')}".strip(" -"))
        qty = float(ln.get("quantity", 0) or 0)
        rate = float(ln.get("unit_price", 0) or 0)
        line_total = float(ln.get("line_total") or (qty * rate))
        uom = _xml_escape(it.get("unit_of_measure", "Nos"))
        line_gst_rate = float(ln.get("gst_rate", 0) or 0)
        line_gst_amt = float(ln.get("gst_amount", 0) or 0)
        # Per-line tax classification — Tally uses these to fan-out the line
        # tax into the correct CGST/SGST/IGST ledger. Inter-state → IGST;
        # else split 50/50 into CGST + SGST.
        tax_class_xml = ""
        if line_gst_rate > 0 and line_gst_amt > 0:
            if is_inter_state:
                tax_class_xml = f"""
            <RATEDETAILS.LIST>
              <GSTRATEDUTYHEAD>IGST</GSTRATEDUTYHEAD>
              <GSTRATEVALUATIONTYPE>Based on Value</GSTRATEVALUATIONTYPE>
              <GSTRATE>{line_gst_rate:.2f}</GSTRATE>
            </RATEDETAILS.LIST>"""
            else:
                half = line_gst_rate / 2.0
                tax_class_xml = f"""
            <RATEDETAILS.LIST>
              <GSTRATEDUTYHEAD>CGST</GSTRATEDUTYHEAD>
              <GSTRATEVALUATIONTYPE>Based on Value</GSTRATEVALUATIONTYPE>
              <GSTRATE>{half:.2f}</GSTRATE>
            </RATEDETAILS.LIST>
            <RATEDETAILS.LIST>
              <GSTRATEDUTYHEAD>SGST/UTGST</GSTRATEDUTYHEAD>
              <GSTRATEVALUATIONTYPE>Based on Value</GSTRATEVALUATIONTYPE>
              <GSTRATE>{half:.2f}</GSTRATE>
            </RATEDETAILS.LIST>"""
        inv_entries.append(f"""
          <ALLINVENTORYENTRIES.LIST>
            <STOCKITEMNAME>{name}</STOCKITEMNAME>
            <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>
            <ISLASTDEEMEDPOSITIVE>No</ISLASTDEEMEDPOSITIVE>
            <ISAUTONEGATE>No</ISAUTONEGATE>
            <RATE>{rate:.2f}/{uom}</RATE>
            <AMOUNT>{line_total:.2f}</AMOUNT>
            <ACTUALQTY> {qty} {uom}</ACTUALQTY>
            <BILLEDQTY> {qty} {uom}</BILLEDQTY>
            <BATCHALLOCATIONS.LIST>
              <GODOWNNAME>Main Location</GODOWNNAME>
              <BATCHNAME>Primary Batch</BATCHNAME>
              <DESTINATIONGODOWNNAME>Main Location</DESTINATIONGODOWNNAME>
              <INDENTNO/>
              <ORDERNO/>
              <TRACKINGNUMBER/>
              <DYNAMICCSTISCLEARED>No</DYNAMICCSTISCLEARED>
              <AMOUNT>{line_total:.2f}</AMOUNT>
              <ACTUALQTY> {qty} {uom}</ACTUALQTY>
              <BILLEDQTY> {qty} {uom}</BILLEDQTY>
            </BATCHALLOCATIONS.LIST>
            <ACCOUNTINGALLOCATIONS.LIST>
              <LEDGERNAME>Purchase Account</LEDGERNAME>
              <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>
              <LEDGERFROMITEM>No</LEDGERFROMITEM>
              <REMOVEZEROENTRIES>No</REMOVEZEROENTRIES>
              <ISPARTYLEDGER>No</ISPARTYLEDGER>
              <AMOUNT>-{line_total:.2f}</AMOUNT>
            </ACCOUNTINGALLOCATIONS.LIST>{tax_class_xml}
          </ALLINVENTORYENTRIES.LIST>""")
    inventory_block = "".join(inv_entries)

    # Ledger entries: Party (credit) + Tax ledgers (debit) + Additional charges
    # (debit). Purchase Account debits are emitted as inline
    # ACCOUNTINGALLOCATIONS inside each ALLINVENTORYENTRIES — adding them here
    # too would double-count.
    ledger_entries = [f"""
        <LEDGERENTRIES.LIST>
          <LEDGERNAME>{party_name}</LEDGERNAME>
          <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>
          <ISPARTYLEDGER>Yes</ISPARTYLEDGER>
          <AMOUNT>{total_amount:.2f}</AMOUNT>
          <BILLALLOCATIONS.LIST>
            <NAME>{inv_no}</NAME>
            <BILLTYPE>New Ref</BILLTYPE>
            <AMOUNT>{total_amount:.2f}</AMOUNT>
          </BILLALLOCATIONS.LIST>
        </LEDGERENTRIES.LIST>"""]
    # Additional charges — each becomes its own ledger debit. Charge name maps
    # directly to a Tally ledger so the user can create matching ledgers like
    # "Freight Charges", "Packaging Charges" etc. in Tally beforehand.
    for charge in additional_charges:
        ch_name = _xml_escape(str(charge.get("name") or "Other Charges"))
        ch_amount = float(charge.get("amount") or 0)
        ch_gst_rate = float(charge.get("gst_rate") or 0)
        ch_tax = float(charge.get("tax_amount") or 0)
        if ch_amount <= 0 and ch_tax <= 0:
            continue
        ledger_entries.append(f"""
        <LEDGERENTRIES.LIST>
          <LEDGERNAME>{ch_name}</LEDGERNAME>
          <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>
          <AMOUNT>-{ch_amount:.2f}</AMOUNT>
          <GSTRATE>{ch_gst_rate:.2f}</GSTRATE>
        </LEDGERENTRIES.LIST>""")
    if is_inter_state:
        if total_igst > 0:
            ledger_entries.append(f"""
        <LEDGERENTRIES.LIST>
          <LEDGERNAME>IGST Input</LEDGERNAME>
          <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>
          <AMOUNT>-{total_igst:.2f}</AMOUNT>
        </LEDGERENTRIES.LIST>""")
    else:
        if total_cgst > 0:
            ledger_entries.append(f"""
        <LEDGERENTRIES.LIST>
          <LEDGERNAME>CGST Input</LEDGERNAME>
          <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>
          <AMOUNT>-{total_cgst:.2f}</AMOUNT>
        </LEDGERENTRIES.LIST>""")
        if total_sgst > 0:
            ledger_entries.append(f"""
        <LEDGERENTRIES.LIST>
          <LEDGERNAME>SGST Input</LEDGERNAME>
          <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>
          <AMOUNT>-{total_sgst:.2f}</AMOUNT>
        </LEDGERENTRIES.LIST>""")
    ledger_block = "".join(ledger_entries)

    # Build a small charges narration so the operator can sanity-check what
    # was attached when reviewing in Tally.
    charges_narration = ""
    if additional_charges and charges_subtotal > 0:
        items_str = ", ".join(
            f"{_xml_escape(str(c.get('name') or ''))} {float(c.get('amount') or 0):.2f}"
            for c in additional_charges
            if float(c.get("amount") or 0) > 0
        )
        if items_str:
            charges_narration = f" | Add'l: {items_str}"

    return f"""
    <TALLYMESSAGE xmlns:UDF="TallyUDF">
      <VOUCHER VCHTYPE="Purchase" ACTION="Create" OBJVIEW="Invoice Voucher View">
        <DATE>{inv_date}</DATE>
        <NARRATION>{narration}{charges_narration}</NARRATION>
        <VOUCHERTYPENAME>Purchase</VOUCHERTYPENAME>
        <VOUCHERNUMBER>{inv_no}</VOUCHERNUMBER>
        <REFERENCE>{inv_no}</REFERENCE>
        <PARTYLEDGERNAME>{party_name}</PARTYLEDGERNAME>
        <PARTYNAME>{party_name}</PARTYNAME>
        <BASICBUYERNAME>{party_name}</BASICBUYERNAME>
        <PARTYGSTIN>{_xml_escape(supplier.get('gstin') or '')}</PARTYGSTIN>
        <BUYERSGSTIN>{_xml_escape(supplier.get('gstin') or '')}</BUYERSGSTIN>
        <STATENAME>{state_name}</STATENAME>
        <PLACEOFSUPPLY>{state_name}</PLACEOFSUPPLY>
        <CONSIGNEEGSTIN>{_xml_escape(supplier.get('gstin') or '')}</CONSIGNEEGSTIN>
        <BASICBUYERADDRESS.LIST>{"".join(f"<BASICBUYERADDRESS>{_xml_escape(str(supplier.get(k, '') or ''))}</BASICBUYERADDRESS>" for k in ("address", "address_line2", "city") if supplier.get(k))}</BASICBUYERADDRESS.LIST>
        <ISINVOICE>Yes</ISINVOICE>
        <EFFECTIVEDATE>{inv_date}</EFFECTIVEDATE>
        {ledger_block}
        {inventory_block}
      </VOUCHER>
    </TALLYMESSAGE>"""

def _wrap_tally_envelope(messages_xml, report_desc="Vouchers", company=None):
    """Wrap a list of <TALLYMESSAGE> blocks in the standard Tally import envelope.

    When `company` is provided, the seller's company name is embedded in
    `<SVCURRENTCOMPANY>` so the import targets the right Tally company file
    automatically (instead of relying on the default `##SVCurrentCompany`
    placeholder).
    """
    comp_name = (company or {}).get("company_name") or "##SVCurrentCompany"
    comp_name_x = _xml_escape(comp_name)
    return f"""<?xml version="1.0" encoding="UTF-8"?>
<ENVELOPE>
  <HEADER>
    <TALLYREQUEST>Import Data</TALLYREQUEST>
  </HEADER>
  <BODY>
    <IMPORTDATA>
      <REQUESTDESC>
        <REPORTNAME>{report_desc}</REPORTNAME>
        <STATICVARIABLES>
          <SVCURRENTCOMPANY>{comp_name_x}</SVCURRENTCOMPANY>
        </STATICVARIABLES>
      </REQUESTDESC>
      <REQUESTDATA>{messages_xml}
      </REQUESTDATA>
    </IMPORTDATA>
  </BODY>
</ENVELOPE>"""

@purchase_invoices_router.get("/{invoice_id}/tally-xml")
async def export_purchase_invoice_to_tally(invoice_id: str, request: Request):
    """Return Tally-compatible XML for a single purchase invoice as a downloadable file."""
    user = await get_current_user(request)
    invoice = await db.purchase_invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    supplier = await db.suppliers.find_one({"id": invoice.get("supplier_id")}, {"_id": 0}) or {}
    company = await db.company_settings.find_one({"type": "company"}, {"_id": 0}) or {}
    is_inter_state = supplier.get("state_code", "") != company.get("state_code", "")
    # Hydrate line items with their item docs
    lines = invoice.get("lines", []) or []
    for ln in lines:
        ln["item"] = await db.items.find_one({"id": ln.get("item_id")}, {"_id": 0}) or {}
    # Emit master messages first (stock items + ledger + UOM), then the
    # voucher. Without the masters, Tally treats unknown stock items as text
    # and unknown ledgers as plain-amount entries — qty/rate/GSTIN are not
    # rendered. The masters use ACTION="Alter" so re-imports are idempotent.
    additional_charges = invoice.get("additional_charges", []) or []
    masters = _build_tally_master_messages(invoice, supplier, lines, additional_charges)
    msg = _build_tally_purchase_voucher_xml(invoice, supplier, company, lines, is_inter_state)
    xml = _wrap_tally_envelope(masters + msg, company=company)
    fname = f"tally_{invoice.get('invoice_number', invoice_id)}.xml"
    return Response(
        content=xml,
        media_type="application/xml",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'}
    )

@purchase_invoices_router.post("/tally-xml-bulk")
async def export_purchase_invoices_bulk_tally(request: Request, payload: dict = Body(...)):
    """Bulk export — send {"invoice_ids": [...]} to download one XML file with all vouchers."""
    await get_current_user(request)
    ids = payload.get("invoice_ids") or []
    if not ids:
        raise HTTPException(status_code=400, detail="invoice_ids required")
    company = await db.company_settings.find_one({"type": "company"}, {"_id": 0}) or {}
    messages = []
    for inv_id in ids:
        invoice = await db.purchase_invoices.find_one({"id": inv_id}, {"_id": 0})
        if not invoice:
            continue
        supplier = await db.suppliers.find_one({"id": invoice.get("supplier_id")}, {"_id": 0}) or {}
        is_inter_state = supplier.get("state_code", "") != company.get("state_code", "")
        lines = invoice.get("lines", []) or []
        for ln in lines:
            ln["item"] = await db.items.find_one({"id": ln.get("item_id")}, {"_id": 0}) or {}
        additional_charges = invoice.get("additional_charges", []) or []
        messages.append(_build_tally_master_messages(invoice, supplier, lines, additional_charges))
        messages.append(_build_tally_purchase_voucher_xml(invoice, supplier, company, lines, is_inter_state))
    xml = _wrap_tally_envelope("".join(messages), company=company)
    return Response(
        content=xml,
        media_type="application/xml",
        headers={"Content-Disposition": f'attachment; filename="tally_purchase_invoices_bulk.xml"'}
    )


# ===== Tally Sales Voucher (Tax Invoice export) ============================
# Mirrors the Purchase Invoice flow above. Sales vouchers debit the customer
# ledger (positive) and credit Sales + output GST ledgers. Tally accepts the
# same envelope format — only the VCHTYPE, ISDEEMEDPOSITIVE signs and ledger
# names change.

def _build_tally_sales_voucher_xml(invoice, customer, company, lines, is_inter_state):
    """Build a single <TALLYMESSAGE> block for a SALES voucher (Tax Invoice).

    Enriched output (per user request 2026-02-17):
    - Full party address block (BUYERADDRESS.LIST) so the bill-to ledger
      page in Tally shows the customer's billing address.
    - Per-line stock item entries include description + discount %
      (Tally's `DISCOUNT` element on ALLINVENTORYENTRIES.LIST).
    - A 'Discount Allowed' ledger entry is added when the invoice has any
      discount so the gross / net values reconcile cleanly in Tally.
    """
    inv_no = _xml_escape(invoice.get("invoice_no") or "")
    inv_date = _tally_date(invoice.get("invoice_date"))
    party_name = _xml_escape(customer.get("name") or invoice.get("customer_name") or "")
    narration = _xml_escape(invoice.get("notes", "") or f"Sale against invoice {inv_no}")

    # Party address block — Tally expects each line of the address as its
    # own <ADDRESS> child wrapped inside <BUYERADDRESS.LIST>.
    addr_lines = []
    addr_source = invoice.get("billing_address") or customer.get("address") or ""
    for raw in addr_source.split("\n"):
        s = (raw or "").strip()
        if s:
            addr_lines.append(s)
    city_state_pin = ", ".join([x for x in [customer.get("city"), customer.get("state"), customer.get("pin_code")] if x])
    if city_state_pin:
        addr_lines.append(city_state_pin)
    if customer.get("gstin"):
        addr_lines.append(f"GSTIN: {customer.get('gstin')}")
    if customer.get("state_code"):
        addr_lines.append(f"State Code: {customer.get('state_code')}")
    addr_block = "".join(f"<ADDRESS>{_xml_escape(a)}</ADDRESS>" for a in addr_lines)

    subtotal = float(invoice.get("net_subtotal") or invoice.get("subtotal") or 0)
    total_discount = sum(
        (float(ln.get("quantity", 0) or 0) * float(ln.get("unit_price") or ln.get("rate") or 0) * float(ln.get("discount_pct") or ln.get("discount") or 0) / 100.0)
        for ln in lines
    )
    total_cgst = float(invoice.get("cgst", 0) or 0)
    total_sgst = float(invoice.get("sgst", 0) or 0)
    total_igst = float(invoice.get("igst", 0) or 0)
    grand_total = float(invoice.get("grand_total", 0) or 0)

    inv_entries = []
    for ln in lines:
        it = ln.get("item") or {}
        # Use the shared resolver so TI lines (no item_id, description-only)
        # still emit a valid STOCKITEMNAME — falls back to description / HSN.
        name = _xml_escape(_stock_item_name_for_line(ln))
        qty = float(ln.get("quantity", 0) or 0)
        rate = float(ln.get("unit_price") or ln.get("rate") or 0)
        disc_pct = float(ln.get("discount_pct") or ln.get("discount") or 0)
        gross = qty * rate
        net_after_disc = gross * (1 - disc_pct / 100.0)
        uom = _xml_escape(it.get("unit_of_measure") or it.get("uom") or ln.get("uom") or "Nos")
        desc = _xml_escape((ln.get("description") or it.get("description") or "").strip())
        # Per-line CGST/SGST/IGST classification — lets Tally re-derive slabs.
        line_gst_rate = float(ln.get("gst_rate") or 0)
        if is_inter_state and line_gst_rate > 0:
            line_tax_class = f"""
            <RATEDETAILS.LIST>
              <GSTRATEDUTYHEAD>IGST</GSTRATEDUTYHEAD>
              <GSTRATEVALUATIONTYPE>Based on Value</GSTRATEVALUATIONTYPE>
              <GSTRATE>{line_gst_rate:.2f}</GSTRATE>
            </RATEDETAILS.LIST>"""
        elif line_gst_rate > 0:
            half = line_gst_rate / 2.0
            line_tax_class = f"""
            <RATEDETAILS.LIST>
              <GSTRATEDUTYHEAD>CGST</GSTRATEDUTYHEAD>
              <GSTRATEVALUATIONTYPE>Based on Value</GSTRATEVALUATIONTYPE>
              <GSTRATE>{half:.2f}</GSTRATE>
            </RATEDETAILS.LIST>
            <RATEDETAILS.LIST>
              <GSTRATEDUTYHEAD>SGST/UTGST</GSTRATEDUTYHEAD>
              <GSTRATEVALUATIONTYPE>Based on Value</GSTRATEVALUATIONTYPE>
              <GSTRATE>{half:.2f}</GSTRATE>
            </RATEDETAILS.LIST>"""
        else:
            line_tax_class = ""
        inv_entries.append(f"""
          <ALLINVENTORYENTRIES.LIST>
            <STOCKITEMNAME>{name}</STOCKITEMNAME>
            <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>
            <ISLASTDEEMEDPOSITIVE>No</ISLASTDEEMEDPOSITIVE>
            <ISAUTONEGATE>No</ISAUTONEGATE>
            <RATE>{rate:.2f}/{uom}</RATE>
            <AMOUNT>{net_after_disc:.2f}</AMOUNT>
            <ACTUALQTY>{qty} {uom}</ACTUALQTY>
            <BILLEDQTY>{qty} {uom}</BILLEDQTY>
            <DISCOUNT>{disc_pct:.2f}</DISCOUNT>
            {('<DESCRIPTION>' + desc + '</DESCRIPTION>') if desc else ''}
            <BATCHALLOCATIONS.LIST>
              <GODOWNNAME>Main Location</GODOWNNAME>
              <BATCHNAME>Primary Batch</BATCHNAME>
              <DESTINATIONGODOWNNAME>Main Location</DESTINATIONGODOWNNAME>
              <INDENTNO/>
              <ORDERNO/>
              <TRACKINGNUMBER/>
              <DYNAMICCSTISCLEARED>No</DYNAMICCSTISCLEARED>
              <AMOUNT>{net_after_disc:.2f}</AMOUNT>
              <ACTUALQTY>{qty} {uom}</ACTUALQTY>
              <BILLEDQTY>{qty} {uom}</BILLEDQTY>
            </BATCHALLOCATIONS.LIST>
            <ACCOUNTINGALLOCATIONS.LIST>
              <LEDGERNAME>Sales Account</LEDGERNAME>
              <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>
              <LEDGERFROMITEM>No</LEDGERFROMITEM>
              <REMOVEZEROENTRIES>No</REMOVEZEROENTRIES>
              <ISPARTYLEDGER>No</ISPARTYLEDGER>
              <AMOUNT>{net_after_disc:.2f}</AMOUNT>
            </ACCOUNTINGALLOCATIONS.LIST>{line_tax_class}
          </ALLINVENTORYENTRIES.LIST>""")
    inventory_block = "".join(inv_entries)

    ledger_entries = [f"""
        <LEDGERENTRIES.LIST>
          <LEDGERNAME>{party_name}</LEDGERNAME>
          <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>
          <ISPARTYLEDGER>Yes</ISPARTYLEDGER>
          <AMOUNT>-{grand_total:.2f}</AMOUNT>
          <BILLALLOCATIONS.LIST>
            <NAME>{inv_no}</NAME>
            <BILLTYPE>New Ref</BILLTYPE>
            <AMOUNT>-{grand_total:.2f}</AMOUNT>
          </BILLALLOCATIONS.LIST>
        </LEDGERENTRIES.LIST>"""]
    if total_discount > 0:
        ledger_entries.append(f"""
        <LEDGERENTRIES.LIST>
          <LEDGERNAME>Discount Allowed</LEDGERNAME>
          <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>
          <AMOUNT>-{total_discount:.2f}</AMOUNT>
        </LEDGERENTRIES.LIST>""")
    if is_inter_state:
        if total_igst > 0:
            ledger_entries.append(f"""
        <LEDGERENTRIES.LIST>
          <LEDGERNAME>IGST Output</LEDGERNAME>
          <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>
          <AMOUNT>{total_igst:.2f}</AMOUNT>
        </LEDGERENTRIES.LIST>""")
    else:
        if total_cgst > 0:
            ledger_entries.append(f"""
        <LEDGERENTRIES.LIST>
          <LEDGERNAME>CGST Output</LEDGERNAME>
          <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>
          <AMOUNT>{total_cgst:.2f}</AMOUNT>
        </LEDGERENTRIES.LIST>""")
        if total_sgst > 0:
            ledger_entries.append(f"""
        <LEDGERENTRIES.LIST>
          <LEDGERNAME>SGST Output</LEDGERNAME>
          <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>
          <AMOUNT>{total_sgst:.2f}</AMOUNT>
        </LEDGERENTRIES.LIST>""")
    ledger_block = "".join(ledger_entries)

    gstin = _xml_escape(customer.get("gstin") or "")
    state_code = _xml_escape(customer.get("state_code") or "")
    buyer_state_name = _xml_escape(_resolve_state_name(customer))

    # ----- Seller (company) details ----------------------------------------
    # Tally usually picks the seller from the loaded company file, but we
    # embed the seller block so imports into a fresh / wrong-company target
    # still carry the correct GSTIN, address, and state code on the voucher.
    seller_name = _xml_escape(company.get("company_name") or "")
    seller_addr_src = (
        company.get("address") or company.get("address_line1") or ""
    )
    seller_addr_lines = [s.strip() for s in seller_addr_src.split("\n") if s.strip()]
    seller_city_state = ", ".join([x for x in [company.get("city"), company.get("state"), company.get("pin_code")] if x])
    if seller_city_state:
        seller_addr_lines.append(seller_city_state)
    if company.get("gstin"):
        seller_addr_lines.append(f"GSTIN: {company.get('gstin')}")
    if company.get("state_code"):
        seller_addr_lines.append(f"State Code: {company.get('state_code')}")
    seller_addr_block = "".join(f"<ADDRESS>{_xml_escape(a)}</ADDRESS>" for a in seller_addr_lines)
    seller_gstin = _xml_escape(company.get("gstin") or "")
    seller_state_code = _xml_escape(company.get("state_code") or "")

    return f"""
    <TALLYMESSAGE xmlns:UDF="TallyUDF">
      <VOUCHER VCHTYPE="Sales" ACTION="Create" OBJVIEW="Invoice Voucher View">
        <DATE>{inv_date}</DATE>
        <NARRATION>{narration}</NARRATION>
        <VOUCHERTYPENAME>Sales</VOUCHERTYPENAME>
        <VOUCHERNUMBER>{inv_no}</VOUCHERNUMBER>
        <REFERENCE>{inv_no}</REFERENCE>
        <PARTYLEDGERNAME>{party_name}</PARTYLEDGERNAME>
        <PARTYNAME>{party_name}</PARTYNAME>
        <BASICBUYERNAME>{party_name}</BASICBUYERNAME>
        <BASICBUYERADDRESS.LIST>{addr_block}</BASICBUYERADDRESS.LIST>
        <BUYERADDRESS.LIST>{addr_block}</BUYERADDRESS.LIST>
        {('<PARTYGSTIN>' + gstin + '</PARTYGSTIN>') if gstin else ''}
        {('<STATENAME>' + buyer_state_name + '</STATENAME>') if buyer_state_name else ''}
        {('<PLACEOFSUPPLY>' + (_xml_escape(invoice.get('place_of_supply') or '') or buyer_state_name) + '</PLACEOFSUPPLY>') if (invoice.get('place_of_supply') or buyer_state_name) else ''}
        <!-- Seller / Company details — embedded so imports carry GSTIN + address -->
        {('<BASICCOMPANYNAME>' + seller_name + '</BASICCOMPANYNAME>') if seller_name else ''}
        {('<BASICCOMPANYFORMALNAME>' + seller_name + '</BASICCOMPANYFORMALNAME>') if seller_name else ''}
        <BASICCOMPANYADDRESS.LIST>{seller_addr_block}</BASICCOMPANYADDRESS.LIST>
        <COMPANYADDRESS.LIST>{seller_addr_block}</COMPANYADDRESS.LIST>
        {('<COMPANYGSTIN>' + seller_gstin + '</COMPANYGSTIN>') if seller_gstin else ''}
        {('<COMPANYSTATENAME>' + seller_state_code + '</COMPANYSTATENAME>') if seller_state_code else ''}
        <ISINVOICE>Yes</ISINVOICE>
        <EFFECTIVEDATE>{inv_date}</EFFECTIVEDATE>
        {ledger_block}
        {inventory_block}
      </VOUCHER>
    </TALLYMESSAGE>"""


async def _hydrate_ti_lines_for_tally(invoice):
    """Attach item docs to each line — Tally needs part_number/name/UOM."""
    for ln in (invoice.get("lines") or []):
        if ln.get("item_id") and "item" not in ln:
            ln["item"] = await db.items.find_one({"id": ln.get("item_id")}, {"_id": 0}) or {}
    return invoice


@crm_router.get("/tax-invoices/{tid}/tally-xml")
async def export_tax_invoice_to_tally(tid: str, request: Request):
    """Single Tax Invoice → Tally Sales voucher XML download."""
    await get_current_user(request)
    invoice = await db.tax_invoices.find_one({"id": tid}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Tax invoice not found")
    customer = await db.customers.find_one({"id": invoice.get("customer_id")}, {"_id": 0}) or {}
    company = await db.company_settings.find_one({"type": "company"}, {"_id": 0}) or {}
    is_inter_state = bool(invoice.get("is_inter_state"))
    await _hydrate_ti_lines_for_tally(invoice)
    # If the customer record was deleted / lookup misses, synthesize a
    # party-like dict from the invoice fields so the Tally Sundry Debtors
    # ledger still carries the snapshot the user printed on the invoice.
    party_for_master = customer if (customer and customer.get("name")) else {
        "name": invoice.get("customer_name") or "",
        "gstin": invoice.get("customer_gstin") or "",
        "address": invoice.get("billing_address") or "",
        "state": invoice.get("customer_state") or invoice.get("place_of_supply") or "",
        "state_code": invoice.get("customer_state_code") or "",
        "country": "India",
    }
    # Emit customer ledger master + stock-item / UOM masters BEFORE the
    # voucher so the Sundry Debtors ledger shows GSTIN/Address AND the
    # voucher view actually renders quantity + rate per line (without
    # masters, Tally treats STOCKITEMNAME as plain text → qty/rate blank).
    inv_date_xml = _tally_date(invoice.get("invoice_date")) or "20170701"
    cust_master = _build_tally_party_ledger_xml(party_for_master, "Sundry Debtors", inv_date_xml)
    inv_masters = _build_tally_inventory_masters(invoice, invoice.get("lines", []), invoice.get("additional_charges", []))
    msg = _build_tally_sales_voucher_xml(invoice, party_for_master, company, invoice.get("lines", []), is_inter_state)
    xml = _wrap_tally_envelope(cust_master + inv_masters + msg, company=company)
    fname = f"tally_{invoice.get('invoice_no', tid)}.xml"
    return Response(
        content=xml,
        media_type="application/xml",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@crm_router.post("/tax-invoices/tally-xml-bulk")
async def export_tax_invoices_bulk_tally(request: Request, payload: dict = Body(...)):
    """Bulk Tax Invoices → single Tally XML. POST {"invoice_ids": [...]}"""
    await get_current_user(request)
    ids = payload.get("invoice_ids") or []
    if not ids:
        raise HTTPException(status_code=400, detail="invoice_ids required")
    company = await db.company_settings.find_one({"type": "company"}, {"_id": 0}) or {}
    messages = []
    seen_customer_masters = set()
    for tid in ids:
        invoice = await db.tax_invoices.find_one({"id": tid}, {"_id": 0})
        if not invoice:
            continue
        customer = await db.customers.find_one({"id": invoice.get("customer_id")}, {"_id": 0}) or {}
        is_inter_state = bool(invoice.get("is_inter_state"))
        await _hydrate_ti_lines_for_tally(invoice)
        # Emit customer ledger master once per unique customer in the batch.
        cust_id = customer.get("id") or customer.get("name") or ""
        if cust_id and cust_id not in seen_customer_masters:
            seen_customer_masters.add(cust_id)
            inv_date_xml = _tally_date(invoice.get("invoice_date")) or "20170701"
            messages.append(_build_tally_party_ledger_xml(customer, "Sundry Debtors", inv_date_xml))
        # Stock-item / UOM masters per invoice — Tally needs these so the
        # voucher view renders BILLEDQTY + RATE for each line.
        messages.append(_build_tally_inventory_masters(invoice, invoice.get("lines", []), invoice.get("additional_charges", [])))
        messages.append(_build_tally_sales_voucher_xml(invoice, customer, company, invoice.get("lines", []), is_inter_state))
    xml = _wrap_tally_envelope("".join(messages), company=company)
    return Response(
        content=xml,
        media_type="application/xml",
        headers={"Content-Disposition": f'attachment; filename="tally_tax_invoices_bulk.xml"'},
    )


# ================== JOB WORK / SUBCONTRACTING ROUTES ==================

@jobwork_router.get("/orders/{sc_id}/dc-lines")
async def get_dc_lines_for_job_os(sc_id: str, request: Request):
    """Expand a Job Card OS SC's job_work_parts for the DC table. Returns one row per Part
    with: qty, charges_per_unit (processing), rm_cost_per_unit (BOM rollup), item_description.

    Charges_per_unit resolution:
      - When the part has a `process_name` (specific outsourced routing op),
        always recompute from `find_routing_cost(item_id, process_name)` so
        stale stored values (legacy data polluted by older save flows) don't
        leak through. This guarantees the DC dialog ALWAYS shows the specific
        op's per-unit cost rather than the combined process cost.
      - Otherwise, fall back to the stored `charges` value.
    """
    await get_current_user(request)
    sc = await db.subcontract_orders.find_one({"id": sc_id}, {"_id": 0})
    if not sc:
        raise HTTPException(status_code=404, detail="SC not found")
    out_lines = []
    for jp in sc.get("job_work_parts", []):
        part_item = await db.items.find_one({"id": jp.get("item_id")}, {"_id": 0}) or {}
        qty = float(jp.get("quantity", 0) or 0)
        process_name = (jp.get("process_name") or "").strip()
        # Stored charges first
        charges_per_unit = float(jp.get("charges", 0) or 0)
        # Override with specific routing cost when process_name is set.
        if process_name:
            specific = await find_routing_cost(jp.get("item_id"), process_name)
            if specific:
                charges_per_unit = float(specific)
        rm_cost_per_unit = float(jp.get("bom_rollup_cost", 0) or 0)
        out_lines.append({
            "type": "part",
            "item_id": jp.get("item_id"),
            "item": part_item,
            "quantity": qty,
            "charges_per_unit": charges_per_unit,
            "rm_cost_per_unit": rm_cost_per_unit,
            "total_charges": round(qty * charges_per_unit, 2),
            "total_amount": round(qty * rm_cost_per_unit, 2),
            "item_description": jp.get("item_description") or "",
            "process_name": process_name,
        })
    return out_lines


@jobwork_router.get("/orders")
async def get_subcontract_orders(request: Request, status: str = None):
    user = await get_current_user(request)
    query = {}
    if status:
        query["status"] = status
    orders = await db.subcontract_orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Batch collect all referenced IDs
    supplier_ids = set()
    wo_ids = set()
    item_ids = set()
    for order in orders:
        if order.get("supplier_id"): supplier_ids.add(order["supplier_id"])
        if order.get("reference_wo_ids"):
            wo_ids.update(order["reference_wo_ids"])
        elif order.get("reference_wo_id"):
            wo_ids.add(order["reference_wo_id"])
        for line in order.get("lines", []):
            if line.get("item_id"): item_ids.add(line["item_id"])
        for part in order.get("job_work_parts", []):
            if part.get("item_id"): item_ids.add(part["item_id"])
    
    # Batch fetch (3 queries instead of N×4 per order)
    suppliers_map = {}
    if supplier_ids:
        async for s in db.suppliers.find({"id": {"$in": list(supplier_ids)}}, {"_id": 0}):
            suppliers_map[s["id"]] = s
    wos_map = {}
    if wo_ids:
        async for w in db.work_orders.find({"id": {"$in": list(wo_ids)}}, {"_id": 0, "id": 1, "wo_number": 1}):
            wos_map[w["id"]] = w.get("wo_number")
    items_map = {}
    if item_ids:
        async for it in db.items.find({"id": {"$in": list(item_ids)}}, {"_id": 0}):
            items_map[it["id"]] = it
    
    # BOM cost cache — avoid recomputing for the same item across multiple SCs
    bom_cost_cache = {}
    bom_total_cache = {}
    
    for order in orders:
        order["supplier"] = suppliers_map.get(order.get("supplier_id"))
        # Include linked MO numbers (single or bulk)
        mo_numbers = []
        if order.get("reference_wo_ids"):
            mo_numbers = [wos_map[w] for w in order["reference_wo_ids"] if w in wos_map]
        elif order.get("reference_wo_id") and order["reference_wo_id"] in wos_map:
            mo_numbers = [wos_map[order["reference_wo_id"]]]
        order["mo_number"] = ", ".join(mo_numbers) if mo_numbers else None
        order["mo_numbers"] = mo_numbers
        
        is_live = order.get("status") in ("draft", "in_progress")
        for line in order.get("lines", []):
            item = items_map.get(line.get("item_id"))
            line["item"] = item
            # Auto-refresh rate from BOM Total/Unit for Part/SA lines on draft/in_progress SCs
            if is_live and item and item.get("category") in ("component", "sub_assembly"):
                iid = line["item_id"]
                if iid not in bom_total_cache:
                    try:
                        bom_total_cache[iid] = await compute_bom_total_unit_cost(iid)
                    except Exception:
                        bom_total_cache[iid] = 0
                _tu = bom_total_cache[iid]
                if _tu > 0:
                    line["rate"] = round(_tu, 2)
        for part in order.get("job_work_parts", []):
            part["item"] = items_map.get(part.get("item_id"))
            # Self-heal `charges` for Job Card OS lines (process_name set) on
            # EVERY response — not gated on is_live — because legacy data
            # polluted with the combined process cost needs to be corrected
            # for display regardless of the SC's current status. This is a
            # display-only override; the DB row is untouched.
            process_name = (part.get("process_name") or "").strip()
            if process_name and part.get("item_id"):
                try:
                    _specific = await find_routing_cost(part.get("item_id"), process_name)
                except Exception:
                    _specific = 0.0
                if _specific:
                    part["charges"] = round(float(_specific), 2)
            # Auto-refresh combined cost for Full MO-SC lines (no process_name)
            # on LIVE SCs only — preserves audit snapshot on completed SCs.
            if is_live and part.get("item_id"):
                iid = part["item_id"]
                if iid not in bom_cost_cache:
                    try:
                        bom_cost_cache[iid] = await compute_bom_costs(iid)
                    except Exception:
                        bom_cost_cache[iid] = None
                _bc = bom_cost_cache[iid]
                if _bc:
                    _fg = round(_bc.get("fg_process_cost", 0) or 0, 2)
                    _total_unit = round((_bc.get("rm_cost", 0) or 0) + (_bc.get("process_cost", 0) or 0), 2)
                    if _fg and not process_name:
                        part["charges"] = _fg
                    if _total_unit:
                        part["bom_rollup_cost"] = _total_unit
                    _names = _bc.get("process_names") or []
                    if _names:
                        part["process_names"] = _names
        # Enrich Job Card OS SCs with the list of Parts (from job_work_parts) — the Part
        # IS what physically goes to the vendor for the outsource operation. Cost = BOM
        # rollup (RM cost per Part). This populates the "RM" column on the SC list with
        # the Part itself, not the underlying raw materials.
        if order.get("subcontract_type") == "without_material" and (order.get("reference_operation_seqs") or order.get("reference_operation_seq")):
            rm_agg = []
            for jp in order.get("job_work_parts", []):
                part_qty = float(jp.get("quantity", 0) or 0)
                part_id = jp.get("item_id")
                part_item = items_map.get(part_id) or (await db.items.find_one({"id": part_id}, {"_id": 0}))
                rm_rate = float(jp.get("bom_rollup_cost", 0) or 0)
                rm_agg.append({
                    "item_id": part_id,
                    "item": part_item,
                    "quantity": part_qty,
                    "rate": rm_rate,
                })
            order["rm_items"] = rm_agg
    return orders

@jobwork_router.post("/orders", status_code=201)
async def create_subcontract_order(data: SubcontractOrderCreate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="job_work", action="create")
    supplier = await db.suppliers.find_one({"id": data.supplier_id})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    count = await db.subcontract_orders.count_documents({})
    order_number = f"JW-{str(count + 1).zfill(6)}"
    
    lines = []
    for line in data.lines:
        item = await db.items.find_one({"id": line.item_id}, {"_id": 0})
        if not item:
            continue
        lines.append({
            "item_id": line.item_id,
            "quantity": line.quantity,
            "sent_quantity": 0,
            "received_quantity": 0,
            "rate": line.rate or 0
        })
    
    # Enrich job_work_parts with BOM-based defaults (process cost, RM rollup, process names)
    enriched_parts = []
    for p in (data.job_work_parts or []):
        bom_costs = await compute_bom_costs(p.item_id)
        # When a specific outsource routing is named (Job Card OS), use its
        # per-op cost. Otherwise default to the combined BOM process cost
        # (Full MO-SC without RM scenario).
        if p.charges:
            charges = p.charges
        elif p.process_name:
            charges = await find_routing_cost(p.item_id, p.process_name) or bom_costs["process_cost"]
        else:
            charges = bom_costs["process_cost"]
        enriched_parts.append({
            "item_id": p.item_id,
            "quantity": p.quantity,
            "charges": charges,
            "received_quantity": 0,
            "bom_rollup_cost": bom_costs["rm_cost"],
            "process_names": bom_costs.get("process_names", []),
            "process_name": p.process_name or "",
            "item_description": p.item_description or "",
        })
    
    order_doc = {
        "id": str(uuid.uuid4()),
        "order_number": order_number,
        "supplier_id": data.supplier_id,
        "lines": lines,
        "job_work_parts": enriched_parts,
        "expected_return_date": data.expected_return_date,
        "processing_charges": data.processing_charges or 0,
        "status": "draft",
        "notes": data.notes or "",
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"]
    }
    
    await db.subcontract_orders.insert_one(order_doc)
    del order_doc["_id"]
    return order_doc

@jobwork_router.put("/orders/{order_id}")
async def update_subcontract_order(order_id: str, data: SubcontractOrderUpdate, request: Request):
    user = await get_current_user(request)
    # Accept either edit OR create permission on job_work (a user who can create
    # an SC should logically also be able to amend it — matches the frontend's
    # canEdit fallback which includes canCreate). inventory_manager added so
    # role-group users with only that role can amend descriptions/quantities.
    if user.get("role") not in ("admin", "production_manager", "inventory_manager"):
        perms = (user.get("permissions") or {}).get("job_work") or []
        if not ({"edit", "create"} & set(perms)):
            raise HTTPException(status_code=403, detail="Insufficient permissions")
    order = await db.subcontract_orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    update_data = {}
    if data.expected_return_date is not None:
        update_data["expected_return_date"] = data.expected_return_date
    if data.processing_charges is not None:
        update_data["processing_charges"] = data.processing_charges
    if data.notes is not None:
        update_data["notes"] = data.notes
    if data.status is not None:
        update_data["status"] = data.status
    # ------------------------------------------------------------------
    # Allow vendor change ONLY while the SC has not yet shipped any
    # material (no DC sent). Once a DC exists, switching supplier mid-
    # flight would orphan the material at the original vendor. The
    # supplier swap also propagates to all linked MO operations so the
    # operation card shows the correct vendor name.
    # ------------------------------------------------------------------
    if data.supplier_id is not None and data.supplier_id != order.get("supplier_id"):
        if order.get("dc_created"):
            raise HTTPException(status_code=400, detail="Cannot change vendor — a Delivery Challan has already been sent for this SC. Short Close the SC and create a fresh one for the new vendor.")
        new_supplier = await db.suppliers.find_one({"id": data.supplier_id}, {"_id": 0})
        if not new_supplier:
            raise HTTPException(status_code=404, detail="Selected supplier not found")
        update_data["supplier_id"] = data.supplier_id
        # Cascade vendor change to linked MO operations
        new_name = new_supplier.get("name", "Outsourced")
        ref_wo_ids = order.get("reference_wo_ids") or ([order.get("reference_wo_id")] if order.get("reference_wo_id") else [])
        for _wid in ref_wo_ids:
            if not _wid:
                continue
            _wo = await db.work_orders.find_one({"id": _wid})
            if not _wo:
                continue
            _ops = _wo.get("operations_status") or []
            _changed = False
            for _op in _ops:
                if _op.get("outsource_sc_order_id") == order_id:
                    _op["job_work_supplier_id"] = data.supplier_id
                    _op["outsource_supplier_name"] = new_name
                    _op["operator"] = f"OS: {new_name}"
                    _changed = True
                    for _r in (_op.get("runs") or []):
                        if (_r.get("operator") or "").startswith("OS: "):
                            _r["operator"] = f"OS: {new_name}"
            if _changed:
                await db.work_orders.update_one({"id": _wid}, {"$set": {"operations_status": _ops}})
    if data.job_work_parts is not None:
        # Preserve existing per-line context (process_name, wo_id, etc.) by
        # matching incoming entries to the original job_work_parts. Matching
        # priority: (item_id, process_name) > (item_id alone if process_name
        # not provided by the client). This is critical for Job Card OS SCs
        # where each line carries a SPECIFIC routing op and its per-op cost
        # — without this preservation, a save-without-change wipes the
        # per-op `charges` and replaces it with the combined process cost.
        existing_jwp = list(order.get("job_work_parts", []) or [])
        # is_job_card_os: SC was auto-created from a Job Card outsource action
        # — has reference_operation_seqs. In that case, each line's charges
        # must be the SPECIFIC routing's cost (not the combined total).
        is_job_card_os = bool(order.get("reference_operation_seqs") or order.get("reference_operation_seq"))

        def _find_existing(item_id_in: str, process_name_in: Optional[str]):
            # Prefer exact (item_id, process_name) match — required when the
            # same item appears in multiple ops on a consolidated Job Card OS.
            if process_name_in:
                for ex in existing_jwp:
                    if ex.get("item_id") == item_id_in and (ex.get("process_name") or "") == process_name_in:
                        return ex
            for ex in existing_jwp:
                if ex.get("item_id") == item_id_in:
                    return ex
            return None

        enriched_jwp = []
        for p in data.job_work_parts:
            ex = _find_existing(p.item_id, p.process_name)
            # Determine process_name: prefer client value, else existing, else
            # blank.
            process_name = p.process_name or (ex.get("process_name") if ex else "") or ""

            bom_costs = await compute_bom_costs(p.item_id)
            # Charges resolution priority:
            #   1. Client-supplied charges (user edited the row)
            #   2. Existing per-op charges if available (preserve auto-created
            #      Job Card OS cost)
            #   3. Specific routing cost via find_routing_cost (Job Card OS or
            #      whenever a process_name is known)
            #   4. Combined BOM process_cost (Full MO-SC without per-op
            #      context)
            if p.charges:
                charges = p.charges
            elif ex and ex.get("charges"):
                charges = ex.get("charges")
            elif process_name:
                charges = await find_routing_cost(p.item_id, process_name)
                if not charges and not is_job_card_os:
                    charges = bom_costs["process_cost"]
            else:
                charges = bom_costs["process_cost"] if not is_job_card_os else 0

            enriched_jwp.append({
                "item_id": p.item_id,
                "quantity": p.quantity,
                "charges": charges,
                "received_quantity": (ex.get("received_quantity", 0) if ex else 0),
                "bom_rollup_cost": bom_costs["rm_cost"],
                "process_names": bom_costs.get("process_names", []),
                "process_name": process_name,
                "item_description": (p.item_description if p.item_description is not None else (ex.get("item_description") if ex else "")) or "",
                # Preserve any auto-creation references
                "wo_id": (ex.get("wo_id") if ex else None),
            })
        update_data["job_work_parts"] = enriched_jwp

        # ------------------------------------------------------------------
        # Fix 4 — Block qty INCREASE: a JW SC line's quantity can only be
        # reduced or kept the same. Increasing would require a new outsource
        # event on the source MO (so material/process flows stay auditable).
        # New lines (no matching wo+process pair in existing_jwp) ARE allowed
        # so a manual SC can grow extra rows.
        # ------------------------------------------------------------------
        def _key(p):
            return (p.get("wo_id") or "", (p.get("process_name") or "").strip())
        existing_qty_by_key = {}
        for ex in existing_jwp:
            k = _key(ex)
            existing_qty_by_key[k] = existing_qty_by_key.get(k, 0) + (ex.get("quantity") or 0)
        new_qty_by_key = {}
        for p in enriched_jwp:
            k = _key(p)
            new_qty_by_key[k] = new_qty_by_key.get(k, 0) + (p.get("quantity") or 0)
        for k, new_qty in new_qty_by_key.items():
            old_qty = existing_qty_by_key.get(k, 0)
            if old_qty > 0 and new_qty > old_qty:
                raise HTTPException(
                    status_code=400,
                    detail=f"Cannot increase quantity on SC line (process '{k[1] or 'N/A'}'). Existing: {old_qty}, requested: {new_qty}. To outsource MORE, start a fresh outsource on the source MO.",
                )

        # ------------------------------------------------------------------
        # Auto-restore source MO operations when a JW line is REMOVED or qty
        # is REDUCED — so the user can re-outsource to a different vendor
        # without manually editing the MO. Compare existing vs new
        # job_work_parts keyed by (wo_id, process_name). Skip lines that
        # already have received_quantity > 0 (those are mid-flight — admin
        # must use the per-operation Short Close button instead).
        # ------------------------------------------------------------------
        for old in existing_jwp:
            k = _key(old)
            if not k[0]:
                continue
            if (old.get("received_quantity") or 0) > 0:
                continue
            old_qty = old.get("quantity") or 0
            new_qty = new_qty_by_key.get(k, 0)
            qty_returned = old_qty - new_qty
            if qty_returned <= 0:
                continue
            wo = await db.work_orders.find_one({"id": k[0]})
            if not wo:
                continue
            ops = wo.get("operations_status") or []
            updated = False
            for op in ops:
                op_name = (op.get("operation_name") or "")
                if isinstance(op_name, dict):
                    op_name = op_name.get("name", "")
                if op.get("outsource_sc_order_id") == order_id and op_name == k[1]:
                    if new_qty <= 0:
                        for f in (
                            "is_job_work", "job_work_supplier_id", "outsource_status",
                            "outsource_supplier_name", "outsource_charges",
                            "outsource_sc_order_id", "outsource_sc_order_number",
                            "actual_start", "operator",
                        ):
                            op.pop(f, None)
                        op["status"] = "pending"
                        op["outsourced_quantity"] = 0
                        runs = op.get("runs") or []
                        op["runs"] = [r for r in runs if not (r.get("operator") or "").startswith("OS: ")]
                    else:
                        # Partial reduction: reduce the OS run's planned qty AND
                        # the op's tracking counter so the freed-up qty becomes
                        # available for an in-house Start or a new outsource.
                        # If the OS run's planned qty drops to 0, remove it
                        # entirely; if no OS runs remain afterward, also clear
                        # the OS metadata + reset status so the op behaves like
                        # a fresh pending one.
                        runs = op.get("runs") or []
                        new_runs = []
                        remaining_to_return = qty_returned
                        for r in runs:
                            if (r.get("operator") or "").startswith("OS: ") and remaining_to_return > 0:
                                planned = r.get("quantity_planned") or 0
                                if planned <= remaining_to_return:
                                    remaining_to_return -= planned
                                    # drop the run entirely
                                    continue
                                r["quantity_planned"] = max(0, planned - remaining_to_return)
                                remaining_to_return = 0
                            new_runs.append(r)
                        op["runs"] = new_runs
                        op["outsourced_quantity"] = max(0, (op.get("outsourced_quantity") or 0) - qty_returned)
                        if not any((r.get("operator") or "").startswith("OS: ") for r in new_runs):
                            for f in (
                                "is_job_work", "job_work_supplier_id", "outsource_status",
                                "outsource_supplier_name", "outsource_charges",
                                "outsource_sc_order_id", "outsource_sc_order_number",
                                "actual_start", "operator",
                            ):
                                op.pop(f, None)
                            op["status"] = "pending"
                    updated = True
                    break
            if updated:
                await db.work_orders.update_one({"id": k[0]}, {"$set": {"operations_status": ops}})

        remaining_wo_ids = {p.get("wo_id") for p in enriched_jwp if p.get("wo_id")}
        if order.get("reference_wo_ids"):
            update_data["reference_wo_ids"] = [w for w in order["reference_wo_ids"] if w in remaining_wo_ids]
        
        # Only auto-recalculate RM lines if this SC was NOT created via create-sc (smart resolution)
        # SC orders with reference_wo_ids have already been smart-resolved — preserve their lines
        if not order.get("reference_wo_ids") and not order.get("reference_wo_id"):
            new_rm_lines = {}
            for part in data.job_work_parts:
                part_bom = await db.boms.find_one({"parent_item_id": part.item_id, "status": "active"}, {"_id": 0})
                if part_bom:
                    for comp in part_bom.get("components", []):
                        if comp.get("is_alternate"):
                            continue
                        comp_item = await db.items.find_one({"id": comp.get("item_id")}, {"_id": 0})
                        if comp_item and comp_item.get("category") in ["raw_material", "component", "sub_assembly"]:
                            cid = comp["item_id"]
                            qty = int(comp.get("quantity", 1) * part.quantity)
                            rate = comp_item.get("unit_cost", 0)
                            if cid in new_rm_lines:
                                new_rm_lines[cid]["quantity"] += qty
                            else:
                                new_rm_lines[cid] = {"item_id": cid, "quantity": qty, "sent_quantity": 0, "received_quantity": 0, "rate": rate}
            update_data["lines"] = list(new_rm_lines.values())
        # For create-sc orders: preserve existing smart-resolved lines (don't recalculate)
    elif data.lines is not None:
        update_data["lines"] = [{"item_id": l.item_id, "quantity": l.quantity, "sent_quantity": 0, "received_quantity": 0, "rate": l.rate or 0} for l in data.lines]
    
    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc)
        await db.subcontract_orders.update_one({"id": order_id}, {"$set": update_data})
    
    # Sync changes to linked draft PO if job_work_parts changed
    if data.job_work_parts is not None and order.get("po_created"):
        draft_po = await db.purchase_orders.find_one({
            "$or": [{"reference_sc_order_id": order_id}, {"reference_sc_order_ids": order_id}],
            "status": "draft"
        })
        if draft_po:
            po_lines = []
            total = 0
            for p in data.job_work_parts:
                item = await db.items.find_one({"id": p.item_id}, {"_id": 0})
                if item:
                    unit_cost = p.charges or item.get("unit_cost", 0)
                    line_total = p.quantity * unit_cost
                    total += line_total
                    po_lines.append({"item_id": p.item_id, "quantity": p.quantity, "unit_price": unit_cost, "total_price": line_total, "received_quantity": 0, "description": f"{item.get('part_number','')} - {item.get('name','')}", "uom": item.get("unit_of_measure","pcs"), "hsn_code": item.get("hsn_code",""), "gst_rate": item.get("gst_rate",18), "discount_type": "percentage", "discount_value": 0, "notes": ""})
            await db.purchase_orders.update_one({"id": draft_po["id"]}, {"$set": {"lines": po_lines, "subtotal": total, "total_amount": total, "updated_at": datetime.now(timezone.utc)}})
    
    return await db.subcontract_orders.find_one({"id": order_id}, {"_id": 0})

@jobwork_router.post("/orders/{order_id}/confirm")
async def confirm_subcontract_order(order_id: str, request: Request):
    user = await get_current_user(request)
    order = await db.subcontract_orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.get("status") != "draft":
        raise HTTPException(status_code=400, detail="Only draft orders can be confirmed")
    await db.subcontract_orders.update_one({"id": order_id}, {"$set": {"status": "confirmed", "confirmed_at": datetime.now(timezone.utc)}})
    return await db.subcontract_orders.find_one({"id": order_id}, {"_id": 0})


@jobwork_router.post("/orders/{order_id}/short-close")
async def short_close_subcontract_order(order_id: str, request: Request):
    """Admin-only: short-close an in-progress JW SC order.

    Use when a subcontract operation cannot be completed as planned and the
    user wants to release the source MO operations so the work can resume
    in-house or be re-outsourced. This is intentionally a manual escalation
    button (sits next to the row actions for admins only) — not an automatic
    side-effect of editing the SC (which is non-destructive per user request).

    Behaviour:
      - SC `status` flipped to `short_closed` (a terminal state).
      - For each linked WO operation that was outsourced to this SC and has
        NOT been GRN'd (received_quantity == 0 on the corresponding line):
          * outsource_* and is_job_work / operator fields cleared.
          * status reverted to `pending`.
          * The OS run row (operator starts with "OS: ") is removed.
      - Operations whose SC line has any received_quantity > 0 are LEFT
        ALONE — they represent partially-delivered work and must be
        reconciled via the GRN reversal flow.
      - Linked draft PO (if any) is also marked cancelled so it doesn't
        accidentally get confirmed afterwards.
    """
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only admin can short-close a subcontract order")
    order = await db.subcontract_orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.get("status") in ("short_closed", "completed", "cancelled"):
        raise HTTPException(status_code=400, detail=f"Cannot short-close — status is already {order.get('status')}")

    # Build a {wo_id -> {process_name -> received_qty}} lookup so we can
    # skip operations that already have GRN'd output against them.
    jwp = order.get("job_work_parts") or []
    grn_lookup: dict = {}
    for p in jwp:
        wid = p.get("wo_id")
        if not wid:
            continue
        pname = (p.get("process_name") or "").strip()
        rq = p.get("received_quantity") or 0
        grn_lookup.setdefault(wid, {})[pname] = grn_lookup.get(wid, {}).get(pname, 0) + rq

    # Iterate over every WO this SC references (handles both
    # reference_wo_id [legacy single] and reference_wo_ids [consolidated]).
    wo_ids = list(order.get("reference_wo_ids") or [])
    legacy = order.get("reference_wo_id")
    if legacy and legacy not in wo_ids:
        wo_ids.append(legacy)
    released = []  # [(wo_number, operation_name)]
    for wid in wo_ids:
        wo = await db.work_orders.find_one({"id": wid})
        if not wo:
            continue
        ops = wo.get("operations_status") or []
        any_updated = False
        for op in ops:
            op_name = op.get("operation_name") or ""
            if isinstance(op_name, dict):
                op_name = op_name.get("name", "")
            if op.get("outsource_sc_order_id") != order_id:
                continue
            received = grn_lookup.get(wid, {}).get(op_name, 0)
            if received > 0:
                # Partially / fully GRN'd — must be reconciled via GRN
                # reversal flow, not by short-close.
                continue
            for f in (
                "is_job_work", "job_work_supplier_id", "outsource_status",
                "outsource_supplier_name", "outsource_charges",
                "outsource_sc_order_id", "outsource_sc_order_number",
                "actual_start", "operator",
            ):
                op.pop(f, None)
            op["status"] = "pending"
            runs = op.get("runs") or []
            op["runs"] = [r for r in runs if not (r.get("operator") or "").startswith("OS: ")]
            any_updated = True
            released.append((wo.get("wo_number"), op_name))
        if any_updated:
            await db.work_orders.update_one({"id": wid}, {"$set": {"operations_status": ops}})

    # Cancel any linked draft PO that hasn't been confirmed yet.
    cancelled_pos = []
    async for draft_po in db.purchase_orders.find(
        {"$or": [{"reference_sc_order_id": order_id}, {"reference_sc_order_ids": order_id}], "status": "draft"},
        {"_id": 0, "id": 1, "po_number": 1},
    ):
        await db.purchase_orders.update_one(
            {"id": draft_po["id"]},
            {"$set": {"status": "cancelled", "cancelled_reason": f"SC {order.get('order_number')} short-closed by admin", "cancelled_at": datetime.now(timezone.utc), "cancelled_by": user["id"]}},
        )
        cancelled_pos.append(draft_po.get("po_number"))

    await db.subcontract_orders.update_one(
        {"id": order_id},
        {"$set": {
            "status": "short_closed",
            "short_closed_at": datetime.now(timezone.utc),
            "short_closed_by": user["id"],
            "updated_at": datetime.now(timezone.utc),
        }},
    )
    return {
        "ok": True,
        "released_operations": [{"wo_number": w, "operation_name": o} for w, o in released],
        "cancelled_pos": cancelled_pos,
    }


@jobwork_router.post("/create-po")
async def create_po_from_sc(request: Request, data: dict = Body(...)):
    """Create a single PO from one or multiple SC orders (same supplier)"""
    user = await get_current_user(request)
    # Support single or multiple SC order IDs
    sc_order_ids = data.get("subcontract_order_ids", [])
    if not sc_order_ids:
        single_id = data.get("subcontract_order_id")
        if single_id:
            sc_order_ids = [single_id]
    
    if not sc_order_ids:
        raise HTTPException(status_code=400, detail="No SC order IDs provided")
    
    # Collect all SC orders and verify same supplier
    sc_orders = []
    supplier_id = None
    for sc_id in sc_order_ids:
        sc = await db.subcontract_orders.find_one({"id": sc_id})
        if not sc:
            raise HTTPException(status_code=404, detail=f"SC order {sc_id} not found")
        # Check if PO already exists
        existing_po = await db.purchase_orders.find_one({"reference_sc_order_id": sc_id})
        if existing_po:
            raise HTTPException(status_code=400, detail=f"PO {existing_po.get('po_number')} already exists for SC {sc.get('order_number')}")
        if supplier_id and sc.get("supplier_id") != supplier_id:
            raise HTTPException(status_code=400, detail="All SC orders must be from the same supplier")
        supplier_id = sc.get("supplier_id")
        sc_orders.append(sc)
    
    # Build consolidated PO lines from all SC orders
    po_lines = []
    total_amount = 0
    sc_refs = []
    
    for sc_order in sc_orders:
        sc_refs.append(sc_order.get("order_number", ""))
        source_items = sc_order.get("job_work_parts", [])
        if not source_items:
            source_items = sc_order.get("lines", [])
        
        for line in source_items:
            item = await db.items.find_one({"id": line["item_id"]}, {"_id": 0})
            if item:
                unit_cost = line.get("charges", 0) or item.get("unit_cost", 0) or line.get("rate", 0)
                qty = line.get("quantity", 0)
                line_total = qty * unit_cost
                total_amount += line_total
                # Merge same item
                found = False
                for pl in po_lines:
                    if pl["item_id"] == line["item_id"]:
                        pl["quantity"] += qty
                        pl["total_price"] += line_total
                        found = True
                        break
                if not found:
                    po_lines.append({
                        "item_id": line["item_id"],
                        "quantity": qty,
                        "unit_price": unit_cost,
                        "total_price": line_total,
                        "received_quantity": 0,
                        "description": f"{item.get('part_number', '')} - {item.get('name', '')}",
                        "uom": item.get("unit_of_measure", "pcs"),
                        "hsn_code": item.get("hsn_code", ""),
                        "gst_rate": item.get("gst_rate", 18),
                        "discount_type": "percentage",
                        "discount_value": 0,
                        "notes": ""
                    })
    
    po_number = await get_next_series_number("po_number")
    # ALL POs created from SC go into DRAFT status — they must be explicitly
    # approved on the Purchase Orders page before dispatch/receipt. This applies
    # to both SC with RM (with_material) and SC without RM / Job OS (without_material).
    po_status = "draft"
    po_doc = {
        "id": str(uuid.uuid4()),
        "po_number": po_number,
        "supplier_id": supplier_id,
        "reference_sc_order_id": sc_order_ids[0],
        "reference_sc_order_ids": sc_order_ids,
        "lines": po_lines,
        "subtotal": total_amount,
        "additional_charges": [],
        "total_amount": total_amount,
        "status": po_status,
        "notes": f"Auto-created from SC Orders: {', '.join(sc_refs)}",
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"]
    }
    await db.purchase_orders.insert_one(po_doc)
    po_doc.pop("_id", None)
    
    # Mark all SC orders as having PO
    for sc_id in sc_order_ids:
        await db.subcontract_orders.update_one({"id": sc_id}, {"$set": {"po_created": True, "po_number": po_number}})
    
    return {"po_number": po_number, "po_id": po_doc["id"], "total_amount": total_amount, "status": po_status}


@jobwork_router.get("/challans")
async def get_delivery_challans(request: Request):
    user = await get_current_user(request)
    challans = await db.delivery_challans.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Batch fetch related SCs, suppliers, items
    sc_ids = {dc.get("subcontract_order_id") for dc in challans if dc.get("subcontract_order_id")}
    scs_map = {}
    if sc_ids:
        async for sc in db.subcontract_orders.find({"id": {"$in": list(sc_ids)}}, {"_id": 0}):
            scs_map[sc["id"]] = sc
    # Collect all supplier ids from both SCs and manual DCs
    supplier_ids = {sc.get("supplier_id") for sc in scs_map.values() if sc.get("supplier_id")}
    supplier_ids |= {dc.get("supplier_id") for dc in challans if dc.get("supplier_id")}
    suppliers_map = {}
    if supplier_ids:
        async for s in db.suppliers.find({"id": {"$in": list(supplier_ids)}}, {"_id": 0}):
            suppliers_map[s["id"]] = s
    item_ids = set()
    for dc in challans:
        for line in dc.get("lines", []):
            if line.get("item_id"): item_ids.add(line["item_id"])
    for sc in scs_map.values():
        for part in sc.get("job_work_parts", []):
            if part.get("item_id"): item_ids.add(part["item_id"])
    items_map = {}
    if item_ids:
        async for it in db.items.find({"id": {"$in": list(item_ids)}}, {"_id": 0}):
            items_map[it["id"]] = it
    # Resolve creator name for each DC so the print template can show the
    # ORIGINAL creator (not the user currently taking the print). Fetch
    # once per user_id.
    creator_ids = {dc.get("created_by") for dc in challans if dc.get("created_by")}
    creators_map = {}
    if creator_ids:
        async for u in db.users.find({"id": {"$in": list(creator_ids)}}, {"_id": 0, "id": 1, "name": 1, "email": 1}):
            creators_map[u["id"]] = u.get("name") or u.get("email") or ""

    for dc in challans:
        order = scs_map.get(dc.get("subcontract_order_id"))
        dc["order"] = order
        dc["fg_item_name"] = order.get("fg_item_name", "") if order else ""
        # Manual DCs carry supplier_id directly; parent-SC DCs inherit from the SC.
        dc_supplier_id = dc.get("supplier_id") or (order.get("supplier_id") if order else None)
        dc["supplier"] = suppliers_map.get(dc_supplier_id) if dc_supplier_id else None
        dc["is_manual"] = bool(dc.get("is_manual"))
        # Surface the creator's display name so the print template doesn't
        # have to fall back to the currently-logged-in user.
        dc["created_by_name"] = creators_map.get(dc.get("created_by"), "")
        # For Job Card OS DCs that are still UNSENT (status=draft), override the
        # stored per-line processing_charges with the SPECIFIC outsourced
        # routing's cost computed from the parent SC's job_work_parts +
        # find_routing_cost. This self-heals draft DCs that were saved while
        # the SC carried a polluted combined `charges` value. Sent/completed
        # DCs are left untouched (their snapshot must be preserved for audit).
        dc_is_draft_job_os = (
            order is not None
            and (dc.get("status") or "").lower() in ("draft", "open", "")
            and order.get("subcontract_type") == "without_material"
            and (order.get("reference_operation_seqs") or order.get("reference_operation_seq"))
        )
        jwp_lookup = {}
        if dc_is_draft_job_os:
            for jp in order.get("job_work_parts", []) or []:
                # Key by item_id only — the typical Job Card OS DC has 1 part
                # per item. For consolidated SCs with the same item under
                # multiple ops, prefer the first matching entry; user can
                # always edit the value in the dialog.
                jwp_lookup.setdefault(jp.get("item_id"), jp)
        for line in dc.get("lines", []):
            line["item"] = items_map.get(line.get("item_id"))
            if dc_is_draft_job_os:
                jp = jwp_lookup.get(line.get("item_id"))
                if jp and jp.get("process_name"):
                    specific = await find_routing_cost(jp.get("item_id"), jp.get("process_name"))
                    if specific:
                        line["processing_charges"] = float(specific)
                        line["_overridden_specific_op"] = True
                # Surface description for the DC table / print template.
                if jp and jp.get("item_description") and not line.get("item_description"):
                    line["item_description"] = jp.get("item_description")
        if order:
            for part in order.get("job_work_parts", []):
                part["item"] = items_map.get(part.get("item_id"))
    return challans

@jobwork_router.get("/manual-dc/open")
async def list_open_manual_dcs(request: Request):
    """List Manual DCs that still have receivable balance — used by the
    Manual GRN dialog's DC picker to auto-fill items + remaining qty + unit
    price. A DC is "open" when status is in (draft, sent, approved) AND
    at least one line has `quantity > received_qty`. Includes the
    supplier_name for the picker label.
    """
    await get_current_user(request)
    dcs = await db.delivery_challans.find(
        {"is_manual": True, "status": {"$in": ["draft", "sent", "approved"]}},
        {"_id": 0},
    ).sort("created_at", -1).to_list(500)
    # Attach supplier_name and per-line received_qty (default 0 if not set).
    sup_ids = list({d.get("supplier_id") for d in dcs if d.get("supplier_id")})
    sup_map = {}
    if sup_ids:
        async for s in db.suppliers.find({"id": {"$in": sup_ids}}, {"_id": 0, "id": 1, "name": 1}):
            sup_map[s["id"]] = s.get("name") or ""
    out = []
    for d in dcs:
        new_lines = []
        any_open = False
        for ln in (d.get("lines") or []):
            rq = float(ln.get("received_qty") or 0)
            qty = float(ln.get("quantity") or 0)
            new_lines.append({**ln, "received_qty": rq})
            if qty - rq > 0:
                any_open = True
        if not any_open:
            continue
        d["lines"] = new_lines
        d["supplier_name"] = sup_map.get(d.get("supplier_id") or "", "")
        out.append(d)
    return out


@jobwork_router.post("/challans/manual", status_code=201)
async def create_manual_delivery_challan(data: ManualDCCreate, request: Request):
    """Create a standalone (manual) DC — not linked to any Subcontract Order.
    Used when goods are shipped outward and a GRN is expected back later (DC→GRN flow).

    Stock is NOT deducted on create/edit — only when the DC is explicitly
    sent via POST /challans/{dc_id}/send. This prevents the historic
    "stock deducted twice" bug where draft creation + send each issued
    their own inventory_transactions row.
    """
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager", "inventory_manager"], module="job_work", action="create")
    supplier = await db.suppliers.find_one({"id": data.supplier_id})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")

    if not data.lines:
        raise HTTPException(status_code=400, detail="At least one line item is required")

    count = await db.delivery_challans.count_documents({})
    dc_number = await get_next_series_number("delivery_challan")

    dc_lines = []
    for line in data.lines:
        item = await db.items.find_one({"id": line.item_id})
        line_doc = {
            "item_id": line.item_id,
            "quantity": float(line.quantity),
            "unit": line.unit or (item.get("unit_of_measure") if item else "pcs"),
            "unit_price": float(line.unit_price or (item.get("unit_cost") if item else 0) or 0),
            "processing_charges": float(line.processing_charges or 0),
            "notes": line.notes or "",
            "item_description": line.item_description or "",
        }
        dc_lines.append(line_doc)

    dc_doc = {
        "id": str(uuid.uuid4()),
        "dc_number": dc_number,
        "subcontract_order_id": None,  # manual DCs are not tied to an SC
        "is_manual": True,
        "supplier_id": data.supplier_id,
        "dc_purpose": data.dc_purpose or "subcontract",
        "warehouse_id": data.warehouse_id or "",
        "lines": dc_lines,
        "notes": data.notes or "",
        "status": "draft",
        # User-pickable DC date (YYYY-MM-DD). Defaults to today when the
        # client doesn't send one. Stored as ISO date string so the
        # printable DC and GRN-from-DC flows can show the original DC
        # date instead of the server timestamp.
        "dc_date": data.dc_date or datetime.now(timezone.utc).date().isoformat(),
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"]
    }
    await db.delivery_challans.insert_one(dc_doc)
    dc_doc.pop("_id", None)
    return dc_doc


@jobwork_router.put("/challans/manual/{dc_id}")
async def update_manual_delivery_challan(dc_id: str, data: ManualDCCreate, request: Request):
    """Update a manual (standalone) DC while it's still in draft. Stock movements
    are recalculated diff-style: items removed from the DC have their stock
    refunded, items added have stock deducted, and quantity changes are netted.

    Only manual DCs in `draft` status can be edited — once a DC is `sent` it's
    treated as an outward shipment and is immutable from this endpoint.
    """
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager", "inventory_manager"], module="job_work", action="edit")

    dc = await db.delivery_challans.find_one({"id": dc_id})
    if not dc:
        raise HTTPException(status_code=404, detail="DC not found")
    if not dc.get("is_manual"):
        raise HTTPException(status_code=400, detail="Only manual DCs are editable via this endpoint")
    if dc.get("status") not in ("draft",):
        raise HTTPException(status_code=400, detail=f"DC in status '{dc.get('status')}' cannot be edited")

    if not data.lines:
        raise HTTPException(status_code=400, detail="At least one line item is required")

    # Stock is NOT touched on edit while DC is still in draft — it will be
    # deducted in a single pass when the DC is sent. This eliminates the
    # historic double-deduct + complicated diff-based refund logic.

    # Rebuild DC lines preserving the same shape as create.
    dc_lines = []
    for line in data.lines:
        item = await db.items.find_one({"id": line.item_id})
        dc_lines.append({
            "item_id": line.item_id,
            "quantity": float(line.quantity),
            "unit": line.unit or (item.get("unit_of_measure") if item else "pcs"),
            "unit_price": float(line.unit_price or (item.get("unit_cost") if item else 0) or 0),
            "processing_charges": float(line.processing_charges or 0),
            "notes": line.notes or "",
            "item_description": line.item_description or "",
        })

    await db.delivery_challans.update_one(
        {"id": dc_id},
        {"$set": {
            "supplier_id": data.supplier_id,
            "dc_purpose": data.dc_purpose or "subcontract",
            "warehouse_id": data.warehouse_id or "",
            "lines": dc_lines,
            "notes": data.notes or "",
            "dc_date": data.dc_date or dc.get("dc_date") or datetime.now(timezone.utc).date().isoformat(),
            "updated_at": datetime.now(timezone.utc),
            "updated_by": user["id"],
        }},
    )
    updated = await db.delivery_challans.find_one({"id": dc_id}, {"_id": 0})
    return updated

@jobwork_router.post("/challans", status_code=201)
async def create_delivery_challan(data: DCCreate, request: Request):
    """Create DC - Send materials to subcontractor. Deducts stock."""
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="job_work", action="create")
    order = await db.subcontract_orders.find_one({"id": data.subcontract_order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Subcontract order not found")
    if order.get("status") not in ["confirmed", "in_progress"]:
        raise HTTPException(status_code=400, detail="Order must be confirmed before sending materials")
    
    count = await db.delivery_challans.count_documents({})
    dc_number = await get_next_series_number("delivery_challan")
    
    skip_deduct = data.skip_stock_deduct or False
    
    # First pass: check ALL items for stock availability (skip for Job Card outsource)
    if not skip_deduct:
        insufficient = []
        for line in data.lines:
            item = await db.items.find_one({"id": line.item_id})
            if not item:
                continue
            current_stock = item.get("current_stock", 0)
            if current_stock < line.quantity:
                insufficient.append({
                    "item": item.get("part_number", ""),
                    "name": item.get("name", ""),
                    "required": line.quantity,
                    "available": current_stock,
                    "shortage": line.quantity - current_stock
                })
        
        if insufficient:
            return JSONResponse(status_code=200, content={
                "success": False,
                "message": "Insufficient stock for DC creation",
                "insufficient_materials": insufficient
            })
    
    dc_lines = []
    for line in data.lines:
        item = await db.items.find_one({"id": line.item_id})
        if not item:
            continue
        
        if not skip_deduct:
            # Deduct stock — read fresh current_stock from this item
            item_current_stock = item.get("current_stock", 0)
            new_stock = item_current_stock - line.quantity
            await db.items.update_one({"id": line.item_id}, {"$set": {"current_stock": new_stock}})
            
            # Create inventory transaction
            tx = {
                "id": str(uuid.uuid4()),
                "item_id": line.item_id,
                "transaction_type": "issue",
                "quantity": line.quantity,
                "reference_type": "job_work_dc",
                "reference_id": dc_number,
                "previous_stock": item_current_stock,
                "new_stock": new_stock,
                "notes": f"Sent to subcontractor - {order.get('order_number')}",
                "created_at": datetime.now(timezone.utc),
                "created_by": user["id"]
            }
            await db.inventory_transactions.insert_one(tx)
        
        dc_lines.append({
            "item_id": line.item_id,
            "quantity": line.quantity,
            "rate": line.rate or 0,
            "processing_charges": line.processing_charges or 0,
            "item_description": line.item_description or "",
            "process_name": line.process_name or "",
        })
        
        # Update sent quantity in order lines
        for ol in order.get("lines", []):
            if ol["item_id"] == line.item_id:
                ol["sent_quantity"] = ol.get("sent_quantity", 0) + line.quantity
        # Also update sent in job_work_parts (for Job OS)
        for jp in order.get("job_work_parts", []):
            if jp.get("item_id") == line.item_id:
                jp["sent_quantity"] = jp.get("sent_quantity", 0) + line.quantity
    
    update_fields = {"lines": order.get("lines", []), "status": "in_progress"}
    if order.get("job_work_parts"):
        update_fields["job_work_parts"] = order["job_work_parts"]
        update_fields["dc_created"] = True
    await db.subcontract_orders.update_one({"id": data.subcontract_order_id}, {"$set": update_fields})
    
    dc_doc = {
        "id": str(uuid.uuid4()),
        "dc_number": dc_number,
        "subcontract_order_id": data.subcontract_order_id,
        "lines": dc_lines,
        "warehouse_id": data.warehouse_id or "",
        "status": "sent",
        "notes": data.notes or "",
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"]
    }
    
    await db.delivery_challans.insert_one(dc_doc)
    del dc_doc["_id"]
    return dc_doc

@jobwork_router.post("/challans/{dc_id}/send")
async def send_draft_dc(dc_id: str, request: Request):
    """Send a draft DC - deducts stock and marks as sent"""
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="job_work", action="create")
    dc = await db.delivery_challans.find_one({"id": dc_id})
    if not dc:
        raise HTTPException(status_code=404, detail="Delivery challan not found")
    if dc.get("status") != "draft":
        raise HTTPException(status_code=400, detail="DC is not in draft status")
    
    order = await db.subcontract_orders.find_one({"id": dc.get("subcontract_order_id")})
    
    # First pass: check ALL items for stock availability
    insufficient = []
    for line in dc.get("lines", []):
        item = await db.items.find_one({"id": line["item_id"]})
        if not item:
            continue
        current_stock = item.get("current_stock", 0)
        qty = line.get("quantity", 0)
        if current_stock < qty:
            insufficient.append({
                "item": item.get("part_number", ""),
                "name": item.get("name", ""),
                "required": qty,
                "available": current_stock,
                "shortage": qty - current_stock
            })
    
    if insufficient:
        return JSONResponse(status_code=200, content={
            "success": False,
            "message": "Insufficient stock for DC send",
            "insufficient_materials": insufficient
        })
    
    # Second pass: deduct stock
    consumed_materials = []
    for line in dc.get("lines", []):
        item = await db.items.find_one({"id": line["item_id"]})
        if not item:
            continue
        current_stock = item.get("current_stock", 0)
        qty = line.get("quantity", 0)
        new_stock = current_stock - qty
        await db.items.update_one({"id": line["item_id"]}, {"$set": {"current_stock": new_stock}})
        consumed_materials.append({
            "item": item.get("part_number", ""),
            "name": item.get("name", ""),
            "quantity": qty,
            "uom": item.get("unit_of_measure", "pcs"),
            "previous_stock": current_stock,
            "new_stock": new_stock
        })
        tx = {
            "id": str(uuid.uuid4()),
            "item_id": line["item_id"],
            "transaction_type": "issue",
            "quantity": qty,
            "reference_type": "job_work_dc",
            "reference_id": dc.get("dc_number"),
            "previous_stock": current_stock,
            "new_stock": new_stock,
            "notes": f"Sent to subcontractor - {order.get('order_number', '') if order else ''}",
            "created_at": datetime.now(timezone.utc),
            "created_by": user["id"]
        }
        await db.inventory_transactions.insert_one(tx)
        if order:
            for ol in order.get("lines", []):
                if ol["item_id"] == line["item_id"]:
                    ol["sent_quantity"] = ol.get("sent_quantity", 0) + qty
    
    if order:
        await db.subcontract_orders.update_one({"id": order["id"]}, {"$set": {"lines": order["lines"], "status": "in_progress"}})
    
    await db.delivery_challans.update_one({"id": dc_id}, {"$set": {"status": "sent", "sent_at": datetime.now(timezone.utc)}})
    return {"message": f"DC {dc.get('dc_number')} sent successfully", "consumed_materials": consumed_materials}


@jobwork_router.post("/receive-grn")
async def receive_grn_from_jw(request: Request, data: dict = Body(...)):
    """Create GRN directly from JW number (SC with RM). Adds FG/SA stock, process cost tracked."""
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager", "inventory_manager"], module="job_work", action="create")
    # Mandatory fields validation first
    supplier_invoice_no = data.get("supplier_invoice_no", "").strip() if data.get("supplier_invoice_no") else ""
    supplier_invoice_date = data.get("supplier_invoice_date")
    if not supplier_invoice_no:
        raise HTTPException(status_code=400, detail="Supplier Invoice No. is mandatory")
    if not supplier_invoice_date:
        raise HTTPException(status_code=400, detail="Supplier Invoice Date is mandatory")
    
    sc_order_id = data.get("subcontract_order_id")
    order = await db.subcontract_orders.find_one({"id": sc_order_id})
    if not order:
        raise HTTPException(status_code=404, detail="SC order not found")
    
    lines = data.get("lines", [])
    if not lines:
        raise HTTPException(status_code=400, detail="No items to receive")
    
    # Generate GRN number
    count = await db.grn.count_documents({})
    grn_number = f"GRN-{str(count + 1).zfill(6)}"
    
    grn_lines = []
    total_process_cost = 0
    # Track per-line how much qty went to each wo_id so we can mark the
    # right outsourced operation as completed below. Indexed by wo_id.
    received_per_wo = {}
    for line in lines:
        item = await db.items.find_one({"id": line["item_id"]}, {"_id": 0})
        if not item:
            continue
        
        recv_qty = line.get("received_quantity", 0)
        process_charges = line.get("process_charges", 0)
        total_process_cost += recv_qty * process_charges
        
        # Add stock
        current_stock = item.get("current_stock", 0)
        new_stock = current_stock + recv_qty
        await db.items.update_one({"id": line["item_id"]}, {"$set": {"current_stock": new_stock}})
        
        # Inventory transaction
        await db.inventory_transactions.insert_one({
            "id": str(uuid.uuid4()),
            "item_id": line["item_id"],
            "transaction_type": "receive",
            "quantity": recv_qty,
            "reference_type": "jw_grn",
            "reference_id": grn_number,
            "previous_stock": current_stock,
            "new_stock": new_stock,
            "notes": f"JW GRN {grn_number} from {order.get('order_number')} - Process cost: {process_charges}/unit",
            "created_at": datetime.now(timezone.utc),
            "created_by": user["id"]
        })
        
        grn_lines.append({
            "item_id": line["item_id"],
            "received_quantity": recv_qty,
            "process_charges": process_charges,
            "total_charges": recv_qty * process_charges,
            "uom": item.get("unit_of_measure", "pcs"),
            "hsn_code": item.get("hsn_code", ""),
        })
        
        # Distribute received qty across ALL matching job_work_parts entries
        # FIFO (by remaining open balance). Consolidated SCs frequently carry
        # multiple parts for the SAME item — one per source MO. Previously we
        # incremented only the first match, leaving downstream MOs stuck.
        remaining = float(recv_qty)
        # Pass 1: parts that still have open balance — fill them FIFO.
        for p in order.get("job_work_parts", []):
            if remaining <= 0:
                break
            if p.get("item_id") != line["item_id"]:
                continue
            open_qty = float(p.get("quantity") or 0) - float(p.get("received_quantity") or 0)
            if open_qty <= 0:
                continue
            take = min(open_qty, remaining)
            p["received_quantity"] = float(p.get("received_quantity") or 0) + take
            remaining -= take
            if p.get("wo_id"):
                received_per_wo[p["wo_id"]] = received_per_wo.get(p["wo_id"], 0.0) + take
        # Pass 2: any leftover goes onto the first matching part (over-receipt).
        if remaining > 0:
            for p in order.get("job_work_parts", []):
                if p.get("item_id") == line["item_id"]:
                    p["received_quantity"] = float(p.get("received_quantity") or 0) + remaining
                    if p.get("wo_id"):
                        received_per_wo[p["wo_id"]] = received_per_wo.get(p["wo_id"], 0.0) + remaining
                    break
    
    # Save GRN
    grn_doc = {
        "id": str(uuid.uuid4()),
        "grn_number": grn_number,
        "jw_order_id": sc_order_id,
        "jw_order_number": order.get("order_number", ""),
        "supplier_id": order.get("supplier_id", ""),
        "supplier_invoice_no": supplier_invoice_no,
        "supplier_invoice_date": supplier_invoice_date,
        "lines": grn_lines,
        "total_process_cost": total_process_cost,
        "notes": f"JW GRN from {order.get('order_number')}",
        "status": "completed",
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"]
    }
    await db.grn.insert_one(grn_doc)
    
    # Check if all parts are fully received
    all_received = True
    for p in order.get("job_work_parts", []):
        if p.get("received_quantity", 0) < p.get("quantity", 0):
            all_received = False
            break
    
    # Update SC order
    sc_update = {
        "job_work_parts": order.get("job_work_parts", []),
        "last_receipt_date": datetime.now(timezone.utc).isoformat(),
        "grn_number": grn_number
    }
    if all_received:
        sc_update["status"] = "completed"
        sc_update["completed_at"] = datetime.now(timezone.utc).isoformat()
    await db.subcontract_orders.update_one({"id": sc_order_id}, {"$set": sc_update})

    # Update MO progress — for EVERY linked WO, regardless of whether the
    # whole SC is fully received. This is the critical fix: a consolidated
    # SC may have one MO fully received while another is still pending;
    # the previously gated `if all_received:` block left the completed
    # MO stuck in "in_progress" because partial receipts never ran the
    # WO update. Now each WO's outsourced op is updated based on the
    # qty actually received against that specific MO (via wo_id linkage).
    all_wo_ids = list(set(filter(None, [
        order.get("reference_wo_id"),
        *(order.get("reference_wo_ids", []))
    ])))
    for ref_wo_id in all_wo_ids:
        ref_wo = await db.work_orders.find_one({"id": ref_wo_id})
        if not ref_wo or ref_wo.get("status") == "completed":
            continue
        mo_qty = float(ref_wo.get("quantity", 0) or 0)
        # Total qty received against THIS specific WO across all matching
        # parts. Falls back to total received for the WO's FG item when
        # wo_id wasn't tracked on the parts (legacy SCs).
        wo_recv = received_per_wo.get(ref_wo_id, 0.0)
        if wo_recv <= 0:
            # Legacy fallback — sum received_quantity from job_work_parts
            # whose wo_id matches OR (when wo_id missing) whose item_id
            # matches the WO's FG item.
            for p in order.get("job_work_parts", []):
                if p.get("wo_id") == ref_wo_id:
                    wo_recv += float(p.get("received_quantity") or 0)
            if wo_recv <= 0:
                for p in order.get("job_work_parts", []):
                    if not p.get("wo_id") and p.get("item_id") == ref_wo.get("item_id"):
                        wo_recv += float(p.get("received_quantity") or 0)
        ops = ref_wo.get("operations_status", [])
        qty_completed = min(wo_recv, mo_qty)
        fully_done_for_wo = qty_completed >= mo_qty and mo_qty > 0
        for op in ops:
            # Mark the outsourced op as received (or completed if all qty
            # is back). Non-outsourced ops are left untouched.
            if op.get("is_job_work"):
                op["outsource_status"] = "received" if fully_done_for_wo else op.get("outsource_status") or "received_partial"
                op["quantity_completed"] = max(qty_completed, float(op.get("quantity_completed") or 0))
                op["quantity_accepted"] = max(qty_completed, float(op.get("quantity_accepted") or 0))
                if fully_done_for_wo:
                    op["status"] = "completed"
                    op["actual_end"] = datetime.now(timezone.utc)
        mo_update = {
            "operations_status": ops,
            "quantity_completed": max(qty_completed, float(ref_wo.get("quantity_completed") or 0)),
            "updated_at": datetime.now(timezone.utc),
        }
        # MO is completed iff every operation is now completed.
        if all((op.get("status") == "completed") for op in ops) and ops:
            mo_update["status"] = "completed"
            mo_update["actual_end"] = datetime.now(timezone.utc)
        await db.work_orders.update_one({"id": ref_wo_id}, {"$set": mo_update})
    
    grn_doc.pop("_id", None)
    return {"grn_number": grn_number, "total_process_cost": total_process_cost, "all_received": all_received}

@jobwork_router.get("/receipts")
async def get_subcontract_receipts(request: Request):
    user = await get_current_user(request)
    receipts = await db.subcontract_receipts.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    sc_ids = {r.get("subcontract_order_id") for r in receipts if r.get("subcontract_order_id")}
    scs_map = {}
    if sc_ids:
        async for sc in db.subcontract_orders.find({"id": {"$in": list(sc_ids)}}, {"_id": 0}):
            scs_map[sc["id"]] = sc
    supplier_ids = {sc.get("supplier_id") for sc in scs_map.values() if sc.get("supplier_id")}
    suppliers_map = {}
    if supplier_ids:
        async for s in db.suppliers.find({"id": {"$in": list(supplier_ids)}}, {"_id": 0}):
            suppliers_map[s["id"]] = s
    item_ids = set()
    for r in receipts:
        for line in r.get("lines", []):
            if line.get("item_id"): item_ids.add(line["item_id"])
    items_map = {}
    if item_ids:
        async for it in db.items.find({"id": {"$in": list(item_ids)}}, {"_id": 0}):
            items_map[it["id"]] = it
    for rec in receipts:
        order = scs_map.get(rec.get("subcontract_order_id"))
        rec["order"] = order
        rec["supplier"] = suppliers_map.get(order.get("supplier_id")) if order else None
        for line in rec.get("lines", []):
            line["item"] = items_map.get(line.get("item_id"))
    return receipts

@jobwork_router.post("/receipts", status_code=201)
async def create_subcontract_receipt(data: SubcontractReceiptCreate, request: Request):
    """Receive materials back from subcontractor. Adds stock for FG/SA items only."""
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="job_work", action="create")
    order = await db.subcontract_orders.find_one({"id": data.subcontract_order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Subcontract order not found")
    
    sc_type = order.get("subcontract_type", "with_material")
    
    # For with_material SC: only allow receiving job_work_parts items (FG/SA/Parts)
    # NOT the RM lines that were sent to vendor
    valid_receipt_item_ids = set()
    if sc_type != "without_material" and order.get("job_work_parts"):
        valid_receipt_item_ids = {jp["item_id"] for jp in order.get("job_work_parts", []) if jp.get("item_id")}
    
    count = await db.subcontract_receipts.count_documents({})
    receipt_number = f"SR-{str(count + 1).zfill(6)}"
    
    rec_lines = []
    total_rework_qty = 0
    total_reject_qty = 0
    for line in data.lines:
        item = await db.items.find_one({"id": line.item_id})
        if not item:
            continue
        
        rework_qty = getattr(line, 'rework_qty', 0) or 0
        reject_qty = line.reject_qty or 0
        accepted_qty = max(0, line.received_quantity - reject_qty - rework_qty)
        total_rework_qty += rework_qty
        total_reject_qty += reject_qty
        
        # Add only accepted stock — but for with_material SC, only for job_work_parts items
        should_add_stock = accepted_qty > 0
        if should_add_stock and valid_receipt_item_ids and line.item_id not in valid_receipt_item_ids:
            # This item is from RM lines, not job_work_parts — do NOT add to stock
            should_add_stock = False
        
        if should_add_stock:
            current_stock = item.get("current_stock", 0)
            new_stock = current_stock + accepted_qty
            await db.items.update_one({"id": line.item_id}, {"$set": {"current_stock": new_stock}})
            
            tx = {
                "id": str(uuid.uuid4()),
                "item_id": line.item_id,
                "transaction_type": "receive",
                "quantity": accepted_qty,
                "reference_type": "job_work_receipt",
                "reference_id": receipt_number,
                "previous_stock": current_stock,
                "new_stock": new_stock,
                "notes": f"Received from subcontractor - {order.get('order_number')}" + (f" (Rework: {rework_qty}, Reject: {reject_qty})" if rework_qty or reject_qty else ""),
                "created_at": datetime.now(timezone.utc),
                "created_by": user["id"]
            }
            await db.inventory_transactions.insert_one(tx)
        
        rec_lines.append({
            "item_id": line.item_id,
            "received_quantity": line.received_quantity,
            "accepted_quantity": accepted_qty,
            "rework_qty": rework_qty,
            "quality_result": line.quality_result or "accept",
            "reject_qty": reject_qty
        })
        
        # Update received quantity in order
        # For "with material": receipt is FG/SA/Parts from job_work_parts
        # For lines/without_material: receipt is from lines
        updated_jw_parts = False
        for jp in order.get("job_work_parts", []):
            if jp.get("item_id") == line.item_id:
                jp["received_quantity"] = jp.get("received_quantity", 0) + accepted_qty
                updated_jw_parts = True
        if not updated_jw_parts:
            for ol in order.get("lines", []):
                if ol["item_id"] == line.item_id:
                    ol["received_quantity"] = ol.get("received_quantity", 0) + accepted_qty
    
    # Check completion based on SC type
    sc_type = order.get("subcontract_type", "with_material")
    has_rework = total_rework_qty > 0
    
    if sc_type != "without_material" and order.get("job_work_parts"):
        # With material: check job_work_parts completion
        all_received = all(jp.get("received_quantity", 0) >= jp.get("quantity", 0) for jp in order.get("job_work_parts", []))
    elif sc_type == "without_material":
        all_received = all(ol.get("received_quantity", 0) >= ol.get("quantity", 0) for ol in order.get("lines", []))
    else:
        all_received = all(ol.get("received_quantity", 0) >= ol.get("sent_quantity", 0) for ol in order.get("lines", []))
    
    new_status = "completed" if all_received and not has_rework else "in_progress"
    
    update_fields = {"lines": order["lines"], "status": new_status, "last_receipt_date": datetime.now(timezone.utc).isoformat()}
    if new_status == "completed":
        update_fields["completed_at"] = datetime.now(timezone.utc).isoformat()
    if order.get("job_work_parts"):
        update_fields["job_work_parts"] = order["job_work_parts"]
    await db.subcontract_orders.update_one({"id": data.subcontract_order_id}, {"$set": update_fields})
    
    # Update MO progress for ALL linked WOs (partial or complete)
    all_wo_ids = list(set(filter(None, [
        order.get("reference_wo_id"),
        *(order.get("reference_wo_ids", []))
    ])))
    
    for ref_wo_id in all_wo_ids:
        ref_wo = await db.work_orders.find_one({"id": ref_wo_id})
        if not ref_wo or ref_wo.get("status") == "completed":
            continue
        
        # Calculate received qty for this MO's item from job_work_parts
        mo_item_id = ref_wo.get("item_id")
        mo_qty = ref_wo.get("quantity", 0)
        received_for_mo = 0
        for jp in order.get("job_work_parts", []):
            if jp.get("item_id") == mo_item_id:
                received_for_mo = jp.get("received_quantity", 0)
                break
        
        # Update MO quantity_completed and progress
        qty_completed = min(received_for_mo, mo_qty)
        ops = ref_wo.get("operations_status", [])
        for op in ops:
            op["quantity_completed"] = qty_completed
            op["quantity_accepted"] = qty_completed
            if qty_completed >= mo_qty:
                op["status"] = "completed"
                op["actual_end"] = datetime.now(timezone.utc)
        
        mo_update = {
            "quantity_completed": qty_completed,
            "operations_status": ops,
            "updated_at": datetime.now(timezone.utc)
        }
        
        if qty_completed >= mo_qty:
            mo_update["status"] = "completed"
            mo_update["actual_end"] = datetime.now(timezone.utc)
            # NOTE: Stock for FG/SA is already added at receipt time (line-by-line above)
            # Do NOT add FG stock again here to prevent double-counting
        
        await db.work_orders.update_one({"id": ref_wo_id}, {"$set": mo_update})
            
    
    rec_doc = {
        "id": str(uuid.uuid4()),
        "receipt_number": receipt_number,
        "subcontract_order_id": data.subcontract_order_id,
        "dc_id": data.dc_id or "",
        "lines": rec_lines,
        "warehouse_id": data.warehouse_id or "",
        "status": "received",
        "notes": data.notes or "",
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"]
    }
    
    await db.subcontract_receipts.insert_one(rec_doc)
    del rec_doc["_id"]
    return rec_doc


# ==================== CRM MODULE ====================
# Two independent pipelines — Marketing (Leads) and Support (Tickets).
# Each record tracks stage + activities[] inline for a simple audit trail.

SLA_HOURS_BY_PRIORITY = {"low": 72, "medium": 24, "high": 8, "urgent": 2}

class CRMActivity(BaseModel):
    note: str
    created_at: Optional[datetime] = None
    created_by: Optional[str] = None
    author_name: Optional[str] = None

class LeadCreate(BaseModel):
    name: str  # Lead / opportunity title e.g. "Website enquiry — ABC Corp"
    customer_id: str  # REQUIRED — lead must link to a real Customer record
    customer_name: Optional[str] = ""  # auto-populated from customer master when omitted
    contact_person: Optional[str] = ""
    email: Optional[str] = ""
    phone: Optional[str] = ""
    source: Optional[str] = "website"  # website / referral / trade_show / cold_call / other
    estimated_value: Optional[float] = 0
    assignee_id: Optional[str] = ""
    next_followup: Optional[datetime] = None
    notes: Optional[str] = ""
    stage: Optional[str] = "enquiry"  # enquiry | quotation | negotiation | won | lost

class LeadUpdate(BaseModel):
    name: Optional[str] = None
    customer_name: Optional[str] = None
    contact_person: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    source: Optional[str] = None
    estimated_value: Optional[float] = None
    assignee_id: Optional[str] = None
    next_followup: Optional[datetime] = None
    notes: Optional[str] = None
    stage: Optional[str] = None
    customer_id: Optional[str] = None  # set after "Convert to Customer"
    lost_reason: Optional[str] = None

class TicketCreate(BaseModel):
    subject: str
    customer_id: str  # tickets always require an existing customer record
    description: Optional[str] = ""
    priority: Optional[str] = "medium"  # low / medium / high / urgent
    assignee_id: Optional[str] = ""
    linked_so_id: Optional[str] = ""  # deprecated — kept for back-compat
    linked_item_id: Optional[str] = ""  # deprecated — kept for back-compat
    product_ids: Optional[List[str]] = []  # items associated with this ticket
    stage: Optional[str] = "complaint"  # complaint | open | in_progress | pending | closed

class TicketUpdate(BaseModel):
    subject: Optional[str] = None
    description: Optional[str] = None
    priority: Optional[str] = None
    assignee_id: Optional[str] = None
    linked_so_id: Optional[str] = None
    linked_item_id: Optional[str] = None
    product_ids: Optional[List[str]] = None
    stage: Optional[str] = None
    resolution: Optional[str] = None

def _compute_sla_due(created_at, priority):
    """Add SLA hours to created_at based on priority. Returns tz-aware UTC datetime."""
    if not created_at:
        return None
    if isinstance(created_at, str):
        try:
            created_at = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
        except Exception:
            return None
    if isinstance(created_at, datetime) and created_at.tzinfo is None:
        created_at = created_at.replace(tzinfo=timezone.utc)
    hours = SLA_HOURS_BY_PRIORITY.get((priority or "medium").lower(), 24)
    return created_at + timedelta(hours=hours)

async def _enrich_lead(lead):
    lead.pop("_id", None)
    # Hydrate assignee name
    if lead.get("assignee_id"):
        a = await db.users.find_one({"id": lead["assignee_id"]}, {"_id": 0, "name": 1, "email": 1})
        lead["assignee"] = a
    if lead.get("customer_id"):
        c = await db.customers.find_one({"id": lead["customer_id"]}, {"_id": 0})
        lead["customer"] = c
    return lead

async def _enrich_ticket(t):
    t.pop("_id", None)
    if t.get("customer_id"):
        c = await db.customers.find_one({"id": t["customer_id"]}, {"_id": 0})
        t["customer"] = c
    if t.get("assignee_id"):
        a = await db.users.find_one({"id": t["assignee_id"]}, {"_id": 0, "name": 1, "email": 1})
        t["assignee"] = a
    if t.get("linked_so_id"):
        so = await db.production_orders.find_one({"id": t["linked_so_id"]}, {"_id": 0, "order_number": 1})
        t["linked_so"] = so
    # Hydrate products
    pids = t.get("product_ids") or []
    if pids:
        prods = await db.items.find({"id": {"$in": pids}}, {"_id": 0, "id": 1, "part_number": 1, "name": 1, "uom": 1}).to_list(200)
        t["products"] = prods
    else:
        t["products"] = []
    # Compute SLA breach flag on the fly
    sla_due = _compute_sla_due(t.get("created_at"), t.get("priority"))
    t["sla_due"] = sla_due
    terminal = t.get("stage") in ("closed",)
    if sla_due and not terminal:
        t["sla_breached"] = datetime.now(timezone.utc) > sla_due
    else:
        t["sla_breached"] = False
    return t

# --------- LEADS (Marketing) ---------
@crm_router.get("/leads")
async def list_leads(request: Request, stage: Optional[str] = None):
    await get_current_user(request)
    q = {}
    if stage:
        q["stage"] = stage
    leads = await db.crm_leads.find(q).sort("created_at", -1).to_list(2000)
    return [await _enrich_lead(l) for l in leads]

@crm_router.post("/leads", status_code=201)
async def create_lead(data: LeadCreate, request: Request):
    user = await get_current_user(request)
    # Validate customer exists and auto-populate enrichment
    cust = await db.customers.find_one({"id": data.customer_id}, {"_id": 0})
    if not cust:
        raise HTTPException(status_code=404, detail="Customer not found — create the customer first")
    payload = data.model_dump(exclude_none=True)
    # Auto-fill derived contact fields from customer if blank
    if not payload.get("customer_name"):
        payload["customer_name"] = cust.get("name") or ""
    if not payload.get("contact_person"):
        payload["contact_person"] = cust.get("contact_person") or ""
    if not payload.get("email"):
        payload["email"] = cust.get("email") or ""
    if not payload.get("phone"):
        payload["phone"] = cust.get("phone") or ""
    count = await db.crm_leads.count_documents({})
    lead_no = f"LEAD-{str(count + 1).zfill(6)}"
    doc = {
        "id": str(uuid.uuid4()),
        "lead_no": lead_no,
        **payload,
        "activities": [{
            "note": f"Lead created in stage: {data.stage or 'enquiry'}",
            "created_at": datetime.now(timezone.utc),
            "created_by": user["id"],
            "author_name": user.get("name")
        }],
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"]
    }
    await db.crm_leads.insert_one(doc)
    return await _enrich_lead(doc)

@crm_router.put("/leads/{lead_id}")
async def update_lead(lead_id: str, data: LeadUpdate, request: Request):
    user = await get_current_user(request)
    existing = await db.crm_leads.find_one({"id": lead_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Lead not found")
    update = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update:
        raise HTTPException(status_code=400, detail="No data to update")
    # If stage changes, append an activity line
    activity_entries = []
    if "stage" in update and update["stage"] != existing.get("stage"):
        activity_entries.append({
            "note": f"Stage changed: {existing.get('stage', '-')} → {update['stage']}",
            "created_at": datetime.now(timezone.utc),
            "created_by": user["id"],
            "author_name": user.get("name")
        })
    update["updated_at"] = datetime.now(timezone.utc)
    mongo_update = {"$set": update}
    if activity_entries:
        mongo_update["$push"] = {"activities": {"$each": activity_entries}}
    await db.crm_leads.update_one({"id": lead_id}, mongo_update)
    return await _enrich_lead(await db.crm_leads.find_one({"id": lead_id}))

@crm_router.post("/leads/{lead_id}/activity")
async def add_lead_activity(lead_id: str, payload: dict = Body(...), request: Request = None):
    user = await get_current_user(request)
    note = (payload.get("note") or "").strip()
    if not note:
        raise HTTPException(status_code=400, detail="Activity note required")
    await db.crm_leads.update_one(
        {"id": lead_id},
        {"$push": {"activities": {"note": note, "created_at": datetime.now(timezone.utc), "created_by": user["id"], "author_name": user.get("name")}}}
    )
    return await _enrich_lead(await db.crm_leads.find_one({"id": lead_id}))

@crm_router.post("/leads/{lead_id}/convert-to-customer")
async def convert_lead_to_customer(lead_id: str, payload: dict = Body(default={}), request: Request = None):
    """Convert a qualified lead's free-text customer_name into a real Customer record."""
    user = await get_current_user(request)
    lead = await db.crm_leads.find_one({"id": lead_id})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    if lead.get("customer_id"):
        raise HTTPException(status_code=400, detail="Lead already linked to a customer")
    # Build a minimal customer record from the lead
    customers_count = await db.customers.count_documents({})
    cust_code = payload.get("customer_code") or f"CUST-{str(customers_count + 1).zfill(6)}"
    cust_doc = {
        "id": str(uuid.uuid4()),
        "code": cust_code,  # Use 'code' to match the unique index
        "name": payload.get("name") or lead.get("customer_name"),
        "contact_person": payload.get("contact_person") or lead.get("contact_person", ""),
        "email": payload.get("email") or lead.get("email", ""),
        "phone": payload.get("phone") or lead.get("phone", ""),
        "address": payload.get("address", ""),
        "gstin": payload.get("gstin", ""),
        "status": "active",
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"]
    }
    await db.customers.insert_one(cust_doc)
    await db.crm_leads.update_one(
        {"id": lead_id},
        {
            "$set": {"customer_id": cust_doc["id"], "updated_at": datetime.now(timezone.utc)},
            "$push": {"activities": {
                "note": f"Converted to Customer: {cust_code} — {cust_doc['name']}",
                "created_at": datetime.now(timezone.utc),
                "created_by": user["id"],
                "author_name": user.get("name")
            }}
        }
    )
    cust_doc.pop("_id", None)
    return {"lead": await _enrich_lead(await db.crm_leads.find_one({"id": lead_id})), "customer": cust_doc}

@crm_router.delete("/leads/{lead_id}")
async def delete_lead(lead_id: str, request: Request):
    await get_current_user(request)
    res = await db.crm_leads.delete_one({"id": lead_id})
    if not res.deleted_count:
        raise HTTPException(status_code=404, detail="Lead not found")
    return {"message": "Lead deleted"}

@crm_router.post("/leads/import")
async def import_leads(payload: dict = Body(...), request: Request = None):
    """Bulk-import leads from an array of rows. Each row must have at least `name` and either
    `customer_id` (existing) OR `customer_name` (creates a minimal customer on the fly with address if provided).
    """
    user = await get_current_user(request)
    rows = payload.get("rows") or []
    if not rows:
        raise HTTPException(status_code=400, detail="No rows to import")
    created = 0
    skipped = []
    for idx, r in enumerate(rows, start=1):
        try:
            name = (r.get("name") or "").strip()
            if not name:
                skipped.append({"row": idx, "reason": "missing lead name"}); continue
            customer_id = r.get("customer_id")
            if not customer_id:
                cust_name = (r.get("customer_name") or "").strip()
                if not cust_name:
                    skipped.append({"row": idx, "reason": "missing customer_id and customer_name"}); continue
                # Try match by name
                existing = await db.customers.find_one({"name": cust_name}, {"_id": 0, "id": 1})
                if existing:
                    customer_id = existing["id"]
                else:
                    # Create minimal customer
                    ccount = await db.customers.count_documents({})
                    cdoc = {
                        "id": str(uuid.uuid4()),
                        "code": r.get("customer_code") or f"CUST-{str(ccount+1).zfill(6)}",
                        "name": cust_name,
                        "gstin": r.get("gstin", ""),
                        "contact_person": r.get("contact_person", ""),
                        "email": r.get("email", ""),
                        "phone": r.get("phone", ""),
                        "address": r.get("address", ""),
                        "status": "active",
                        "created_at": datetime.now(timezone.utc),
                        "created_by": user["id"],
                    }
                    await db.customers.insert_one(cdoc)
                    customer_id = cdoc["id"]
            # Build lead
            count = await db.crm_leads.count_documents({})
            lead_no = f"LEAD-{str(count+1).zfill(6)}"
            lead_doc = {
                "id": str(uuid.uuid4()),
                "lead_no": lead_no,
                "name": name,
                "customer_id": customer_id,
                "customer_name": r.get("customer_name", ""),
                "contact_person": r.get("contact_person", ""),
                "email": r.get("email", ""),
                "phone": r.get("phone", ""),
                "source": r.get("source", "website"),
                "estimated_value": float(r.get("estimated_value") or 0),
                "notes": r.get("notes", ""),
                "stage": r.get("stage") or "enquiry",
                "activities": [{"note": "Lead imported via bulk import", "created_at": datetime.now(timezone.utc), "created_by": user["id"], "author_name": user.get("name")}],
                "created_at": datetime.now(timezone.utc),
                "created_by": user["id"],
            }
            await db.crm_leads.insert_one(lead_doc)
            created += 1
        except Exception as e:
            skipped.append({"row": idx, "reason": str(e)})
    return {"created": created, "skipped": skipped, "total": len(rows)}

# --------- TICKETS (Support) ---------
@crm_router.get("/tickets")
async def list_tickets(request: Request, stage: Optional[str] = None):
    await get_current_user(request)
    q = {}
    if stage:
        q["stage"] = stage
    tickets = await db.crm_tickets.find(q).sort("created_at", -1).to_list(2000)
    return [await _enrich_ticket(t) for t in tickets]

@crm_router.post("/tickets", status_code=201)
async def create_ticket(data: TicketCreate, request: Request):
    user = await get_current_user(request)
    customer = await db.customers.find_one({"id": data.customer_id})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    count = await db.crm_tickets.count_documents({})
    ticket_no = f"TKT-{str(count + 1).zfill(6)}"
    doc = {
        "id": str(uuid.uuid4()),
        "ticket_no": ticket_no,
        **data.model_dump(exclude_none=True),
        "activities": [{
            "note": f"Ticket created — priority: {data.priority or 'medium'}",
            "created_at": datetime.now(timezone.utc),
            "created_by": user["id"],
            "author_name": user.get("name")
        }],
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"]
    }
    await db.crm_tickets.insert_one(doc)
    return await _enrich_ticket(doc)

@crm_router.put("/tickets/{ticket_id}")
async def update_ticket(ticket_id: str, data: TicketUpdate, request: Request):
    user = await get_current_user(request)
    existing = await db.crm_tickets.find_one({"id": ticket_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Ticket not found")
    update = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update:
        raise HTTPException(status_code=400, detail="No data to update")
    activity_entries = []
    if "stage" in update and update["stage"] != existing.get("stage"):
        activity_entries.append({
            "note": f"Stage changed: {existing.get('stage', '-')} → {update['stage']}",
            "created_at": datetime.now(timezone.utc),
            "created_by": user["id"],
            "author_name": user.get("name")
        })
    if "priority" in update and update["priority"] != existing.get("priority"):
        activity_entries.append({
            "note": f"Priority changed: {existing.get('priority', '-')} → {update['priority']}",
            "created_at": datetime.now(timezone.utc),
            "created_by": user["id"],
            "author_name": user.get("name")
        })
    update["updated_at"] = datetime.now(timezone.utc)
    mongo_update = {"$set": update}
    if activity_entries:
        mongo_update["$push"] = {"activities": {"$each": activity_entries}}
    await db.crm_tickets.update_one({"id": ticket_id}, mongo_update)
    return await _enrich_ticket(await db.crm_tickets.find_one({"id": ticket_id}))

@crm_router.post("/tickets/{ticket_id}/activity")
async def add_ticket_activity(ticket_id: str, payload: dict = Body(...), request: Request = None):
    user = await get_current_user(request)
    note = (payload.get("note") or "").strip()
    if not note:
        raise HTTPException(status_code=400, detail="Activity note required")
    await db.crm_tickets.update_one(
        {"id": ticket_id},
        {"$push": {"activities": {"note": note, "created_at": datetime.now(timezone.utc), "created_by": user["id"], "author_name": user.get("name")}}}
    )
    return await _enrich_ticket(await db.crm_tickets.find_one({"id": ticket_id}))

@crm_router.delete("/tickets/{ticket_id}")
async def delete_ticket(ticket_id: str, request: Request):
    await get_current_user(request)
    res = await db.crm_tickets.delete_one({"id": ticket_id})
    if not res.deleted_count:
        raise HTTPException(status_code=404, detail="Ticket not found")
    return {"message": "Ticket deleted"}


# --------- PIPELINE CONFIGURATION (customizable stages) ---------
DEFAULT_PIPELINE_STAGES = {
    "marketing": [
        {"key": "enquiry", "label": "Enquiry", "color": "bg-[#E1EFFE] text-[#1E429F]", "order": 1},
        {"key": "quotation", "label": "Quotation", "color": "bg-[#FEF3C7] text-[#92400E]", "order": 2},
        {"key": "negotiation", "label": "Negotiation", "color": "bg-[#FCE7F3] text-[#9D174D]", "order": 3},
        {"key": "won", "label": "Won", "color": "bg-[#DEF7EC] text-[#03543F]", "order": 4},
        {"key": "lost", "label": "Lost", "color": "bg-[#FDE8E8] text-[#9B1C1C]", "order": 5},
    ],
    "support": [
        {"key": "complaint", "label": "Complaint", "color": "bg-[#E1EFFE] text-[#1E429F]", "order": 1},
        {"key": "open", "label": "Open / Assigned", "color": "bg-[#FDF6B2] text-[#723B13]", "order": 2},
        {"key": "in_progress", "label": "In Progress", "color": "bg-[#FEF3C7] text-[#92400E]", "order": 3},
        {"key": "pending", "label": "Pending", "color": "bg-[#FCE7F3] text-[#9D174D]", "order": 4},
        {"key": "closed", "label": "Closed", "color": "bg-[#DEF7EC] text-[#03543F]", "order": 5},
    ],
}

class CRMPipelineStage(BaseModel):
    key: str
    label: str
    color: Optional[str] = "bg-[#F3F4F6] text-[#4B5563]"
    order: Optional[int] = 0

class CRMPipelineConfigUpdate(BaseModel):
    stages: List[CRMPipelineStage]

@crm_router.get("/pipeline-config/{pipeline_type}")
async def get_pipeline_config(pipeline_type: str, request: Request):
    await get_current_user(request)
    if pipeline_type not in ("marketing", "support"):
        raise HTTPException(status_code=400, detail="Invalid pipeline type")
    doc = await db.crm_pipeline_configs.find_one({"pipeline_type": pipeline_type}, {"_id": 0})
    if not doc:
        return {"pipeline_type": pipeline_type, "stages": DEFAULT_PIPELINE_STAGES[pipeline_type]}
    # Sort stages by order
    stages = sorted(doc.get("stages") or [], key=lambda s: s.get("order", 0))
    return {"pipeline_type": pipeline_type, "stages": stages}


# ---- Marketing config (default Quotation T&C, lead defaults, etc.) ---------
class MarketingConfigUpdate(BaseModel):
    default_quotation_terms: Optional[str] = None
    default_quotation_notes: Optional[str] = None

@crm_router.get("/marketing-config")
async def get_marketing_config(request: Request):
    """Returns the singleton marketing config doc. Auto-seeds an empty
    default record on first read so the frontend always has a stable shape."""
    await get_current_user(request)
    doc = await db.crm_marketing_config.find_one({"_id": "singleton"}, {"_id": 0}) or {}
    return {
        "default_quotation_terms": doc.get("default_quotation_terms", ""),
        "default_quotation_notes": doc.get("default_quotation_notes", ""),
    }

@crm_router.put("/marketing-config")
async def update_marketing_config(data: MarketingConfigUpdate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin"], module="marketing_configuration", action="edit")
    payload = {k: v for k, v in data.model_dump(exclude_none=True).items()}
    if not payload:
        raise HTTPException(status_code=400, detail="No fields to update")
    payload["updated_at"] = datetime.now(timezone.utc)
    await db.crm_marketing_config.update_one(
        {"_id": "singleton"},
        {"$set": payload},
        upsert=True,
    )
    return await get_marketing_config(request)


# ---- Additional Charges master (Packing/Forwarding/Insurance/...) ---------
# These are user-defined named charges with HSN + GST% that can be selected
# on Quotations/Proformas/Tax Invoices after the global discount. The charge
# master is shared across all three CRM doc types.
class AdditionalChargeCreate(BaseModel):
    name: str
    hsn_code: Optional[str] = ""
    gst_rate: Optional[float] = 18.0
    is_active: Optional[bool] = True
    sort_order: Optional[int] = 100

class AdditionalChargeUpdate(BaseModel):
    name: Optional[str] = None
    hsn_code: Optional[str] = None
    gst_rate: Optional[float] = None
    is_active: Optional[bool] = None
    sort_order: Optional[int] = None

@crm_router.get("/additional-charges")
async def list_additional_charges(request: Request):
    await get_current_user(request)
    docs = await db.crm_additional_charges.find({}, {"_id": 0}).sort([("sort_order", 1), ("name", 1)]).to_list(500)
    return docs

@crm_router.post("/additional-charges", status_code=201)
async def create_additional_charge(data: AdditionalChargeCreate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin"], module="marketing_configuration", action="create")
    name = (data.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Name is required")
    doc = {
        "id": str(uuid.uuid4()),
        "name": name,
        "hsn_code": (data.hsn_code or "").strip(),
        "gst_rate": float(data.gst_rate or 0),
        "is_active": bool(data.is_active) if data.is_active is not None else True,
        "sort_order": int(data.sort_order or 100),
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"],
    }
    await db.crm_additional_charges.insert_one(doc)
    doc.pop("_id", None)
    return doc

@crm_router.put("/additional-charges/{cid}")
async def update_additional_charge(cid: str, data: AdditionalChargeUpdate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin"], module="marketing_configuration", action="edit")
    existing = await db.crm_additional_charges.find_one({"id": cid})
    if not existing:
        raise HTTPException(status_code=404, detail="Additional charge not found")
    update = {k: v for k, v in data.model_dump(exclude_none=True).items()}
    if "name" in update:
        update["name"] = (update["name"] or "").strip()
        if not update["name"]:
            raise HTTPException(status_code=400, detail="Name cannot be blank")
    if "hsn_code" in update:
        update["hsn_code"] = (update["hsn_code"] or "").strip()
    if "gst_rate" in update:
        update["gst_rate"] = float(update["gst_rate"] or 0)
    update["updated_at"] = datetime.now(timezone.utc)
    await db.crm_additional_charges.update_one({"id": cid}, {"$set": update})
    fresh = await db.crm_additional_charges.find_one({"id": cid}, {"_id": 0})
    return fresh

@crm_router.delete("/additional-charges/{cid}")
async def delete_additional_charge(cid: str, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin"], module="marketing_configuration", action="delete")
    res = await db.crm_additional_charges.delete_one({"id": cid})
    if not res.deleted_count:
        raise HTTPException(status_code=404, detail="Additional charge not found")
    return {"message": "Additional charge deleted"}


@crm_router.put("/pipeline-config/{pipeline_type}")
async def update_pipeline_config(pipeline_type: str, data: CRMPipelineConfigUpdate, request: Request):
    await get_current_user(request)
    if pipeline_type not in ("marketing", "support"):
        raise HTTPException(status_code=400, detail="Invalid pipeline type")
    # Validate non-empty + unique keys
    keys = [s.key for s in data.stages]
    if not keys:
        raise HTTPException(status_code=400, detail="At least one stage is required")
    if len(keys) != len(set(keys)):
        raise HTTPException(status_code=400, detail="Stage keys must be unique")
    stages = [s.model_dump() for s in data.stages]
    await db.crm_pipeline_configs.update_one(
        {"pipeline_type": pipeline_type},
        {"$set": {"pipeline_type": pipeline_type, "stages": stages, "updated_at": datetime.now(timezone.utc)}},
        upsert=True,
    )
    return {"pipeline_type": pipeline_type, "stages": stages}

@crm_router.post("/pipeline-config/{pipeline_type}/reset")
async def reset_pipeline_config(pipeline_type: str, request: Request):
    await get_current_user(request)
    if pipeline_type not in ("marketing", "support"):
        raise HTTPException(status_code=400, detail="Invalid pipeline type")
    await db.crm_pipeline_configs.delete_one({"pipeline_type": pipeline_type})
    return {"pipeline_type": pipeline_type, "stages": DEFAULT_PIPELINE_STAGES[pipeline_type]}


# --------- ACTIVITY LOG AGGREGATION ---------
@crm_router.get("/activities")
async def list_activities(request: Request, type: Optional[str] = None, limit: int = 500):
    """Aggregate activity log across all leads (type=marketing) or all tickets (type=support).
    If type is not provided, returns both streams combined & sorted desc by created_at.
    """
    await get_current_user(request)
    out = []
    if type in (None, "marketing"):
        leads = await db.crm_leads.find({}, {"_id": 0, "id": 1, "lead_no": 1, "name": 1, "customer_name": 1, "stage": 1, "activities": 1}).to_list(2000)
        for l in leads:
            for a in (l.get("activities") or []):
                out.append({
                    "source_type": "lead",
                    "entity_id": l["id"],
                    "entity_no": l.get("lead_no"),
                    "entity_title": l.get("name"),
                    "customer_name": l.get("customer_name"),
                    "stage": l.get("stage"),
                    "note": a.get("note"),
                    "author_name": a.get("author_name"),
                    "created_at": a.get("created_at"),
                })
    if type in (None, "support"):
        tickets = await db.crm_tickets.find({}, {"_id": 0, "id": 1, "ticket_no": 1, "subject": 1, "customer_id": 1, "stage": 1, "priority": 1, "activities": 1}).to_list(2000)
        cust_ids = [t.get("customer_id") for t in tickets if t.get("customer_id")]
        cust_map = {}
        if cust_ids:
            cs = await db.customers.find({"id": {"$in": cust_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(2000)
            cust_map = {c["id"]: c["name"] for c in cs}
        for t in tickets:
            for a in (t.get("activities") or []):
                out.append({
                    "source_type": "ticket",
                    "entity_id": t["id"],
                    "entity_no": t.get("ticket_no"),
                    "entity_title": t.get("subject"),
                    "customer_name": cust_map.get(t.get("customer_id"), ""),
                    "stage": t.get("stage"),
                    "priority": t.get("priority"),
                    "note": a.get("note"),
                    "author_name": a.get("author_name"),
                    "created_at": a.get("created_at"),
                })
    # Sort newest first
    def _k(x):
        v = x.get("created_at")
        if isinstance(v, datetime):
            return v.replace(tzinfo=timezone.utc) if v.tzinfo is None else v
        return datetime.min.replace(tzinfo=timezone.utc)
    out.sort(key=_k, reverse=True)
    return out[:limit]


# --------- QUOTATIONS (CRM — Enquiry → Quotation → SO) ---------
class QuotationLine(BaseModel):
    line_no: Optional[int] = None
    item_id: Optional[str] = ""   # optional; free-text only quotes allowed
    description: Optional[str] = ""
    hsn_code: Optional[str] = ""  # HSN/SAC code (printed on quotation PDF)
    quantity: float
    uom: Optional[str] = "Nos"
    rate: float
    discount_pct: Optional[float] = 0.0   # % off line amount before GST
    gst_rate: Optional[float] = 18.0
    amount: Optional[float] = 0.0  # (qty * rate) * (1 - discount/100), excl. GST

class AdditionalChargeLine(BaseModel):
    """One additional charge attached to a Quotation / PI / TI.

    `charge_id` (optional) is a soft FK to crm_additional_charges; the master
    name/hsn/gst is snapshotted onto the doc so future edits to the master
    don't retroactively change historical docs. `amount` is the resolved
    pre-GST value used in all totals math. `value_type` + `value` are the
    INPUT mode (₹ flat or % of post-line-discount subtotal) that the user
    selected — persisted so the form can re-open in the original mode.
    GST on the charge is computed at compute-time using the snapshotted
    `gst_rate`.
    """
    charge_id: Optional[str] = ""
    name: str                                  # e.g. "Packing & Forwarding"
    hsn_code: Optional[str] = ""               # printed in the totals row
    gst_rate: Optional[float] = 18.0
    value_type: Optional[str] = "amount"       # "amount" | "percent"
    value: Optional[float] = 0.0               # raw input (₹ or %)
    amount: float = 0.0                        # resolved pre-GST charge amount

class QuotationCreate(BaseModel):
    lead_id: Optional[str] = ""
    customer_id: Optional[str] = ""
    customer_name: str  # required display name
    contact_person: Optional[str] = ""
    email: Optional[str] = ""
    phone: Optional[str] = ""
    quotation_date: Optional[datetime] = None
    valid_until: Optional[datetime] = None
    lines: List[QuotationLine]
    notes: Optional[str] = ""
    terms: Optional[str] = ""
    status: Optional[str] = "draft"   # draft | sent | accepted | rejected | converted
    currency: Optional[str] = "INR"   # INR (default), USD, EUR, GBP, AED — non-INR = export (no GST)
    # Global (footer) discount applied AFTER line subtotal, BEFORE GST. Only one
    # of the two should be non-zero in practice — the UI enforces a single mode
    # per document. type="percent" interprets value as % of subtotal.
    global_discount_type: Optional[str] = "amount"   # "amount" | "percent"
    global_discount_value: Optional[float] = 0.0
    # User-defined additional charges (Packing/Forwarding/Insurance/etc.) added
    # AFTER global discount, BEFORE GST. Each charge carries its own HSN + GST%.
    additional_charges: Optional[List[AdditionalChargeLine]] = []

class QuotationUpdate(BaseModel):
    lead_id: Optional[str] = None
    customer_id: Optional[str] = None
    customer_name: Optional[str] = None
    contact_person: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    quotation_date: Optional[datetime] = None
    valid_until: Optional[datetime] = None
    lines: Optional[List[QuotationLine]] = None
    notes: Optional[str] = None
    terms: Optional[str] = None
    status: Optional[str] = None
    currency: Optional[str] = None
    global_discount_type: Optional[str] = None
    global_discount_value: Optional[float] = None
    additional_charges: Optional[List[AdditionalChargeLine]] = None

def _compute_quotation_totals(lines, global_discount_type: str = "amount", global_discount_value: float = 0.0):
    subtotal = 0.0
    total_discount = 0.0
    total_gst = 0.0
    # First pass — compute per-line net (after line-level discount) but DEFER GST
    # so that we can apply the global discount proportionally before tax.
    line_nets = []
    for idx, l in enumerate(lines, start=1):
        qty = float(l.get("quantity") or 0)
        rate = float(l.get("rate") or 0)
        discount_pct = float(l.get("discount_pct") or 0)
        gross = qty * rate
        discount = gross * discount_pct / 100.0
        net = gross - discount
        l["line_no"] = idx
        l["amount"] = round(net, 2)
        subtotal += net
        total_discount += discount
        line_nets.append((l, net, float(l.get("gst_rate") or 0)))
    # Resolve global discount → absolute amount on the post-line-discount subtotal.
    gd_type = (global_discount_type or "amount").lower()
    gd_val = float(global_discount_value or 0)
    if gd_type == "percent":
        global_discount_amt = round(subtotal * gd_val / 100.0, 2)
    else:
        global_discount_amt = round(gd_val, 2)
    # Clamp so a too-large global discount can't drive subtotal negative.
    global_discount_amt = max(0.0, min(global_discount_amt, subtotal))
    net_subtotal = subtotal - global_discount_amt
    # Spread the global discount proportionally across lines so per-line GST is
    # accurate. Each line's effective taxable = net * (net_subtotal / subtotal).
    factor = (net_subtotal / subtotal) if subtotal > 0 else 1.0
    for l, net, gst_rate in line_nets:
        taxable = net * factor
        total_gst += taxable * gst_rate / 100.0
    return {
        "subtotal": round(subtotal, 2),
        "total_discount": round(total_discount, 2),
        "global_discount_amount": round(global_discount_amt, 2),
        "net_subtotal": round(net_subtotal, 2),
        "total_gst": round(total_gst, 2),
        "grand_total": round(net_subtotal + total_gst, 2),
    }


def _normalize_additional_charges(charges):
    """Sanitise the additional_charges list, dropping rows with non-positive
    amounts. Returns the cleaned list of plain dicts (suitable for storage).
    Preserves value_type / value if the frontend supplied them so the form
    can re-open in the original ₹/% mode."""
    out = []
    for c in (charges or []):
        if hasattr(c, "model_dump"):
            c = c.model_dump()
        if not isinstance(c, dict):
            continue
        amt = float(c.get("amount") or 0)
        if amt <= 0:
            continue
        out.append({
            "charge_id": (c.get("charge_id") or "").strip(),
            "name": (c.get("name") or "").strip() or "Additional Charge",
            "hsn_code": (c.get("hsn_code") or "").strip(),
            "gst_rate": float(c.get("gst_rate") or 0),
            "value_type": (c.get("value_type") or "amount") if c.get("value_type") in ("amount", "percent") else "amount",
            "value": float(c.get("value") or amt),
            "amount": round(amt, 2),
        })
    return out


def _apply_additional_charges(totals, charges, is_inter_state, is_export=False):
    """Merge additional-charge amounts and their GST into the running totals.

    - `totals` MUST contain `net_subtotal`, `total_gst`, `grand_total` (mutated in place).
    - `charges` is a list of pre-normalised dicts (see _normalize_additional_charges).
    - `is_inter_state` decides CGST/SGST vs IGST split for the charge GST.
    - When `is_export=True`, GST on charges is forced to zero (exports are tax-free).

    Returns the per-charge GST split dict {"cgst": x, "sgst": y, "igst": z, "total_gst": w,
    "charges_total": t} so callers can fold it into the document-level GST split.
    """
    charges_total = 0.0
    charges_cgst = 0.0
    charges_sgst = 0.0
    charges_igst = 0.0
    for c in (charges or []):
        amt = float(c.get("amount") or 0)
        if amt <= 0:
            continue
        rate = 0.0 if is_export else float(c.get("gst_rate") or 0)
        tax = round(amt * rate / 100.0, 2)
        charges_total += amt
        if is_export or rate <= 0:
            pass
        elif is_inter_state:
            charges_igst += tax
        else:
            half = round(tax / 2.0, 2)
            charges_cgst += half
            charges_sgst += half
    charges_total = round(charges_total, 2)
    charges_cgst = round(charges_cgst, 2)
    charges_sgst = round(charges_sgst, 2)
    charges_igst = round(charges_igst, 2)
    charges_total_gst = round(charges_cgst + charges_sgst + charges_igst, 2)
    # Fold into the running totals.
    totals["additional_charges_total"] = charges_total
    totals["additional_charges_gst"] = charges_total_gst
    totals["total_gst"] = round(float(totals.get("total_gst") or 0) + charges_total_gst, 2)
    totals["grand_total"] = round(
        float(totals.get("net_subtotal") or 0) + charges_total + float(totals.get("total_gst") or 0),
        2,
    )
    return {
        "charges_total": charges_total,
        "cgst": charges_cgst,
        "sgst": charges_sgst,
        "igst": charges_igst,
        "total_gst": charges_total_gst,
    }


async def _lookup_creator(created_by):
    """Resolve a document's created_by id to a slim creator profile for print.

    Returns `{name, email, signature_url}` so the print template can stamp the
    document in the CREATOR's name/signature (not whoever is currently logged
    in). None if the user was deleted or the id is missing/invalid.
    """
    if not created_by:
        return None
    try:
        u = await db.users.find_one(
            {"_id": ObjectId(created_by)},
            {"_id": 0, "name": 1, "email": 1, "signature_url": 1},
        )
    except Exception:
        u = None
    if not u:
        return None
    return {
        "name": u.get("name") or u.get("email") or "",
        "email": u.get("email") or "",
        "signature_url": u.get("signature_url") or "",
    }

async def _enrich_quotation(q):
    q.pop("_id", None)
    if q.get("customer_id"):
        c = await db.customers.find_one({"id": q["customer_id"]}, {"_id": 0})
        q["customer"] = c
    if q.get("lead_id"):
        l = await db.crm_leads.find_one({"id": q["lead_id"]}, {"_id": 0, "lead_no": 1, "name": 1, "stage": 1})
        q["lead"] = l
    if q.get("converted_so_id"):
        so = await db.production_orders.find_one({"id": q["converted_so_id"]}, {"_id": 0, "order_number": 1, "status": 1})
        q["converted_so"] = so
    # Attach document creator so the printed doc signs in the creator's name
    # (not the current logged-in user's).
    q["created_by_user"] = await _lookup_creator(q.get("created_by"))
    # Hydrate item details on each line.
    # IMPORTANT: include hsn_code + gst_rate so the print template's
    # `l.item?.hsn_code` fallback resolves on legacy lines that were
    # saved before the line-level hsn_code field was added. Without
    # this, the HSN column rendered as '-' on every printed line.
    for ln in (q.get("lines") or []):
        if ln.get("item_id"):
            it = await db.items.find_one({"id": ln["item_id"]}, {"_id": 0, "part_number": 1, "name": 1, "uom": 1, "unit_of_measure": 1, "hsn_code": 1, "gst_rate": 1})
            if it:
                ln["item"] = it
    # Back-fill GST split for legacy quotations saved before _compute_gst_split was
    # wired into create/update. The print template reads doc.is_inter_state /
    # doc.cgst / doc.sgst / doc.igst directly; without this back-fill they would
    # all be zero on historical docs.
    if "is_inter_state" not in q or q.get("cgst") is None:
        try:
            split = await _compute_gst_split(q.get("customer_id") or None, q.get("lines") or [])
            q["is_inter_state"] = split["is_inter_state"]
            q["cgst"] = split["cgst"]
            q["sgst"] = split["sgst"]
            q["igst"] = split["igst"]
            q["hsn_summary"] = split["hsn_summary"]
        except Exception:
            # Non-fatal — print will just fall back to total_gst in Grand Total math.
            pass
    # Compute lock flag
    q["is_locked"] = await _quotation_is_locked(q)
    return q

@crm_router.get("/quotations")
async def list_quotations(request: Request, status: Optional[str] = None, lead_id: Optional[str] = None):
    await get_current_user(request)
    q = {}
    if status:
        q["status"] = status
    if lead_id:
        q["lead_id"] = lead_id
    docs = await db.crm_quotations.find(q).sort("created_at", -1).to_list(2000)
    return [await _enrich_quotation(d) for d in docs]

@crm_router.post("/quotations", status_code=201)
async def create_quotation(data: QuotationCreate, request: Request):
    user = await get_current_user(request)
    if not data.lines:
        raise HTTPException(status_code=400, detail="At least one line item is required")
    q_no = await _get_next_number("quotation")
    # Apply default T&C / Notes from marketing config if blank.
    mk_cfg = await db.crm_marketing_config.find_one({"_id": "singleton"}, {"_id": 0}) or {}
    if not (data.terms or "").strip() and mk_cfg.get("default_quotation_terms"):
        data.terms = mk_cfg["default_quotation_terms"]
    if not (data.notes or "").strip() and mk_cfg.get("default_quotation_notes"):
        data.notes = mk_cfg["default_quotation_notes"]
    lines = [l.model_dump() for l in data.lines]
    add_charges = _normalize_additional_charges(data.additional_charges)
    totals = _compute_quotation_totals(lines, data.global_discount_type or "amount", data.global_discount_value or 0)
    currency = (data.currency or "INR").upper()
    # Non-INR (export) → GST is not applicable. Otherwise compute CGST/SGST vs IGST.
    if currency != "INR":
        gst_split = _zero_gst_split_for_export(lines)
        # Override the quotation totals so grand_total drops GST too.
        totals["total_gst"] = 0
        totals["grand_total"] = totals["net_subtotal"]
    else:
        gst_split = await _compute_gst_split(data.customer_id or None, lines)
        # Scale CGST/SGST/IGST split to honour the global discount (the split
        # helper computed tax on raw line amounts; we proportionally reduce it
        # so split + grand_total stay consistent).
        if totals["subtotal"] > 0 and totals["net_subtotal"] != totals["subtotal"]:
            scale = totals["net_subtotal"] / totals["subtotal"]
            gst_split["cgst"] = round(gst_split["cgst"] * scale, 2)
            gst_split["sgst"] = round(gst_split["sgst"] * scale, 2)
            gst_split["igst"] = round(gst_split["igst"] * scale, 2)
            gst_split["total_gst"] = round(gst_split["cgst"] + gst_split["sgst"] + gst_split["igst"], 2)
            for hsn in gst_split.get("hsn_summary", []):
                hsn["cgst"] = round(hsn.get("cgst", 0) * scale, 2)
                hsn["sgst"] = round(hsn.get("sgst", 0) * scale, 2)
                hsn["igst"] = round(hsn.get("igst", 0) * scale, 2)
                hsn["taxable"] = round(hsn.get("taxable", 0) * scale, 2)
    # Fold additional charges (amount + their own GST) into totals + split.
    charges_split = _apply_additional_charges(
        totals,
        add_charges,
        bool(gst_split.get("is_inter_state")),
        is_export=(currency != "INR"),
    )
    gst_split["cgst"] = round(float(gst_split.get("cgst") or 0) + charges_split["cgst"], 2)
    gst_split["sgst"] = round(float(gst_split.get("sgst") or 0) + charges_split["sgst"], 2)
    gst_split["igst"] = round(float(gst_split.get("igst") or 0) + charges_split["igst"], 2)
    gst_split["total_gst"] = round(gst_split["cgst"] + gst_split["sgst"] + gst_split["igst"], 2)
    doc = {
        "id": str(uuid.uuid4()),
        "quotation_no": q_no,
        **data.model_dump(exclude={"lines", "additional_charges"}, exclude_none=False),
        "lines": lines,
        "additional_charges": add_charges,
        **totals,
        "is_inter_state": gst_split["is_inter_state"],
        "cgst": gst_split["cgst"],
        "sgst": gst_split["sgst"],
        "igst": gst_split["igst"],
        "hsn_summary": gst_split["hsn_summary"],
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"],
    }
    if not doc.get("quotation_date"):
        doc["quotation_date"] = datetime.now(timezone.utc)
    await db.crm_quotations.insert_one(doc)
    # If linked to a lead, bump lead stage to 'quotation' and add activity
    if data.lead_id:
        await db.crm_leads.update_one(
            {"id": data.lead_id},
            {
                "$set": {"stage": "quotation", "updated_at": datetime.now(timezone.utc)},
                "$push": {"activities": {
                    "note": f"Quotation {q_no} created (₹{totals['grand_total']:.2f})",
                    "created_at": datetime.now(timezone.utc),
                    "created_by": user["id"],
                    "author_name": user.get("name"),
                }},
            },
        )
    return await _enrich_quotation(doc)

async def _quotation_is_locked(q):
    """Quotation is locked for edit/delete only when its linked SO is still active.
    If the linked SO has been cancelled, the quotation becomes unlocked again.
    """
    if q.get("status") != "converted":
        return False
    so_id = q.get("converted_so_id")
    if not so_id:
        return True  # marked converted but no SO reference — still lock
    so = await db.production_orders.find_one({"id": so_id}, {"_id": 0, "status": 1})
    if not so:
        return False  # SO vanished — unlock
    return so.get("status") not in ("cancelled",)


async def _compute_quotation_balance(q):
    """For each line of a quotation, compute the qty already consumed by
    Sales Orders sourced from this quotation (excluding cancelled SOs) and
    return the per-line balance.

    Returns: list of dicts: { line_no, item_id, description, original_qty,
                              consumed_qty, balance_qty, hsn_code }
    """
    qid = q.get("id")
    lines = q.get("lines") or []
    # Pull every SO that references this quotation as its source.
    consumed_map = {}  # line_no -> consumed qty
    if qid:
        cur = db.production_orders.find(
            {"source_quotation_id": qid, "status": {"$ne": "cancelled"}},
            {"_id": 0, "lines": 1}
        )
        async for so in cur:
            for sln in so.get("lines") or []:
                sqln = sln.get("source_quotation_line_no")
                if sqln is None:
                    continue
                try:
                    sqln_int = int(sqln)
                except (TypeError, ValueError):
                    continue
                consumed_map[sqln_int] = consumed_map.get(sqln_int, 0) + int(sln.get("quantity") or 0)
    out = []
    for ln in lines:
        lno = ln.get("line_no")
        try:
            lno_int = int(lno) if lno is not None else None
        except (TypeError, ValueError):
            lno_int = None
        original_qty = float(ln.get("quantity") or 0)
        consumed = consumed_map.get(lno_int, 0) if lno_int is not None else 0
        balance = max(0.0, original_qty - consumed)
        out.append({
            "line_no": lno_int,
            "item_id": ln.get("item_id") or "",
            "description": ln.get("description") or "",
            "hsn_code": ln.get("hsn_code") or "",
            "uom": ln.get("uom") or "",
            "rate": float(ln.get("rate") or 0),
            "original_qty": original_qty,
            "consumed_qty": consumed,
            "balance_qty": balance,
        })
    return out


async def _refresh_quotation_conversion_status(qid: str, so_id: str, so_no: str):
    """Called after a new SO is created with source_quotation_id set.
    Marks the quotation 'converted' once every line's balance is zero;
    otherwise leaves status untouched but always records the first/last
    converted_so_id reference and appends to converted_so_ids[].
    """
    q = await db.crm_quotations.find_one({"id": qid})
    if not q:
        return
    balance_rows = await _compute_quotation_balance(q)
    total_balance = sum(r["balance_qty"] for r in balance_rows)
    update = {
        "updated_at": datetime.now(timezone.utc),
    }
    # Maintain a list of every SO sourced from this quotation.
    converted_ids = list(q.get("converted_so_ids") or [])
    converted_nos = list(q.get("converted_so_nos") or [])
    if so_id and so_id not in converted_ids:
        converted_ids.append(so_id)
        converted_nos.append(so_no or "")
    update["converted_so_ids"] = converted_ids
    update["converted_so_nos"] = converted_nos
    if total_balance <= 0:
        # Fully consumed — flip to converted and set the primary SO link
        # (mirrors convert_quotation_to_so for backward compat).
        update["status"] = "converted"
        update["converted_so_id"] = so_id
        update["converted_so_no"] = so_no or ""
        update["converted_at"] = datetime.now(timezone.utc)
    await db.crm_quotations.update_one({"id": qid}, {"$set": update})


@crm_router.get("/quotations/{qid}/balance")
async def get_quotation_balance(qid: str, request: Request):
    """Return per-line balance qty for a quotation (after deducting qty
    already issued via prior Sales Orders). Used by the Production page
    "From Quotation" picker to pre-fill the SO form with balance qty.
    """
    await get_current_user(request)
    q = await db.crm_quotations.find_one({"id": qid})
    if not q:
        raise HTTPException(status_code=404, detail="Quotation not found")
    balance = await _compute_quotation_balance(q)
    # Hydrate item details + active BOM id per balance row so the SO form
    # can pre-fill line.bom_id without an extra round-trip.
    for row in balance:
        if row["item_id"]:
            it = await db.items.find_one({"id": row["item_id"]}, {"_id": 0, "part_number": 1, "name": 1, "uom": 1})
            row["item"] = it
            bom = await db.boms.find_one({"parent_item_id": row["item_id"], "status": "active"}, {"_id": 0, "id": 1, "revision": 1, "name": 1})
            if not bom:
                bom = await db.boms.find_one({"parent_item_id": row["item_id"]}, {"_id": 0, "id": 1, "revision": 1, "name": 1})
            row["bom"] = bom
    customer = None
    if q.get("customer_id"):
        customer = await db.customers.find_one({"id": q["customer_id"]}, {"_id": 0, "id": 1, "name": 1})
    return {
        "id": q.get("id"),
        "quotation_no": q.get("quotation_no"),
        "status": q.get("status"),
        "customer_id": q.get("customer_id") or "",
        "customer": customer,
        "customer_name": q.get("customer_name") or (customer or {}).get("name") or "",
        "currency": q.get("currency") or "INR",
        "lines": balance,
    }


@crm_router.put("/quotations/{qid}")
async def update_quotation(qid: str, data: QuotationUpdate, request: Request):
    user = await get_current_user(request)
    existing = await db.crm_quotations.find_one({"id": qid})
    if not existing:
        raise HTTPException(status_code=404, detail="Quotation not found")
    locked = await _quotation_is_locked(existing)
    if locked:
        raise HTTPException(status_code=400, detail="Cannot edit a quotation with an active linked Sales Order. Cancel the SO first.")

    update = {k: v for k, v in data.model_dump(exclude_none=True).items()}
    # Resolve effective global discount (use update value if provided, else existing).
    eff_gd_type = update.get("global_discount_type") or existing.get("global_discount_type") or "amount"
    eff_gd_value = update.get("global_discount_value") if "global_discount_value" in update else existing.get("global_discount_value", 0)
    needs_recompute = (
        "lines" in update or "customer_id" in update or "currency" in update
        or "global_discount_type" in update or "global_discount_value" in update
        or "additional_charges" in update
    )
    if "lines" in update:
        lines = [l if isinstance(l, dict) else l.model_dump() for l in update["lines"]]
    elif needs_recompute:
        lines = existing.get("lines") or []
    else:
        lines = None
    if lines is not None:
        # Resolve effective additional charges (update value if provided, else existing).
        raw_charges = update.get("additional_charges") if "additional_charges" in update else existing.get("additional_charges", [])
        add_charges = _normalize_additional_charges(raw_charges)
        update["additional_charges"] = add_charges
        totals = _compute_quotation_totals(lines, eff_gd_type, eff_gd_value or 0)
        update["lines"] = lines
        cust_id = update.get("customer_id") if "customer_id" in update else existing.get("customer_id")
        currency = (update.get("currency") or existing.get("currency") or "INR").upper()
        if currency != "INR":
            gst_split = _zero_gst_split_for_export(lines)
            totals["total_gst"] = 0
            totals["grand_total"] = totals["net_subtotal"]
        else:
            gst_split = await _compute_gst_split(cust_id or None, lines)
            # Scale CGST/SGST/IGST split for the global discount (mirror the create path).
            if totals["subtotal"] > 0 and totals["net_subtotal"] != totals["subtotal"]:
                scale = totals["net_subtotal"] / totals["subtotal"]
                gst_split["cgst"] = round(gst_split["cgst"] * scale, 2)
                gst_split["sgst"] = round(gst_split["sgst"] * scale, 2)
                gst_split["igst"] = round(gst_split["igst"] * scale, 2)
                gst_split["total_gst"] = round(gst_split["cgst"] + gst_split["sgst"] + gst_split["igst"], 2)
                for hsn in gst_split.get("hsn_summary", []):
                    hsn["cgst"] = round(hsn.get("cgst", 0) * scale, 2)
                    hsn["sgst"] = round(hsn.get("sgst", 0) * scale, 2)
                    hsn["igst"] = round(hsn.get("igst", 0) * scale, 2)
                    hsn["taxable"] = round(hsn.get("taxable", 0) * scale, 2)
        # Fold additional charges into totals + split.
        charges_split = _apply_additional_charges(
            totals,
            add_charges,
            bool(gst_split.get("is_inter_state")),
            is_export=(currency != "INR"),
        )
        gst_split["cgst"] = round(float(gst_split.get("cgst") or 0) + charges_split["cgst"], 2)
        gst_split["sgst"] = round(float(gst_split.get("sgst") or 0) + charges_split["sgst"], 2)
        gst_split["igst"] = round(float(gst_split.get("igst") or 0) + charges_split["igst"], 2)
        gst_split["total_gst"] = round(gst_split["cgst"] + gst_split["sgst"] + gst_split["igst"], 2)
        update.update(totals)
        update["is_inter_state"] = gst_split["is_inter_state"]
        update["cgst"] = gst_split["cgst"]
        update["sgst"] = gst_split["sgst"]
        update["igst"] = gst_split["igst"]
        update["hsn_summary"] = gst_split["hsn_summary"]

    new_status = update.get("status")
    # Auto-convert to SO when status transitions to 'accepted'
    auto_convert = (
        new_status == "accepted"
        and existing.get("status") != "accepted"
        and not existing.get("converted_so_id")
    )

    # If quotation was previously converted and SO was cancelled, resetting to draft/sent allows re-use
    if existing.get("status") == "converted" and not locked and new_status in (None, "draft", "sent"):
        # Clear the converted link so the quotation becomes freshly editable
        update.setdefault("converted_so_id", "")
        update.setdefault("converted_so_no", "")

    update["updated_at"] = datetime.now(timezone.utc)
    await db.crm_quotations.update_one({"id": qid}, {"$set": update})

    if auto_convert:
        # Fire the same convert flow used by the explicit action
        try:
            await convert_quotation_to_so(qid, payload={}, request=request)
        except HTTPException as he:
            # Roll back the status change so the user sees the failure
            await db.crm_quotations.update_one({"id": qid}, {"$set": {"status": existing.get("status") or "draft"}})
            raise he

    refreshed = await db.crm_quotations.find_one({"id": qid})
    return await _enrich_quotation(refreshed)

@crm_router.delete("/quotations/{qid}")
async def delete_quotation(qid: str, request: Request):
    await get_current_user(request)
    existing = await db.crm_quotations.find_one({"id": qid})
    if not existing:
        raise HTTPException(status_code=404, detail="Quotation not found")
    if await _quotation_is_locked(existing):
        raise HTTPException(status_code=400, detail="Cannot delete a quotation with an active linked Sales Order. Cancel the SO first.")
    await db.crm_quotations.delete_one({"id": qid})
    return {"message": "Quotation deleted"}


@crm_router.post("/quotations/{qid}/revise", status_code=201)
async def revise_quotation(qid: str, request: Request):
    """Clone a quotation as a new revision.

    The original becomes read-only (status=superseded). The new quotation:
      - shares the same `root_quotation_no` (defaulting to the original's quotation_no)
      - keeps the same base prefix and appends `-R<n>` to the number (e.g. QUO-000001 → QUO-000001-R1)
      - starts as status=draft so the salesperson can edit it
    """
    user = await get_current_user(request)
    original = await db.crm_quotations.find_one({"id": qid})
    if not original:
        raise HTTPException(status_code=404, detail="Quotation not found")

    root_no = original.get("root_quotation_no") or original.get("quotation_no")
    # Count existing revisions sharing this root and bump.
    existing_revs = await db.crm_quotations.count_documents({"root_quotation_no": root_no}) if root_no else 0
    # If root has never been set, the original counts as rev 0 — so first revision is 1.
    new_revision = max(int(original.get("revision") or 0) + 1, existing_revs + 1)
    new_quotation_no = f"{root_no}-R{new_revision}"

    clone = {k: v for k, v in original.items() if k != "_id"}
    clone.update({
        "id": str(uuid.uuid4()),
        "quotation_no": new_quotation_no,
        "root_quotation_no": root_no,
        "revision": new_revision,
        "previous_revision_id": qid,
        "status": "draft",
        "converted_so_id": "",
        "converted_so_no": "",
        "proforma_id": "",
        "proforma_no": "",
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"],
    })
    await db.crm_quotations.insert_one(clone)
    # Mark original with root + link + bump parent status to 'superseded' if not accepted/converted
    update_original = {"root_quotation_no": root_no, "superseded_by_id": clone["id"], "superseded_by_no": new_quotation_no, "updated_at": datetime.now(timezone.utc)}
    if original.get("status") not in ("accepted", "converted"):
        update_original["status"] = "superseded"
    await db.crm_quotations.update_one({"id": qid}, {"$set": update_original})
    return await _enrich_quotation(clone)

@crm_router.post("/quotations/{qid}/convert-to-so")
async def convert_quotation_to_so(qid: str, payload: dict = Body(default={}), request: Request = None):
    """Convert a quotation into a multi-line Sales Order (Production Order).
    Each quotation line must reference an item that has an active BOM.
    Payload can override order_type per line: {line_order_types: {line_no: 'auto|mts|mto'}, due_date: ISO}.
    """
    user = await get_current_user(request)
    q = await db.crm_quotations.find_one({"id": qid})
    if not q:
        raise HTTPException(status_code=404, detail="Quotation not found")
    if q.get("status") == "converted":
        raise HTTPException(status_code=400, detail="Quotation already converted", )
    lines = q.get("lines") or []
    if not lines:
        raise HTTPException(status_code=400, detail="Quotation has no lines to convert")
    line_order_types = payload.get("line_order_types") or {}
    default_order_type = payload.get("order_type") or "mts"
    due_date_raw = payload.get("due_date")
    due_date = None
    if due_date_raw:
        try:
            due_date = datetime.fromisoformat(str(due_date_raw).replace("Z", "+00:00"))
        except Exception:
            due_date = None

    # Resolve each line's BOM
    so_lines = []
    for ln in lines:
        item_id = ln.get("item_id")
        if not item_id:
            raise HTTPException(status_code=400, detail=f"Line {ln.get('line_no')}: no item selected — cannot convert to SO. Attach an item (with BOM) to each line.")
        bom = await db.boms.find_one({"parent_item_id": item_id, "status": "active"}, {"_id": 0, "id": 1})
        if not bom:
            # fallback to any BOM for that item
            bom = await db.boms.find_one({"parent_item_id": item_id}, {"_id": 0, "id": 1})
        if not bom:
            it = await db.items.find_one({"id": item_id}, {"_id": 0, "part_number": 1, "name": 1})
            pn = (it or {}).get("part_number", item_id)
            raise HTTPException(status_code=400, detail=f"No BOM found for item {pn}. Create a BOM before converting.")
        lno = ln.get("line_no")
        otype = line_order_types.get(str(lno)) or line_order_types.get(lno) or default_order_type
        so_lines.append({
            "line_id": str(uuid.uuid4()),
            "line_no": lno,
            "bom_id": bom["id"],
            "quantity": int(ln.get("quantity") or 0),
            "due_date": due_date,
            "order_type": otype,
            "notes": ln.get("description") or "",
            "reserved_qty": 0,
            "mo_qty": 0,
            "status": "draft",
        })
    total_qty = sum(l["quantity"] for l in so_lines)
    # Build the SO doc — mirrors existing create_production_order shape for compat
    count = await db.production_orders.count_documents({})
    order_number = f"SO-{str(count + 1).zfill(6)}"
    so_doc = {
        "id": str(uuid.uuid4()),
        "order_number": order_number,
        "lines": so_lines,
        "customer_id": q.get("customer_id") or "",
        "bom_id": so_lines[0]["bom_id"],
        "quantity": total_qty,
        "due_date": due_date,
        "priority": "medium",
        "notes": f"Generated from Quotation {q['quotation_no']}",
        "status": "draft",
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"],
        "source_quotation_id": qid,
        "source_quotation_no": q["quotation_no"],
    }
    await db.production_orders.insert_one(so_doc)
    so_doc.pop("_id", None)
    # Mark quotation as converted
    await db.crm_quotations.update_one(
        {"id": qid},
        {"$set": {
            "status": "converted",
            "converted_so_id": so_doc["id"],
            "converted_so_no": order_number,
            "converted_at": datetime.now(timezone.utc),
            "updated_at": datetime.now(timezone.utc),
        }},
    )
    # If linked lead exists, mark it won (quotation accepted path)
    if q.get("lead_id"):
        await db.crm_leads.update_one(
            {"id": q["lead_id"]},
            {
                "$set": {"stage": "won", "updated_at": datetime.now(timezone.utc)},
                "$push": {"activities": {
                    "note": f"Quotation {q['quotation_no']} converted to SO {order_number}",
                    "created_at": datetime.now(timezone.utc),
                    "created_by": user["id"],
                    "author_name": user.get("name"),
                }},
            },
        )
    return {"quotation": await _enrich_quotation(await db.crm_quotations.find_one({"id": qid})), "sales_order": so_doc}


# --------- NUMBER SERIES CONFIGURATION ---------
# Local alias used by CRM/Sales functions only. The Settings-page endpoints at
# line ~6086 use MASTER_NUMBER_SERIES + CRM_DEFAULT_NUMBER_SERIES (merged view).
_CRM_SERIES_LOCAL = CRM_DEFAULT_NUMBER_SERIES
NUMBER_SERIES_TYPES = list(_CRM_SERIES_LOCAL.keys())

async def _get_next_number(doc_type: str) -> str:
    """Atomic-ish next-number generation. Returns formatted number e.g. 'PI-000001'."""
    doc = await db.number_series.find_one({"doc_type": doc_type}, {"_id": 0})
    if not doc:
        doc = {"doc_type": doc_type, **_CRM_SERIES_LOCAL.get(doc_type, {"prefix": f"{doc_type.upper()}-", "padding": 6, "next_number": 1, "reset_yearly": False})}
        await db.number_series.insert_one(doc)
    prefix = doc.get("prefix") or f"{doc_type.upper()}"
    padding = int(doc.get("padding") or 6)
    next_num = int(doc.get("next_number") or 1)
    year_part = ""
    if doc.get("reset_yearly"):
        # Indian FY (Apr-Mar). Compact format '2627' (last-2 of start + end years)
        today = datetime.now(timezone.utc)
        fy_start = today.year if today.month >= 4 else today.year - 1
        fy_str = f"{str(fy_start)[-2:]}{str(fy_start + 1)[-2:]}"
        stored_fy = doc.get("current_fy")
        if stored_fy and stored_fy != fy_str:
            next_num = 1
            await db.number_series.update_one({"doc_type": doc_type}, {"$set": {"current_fy": fy_str, "next_number": 1}})
        elif not stored_fy:
            # First activation of reset_yearly — stamp the FY, keep user-configured next_number
            await db.number_series.update_one({"doc_type": doc_type}, {"$set": {"current_fy": fy_str}})
        year_part = fy_str
    await db.number_series.update_one({"doc_type": doc_type}, {"$set": {"next_number": next_num + 1}})
    return f"{prefix}{year_part}{str(next_num).zfill(padding)}"

@crm_router.get("/number-series")
async def list_number_series(request: Request):
    await get_current_user(request)
    out = []
    for t in NUMBER_SERIES_TYPES:
        doc = await db.number_series.find_one({"doc_type": t}, {"_id": 0})
        if not doc:
            doc = {"doc_type": t, **_CRM_SERIES_LOCAL[t]}
        out.append(doc)
    return out

@crm_router.put("/number-series/{doc_type}")
async def update_crm_number_series(doc_type: str, data: NumberSeriesUpdate, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    if doc_type not in NUMBER_SERIES_TYPES:
        raise HTTPException(status_code=400, detail=f"Unknown doc_type. Allowed: {NUMBER_SERIES_TYPES}")
    update = {k: v for k, v in data.model_dump(exclude_none=True).items()}
    if not update:
        raise HTTPException(status_code=400, detail="Nothing to update")
    update["doc_type"] = doc_type
    await db.number_series.update_one({"doc_type": doc_type}, {"$set": update}, upsert=True)
    return await db.number_series.find_one({"doc_type": doc_type}, {"_id": 0})


# --------- GST SPLIT HELPER ---------
async def _build_upi_qr_payload(amount: float, invoice_no: str) -> str:
    """Build a UPI payment QR URI from company settings. Falls back gracefully when UPI not configured."""
    company = await db.company_settings.find_one({"type": "company"}, {"_id": 0}) or {}
    upi_id = (company.get("bank_upi") or "").strip() or "na@upi"
    payee_name = (company.get("company_name") or "Company").strip().replace(" ", "")[:30] or "Company"
    # Escape & avoid breaking query string
    safe_tn = invoice_no.replace("/", "-")
    return f"upi://pay?pa={upi_id}&pn={payee_name}&am={amount:.2f}&tn={safe_tn}&cu=INR"


async def _compute_gst_split(customer_id: Optional[str], lines: List[dict], place_of_supply: Optional[str] = None) -> dict:
    """Compute CGST/SGST/IGST split based on company-state vs POS (falls back to customer-state).
    `place_of_supply` (state code, e.g. "27") takes precedence over the customer master state.
    Returns: {is_inter_state, cgst, sgst, igst, total_gst, hsn_summary[]}
    """
    company = await db.company_settings.find_one({"type": "company"}, {"_id": 0}) or {}
    company_state = (company.get("state_code") or "").strip()
    customer_state = ""
    if customer_id:
        c = await db.customers.find_one({"id": customer_id}, {"_id": 0, "state_code": 1, "state": 1}) or {}
        customer_state = (c.get("state_code") or "").strip()
    # POS override: if user explicitly set place_of_supply, use that for tax determination
    effective_state = (place_of_supply or "").strip() or customer_state
    is_inter_state = bool(company_state and effective_state and company_state != effective_state)
    cgst = 0.0
    sgst = 0.0
    igst = 0.0
    # Aggregate HSN-wise for tax breakup
    hsn_bucket = {}  # key: (hsn, rate) -> {taxable, cgst, sgst, igst}
    for ln in lines:
        taxable = float(ln.get("amount") or 0)  # already net after discount
        rate = float(ln.get("gst_rate") or 0)
        tax_amt = taxable * rate / 100.0
        ln["taxable_value"] = round(taxable, 2)
        if is_inter_state:
            ln["igst_rate"] = rate
            ln["igst_amt"] = round(tax_amt, 2)
            ln["cgst_rate"] = 0
            ln["cgst_amt"] = 0
            ln["sgst_rate"] = 0
            ln["sgst_amt"] = 0
            igst += tax_amt
        else:
            half = round(tax_amt / 2.0, 2)
            ln["igst_rate"] = 0
            ln["igst_amt"] = 0
            ln["cgst_rate"] = rate / 2.0
            ln["cgst_amt"] = half
            ln["sgst_rate"] = rate / 2.0
            ln["sgst_amt"] = half
            cgst += half
            sgst += half
        hsn = (ln.get("hsn_code") or "").strip() or "-"
        key = (hsn, rate)
        if key not in hsn_bucket:
            hsn_bucket[key] = {"hsn": hsn, "rate": rate, "taxable": 0.0, "cgst": 0.0, "sgst": 0.0, "igst": 0.0}
        hsn_bucket[key]["taxable"] += taxable
        if is_inter_state:
            hsn_bucket[key]["igst"] += tax_amt
        else:
            hsn_bucket[key]["cgst"] += tax_amt / 2.0
            hsn_bucket[key]["sgst"] += tax_amt / 2.0
    hsn_summary = [{
        "hsn": v["hsn"], "rate": v["rate"],
        "taxable": round(v["taxable"], 2),
        "cgst": round(v["cgst"], 2),
        "sgst": round(v["sgst"], 2),
        "igst": round(v["igst"], 2),
    } for v in hsn_bucket.values()]
    return {
        "is_inter_state": is_inter_state,
        "company_state": company_state,
        "customer_state": customer_state,
        "cgst": round(cgst, 2),
        "sgst": round(sgst, 2),
        "igst": round(igst, 2),
        "total_gst": round(cgst + sgst + igst, 2),
        "hsn_summary": hsn_summary,
    }


def _zero_gst_split_for_export(lines: List[dict]) -> dict:
    """For non-INR (export/import) documents, GST is not applicable. Zero out the
    per-line tax fields and return a zero-split aggregate with HSN bucket carrying
    only taxable amounts (rate=0)."""
    hsn_bucket = {}
    for ln in lines:
        taxable = float(ln.get("amount") or 0)
        ln["taxable_value"] = round(taxable, 2)
        ln["gst_rate"] = 0
        ln["igst_rate"] = 0
        ln["igst_amt"] = 0
        ln["cgst_rate"] = 0
        ln["cgst_amt"] = 0
        ln["sgst_rate"] = 0
        ln["sgst_amt"] = 0
        hsn = (ln.get("hsn_code") or "").strip() or "-"
        key = (hsn, 0)
        if key not in hsn_bucket:
            hsn_bucket[key] = {"hsn": hsn, "rate": 0, "taxable": 0.0, "cgst": 0.0, "sgst": 0.0, "igst": 0.0}
        hsn_bucket[key]["taxable"] += taxable
    hsn_summary = [{
        "hsn": v["hsn"], "rate": v["rate"],
        "taxable": round(v["taxable"], 2),
        "cgst": 0, "sgst": 0, "igst": 0,
    } for v in hsn_bucket.values()]
    return {
        "is_inter_state": False,
        "company_state": "",
        "customer_state": "",
        "cgst": 0,
        "sgst": 0,
        "igst": 0,
        "total_gst": 0,
        "hsn_summary": hsn_summary,
    }


# --------- PROFORMA INVOICE ---------
class ProformaLine(BaseModel):
    line_no: Optional[int] = None
    item_id: Optional[str] = ""
    description: Optional[str] = ""
    hsn_code: Optional[str] = ""
    quantity: float
    uom: Optional[str] = "Nos"
    rate: float
    discount_pct: Optional[float] = 0.0
    gst_rate: Optional[float] = 18.0
    amount: Optional[float] = 0.0

class ProformaCreate(BaseModel):
    quotation_id: Optional[str] = ""
    customer_id: Optional[str] = ""
    customer_name: str
    contact_person: Optional[str] = ""
    email: Optional[str] = ""
    phone: Optional[str] = ""
    billing_address: Optional[str] = ""
    shipping_address: Optional[str] = ""
    proforma_date: Optional[datetime] = None
    valid_until: Optional[datetime] = None
    advance_percentage: Optional[float] = 0.0
    lines: List[ProformaLine]
    notes: Optional[str] = ""
    terms: Optional[str] = ""
    status: Optional[str] = "draft"  # draft / sent / paid / cancelled / converted
    currency: Optional[str] = "INR"  # INR (default), USD, EUR, GBP, AED — non-INR = export (no GST)
    additional_charges: Optional[List[AdditionalChargeLine]] = []

class ProformaUpdate(BaseModel):
    status: Optional[str] = None
    lines: Optional[List[ProformaLine]] = None
    notes: Optional[str] = None
    terms: Optional[str] = None
    advance_percentage: Optional[float] = None
    valid_until: Optional[datetime] = None
    billing_address: Optional[str] = None
    shipping_address: Optional[str] = None
    customer_name: Optional[str] = None
    contact_person: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    currency: Optional[str] = None
    additional_charges: Optional[List[AdditionalChargeLine]] = None

async def _finalize_invoice_lines(lines: List[dict]) -> dict:
    """Recompute per-line amount (net after discount) and aggregate totals.
    Returns dict with subtotal, total_discount, grand_total (pre-GST).
    """
    subtotal = 0.0
    total_discount = 0.0
    for idx, l in enumerate(lines, start=1):
        qty = float(l.get("quantity") or 0)
        rate = float(l.get("rate") or 0)
        disc_pct = float(l.get("discount_pct") or 0)
        gross = qty * rate
        disc = gross * disc_pct / 100.0
        net = gross - disc
        l["line_no"] = idx
        l["amount"] = round(net, 2)
        subtotal += net
        total_discount += disc
    return {"subtotal": round(subtotal, 2), "total_discount": round(total_discount, 2)}

async def _enrich_proforma(p):
    p.pop("_id", None)
    if p.get("customer_id"):
        c = await db.customers.find_one({"id": p["customer_id"]}, {"_id": 0})
        p["customer"] = c
    if p.get("quotation_id"):
        q = await db.crm_quotations.find_one({"id": p["quotation_id"]}, {"_id": 0, "quotation_no": 1})
        p["quotation"] = q
    if p.get("converted_tax_invoice_id"):
        ti = await db.tax_invoices.find_one({"id": p["converted_tax_invoice_id"]}, {"_id": 0, "invoice_no": 1, "status": 1})
        p["tax_invoice"] = ti
    for ln in (p.get("lines") or []):
        if ln.get("item_id"):
            it = await db.items.find_one({"id": ln["item_id"]}, {"_id": 0, "part_number": 1, "name": 1, "uom": 1, "unit_of_measure": 1, "hsn_code": 1, "gst_rate": 1})
            if it:
                ln["item"] = it
    # Attach document creator for signature stamping on the printed PI.
    p["created_by_user"] = await _lookup_creator(p.get("created_by"))
    return p

@crm_router.get("/proformas")
async def list_proformas(request: Request, status: Optional[str] = None):
    await get_current_user(request)
    q = {}
    if status:
        q["status"] = status
    docs = await db.proforma_invoices.find(q).sort("created_at", -1).to_list(2000)
    return [await _enrich_proforma(d) for d in docs]

@crm_router.post("/proformas", status_code=201)
async def create_proforma(data: ProformaCreate, request: Request):
    user = await get_current_user(request)
    if not data.lines:
        raise HTTPException(status_code=400, detail="At least one line is required")
    proforma_no = await _get_next_number("proforma")
    lines = [l.model_dump() for l in data.lines]
    add_charges = _normalize_additional_charges(data.additional_charges)
    base = await _finalize_invoice_lines(lines)
    currency = (data.currency or "INR").upper()
    if currency != "INR":
        gst = _zero_gst_split_for_export(lines)
    else:
        gst = await _compute_gst_split(data.customer_id, lines)
    # Fold additional charges + their GST.
    totals = {"net_subtotal": base["subtotal"], "total_gst": gst["total_gst"], "grand_total": 0.0}
    charges_split = _apply_additional_charges(
        totals, add_charges, bool(gst.get("is_inter_state")), is_export=(currency != "INR"),
    )
    gst["cgst"] = round(float(gst.get("cgst") or 0) + charges_split["cgst"], 2)
    gst["sgst"] = round(float(gst.get("sgst") or 0) + charges_split["sgst"], 2)
    gst["igst"] = round(float(gst.get("igst") or 0) + charges_split["igst"], 2)
    gst["total_gst"] = round(gst["cgst"] + gst["sgst"] + gst["igst"], 2)
    doc = {
        "id": str(uuid.uuid4()),
        "proforma_no": proforma_no,
        **data.model_dump(exclude={"lines", "additional_charges"}, exclude_none=False),
        "lines": lines,
        "additional_charges": add_charges,
        **base,
        "additional_charges_total": totals.get("additional_charges_total", 0),
        "additional_charges_gst": totals.get("additional_charges_gst", 0),
        "cgst": gst["cgst"], "sgst": gst["sgst"], "igst": gst["igst"],
        "total_gst": gst["total_gst"],
        "is_inter_state": gst["is_inter_state"],
        "hsn_summary": gst["hsn_summary"],
        "grand_total": totals["grand_total"],
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"],
    }
    if not doc.get("proforma_date"):
        doc["proforma_date"] = datetime.now(timezone.utc)
    await db.proforma_invoices.insert_one(doc)
    return await _enrich_proforma(doc)

@crm_router.put("/proformas/{pid}")
async def update_proforma(pid: str, data: ProformaUpdate, request: Request):
    await get_current_user(request)
    existing = await db.proforma_invoices.find_one({"id": pid})
    if not existing:
        raise HTTPException(status_code=404, detail="Proforma not found")
    if existing.get("status") == "converted":
        raise HTTPException(status_code=400, detail="Cannot edit a Proforma that has been converted to Tax Invoice")
    update = {k: v for k, v in data.model_dump(exclude_none=True).items()}
    need_recompute = ("lines" in update) or ("currency" in update) or ("additional_charges" in update)
    if need_recompute:
        if "lines" in update:
            lines = [l if isinstance(l, dict) else l.model_dump() for l in update["lines"]]
        else:
            lines = existing.get("lines") or []
        raw_charges = update.get("additional_charges") if "additional_charges" in update else existing.get("additional_charges", [])
        add_charges = _normalize_additional_charges(raw_charges)
        update["additional_charges"] = add_charges
        base = await _finalize_invoice_lines(lines)
        currency = (update.get("currency") or existing.get("currency") or "INR").upper()
        if currency != "INR":
            gst = _zero_gst_split_for_export(lines)
        else:
            gst = await _compute_gst_split(existing.get("customer_id"), lines)
        totals = {"net_subtotal": base["subtotal"], "total_gst": gst["total_gst"], "grand_total": 0.0}
        charges_split = _apply_additional_charges(
            totals, add_charges, bool(gst.get("is_inter_state")), is_export=(currency != "INR"),
        )
        gst["cgst"] = round(float(gst.get("cgst") or 0) + charges_split["cgst"], 2)
        gst["sgst"] = round(float(gst.get("sgst") or 0) + charges_split["sgst"], 2)
        gst["igst"] = round(float(gst.get("igst") or 0) + charges_split["igst"], 2)
        gst["total_gst"] = round(gst["cgst"] + gst["sgst"] + gst["igst"], 2)
        update["lines"] = lines
        update.update(base)
        update.update({
            "additional_charges_total": totals.get("additional_charges_total", 0),
            "additional_charges_gst": totals.get("additional_charges_gst", 0),
            "cgst": gst["cgst"], "sgst": gst["sgst"], "igst": gst["igst"],
            "total_gst": gst["total_gst"], "is_inter_state": gst["is_inter_state"],
            "hsn_summary": gst["hsn_summary"],
            "grand_total": totals["grand_total"],
        })
    update["updated_at"] = datetime.now(timezone.utc)
    await db.proforma_invoices.update_one({"id": pid}, {"$set": update})
    return await _enrich_proforma(await db.proforma_invoices.find_one({"id": pid}))

@crm_router.delete("/proformas/{pid}")
async def delete_proforma(pid: str, request: Request):
    await get_current_user(request)
    existing = await db.proforma_invoices.find_one({"id": pid})
    if not existing:
        raise HTTPException(status_code=404, detail="Proforma not found")
    if existing.get("status") == "converted":
        raise HTTPException(status_code=400, detail="Cannot delete a Proforma that has been converted to Tax Invoice")
    await db.proforma_invoices.delete_one({"id": pid})
    return {"message": "Proforma deleted"}

@crm_router.post("/quotations/{qid}/convert-to-proforma")
async def convert_quotation_to_proforma(qid: str, payload: dict = Body(default={}), request: Request = None):
    user = await get_current_user(request)
    q = await db.crm_quotations.find_one({"id": qid})
    if not q:
        raise HTTPException(status_code=404, detail="Quotation not found")
    if not q.get("lines"):
        raise HTTPException(status_code=400, detail="Quotation has no line items")
    # Build proforma payload from quotation
    pf_lines = []
    for ln in q["lines"]:
        pf_lines.append({
            "item_id": ln.get("item_id", ""),
            "description": ln.get("description", ""),
            "hsn_code": "",  # carried by item lookup later if needed
            "quantity": ln.get("quantity", 0),
            "uom": ln.get("uom", "Nos"),
            "rate": ln.get("rate", 0),
            "discount_pct": ln.get("discount_pct", 0),
            "gst_rate": ln.get("gst_rate", 18),
        })
    # hydrate hsn_code from items master
    for ln in pf_lines:
        if ln["item_id"]:
            it = await db.items.find_one({"id": ln["item_id"]}, {"_id": 0, "hsn_code": 1})
            if it and it.get("hsn_code"):
                ln["hsn_code"] = it["hsn_code"]
    proforma_no = await _get_next_number("proforma")
    base = await _finalize_invoice_lines(pf_lines)
    # Carry currency from the source quotation; non-INR ⇒ no GST.
    q_currency = (q.get("currency") or "INR").upper()
    if q_currency != "INR":
        gst = _zero_gst_split_for_export(pf_lines)
    else:
        gst = await _compute_gst_split(q.get("customer_id"), pf_lines)
    # Pull billing/shipping address from customer
    billing = shipping = ""
    if q.get("customer_id"):
        c = await db.customers.find_one({"id": q["customer_id"]}, {"_id": 0})
        if c:
            billing = c.get("address", "") or ""
            shipping = c.get("address", "") or ""
    # Carry additional charges from the source quotation so PI inherits
    # Packing/Forwarding/Insurance/etc. without forcing the user to re-enter.
    add_charges = _normalize_additional_charges(q.get("additional_charges") or [])
    totals = {"net_subtotal": base["subtotal"], "total_gst": gst["total_gst"], "grand_total": 0.0}
    charges_split = _apply_additional_charges(
        totals, add_charges, bool(gst.get("is_inter_state")), is_export=(q_currency != "INR"),
    )
    gst["cgst"] = round(float(gst.get("cgst") or 0) + charges_split["cgst"], 2)
    gst["sgst"] = round(float(gst.get("sgst") or 0) + charges_split["sgst"], 2)
    gst["igst"] = round(float(gst.get("igst") or 0) + charges_split["igst"], 2)
    gst["total_gst"] = round(gst["cgst"] + gst["sgst"] + gst["igst"], 2)
    doc = {
        "id": str(uuid.uuid4()),
        "proforma_no": proforma_no,
        "quotation_id": qid,
        "customer_id": q.get("customer_id", ""),
        "customer_name": q.get("customer_name", ""),
        "contact_person": q.get("contact_person", ""),
        "email": q.get("email", ""),
        "phone": q.get("phone", ""),
        "billing_address": billing,
        "shipping_address": shipping,
        "proforma_date": datetime.now(timezone.utc),
        "valid_until": q.get("valid_until"),
        "advance_percentage": float(payload.get("advance_percentage") or 0),
        "lines": pf_lines,
        "additional_charges": add_charges,
        "notes": q.get("notes", ""),
        "terms": q.get("terms", ""),
        **base,
        "additional_charges_total": totals.get("additional_charges_total", 0),
        "additional_charges_gst": totals.get("additional_charges_gst", 0),
        "cgst": gst["cgst"], "sgst": gst["sgst"], "igst": gst["igst"],
        "total_gst": gst["total_gst"],
        "is_inter_state": gst["is_inter_state"],
        "hsn_summary": gst["hsn_summary"],
        "grand_total": totals["grand_total"],
        "currency": q_currency,
        "status": "draft",
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"],
    }
    await db.proforma_invoices.insert_one(doc)
    # Flag the quotation so it shows the link
    await db.crm_quotations.update_one(
        {"id": qid},
        {"$set": {"proforma_id": doc["id"], "proforma_no": proforma_no, "updated_at": datetime.now(timezone.utc)}}
    )
    return await _enrich_proforma(doc)


# --------- TAX INVOICE ---------
class TaxInvoiceLine(BaseModel):
    line_no: Optional[int] = None
    item_id: Optional[str] = ""
    description: Optional[str] = ""
    hsn_code: Optional[str] = ""
    quantity: float
    uom: Optional[str] = "Nos"
    rate: float
    discount_pct: Optional[float] = 0.0
    gst_rate: Optional[float] = 18.0
    amount: Optional[float] = 0.0

class TaxInvoiceCreate(BaseModel):
    proforma_id: Optional[str] = ""
    sales_order_id: Optional[str] = ""
    customer_po_number: Optional[str] = ""  # Customer's PO reference — used when TI is created without SO
    customer_id: Optional[str] = ""
    customer_name: str
    contact_person: Optional[str] = ""
    email: Optional[str] = ""
    phone: Optional[str] = ""
    billing_address: Optional[str] = ""
    shipping_address: Optional[str] = ""
    invoice_date: Optional[datetime] = None
    due_date: Optional[datetime] = None
    place_of_supply: Optional[str] = ""  # state code
    lines: List[TaxInvoiceLine]
    notes: Optional[str] = ""
    terms: Optional[str] = ""
    status: Optional[str] = "draft"  # draft / issued / paid / cancelled
    currency: Optional[str] = "INR"  # INR (default), USD, EUR, GBP, AED — non-INR = export (no GST)
    ship_from_warehouse_id: Optional[str] = ""  # Source store — stock is decremented from this warehouse on save
    additional_charges: Optional[List[AdditionalChargeLine]] = []

class TaxInvoiceUpdate(BaseModel):
    status: Optional[str] = None
    lines: Optional[List[TaxInvoiceLine]] = None
    notes: Optional[str] = None
    terms: Optional[str] = None
    due_date: Optional[datetime] = None
    place_of_supply: Optional[str] = None
    billing_address: Optional[str] = None
    shipping_address: Optional[str] = None
    customer_po_number: Optional[str] = None
    currency: Optional[str] = None
    ship_from_warehouse_id: Optional[str] = None
    additional_charges: Optional[List[AdditionalChargeLine]] = None

async def _enrich_tax_invoice(t):
    t.pop("_id", None)
    if t.get("customer_id"):
        c = await db.customers.find_one({"id": t["customer_id"]}, {"_id": 0})
        t["customer"] = c
    if t.get("proforma_id"):
        p = await db.proforma_invoices.find_one({"id": t["proforma_id"]}, {"_id": 0, "proforma_no": 1})
        t["proforma"] = p
    if t.get("sales_order_id"):
        so = await db.production_orders.find_one({"id": t["sales_order_id"]}, {"_id": 0, "order_number": 1})
        if so:
            t["sales_order"] = so
    if t.get("ship_from_warehouse_id"):
        w = await db.warehouses.find_one({"id": t["ship_from_warehouse_id"]}, {"_id": 0, "name": 1, "code": 1})
        if w:
            t["ship_from_warehouse"] = w
    for ln in (t.get("lines") or []):
        if ln.get("item_id"):
            it = await db.items.find_one({"id": ln["item_id"]}, {"_id": 0, "part_number": 1, "name": 1, "uom": 1, "hsn_code": 1})
            if it:
                ln["item"] = it
    # Attach document creator for signature stamping on the printed TI.
    t["created_by_user"] = await _lookup_creator(t.get("created_by"))
    return t


async def _consume_tax_invoice_stock(invoice_doc, user_id):
    """Decrement on-hand stock + record inventory_transactions for every
    physical-item line on a Tax Invoice. Idempotent — sets
    invoice_doc['stock_consumed']=True so a second call is a no-op.

    Stock is reduced from `current_stock` on the item master. When a
    `ship_from_warehouse_id` is specified, the warehouse is recorded on the
    inventory_transaction row for audit; future iterations can add
    per-warehouse stock breakdowns without changing this contract.
    """
    if invoice_doc.get("stock_consumed"):
        return
    warehouse_id = invoice_doc.get("ship_from_warehouse_id") or ""
    inv_no = invoice_doc.get("invoice_no", "")
    # ---------- Pre-flight stock check ---------------------------------
    # Reject the consume if ANY line would push the item's current_stock
    # negative. Without this guard the issue flow silently went into the
    # red — finance teams then can't reconcile WIP / inventory ledgers.
    # We aggregate quantities per item so a TI that has the same item on
    # multiple lines is checked against the SUM, not each line in turn.
    needed = {}
    for ln in (invoice_doc.get("lines") or []):
        item_id = ln.get("item_id")
        qty = float(ln.get("quantity") or 0)
        if not item_id or qty <= 0:
            continue
        needed[item_id] = needed.get(item_id, 0.0) + qty
    shortages = []
    if needed:
        items_now = await db.items.find({"id": {"$in": list(needed.keys())}}, {"_id": 0, "id": 1, "part_number": 1, "name": 1, "current_stock": 1, "unit_of_measure": 1}).to_list(2000)
        items_map = {it["id"]: it for it in items_now}
        for iid, req_qty in needed.items():
            it = items_map.get(iid) or {}
            avail = float(it.get("current_stock") or 0)
            if req_qty > avail:
                shortages.append({
                    "item_id": iid,
                    "part_number": it.get("part_number") or iid,
                    "name": it.get("name") or "",
                    "required": req_qty,
                    "available": avail,
                    "short_by": req_qty - avail,
                    "uom": it.get("unit_of_measure") or "pcs",
                })
    if shortages:
        # 422 lets the frontend show a structured dialog with the per-item
        # shortage breakdown instead of a generic error toast.
        raise HTTPException(status_code=422, detail={
            "error": "insufficient_stock",
            "message": f"Cannot issue invoice — {len(shortages)} item(s) would go to negative stock.",
            "shortages": shortages,
        })
    consumed_any = False
    for ln in (invoice_doc.get("lines") or []):
        item_id = ln.get("item_id")
        qty = float(ln.get("quantity") or 0)
        if not item_id or qty <= 0:
            continue
        # Fetch BEFORE-stock for the audit trail (snapshot of running stock).
        item = await db.items.find_one({"id": item_id}, {"_id": 0, "current_stock": 1})
        prev_stock = float((item or {}).get("current_stock") or 0)
        new_stock = prev_stock - qty
        await db.items.update_one({"id": item_id}, {"$set": {"current_stock": new_stock}})
        await db.inventory_transactions.insert_one({
            "id": str(uuid.uuid4()),
            "item_id": item_id,
            "transaction_type": "dispatch",
            "quantity": -qty,
            "reference_type": "tax_invoice",
            "reference_id": inv_no,
            "warehouse_id": warehouse_id,
            "previous_stock": prev_stock,
            "new_stock": new_stock,
            "notes": f"Dispatched via Tax Invoice {inv_no}",
            "created_at": datetime.now(timezone.utc),
            "created_by": user_id,
        })
        consumed_any = True
    if consumed_any:
        await db.tax_invoices.update_one({"id": invoice_doc["id"]}, {"$set": {"stock_consumed": True, "stock_consumed_at": datetime.now(timezone.utc)}})


async def _restore_tax_invoice_stock(invoice_doc, user_id):
    """Reverse of _consume_tax_invoice_stock — used when an issued TI is
    cancelled. Idempotent: only runs if `stock_consumed` is True.
    """
    if not invoice_doc.get("stock_consumed"):
        return
    inv_no = invoice_doc.get("invoice_no", "")
    warehouse_id = invoice_doc.get("ship_from_warehouse_id") or ""
    for ln in (invoice_doc.get("lines") or []):
        item_id = ln.get("item_id")
        qty = float(ln.get("quantity") or 0)
        if not item_id or qty <= 0:
            continue
        item = await db.items.find_one({"id": item_id}, {"_id": 0, "current_stock": 1})
        prev_stock = float((item or {}).get("current_stock") or 0)
        new_stock = prev_stock + qty
        await db.items.update_one({"id": item_id}, {"$set": {"current_stock": new_stock}})
        await db.inventory_transactions.insert_one({
            "id": str(uuid.uuid4()),
            "item_id": item_id,
            "transaction_type": "dispatch_reversal",
            "quantity": qty,
            "reference_type": "tax_invoice",
            "reference_id": inv_no,
            "warehouse_id": warehouse_id,
            "previous_stock": prev_stock,
            "new_stock": new_stock,
            "notes": f"Restored — Tax Invoice {inv_no} cancelled",
            "created_at": datetime.now(timezone.utc),
            "created_by": user_id,
        })
    await db.tax_invoices.update_one({"id": invoice_doc["id"]}, {"$set": {"stock_consumed": False, "stock_restored_at": datetime.now(timezone.utc)}})

@crm_router.get("/tax-invoices")
async def list_tax_invoices(
    request: Request,
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    await get_current_user(request)
    q = {}
    if status:
        q["status"] = status
    if date_from or date_to:
        rng = {}
        if date_from:
            try:
                rng["$gte"] = datetime.fromisoformat(date_from.replace("Z", "+00:00"))
            except Exception:
                pass
        if date_to:
            try:
                end = datetime.fromisoformat(date_to.replace("Z", "+00:00"))
                if end.hour == 0 and end.minute == 0 and end.second == 0:
                    end = end.replace(hour=23, minute=59, second=59)
                rng["$lte"] = end
            except Exception:
                pass
        if rng:
            q["invoice_date"] = rng
    docs = await db.tax_invoices.find(q).sort([("invoice_date", -1), ("created_at", -1)]).to_list(2000)
    return [await _enrich_tax_invoice(d) for d in docs]

@crm_router.post("/tax-invoices", status_code=201)
async def create_tax_invoice(data: TaxInvoiceCreate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="tax_invoices", action="create")
    if not data.lines:
        raise HTTPException(status_code=400, detail="At least one line is required")
    invoice_no = await _get_next_number("tax_invoice")
    lines = [l.model_dump() for l in data.lines]
    add_charges = _normalize_additional_charges(data.additional_charges)
    base = await _finalize_invoice_lines(lines)
    currency = (data.currency or "INR").upper()
    if currency != "INR":
        gst = _zero_gst_split_for_export(lines)
    else:
        gst = await _compute_gst_split(data.customer_id, lines, place_of_supply=data.place_of_supply)
    # Fold additional charges + their GST into totals.
    totals = {"net_subtotal": base["subtotal"], "total_gst": gst["total_gst"], "grand_total": 0.0}
    charges_split = _apply_additional_charges(
        totals, add_charges, bool(gst.get("is_inter_state")), is_export=(currency != "INR"),
    )
    gst["cgst"] = round(float(gst.get("cgst") or 0) + charges_split["cgst"], 2)
    gst["sgst"] = round(float(gst.get("sgst") or 0) + charges_split["sgst"], 2)
    gst["igst"] = round(float(gst.get("igst") or 0) + charges_split["igst"], 2)
    gst["total_gst"] = round(gst["cgst"] + gst["sgst"] + gst["igst"], 2)
    doc = {
        "id": str(uuid.uuid4()),
        "invoice_no": invoice_no,
        **data.model_dump(exclude={"lines", "additional_charges"}, exclude_none=False),
        "lines": lines,
        "additional_charges": add_charges,
        **base,
        "additional_charges_total": totals.get("additional_charges_total", 0),
        "additional_charges_gst": totals.get("additional_charges_gst", 0),
        "cgst": gst["cgst"], "sgst": gst["sgst"], "igst": gst["igst"],
        "total_gst": gst["total_gst"],
        "is_inter_state": gst["is_inter_state"],
        "hsn_summary": gst["hsn_summary"],
        "grand_total": totals["grand_total"],
        "qr_code": f"UPI://pay?pa=machineworks@upi&pn=MachineWorksERP&am={totals['grand_total']}&tn={invoice_no}",
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"],
    }
    if not doc.get("invoice_date"):
        doc["invoice_date"] = datetime.now(timezone.utc)
    # Build dynamic UPI QR from company settings (bank_upi + company_name)
    # Only generate QR for INR invoices (domestic); export invoices don't need UPI QR.
    if currency == "INR":
        doc["qr_code"] = await _build_upi_qr_payload(doc["grand_total"], invoice_no)
    else:
        doc["qr_code"] = ""
    await db.tax_invoices.insert_one(doc)
    # Stock consumption: only run when the invoice is being created in an
    # ALREADY-committed state (not draft / cancelled). The reusable helper
    # tracks `stock_consumed` so subsequent updates won't double-consume.
    inv_status = (doc.get("status") or "issued").lower()
    if inv_status not in ("draft", "cancelled"):
        await _consume_tax_invoice_stock(doc, user["id"])
    # Re-fetch so the response reflects stock_consumed and any other server-
    # side mutations the helper may have done (audit timestamps).
    fresh_doc = await db.tax_invoices.find_one({"id": doc["id"]})
    return await _enrich_tax_invoice(fresh_doc or doc)

@crm_router.put("/tax-invoices/{tid}")
async def update_tax_invoice(tid: str, data: TaxInvoiceUpdate, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin", "production_manager"], module="tax_invoices", action="edit")
    existing = await db.tax_invoices.find_one({"id": tid})
    if not existing:
        raise HTTPException(status_code=404, detail="Tax Invoice not found")
    # NOTE: Previously this blocked ANY line-or-address edit once status became "issued".
    # That silently dropped user-edited T&C too because the frontend sends a full payload.
    # Now we allow edits at any status — users can always tweak T&C / notes / billing address.
    # Only a paid invoice is fully locked.
    if existing.get("status") == "paid":
        if data.status is None and any(v is not None for v in (data.lines, data.billing_address, data.shipping_address, data.customer_po_number, data.due_date, data.place_of_supply)):
            raise HTTPException(status_code=400, detail="A paid invoice is locked — only status changes and notes/terms are allowed")
    update = {k: v for k, v in data.model_dump(exclude_none=True).items()}
    # Recompute GST if lines OR place_of_supply OR currency OR additional_charges changed
    needs_gst_recompute = (
        "lines" in update or "place_of_supply" in update or "currency" in update
        or "additional_charges" in update
    )
    if needs_gst_recompute:
        lines = [l if isinstance(l, dict) else l.model_dump() for l in (update.get("lines") or existing.get("lines") or [])]
        raw_charges = update.get("additional_charges") if "additional_charges" in update else existing.get("additional_charges", [])
        add_charges = _normalize_additional_charges(raw_charges)
        update["additional_charges"] = add_charges
        base = await _finalize_invoice_lines(lines)
        effective_pos = update.get("place_of_supply", existing.get("place_of_supply", ""))
        currency = (update.get("currency") or existing.get("currency") or "INR").upper()
        if currency != "INR":
            gst = _zero_gst_split_for_export(lines)
        else:
            gst = await _compute_gst_split(existing.get("customer_id"), lines, place_of_supply=effective_pos)
        totals = {"net_subtotal": base["subtotal"], "total_gst": gst["total_gst"], "grand_total": 0.0}
        charges_split = _apply_additional_charges(
            totals, add_charges, bool(gst.get("is_inter_state")), is_export=(currency != "INR"),
        )
        gst["cgst"] = round(float(gst.get("cgst") or 0) + charges_split["cgst"], 2)
        gst["sgst"] = round(float(gst.get("sgst") or 0) + charges_split["sgst"], 2)
        gst["igst"] = round(float(gst.get("igst") or 0) + charges_split["igst"], 2)
        gst["total_gst"] = round(gst["cgst"] + gst["sgst"] + gst["igst"], 2)
        update["lines"] = lines
        update.update(base)
        update.update({
            "additional_charges_total": totals.get("additional_charges_total", 0),
            "additional_charges_gst": totals.get("additional_charges_gst", 0),
            "cgst": gst["cgst"], "sgst": gst["sgst"], "igst": gst["igst"],
            "total_gst": gst["total_gst"], "is_inter_state": gst["is_inter_state"],
            "hsn_summary": gst["hsn_summary"],
            "grand_total": totals["grand_total"],
        })
        # Refresh UPI QR with current company settings + new grand total (only for INR)
        if currency == "INR":
            update["qr_code"] = await _build_upi_qr_payload(update["grand_total"], existing.get("invoice_no", ""))
        else:
            update["qr_code"] = ""
    update["updated_at"] = datetime.now(timezone.utc)
    # Handle stock-consumption transitions based on status change:
    #   - draft → issued/paid: consume stock (one-time, gated by stock_consumed flag)
    #   - issued/paid → cancelled: restore stock
    # IMPORTANT: Check stock BEFORE updating the database to prevent status change
    # when stock is insufficient. The _consume_tax_invoice_stock function raises
    # HTTPException 422 if stock is insufficient.
    prev_status = (existing.get("status") or "draft").lower()
    new_status = (update.get("status") or prev_status).lower()
    # Pre-flight stock check: if transitioning to issued/paid and stock not yet consumed,
    # verify stock availability BEFORE updating the invoice status
    if new_status not in ("draft", "cancelled") and prev_status in ("draft",) and not existing.get("stock_consumed"):
        # Build a temporary doc with the new status to check stock
        temp_doc = {**existing, **update}
        await _consume_tax_invoice_stock(temp_doc, user["id"])
        # If we get here, stock was consumed successfully - mark it in the update
        update["stock_consumed"] = True
    # Now safe to update the database
    await db.tax_invoices.update_one({"id": tid}, {"$set": update})
    fresh = await db.tax_invoices.find_one({"id": tid})
    if fresh:
        # Handle cancellation - restore stock if previously consumed
        if new_status == "cancelled" and fresh.get("stock_consumed"):
            await _restore_tax_invoice_stock(fresh, user["id"])
    return await _enrich_tax_invoice(await db.tax_invoices.find_one({"id": tid}))

@crm_router.delete("/tax-invoices/{tid}")
async def delete_tax_invoice(tid: str, request: Request):
    user = await get_current_user(request)
    _require_access(user, ["admin"], module="tax_invoices", action="delete")
    existing = await db.tax_invoices.find_one({"id": tid})
    if not existing:
        raise HTTPException(status_code=404, detail="Tax Invoice not found")
    if existing.get("status") in ("paid", "issued"):
        raise HTTPException(status_code=400, detail="Cannot delete an issued/paid invoice — cancel it first")
    await db.tax_invoices.delete_one({"id": tid})
    return {"message": "Tax Invoice deleted"}

# ============================================================================
#  GST e-Invoice IRN Generation (NIC Sandbox / GSP adapter)
#  Reads GSP credentials from company_settings.gst_einvoice_* fields.
#  Returns a clear setup-instruction error when credentials are not configured.
# ============================================================================

# In-process token cache: {provider_key: (token, expires_at_utc)}
_GST_TOKEN_CACHE: Dict[str, Tuple[str, datetime]] = {}

def _build_einvoice_json(ti: dict, company: dict, customer: dict) -> dict:
    """Build the GST e-Invoice v1.1 JSON payload from a Tax Invoice doc."""
    inv_date = ti.get("invoice_date") or datetime.now(timezone.utc)
    if isinstance(inv_date, str):
        try:
            inv_date = datetime.fromisoformat(inv_date.replace("Z", "+00:00"))
        except Exception:
            inv_date = datetime.now(timezone.utc)
    dt_str = inv_date.strftime("%d/%m/%Y")

    seller_state = (company.get("state_code") or "").strip()
    buyer_state = (ti.get("place_of_supply") or (customer or {}).get("state_code") or "").strip()
    buyer_gstin = (customer or {}).get("gstin", "")

    item_list = []
    for idx, ln in enumerate(ti.get("lines") or [], 1):
        qty = float(ln.get("quantity") or 0)
        rate = float(ln.get("rate") or 0)
        disc_pct = float(ln.get("discount_pct") or 0)
        gross = qty * rate
        disc = round(gross * disc_pct / 100.0, 2)
        net = round(gross - disc, 2)
        gst_rate = float(ln.get("gst_rate") or 0)
        cgst_amt = float(ln.get("cgst_amt") or 0)
        sgst_amt = float(ln.get("sgst_amt") or 0)
        igst_amt = float(ln.get("igst_amt") or 0)
        item_list.append({
            "SlNo": str(idx),
            "IsServc": "N",
            "HsnCd": (ln.get("hsn_code") or "").strip() or "0000",
            "Qty": qty,
            "Unit": (ln.get("uom") or "NOS")[:3].upper(),
            "UnitPrice": rate,
            "TotAmt": round(gross, 2),
            "Discount": disc,
            "PreTaxVal": net,
            "AssAmt": net,
            "GstRt": gst_rate,
            "IgstAmt": igst_amt,
            "CgstAmt": cgst_amt,
            "SgstAmt": sgst_amt,
            "CesAmt": 0.0,
            "CesNonAdvlAmt": 0.0,
            "StateCesAmt": 0.0,
            "StateCesNonAdvlAmt": 0.0,
            "OthChrg": 0.0,
            "TotItemVal": round(net + cgst_amt + sgst_amt + igst_amt, 2),
        })

    ass_val = float(ti.get("subtotal") or 0)
    grand = float(ti.get("grand_total") or 0)
    return {
        "Version": "1.1",
        "TranDtls": {
            "TaxSch": "GST",
            "SupTyp": "B2B" if buyer_gstin else "B2C",
            "RegRev": "N",
            "IgstOnIntra": "N",
        },
        "DocDtls": {
            "Typ": "INV",
            "No": ti.get("invoice_no", ""),
            "Dt": dt_str,
        },
        "SellerDtls": {
            "Gstin": company.get("gstin", ""),
            "LglNm": company.get("company_name", "")[:100],
            "Addr1": (company.get("address") or "")[:100] or "NA",
            "Loc": (company.get("city") or "")[:50] or "NA",
            "Pin": int((company.get("pin_code") or "000000")[:6] or 0) or 110001,
            "Stcd": seller_state or "00",
            "Ph": (company.get("phone") or "")[:12],
            "Em": (company.get("email") or "")[:100],
        },
        "BuyerDtls": {
            "Gstin": buyer_gstin or "URP",  # URP = Unregistered person
            "LglNm": (ti.get("customer_name") or "NA")[:100],
            "Pos": buyer_state or seller_state or "00",
            "Addr1": (ti.get("billing_address") or (customer or {}).get("address") or "NA")[:100],
            "Loc": ((customer or {}).get("city") or "NA")[:50],
            "Pin": int(((customer or {}).get("pin_code") or "000000")[:6] or 0) or 110001,
            "Stcd": buyer_state or "00",
            "Ph": (ti.get("phone") or "")[:12],
            "Em": (ti.get("email") or "")[:100],
        },
        "ItemList": item_list,
        "ValDtls": {
            "AssVal": round(ass_val, 2),
            "CgstVal": round(float(ti.get("cgst") or 0), 2),
            "SgstVal": round(float(ti.get("sgst") or 0), 2),
            "IgstVal": round(float(ti.get("igst") or 0), 2),
            "CesVal": 0.0,
            "StCesVal": 0.0,
            "Discount": 0.0,
            "OthChrg": 0.0,
            "RndOffAmt": 0.0,
            "TotInvVal": round(grand, 2),
        },
    }


async def _gst_get_token(provider: str, endpoint: str, username: str, password: str, gstin: str, api_key: Optional[str] = None) -> str:
    """Get/refresh authentication token for the GSP. Cached for ~6 hours."""
    cache_key = f"{provider}|{endpoint}|{username}|{gstin}"
    cached = _GST_TOKEN_CACHE.get(cache_key)
    if cached and cached[1] > datetime.now(timezone.utc):
        return cached[0]
    import httpx
    async with httpx.AsyncClient(timeout=30.0) as client:
        # NIC-style auth (v0.03) — other GSPs (MasterGST) have slightly different schemas.
        auth_url = f"{endpoint.rstrip('/')}/api/v0.03/common/gettoken"
        headers = {"Content-Type": "application/json"}
        if api_key:
            headers["api-key"] = api_key
        r = await client.post(auth_url, json={"username": username, "password": password, "gstin": gstin}, headers=headers)
        try:
            body = r.json()
        except Exception:
            raise HTTPException(status_code=502, detail=f"GSP auth returned non-JSON response (HTTP {r.status_code})")
        if r.status_code != 200 or body.get("status") not in (1, "1", True, "success"):
            detail = body.get("message") or body.get("error") or body
            raise HTTPException(status_code=502, detail=f"GSP auth failed: {detail}")
        token = (body.get("data") or {}).get("token") or body.get("token")
        if not token:
            raise HTTPException(status_code=502, detail="GSP returned no auth token")
        _GST_TOKEN_CACHE[cache_key] = (token, datetime.now(timezone.utc) + timedelta(hours=5, minutes=45))
        return token


@crm_router.post("/tax-invoices/{tid}/generate-irn")
async def generate_tax_invoice_irn(tid: str, request: Request):
    """Generate IRN + signed QR code for a Tax Invoice via configured GSP.

    Pre-requisites: admin must configure GSP credentials in Settings → GST e-Invoice.
    Stores irn, ack_no, ack_dt, signed_invoice, signed_qr_code, irn_generated_at.
    """
    await get_current_user(request)
    ti = await db.tax_invoices.find_one({"id": tid})
    if not ti:
        raise HTTPException(status_code=404, detail="Tax Invoice not found")
    if ti.get("irn"):
        raise HTTPException(status_code=400, detail=f"IRN already generated: {ti.get('irn')}")

    company = await db.company_settings.find_one({"type": "company"}, {"_id": 0}) or {}
    if not company.get("gst_einvoice_enabled"):
        raise HTTPException(
            status_code=400,
            detail="GST e-Invoice is not enabled. Go to Settings → Company → GST e-Invoice and configure NIC Sandbox or GSP credentials.",
        )
    endpoint = (company.get("gst_einvoice_endpoint") or "").strip()
    username = (company.get("gst_einvoice_username") or "").strip()
    password = (company.get("gst_einvoice_password") or "").strip()
    gstin = (company.get("gstin") or "").strip()
    api_key = (company.get("gst_einvoice_api_key") or "").strip() or None
    if not all([endpoint, username, password, gstin]):
        raise HTTPException(
            status_code=400,
            detail="Missing GSP configuration. Please fill Endpoint, Username, Password and ensure Company GSTIN is set.",
        )
    if not company.get("state_code"):
        raise HTTPException(status_code=400, detail="Company state_code missing — required for IRN generation.")

    # Resolve buyer
    customer = {}
    if ti.get("customer_id"):
        customer = await db.customers.find_one({"id": ti["customer_id"]}, {"_id": 0}) or {}

    # Validate HSN on all lines (required by IRP)
    for ln in (ti.get("lines") or []):
        if not (ln.get("hsn_code") or "").strip():
            raise HTTPException(status_code=400, detail=f"Line '{ln.get('description') or ln.get('item_id')}' is missing HSN code — required for IRN.")

    try:
        token = await _gst_get_token(
            company.get("gst_einvoice_provider", "nic_sandbox"),
            endpoint, username, password, gstin, api_key,
        )
        payload = _build_einvoice_json(ti, company, customer)
        import httpx
        async with httpx.AsyncClient(timeout=45.0) as client:
            irn_url = f"{endpoint.rstrip('/')}/api/v0.03/invoices"
            headers = {
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json",
                "Gstin": gstin,
            }
            if api_key:
                headers["api-key"] = api_key
            r = await client.post(irn_url, json=payload, headers=headers)
        try:
            body = r.json()
        except Exception:
            raise HTTPException(status_code=502, detail=f"GSP returned non-JSON response (HTTP {r.status_code}): {r.text[:300]}")
        if r.status_code != 200 or body.get("status") not in (1, "1", True, "success"):
            detail = body.get("message") or body.get("error") or body
            raise HTTPException(status_code=502, detail=f"IRN generation failed: {detail}")
        data = body.get("data") or body
        irn_fields = {
            "irn": data.get("Irn") or data.get("IRN") or "",
            "ack_no": data.get("AckNo") or data.get("ack_no") or "",
            "ack_dt": data.get("AckDt") or data.get("ack_dt") or "",
            "signed_invoice": data.get("SignedInvoice") or "",
            "signed_qr_code": data.get("SignedQRCode") or "",
            "irn_generated_at": datetime.now(timezone.utc),
            "irn_status": "generated",
            "status": "issued",
        }
        await db.tax_invoices.update_one({"id": tid}, {"$set": irn_fields})
        updated = await db.tax_invoices.find_one({"id": tid})
        return await _enrich_tax_invoice(updated)
    except HTTPException:
        raise
    except httpx.RequestError as e:
        raise HTTPException(status_code=502, detail=f"Could not reach GSP endpoint: {e}")
    except Exception as e:
        logger.error(f"IRN generation unexpected error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"IRN generation error: {e}")


@crm_router.get("/tax-invoices/{tid}/einvoice-payload")
async def preview_einvoice_payload(tid: str, request: Request):
    """Return the e-invoice JSON that WOULD be sent to the IRP — for pre-flight validation."""
    await get_current_user(request)
    ti = await db.tax_invoices.find_one({"id": tid})
    if not ti:
        raise HTTPException(status_code=404, detail="Tax Invoice not found")
    company = await db.company_settings.find_one({"type": "company"}, {"_id": 0}) or {}
    customer = {}
    if ti.get("customer_id"):
        customer = await db.customers.find_one({"id": ti["customer_id"]}, {"_id": 0}) or {}
    return _build_einvoice_json(ti, company, customer)




@crm_router.post("/proformas/{pid}/convert-to-tax-invoice")
async def convert_proforma_to_tax_invoice(pid: str, payload: dict = Body(default={}), request: Request = None):
    user = await get_current_user(request)
    p = await db.proforma_invoices.find_one({"id": pid})
    if not p:
        raise HTTPException(status_code=404, detail="Proforma not found")
    if p.get("status") == "converted":
        raise HTTPException(status_code=400, detail="Proforma already converted to a Tax Invoice")
    lines = []
    for ln in p.get("lines", []):
        lines.append({
            "item_id": ln.get("item_id", ""),
            "description": ln.get("description", ""),
            "hsn_code": ln.get("hsn_code", ""),
            "quantity": ln.get("quantity", 0),
            "uom": ln.get("uom", "Nos"),
            "rate": ln.get("rate", 0),
            "discount_pct": ln.get("discount_pct", 0),
            "gst_rate": ln.get("gst_rate", 18),
        })
    invoice_no = await _get_next_number("tax_invoice")
    base = await _finalize_invoice_lines(lines)
    # Carry currency forward from the proforma; non-INR ⇒ no GST.
    p_currency = (p.get("currency") or "INR").upper()
    if p_currency != "INR":
        gst = _zero_gst_split_for_export(lines)
    else:
        gst = await _compute_gst_split(p.get("customer_id"), lines)
    # Carry additional charges forward from the proforma.
    add_charges = _normalize_additional_charges(p.get("additional_charges") or [])
    totals = {"net_subtotal": base["subtotal"], "total_gst": gst["total_gst"], "grand_total": 0.0}
    charges_split = _apply_additional_charges(
        totals, add_charges, bool(gst.get("is_inter_state")), is_export=(p_currency != "INR"),
    )
    gst["cgst"] = round(float(gst.get("cgst") or 0) + charges_split["cgst"], 2)
    gst["sgst"] = round(float(gst.get("sgst") or 0) + charges_split["sgst"], 2)
    gst["igst"] = round(float(gst.get("igst") or 0) + charges_split["igst"], 2)
    gst["total_gst"] = round(gst["cgst"] + gst["sgst"] + gst["igst"], 2)
    doc = {
        "id": str(uuid.uuid4()),
        "invoice_no": invoice_no,
        "proforma_id": pid,
        "customer_id": p.get("customer_id", ""),
        "customer_name": p.get("customer_name", ""),
        "contact_person": p.get("contact_person", ""),
        "email": p.get("email", ""),
        "phone": p.get("phone", ""),
        "billing_address": p.get("billing_address", ""),
        "shipping_address": p.get("shipping_address", ""),
        "invoice_date": datetime.now(timezone.utc),
        "place_of_supply": gst.get("customer_state", ""),
        "lines": lines,
        "additional_charges": add_charges,
        "notes": p.get("notes", ""),
        "terms": p.get("terms", ""),
        **base,
        "additional_charges_total": totals.get("additional_charges_total", 0),
        "additional_charges_gst": totals.get("additional_charges_gst", 0),
        "cgst": gst["cgst"], "sgst": gst["sgst"], "igst": gst["igst"],
        "total_gst": gst["total_gst"],
        "is_inter_state": gst["is_inter_state"],
        "hsn_summary": gst["hsn_summary"],
        "grand_total": totals["grand_total"],
        "qr_code": f"UPI://pay?pa=machineworks@upi&pn=MachineWorksERP&am={totals['grand_total']}&tn={invoice_no}",
        "currency": p_currency,
        "status": "issued",
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"],
    }
    doc["qr_code"] = await _build_upi_qr_payload(doc["grand_total"], invoice_no) if p_currency == "INR" else ""
    await db.tax_invoices.insert_one(doc)
    # Decrement stock for every line item (proforma→tax invoice path also delivers goods).
    for ln in lines:
        item_id = ln.get("item_id")
        qty = float(ln.get("quantity") or 0)
        if item_id and qty > 0:
            await db.items.update_one(
                {"id": item_id},
                {"$inc": {"current_stock": -qty}},
            )
    await db.proforma_invoices.update_one(
        {"id": pid},
        {"$set": {
            "status": "converted",
            "converted_tax_invoice_id": doc["id"],
            "converted_tax_invoice_no": invoice_no,
            "updated_at": datetime.now(timezone.utc),
        }}
    )
    return await _enrich_tax_invoice(doc)


@crm_router.post("/tax-invoices/from-sales-order/{so_id}", status_code=201)
async def create_tax_invoice_from_sales_order(so_id: str, request: Request):
    """Auto-generate a Tax Invoice draft from a confirmed Sales Order.

    Pulls customer + all SO lines, resolves FG/item pricing from each line's BOM parent_item
    (uses item.sale_price if set, else item.purchase_price, else line rate fallback),
    applies the item's hsn_code and the standard GST slab, and returns the enriched invoice.
    """
    user = await get_current_user(request)
    so = await db.production_orders.find_one({"id": so_id})
    if not so:
        raise HTTPException(status_code=404, detail="Sales Order not found")
    if so.get("status") in ("draft", "cancelled"):
        raise HTTPException(status_code=400, detail=f"Cannot invoice a {so.get('status')} Sales Order — confirm it first")

    customer_id = so.get("customer_id") or ""
    customer = None
    if customer_id:
        customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})

    # Build line items from SO lines (multi-line) or fall back to legacy single-line
    so_lines = so.get("lines") or []
    if not so_lines and so.get("bom_id"):
        so_lines = [{
            "bom_id": so["bom_id"],
            "quantity": so.get("quantity", 0),
            "due_date": so.get("due_date"),
            "order_type": "auto",
        }]
    if not so_lines:
        raise HTTPException(status_code=400, detail="Sales Order has no lines to invoice")

    ti_lines = []
    for ln in so_lines:
        bom = await db.boms.find_one({"id": ln.get("bom_id")}, {"_id": 0})
        if not bom:
            continue
        item = await db.items.find_one({"id": bom.get("parent_item_id")}, {"_id": 0}) or {}
        rate = float(item.get("sale_price") or 0) or float(item.get("purchase_price") or 0) or float(item.get("unit_cost") or 0)
        ti_lines.append({
            "item_id": item.get("id", ""),
            "description": item.get("name", ""),
            "hsn_code": item.get("hsn_code", ""),
            "quantity": float(ln.get("quantity", 0)),
            "uom": item.get("uom", "Nos"),
            "rate": rate,
            "discount_pct": 0.0,
            "gst_rate": float(item.get("gst_rate") or 18.0),
        })

    if not ti_lines:
        raise HTTPException(status_code=400, detail="Could not resolve any line items from this Sales Order")

    invoice_no = await _get_next_number("tax_invoice")
    base = await _finalize_invoice_lines(ti_lines)
    gst = await _compute_gst_split(customer_id, ti_lines)

    billing = (customer or {}).get("address", "") if customer else ""
    doc = {
        "id": str(uuid.uuid4()),
        "invoice_no": invoice_no,
        "sales_order_id": so_id,
        "sales_order_number": so.get("order_number", ""),
        "customer_id": customer_id,
        "customer_name": (customer or {}).get("name", ""),
        "contact_person": (customer or {}).get("contact_person", ""),
        "email": (customer or {}).get("email", ""),
        "phone": (customer or {}).get("phone", ""),
        "billing_address": billing,
        "shipping_address": billing,
        "invoice_date": datetime.now(timezone.utc),
        "place_of_supply": gst.get("customer_state", ""),
        "lines": ti_lines,
        "notes": "",
        "terms": "",
        **base,
        "cgst": gst["cgst"], "sgst": gst["sgst"], "igst": gst["igst"],
        "total_gst": gst["total_gst"],
        "is_inter_state": gst["is_inter_state"],
        "hsn_summary": gst["hsn_summary"],
        "grand_total": round(base["subtotal"] + gst["total_gst"], 2),
        "qr_code": f"UPI://pay?pa=machineworks@upi&pn=MachineWorksERP&am={round(base['subtotal'] + gst['total_gst'], 2)}&tn={invoice_no}",
        "status": "draft",
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"],
    }
    doc["qr_code"] = await _build_upi_qr_payload(doc["grand_total"], invoice_no)
    await db.tax_invoices.insert_one(doc)
    return await _enrich_tax_invoice(doc)


# ============================================================================
#  PACKING LIST
#  Created from a Tax Invoice. Each TI line can be EXPANDED to its BOM first-level
#  components (with qty × invoice_qty) or left as-is (FG). Saved + printable.
# ============================================================================
class PackingListLineComponent(BaseModel):
    item_id: str = ""
    part_number: str = ""
    name: str = ""
    uom: str = "Nos"
    qty_per_unit: float = 0
    total_qty: float = 0

class PackingListLine(BaseModel):
    source_line_index: int  # index into the source tax_invoice.lines
    item_id: Optional[str] = ""
    item_name: str
    description: Optional[str] = ""
    uom: str = "Nos"
    invoice_qty: float
    expanded: bool = False
    components: List[PackingListLineComponent] = []

class PackingListCreate(BaseModel):
    tax_invoice_id: str
    lines: List[PackingListLine]
    notes: Optional[str] = ""
    packed_by: Optional[str] = ""
    packed_by_user_id: Optional[str] = ""
    dispatch_date: Optional[datetime] = None

class PackingListUpdate(BaseModel):
    status: Optional[str] = None
    notes: Optional[str] = None
    packed_by: Optional[str] = None
    packed_by_user_id: Optional[str] = None
    dispatch_date: Optional[datetime] = None

async def _enrich_packing_list(pl):
    pl.pop("_id", None)
    if pl.get("tax_invoice_id"):
        ti = await db.tax_invoices.find_one({"id": pl["tax_invoice_id"]}, {"_id": 0, "invoice_no": 1, "customer_name": 1, "customer_id": 1, "billing_address": 1, "shipping_address": 1, "sales_order_id": 1})
        if ti:
            pl["tax_invoice"] = ti
            if ti.get("customer_id"):
                c = await db.customers.find_one({"id": ti["customer_id"]}, {"_id": 0})
                if c:
                    pl["customer"] = c
    if pl.get("packed_by_user_id"):
        u = await db.users.find_one({"id": pl["packed_by_user_id"]}, {"_id": 0, "name": 1, "email": 1, "signature_url": 1})
        if u:
            pl["packed_by_user"] = u
    return pl

@crm_router.get("/packing-lists")
async def list_packing_lists(request: Request):
    await get_current_user(request)
    docs = await db.packing_lists.find({}, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return [await _enrich_packing_list(d) for d in docs]

@crm_router.get("/packing-lists/{pl_id}")
async def get_packing_list(pl_id: str, request: Request):
    await get_current_user(request)
    doc = await db.packing_lists.find_one({"id": pl_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Packing List not found")
    return await _enrich_packing_list(doc)

@crm_router.post("/packing-lists", status_code=201)
async def create_packing_list(data: PackingListCreate, request: Request):
    user = await get_current_user(request)
    ti = await db.tax_invoices.find_one({"id": data.tax_invoice_id})
    if not ti:
        raise HTTPException(status_code=404, detail="Tax Invoice not found")
    # Block duplicate Packing Lists per Tax Invoice. Backend enforcement is
    # defense-in-depth — frontend also disables the action button. Check the
    # AUTHORITATIVE source first (packing_lists collection) so legacy TIs
    # whose back-link was never populated still get blocked when a PL exists.
    existing_pl = await db.packing_lists.find_one(
        {"tax_invoice_id": data.tax_invoice_id},
        {"_id": 0, "packing_list_no": 1, "id": 1},
    )
    if existing_pl or ti.get("packing_list_id") or ti.get("packing_list_no"):
        existing_no = (existing_pl or {}).get("packing_list_no") or ti.get("packing_list_no") or "(existing)"
        # If we found a PL but the TI back-link is missing, repair it so the
        # FE will see the "already created" state on next refresh.
        if existing_pl and not ti.get("packing_list_id"):
            await db.tax_invoices.update_one(
                {"id": data.tax_invoice_id},
                {"$set": {"packing_list_id": existing_pl.get("id"), "packing_list_no": existing_no, "updated_at": datetime.now(timezone.utc)}},
            )
        raise HTTPException(
            status_code=400,
            detail=f"Packing List {existing_no} already exists for this Tax Invoice. Delete it first to regenerate."
        )
    pl_no = await _get_next_number("packing_list")
    doc = {
        "id": str(uuid.uuid4()),
        "packing_list_no": pl_no,
        "tax_invoice_id": data.tax_invoice_id,
        "tax_invoice_no": ti.get("invoice_no", ""),
        "customer_id": ti.get("customer_id", ""),
        "customer_name": ti.get("customer_name", ""),
        "billing_address": ti.get("billing_address", ""),
        "shipping_address": ti.get("shipping_address", ""),
        "lines": [ln.model_dump() for ln in data.lines],
        "notes": data.notes or "",
        "packed_by": data.packed_by or "",
        "packed_by_user_id": data.packed_by_user_id or user["id"],
        "dispatch_date": data.dispatch_date,
        "status": "draft",
        "created_at": datetime.now(timezone.utc),
        "created_by": user["id"],
    }
    await db.packing_lists.insert_one(doc)
    # Back-link on the Tax Invoice so the TI panel can show a "Packing List Created" badge.
    await db.tax_invoices.update_one(
        {"id": data.tax_invoice_id},
        {"$set": {"packing_list_id": doc["id"], "packing_list_no": pl_no, "updated_at": datetime.now(timezone.utc)}}
    )
    return await _enrich_packing_list(doc)

@crm_router.put("/packing-lists/{pl_id}")
async def update_packing_list(pl_id: str, data: PackingListUpdate, request: Request):
    await get_current_user(request)
    existing = await db.packing_lists.find_one({"id": pl_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Packing List not found")
    update = {k: v for k, v in data.model_dump().items() if v is not None}
    if update:
        update["updated_at"] = datetime.now(timezone.utc)
        await db.packing_lists.update_one({"id": pl_id}, {"$set": update})
    return await _enrich_packing_list(await db.packing_lists.find_one({"id": pl_id}, {"_id": 0}))

@crm_router.delete("/packing-lists/{pl_id}")
async def delete_packing_list(pl_id: str, request: Request):
    user = await get_current_user(request)
    existing = await db.packing_lists.find_one({"id": pl_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Packing List not found")
    if existing.get("status") in ("dispatched", "received") and user.get("role") != "admin":
        raise HTTPException(status_code=400, detail="Only admins can delete dispatched/received packing lists")
    await db.packing_lists.delete_one({"id": pl_id})
    # Clear the back-link on the source Tax Invoice so a fresh PL can be generated.
    if existing.get("tax_invoice_id"):
        await db.tax_invoices.update_one(
            {"id": existing["tax_invoice_id"]},
            {"$unset": {"packing_list_id": "", "packing_list_no": ""}, "$set": {"updated_at": datetime.now(timezone.utc)}}
        )
    return {"message": "Packing List deleted"}

@crm_router.post("/packing-lists/preview/{ti_id}")
async def preview_packing_list_lines(ti_id: str, payload: dict = Body(default={}), request: Request = None):
    """Resolve BOM first-level components for the requested lines.

    payload = { "expand": {"<source_line_index>": true/false, ...} }
    Returns: [{source_line_index, item_name, invoice_qty, has_bom, expanded, components: [...]}]
    If `expand[idx]` is omitted, default = True when the item has an active BOM.
    """
    await get_current_user(request)
    ti = await db.tax_invoices.find_one({"id": ti_id}, {"_id": 0})
    if not ti:
        raise HTTPException(status_code=404, detail="Tax Invoice not found")
    expand_map = (payload or {}).get("expand") or {}
    out = []
    for idx, ln in enumerate(ti.get("lines", [])):
        item_id = ln.get("item_id") or ""
        item_name = ln.get("description") or ""
        uom = ln.get("uom") or "Nos"
        invoice_qty = float(ln.get("quantity") or 0)
        bom = None
        has_bom = False
        if item_id:
            item = await db.items.find_one({"id": item_id}, {"_id": 0, "name": 1, "part_number": 1})
            if item and not item_name:
                item_name = item.get("name", "")
            bom = await db.boms.find_one({"parent_item_id": item_id, "status": {"$in": ["active", "approved"]}}, {"_id": 0}, sort=[("created_at", -1)])
            has_bom = bool(bom)
        explicit = expand_map.get(str(idx))
        expand = (has_bom if explicit is None else bool(explicit) and has_bom)
        components = []
        if expand and bom:
            for comp in (bom.get("components") or []):
                comp_item = await db.items.find_one({"id": comp.get("item_id")}, {"_id": 0, "part_number": 1, "name": 1, "uom": 1})
                qty_per = float(comp.get("quantity") or 0)
                components.append({
                    "item_id": comp.get("item_id", ""),
                    "part_number": (comp_item or {}).get("part_number", ""),
                    "name": (comp_item or {}).get("name", comp.get("name", "")),
                    "uom": (comp_item or {}).get("uom", "Nos"),
                    "qty_per_unit": qty_per,
                    "total_qty": round(qty_per * invoice_qty, 4),
                })
        out.append({
            "source_line_index": idx,
            "item_id": item_id,
            "item_name": item_name,
            "description": ln.get("description") or "",
            "uom": uom,
            "invoice_qty": invoice_qty,
            "has_bom": has_bom,
            "expanded": expand,
            "components": components,
        })
    return out


# ============================================================================
#  PUBLIC SHARE ENDPOINTS (no auth) — used for WhatsApp shareable links.
# ============================================================================
public_router = APIRouter(prefix="/public")

@public_router.get("/quotation/{qid}")
async def public_quotation(qid: str):
    doc = await db.crm_quotations.find_one({"id": qid}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Not found")
    return await _enrich_quotation(doc)

@public_router.get("/proforma/{pid}")
async def public_proforma(pid: str):
    doc = await db.proforma_invoices.find_one({"id": pid}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Not found")
    return await _enrich_proforma(doc)

@public_router.get("/tax-invoice/{tid}")
async def public_tax_invoice(tid: str):
    doc = await db.tax_invoices.find_one({"id": tid}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Not found")
    return await _enrich_tax_invoice(doc)

@public_router.get("/packing-list/{pl_id}")
async def public_packing_list(pl_id: str):
    doc = await db.packing_lists.find_one({"id": pl_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Not found")
    return await _enrich_packing_list(doc)

@public_router.get("/company")
async def public_company():
    doc = await db.company_settings.find_one({"type": "company"}, {"_id": 0}) or {}
    return doc


# ============================================================================
#  DATA BACKUP & RESTORE — admin-only. Dumps every collection to a single JSON
#  blob (ObjectId + datetime serialised as strings); restore wipes + reloads.
# ============================================================================
BACKUP_COLLECTIONS = [
    "users", "role_groups", "company_settings", "number_series",
    "items", "boms", "bom_revisions", "suppliers", "customers",
    "warehouses", "inventory", "stock_movements", "production_orders",
    "purchase_orders", "grns", "purchase_invoices", "quality_inspections",
    "job_work_orders", "delivery_challans", "crm_leads", "crm_tickets",
    "crm_quotations", "proforma_invoices", "tax_invoices", "packing_lists",
    "pipeline_config", "activity_logs",
]

def _jsonable(doc):
    """Convert ObjectId + datetime recursively so the entire dump is JSON-serialisable."""
    from bson import ObjectId
    if isinstance(doc, dict):
        return {k: _jsonable(v) for k, v in doc.items()}
    if isinstance(doc, list):
        return [_jsonable(v) for v in doc]
    if isinstance(doc, ObjectId):
        return str(doc)
    if isinstance(doc, datetime):
        return doc.isoformat()
    return doc

@settings_router.get("/backup")
async def backup_database(request: Request):
    """Return a JSON dump of every ERP collection — admin-only."""
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only admin can download a backup")
    out = {
        "backup_version": 1,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "generated_by": user.get("email", ""),
        "collections": {},
    }
    for coll in BACKUP_COLLECTIONS:
        docs = await db[coll].find({}).to_list(200000)
        out["collections"][coll] = [_jsonable(d) for d in docs]
    return out

@settings_router.post("/restore")
async def restore_database(payload: dict = Body(...), request: Request = None):
    """Restore from a backup JSON. Admin-only. **Wipes each listed collection** before inserting.

    Expected payload shape: {"backup_version": 1, "collections": { "<name>": [docs...] }}
    Skips `_id` fields (Mongo will generate new ones); preserves the app-level `id` field.
    """
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only admin can restore a backup")
    if payload.get("backup_version") != 1:
        raise HTTPException(status_code=400, detail="Unsupported backup version — expected version 1")
    collections = payload.get("collections") or {}
    if not isinstance(collections, dict):
        raise HTTPException(status_code=400, detail="Backup has no `collections` object")

    summary = {}
    for name in BACKUP_COLLECTIONS:
        if name not in collections:
            continue
        rows = collections[name] or []
        # Wipe + reload
        await db[name].delete_many({})
        if rows:
            cleaned = []
            for r in rows:
                r = dict(r or {})
                r.pop("_id", None)
                # Parse datetime-looking ISO strings back (best-effort)
                for k, v in list(r.items()):
                    if isinstance(v, str) and len(v) >= 20 and v[4] == "-" and "T" in v:
                        try:
                            r[k] = datetime.fromisoformat(v.replace("Z", "+00:00"))
                        except ValueError:
                            pass
                cleaned.append(r)
            await db[name].insert_many(cleaned)
        summary[name] = len(rows)
    return {"message": "Restore complete", "summary": summary}


# Include routers
api_router.include_router(auth_router)
api_router.include_router(items_router)
api_router.include_router(item_groups_router)
api_router.include_router(bom_router)
api_router.include_router(mrp_router)
api_router.include_router(quality_router)
api_router.include_router(inventory_router)
api_router.include_router(production_router)
api_router.include_router(users_router)
api_router.include_router(dashboard_router)
api_router.include_router(suppliers_router)
api_router.include_router(purchase_orders_router)
api_router.include_router(warehouses_router)
api_router.include_router(work_centers_router)
api_router.include_router(routings_router)
api_router.include_router(work_orders_router)
api_router.include_router(settings_router)
api_router.include_router(customers_router)
api_router.include_router(grn_router)
api_router.include_router(purchase_invoices_router)
api_router.include_router(jobwork_router)
api_router.include_router(crm_router)
api_router.include_router(public_router)

@api_router.get("/health")
async def health_check():
    return {"status": "healthy", "service": "machinery-erp"}


# ─── Server-side HTML → PDF (Chromium headless) ────────────────────────
# Client-side html2pdf / html2canvas rasterisation cannot reliably
# reproduce Chrome's print engine output (column widths shift, logos
# overflow, repeating headers double up). To give users a real
# downloadable PDF that is byte-identical to "Print → Save as PDF", we
# run a Chromium headless instance server-side and pipe the HTML
# through page.pdf({format: 'A4', printBackground: true}).
#
# Trade-off: ~600ms extra latency vs in-browser; in exchange the user
# gets a true Chrome-quality PDF file download.
@api_router.post("/print/html-to-pdf")
async def print_html_to_pdf(payload: dict = Body(...), user: dict = Depends(get_current_user)):
    """Server-side Chromium HTML→PDF. Auto-installs the browser binary on
    first call so production deployments don't need a separate provisioning
    step.
    """
    html = (payload or {}).get("html") or ""
    filename = ((payload or {}).get("filename") or "document") + ""
    if not filename.endswith(".pdf"):
        filename = filename + ".pdf"
    if not html or len(html) < 30:
        raise HTTPException(status_code=400, detail="Missing 'html' in payload")
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        raise HTTPException(status_code=500, detail="Playwright not installed on server. Run: pip install playwright")
    # Detect missing browser binary; auto-install once if needed.
    async def _try_pdf(executable_path=None):
        async with async_playwright() as pw:
            launch_args = {"args": ["--no-sandbox", "--disable-dev-shm-usage"]}
            if executable_path:
                launch_args["executable_path"] = executable_path
            browser = await pw.chromium.launch(**launch_args)
            ctx = await browser.new_context()
            page = await ctx.new_page()
            await page.set_content(html, wait_until="networkidle", timeout=20000)
            # Page X of Y footer — Playwright's pdf() does NOT render
            # @page @bottom-right margin-boxes from CSS, so we have to
            # supply a Chromium footer_template. The page number tokens
            # `<span class="pageNumber"></span>` / `<span class="totalPages"></span>`
            # are special — Chrome substitutes them at print time.
            # display_header_footer with margins ensures the footer doesn't
            # crop the body content; top margin matches the existing CSS @page.
            footer_html = (
                '<div style="font-family:Helvetica,Arial,sans-serif;font-size:8px;'
                'color:#64748b;width:100%;text-align:right;padding:0 8mm 0 0;">'
                'Page <span class="pageNumber"></span> of <span class="totalPages"></span>'
                '</div>'
            )
            data = await page.pdf(
                format="A4",
                print_background=True,
                prefer_css_page_size=True,
                display_header_footer=True,
                # Empty header so only the footer shows.
                header_template='<div></div>',
                footer_template=footer_html,
                # Override CSS margins so the footer has room (the CRM @page
                # rule uses 4mm/4mm/6mm/4mm which is too tight for footer).
                margin={"top": "4mm", "right": "4mm", "bottom": "12mm", "left": "4mm"},
            )
            await browser.close()
            return data

    try:
        pdf_bytes = await _try_pdf()
    except Exception as e1:
        err_str = str(e1)
        # Last-ditch: use system Chrome/Chromium if installed (Docker
        # images frequently have google-chrome at /usr/bin/google-chrome
        # or chromium at /usr/bin/chromium).
        system_chrome = None
        for cand in ("/usr/bin/google-chrome", "/usr/bin/chromium", "/usr/bin/chromium-browser", "/root/bin/chromium"):
            import os as _os
            if _os.path.exists(cand):
                system_chrome = cand
                break
        if "Executable doesn't exist" in err_str or "playwright install" in err_str.lower():
            # Try (1) playwright install, (2) system chrome, in that order.
            import subprocess
            try:
                subprocess.run(["playwright", "install", "chromium"], check=True, timeout=300)
                pdf_bytes = await _try_pdf()
            except Exception as e2:
                if system_chrome:
                    try:
                        pdf_bytes = await _try_pdf(executable_path=system_chrome)
                    except Exception as e3:
                        raise HTTPException(status_code=500, detail=f"PDF gen failed (playwright auto-install + system chrome both failed): {e3}")
                else:
                    raise HTTPException(status_code=500, detail=f"Chromium auto-install failed and no system Chrome found. Run on server: PLAYWRIGHT_BROWSERS_PATH=/pw-browsers playwright install chromium. Error: {e2}")
        else:
            raise HTTPException(status_code=500, detail=f"PDF generation failed: {e1}")
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return Response(content=pdf_bytes, media_type="application/pdf", headers=headers)


app.include_router(api_router)

# CORS
# Note: when frontend sends credentials (cookies), browsers REJECT the wildcard '*'
# Access-Control-Allow-Origin. We must echo back the specific origin.
# `allow_origin_regex=".*"` tells Starlette to match any origin AND respond with that
# exact origin in the header, which is compatible with `allow_credentials=True`.
_cors_env = os.environ.get('CORS_ORIGINS', '*')
if _cors_env == '*':
    app.add_middleware(
        CORSMiddleware,
        allow_origin_regex=".*",
        allow_credentials=True,
        allow_methods=["*"],
        allow_headers=["*"],
        expose_headers=["Content-Disposition"],
    )
else:
    app.add_middleware(
        CORSMiddleware,
        allow_origins=_cors_env.split(','),
        allow_credentials=True,
        allow_methods=["*"],
        allow_headers=["*"],
        expose_headers=["Content-Disposition"],
    )


# Global exception handler — ensures even unhandled errors return a JSON response
# WITH CORS headers (so browser reports the real error instead of masking it as CORS).
from fastapi.requests import Request as _FRequest

@app.exception_handler(Exception)
async def _global_exception_handler(request: _FRequest, exc: Exception):
    logger.error(f"[unhandled] {request.method} {request.url.path} → {type(exc).__name__}: {exc}", exc_info=True)
    # Echo the request's Origin back so credentialed requests don't fail CORS on the error path
    origin = request.headers.get("origin", "*")
    cors_origin = origin if origin != "*" else "*"
    return JSONResponse(
        status_code=500,
        content={"detail": f"Server error: {type(exc).__name__}: {exc}"},
        headers={
            "Access-Control-Allow-Origin": cors_origin,
            "Access-Control-Allow-Credentials": "true" if origin != "*" else "false",
            "Vary": "Origin",
        },
    )
